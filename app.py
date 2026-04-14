import os
import re
import json
import urllib.request
import urllib.parse
import logging
from flask import Flask, request, abort, jsonify
from linebot.v3 import WebhookHandler
from linebot.v3.messaging import (
    Configuration, ApiClient, MessagingApi, MessagingApiBlob,
    ReplyMessageRequest, TextMessage, FlexMessage, FlexContainer,
    QuickReply, QuickReplyItem, MessageAction, PushMessageRequest,
    MulticastRequest,
)
try:
    from linebot.v3.messaging import (
        TemplateMessage, ConfirmTemplate, ButtonsTemplate,
        CarouselTemplate, CarouselColumn,
        PostbackAction, URIAction as MsgURIAction,
        ImagemapMessage, ImagemapBaseSize, ImagemapArea,
        MessageImagemapAction, URIImagemapAction,
        BroadcastRequest,
    )
except ImportError:
    TemplateMessage = None
    BroadcastRequest = None
try:
    from linebot.v3.messaging import ShowLoadingAnimationRequest
except ImportError:
    ShowLoadingAnimationRequest = None
try:
    from linebot.v3.messaging import MarkMessagesAsReadRequest
except ImportError:
    MarkMessagesAsReadRequest = None
try:
    from linebot.v3.messaging import (
        DatetimePickerAction as MsgDatetimePickerAction,
        CameraAction as MsgCameraAction,
        CameraRollAction as MsgCameraRollAction,
        LocationAction as MsgLocationAction,
        ClipboardAction as MsgClipboardAction,
    )
except ImportError:
    MsgDatetimePickerAction = None
    MsgCameraAction = None
    MsgCameraRollAction = None
    MsgLocationAction = None
    MsgClipboardAction = None
try:
    from linebot.v3.messaging import ValidateMessageRequest
except ImportError:
    ValidateMessageRequest = None
try:
    from linebot.v3.messaging import Sender as MessageSender
except ImportError:
    MessageSender = None
try:
    from linebot.v3.messaging import (
        RichMenuRequest, RichMenuArea, RichMenuBounds, RichMenuSize,
        CreateRichMenuAliasRequest, URIAction,
    )
except ImportError:
    RichMenuRequest = None
from linebot.v3.webhooks import MessageEvent, TextMessageContent, ImageMessageContent, AudioMessageContent
try:
    from linebot.v3.webhooks import VideoMessageContent
except ImportError:
    VideoMessageContent = None
try:
    from linebot.v3.webhooks import FileMessageContent
except ImportError:
    FileMessageContent = None
try:
    from linebot.v3.webhooks import LocationMessageContent
except ImportError:
    LocationMessageContent = None
try:
    from linebot.v3.webhooks import StickerMessageContent
except ImportError:
    StickerMessageContent = None
try:
    from linebot.v3.webhooks import JoinEvent
except ImportError:
    JoinEvent = None
try:
    from linebot.v3.webhooks import MemberJoinedEvent
except ImportError:
    MemberJoinedEvent = None
try:
    from linebot.v3.webhooks import MemberLeftEvent
except ImportError:
    MemberLeftEvent = None
try:
    from linebot.v3.webhooks import FollowEvent, UnfollowEvent
except ImportError:
    FollowEvent = None
    UnfollowEvent = None
try:
    from linebot.v3.webhooks import LeaveEvent as BotLeaveEvent
except ImportError:
    BotLeaveEvent = None
try:
    from linebot.v3.webhooks import PostbackEvent
except ImportError:
    PostbackEvent = None
try:
    from linebot.v3.webhooks import UnsendEvent
except ImportError:
    UnsendEvent = None
try:
    from linebot.v3.webhooks import VideoPlayCompleteEvent
except ImportError:
    VideoPlayCompleteEvent = None
from linebot.v3.exceptions import InvalidSignatureError
from openai import OpenAI
import base64
import tempfile
import time
import uuid

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

VERSION = "v2.8-0415b"

LINE_TOKEN = os.environ.get("LINE_CHANNEL_ACCESS_TOKEN", "")
LINE_SECRET = os.environ.get("LINE_CHANNEL_SECRET", "")
OPENAI_KEY = os.environ.get("OPENAI_API_KEY", "")
ADMIN_KEY = os.environ.get("ADMIN_KEY", "changeme")
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = "onerkk/line-translator-bot"

configuration = Configuration(access_token=LINE_TOKEN)
handler = WebhookHandler(LINE_SECRET)
oai = OpenAI(api_key=OPENAI_KEY) if OPENAI_KEY else None

group_settings = {}
# Target language for Chinese translation per group, default "id"
group_target_lang = {}
# Image translation toggle per group, default True
group_img_settings = {}
# Audio/voice translation toggle per group, default True
group_audio_settings = {}
# Work order photo detection toggle per group, default True
group_wo_settings = {}
# Per-group command toggles: {group_id: {"pw1": bool, "pw2": bool, ...}}
group_cmd_enabled = {}
# Command definitions: (key, emoji+label_on, emoji+label_off, default)
CMD_DEFS = [
    ("pw1",    "🔑密碼1開", "🔑密碼1關", True),
    ("pw2",    "🏭密碼2開", "🏭密碼2關", True),
    ("pkg",    "📦包裝開",  "📦包裝關",  True),
    ("scrap",  "🎨廢料開",  "🎨廢料關",  True),
    ("qry",    "🔍儲區開",  "🔍儲區關",  True),
    ("notice", "📢公告開",  "📢公告關",  True),
]

def is_cmd_enabled(group_id, cmd_key):
    """Check if a command is enabled for a group."""
    cmds = group_cmd_enabled.get(group_id, {})
    # Find default from CMD_DEFS
    for key, _, _, default in CMD_DEFS:
        if key == cmd_key:
            return cmds.get(cmd_key, default)
    return True

# Skip list: set of user_ids per group whose messages won't be translated
group_skip_users = {}
# Track user display names per group: {group_id: {user_id: display_name}}
group_user_names = {}
# Group tracking: {group_id: {"name": str, "joined_at": float}}
group_tracking = {}

# DM (private message) target language per user, default "id"
dm_target_lang = {}
# DM master toggle (global on/off for all DM)
dm_master_enabled = True
# DM whitelist: set of user_ids allowed to DM when master is off
dm_whitelist = set()
# DM known users: {user_id: display_name} for anyone who has DM'd the bot
dm_known_users = {}

# Bot start time for uptime tracking
bot_start_time = time.time()

# Stats counters (resets on restart — Render free tier)
bot_stats = {
    "text_translations": 0,
    "image_translations": 0,
    "voice_translations": 0,
    "work_order_detections": 0,
    "commands": 0,
    "tokens_prompt": 0,
    "tokens_completion": 0,
}


def track_tokens(response):
    """Track token usage from OpenAI API response."""
    try:
        if response and hasattr(response, 'usage') and response.usage:
            bot_stats["tokens_prompt"] += response.usage.prompt_tokens or 0
            bot_stats["tokens_completion"] += response.usage.completion_tokens or 0
    except Exception:
        pass


def track_group_usage(group_id, before_prompt, before_completion):
    """Calculate token diff since snapshot and attribute to group."""
    dp = bot_stats.get("tokens_prompt", 0) - before_prompt
    dc = bot_stats.get("tokens_completion", 0) - before_completion
    if group_id and (dp > 0 or dc > 0):
        if group_id not in group_api_usage:
            group_api_usage[group_id] = {"tokens_prompt": 0, "tokens_completion": 0}
        group_api_usage[group_id]["tokens_prompt"] += dp
        group_api_usage[group_id]["tokens_completion"] += dc


def calc_group_cost_twd(group_id):
    """Calculate cost in TWD for a group."""
    u = group_api_usage.get(group_id, {})
    tp = u.get("tokens_prompt", 0)
    tc = u.get("tokens_completion", 0)
    usd = (tp * 0.00000015) + (tc * 0.0000006)
    return round(usd * USD_TO_TWD, 2)

# Admin users tracking: {user_id: {"is_admin": bool}}
admin_users = {}

# User language cache from LINE profile: {user_id: "id"|"zh-TW"|"en"|...}
user_languages = {}

# Per-group API usage tracking: {group_id: {"tokens_prompt": int, "tokens_completion": int}}
group_api_usage = {}

# ── Admin-controllable feature settings ──
# Welcome message: {enabled: bool, text_zh: str, text_id: str}
welcome_settings = {
    "enabled": True,
    "text_zh": "👋 歡迎新成員加入！\n本群組有 AI 翻譯助手，中文和印尼文會自動互譯。",
    "text_id": "👋 Selamat datang!\nGrup ini memiliki asisten penerjemah AI, bahasa Mandarin dan Indonesia akan diterjemahkan otomatis.",
}
# Flex message ON/OFF (True = Flex card, False = plain text)
flex_enabled = True
# Quick Reply buttons ON/OFF
quick_reply_enabled = True
# Silent mode: translation messages don't buzz the phone
silent_mode = False
# Video OCR translation ON/OFF
video_ocr_enabled = True
# Location translation ON/OFF
location_translate_enabled = True
# Mark-as-read ON/OFF (shows 'read' indicator in chat)
mark_read_enabled = True
# X-Line-Retry-Key ON/OFF (idempotent message sending)
retry_key_enabled = True
# Camera Quick Reply button ON/OFF
camera_qr_enabled = True
# Clipboard Quick Reply button ON/OFF (copy storage zone etc.)
clipboard_qr_enabled = False
# Camera Roll Quick Reply button ON/OFF
camera_roll_qr_enabled = False
# Location Quick Reply button ON/OFF
location_qr_enabled = False
# Per-group feature overrides (group_id -> bool), global values above are defaults
group_flex_settings = {}      # per-group flex card toggle
group_qr_settings = {}        # per-group quick reply toggle
group_silent_settings = {}    # per-group silent mode toggle
group_video_settings = {}     # per-group video OCR toggle
group_location_settings = {}  # per-group location translate toggle
group_mark_read_settings = {} # per-group mark-as-read toggle
group_retry_key_settings = {} # per-group retry key toggle
group_camera_qr_settings = {} # per-group camera QR button toggle
group_clipboard_qr_settings = {} # per-group clipboard QR button toggle
group_camera_roll_qr_settings = {} # per-group camera roll QR button toggle
group_location_qr_settings = {} # per-group location QR button toggle
group_welcome_settings = {}   # per-group welcome: {group_id: {"enabled": bool, "text_zh": str, "text_id": str}}
# Translation tone settings
TONE_PRESETS = {
    "casual": "Translate casually like real people talk at work. Use everyday slang and informal language.",
    "natural": "Translate like a native speaker would naturally say it in daily factory conversation. Use the most natural, fluent, mother-tongue phrasing. Prefer colloquial expressions over textbook ones (e.g. Indonesian: prefer 'belum' over 'tidak' for not-yet-done actions, prefer 'udah' over 'sudah').",
    "formal": "Translate in formal, polite, professional language suitable for official announcements or documents.",
}
translation_tone = "casual"       # global default: casual / natural / formal
translation_tone_custom = ""      # global custom tone text (overrides preset if non-empty)
group_tone_settings = {}          # per-group: {gid: {"tone": str, "custom": str}}

# Model auto-switch: use gpt-4o for long messages, gpt-4o-mini for short
model_default = "gpt-4o-mini"     # model for short messages
model_upgrade = "gpt-4o"          # model for long messages
model_threshold = 0               # char count threshold (0 = always use default, no auto-switch)

import threading as _threading
_tl = _threading.local()          # thread-local for passing tone into translate_openai

def get_group_tone(group_id):
    """Return (preset, custom_text) for a group."""
    if group_id and group_id in group_tone_settings:
        gs = group_tone_settings[group_id]
        return gs.get("tone", translation_tone), gs.get("custom", "")
    return translation_tone, translation_tone_custom


def pick_model(text):
    """Pick OpenAI model based on text length and threshold setting."""
    if model_threshold > 0 and len(text) >= model_threshold:
        return model_upgrade
    return model_default

# Custom sender name/icon for translation messages
sender_name = "翻譯小助手"
sender_icon = ""  # URL to icon image, empty = default
# User profile pictures cache: {user_id: url}
user_pictures = {}

# ── Password settings (editable from admin) ──
pw1_text = "班長工號密碼：(尚未設定)\nPassword shift leader: (not set)"
pw2_text = "儲運工號密碼：(尚未設定)\nPassword gudang: (not set)"

# ── Scrap color text ──
scrap_text = (
    "🎨 廢料鋼種顏色 / Warna Scrap\n"
    "==================\n"
    "U物料(廢料) / U Material:\n"
    "  303 → 白/Putih\n"
    "  304 → 黃/Kuning\n"
    "  316 → 桃/Pink\n"
    "  209 → 特藍/Biru Khusus\n"
    "  174 → 紫羅蘭/Ungu\n"
    "  400系列 → 紅/Merah\n"
    "\n"
    "委外代工 / Outsource:\n"
    "  303 → 白/Putih\n"
    "  304 → 黃/Kuning\n"
    "  316 → 桃/Pink\n"
    "  403 → 紅/Merah\n"
    "=================="
)

# ── Packaging code lookup ──
PACKAGING_LOOKUP = {}

# USD to TWD rate (approximate)
USD_TO_TWD = 32.0


def get_group_feature(group_id, feature):
    """Get per-group feature setting with global fallback."""
    _map = {
        'flex': (group_flex_settings, 'flex_enabled'),
        'quick_reply': (group_qr_settings, 'quick_reply_enabled'),
        'silent': (group_silent_settings, 'silent_mode'),
        'video_ocr': (group_video_settings, 'video_ocr_enabled'),
        'location': (group_location_settings, 'location_translate_enabled'),
        'mark_read': (group_mark_read_settings, 'mark_read_enabled'),
        'retry_key': (group_retry_key_settings, 'retry_key_enabled'),
        'camera_qr': (group_camera_qr_settings, 'camera_qr_enabled'),
        'clipboard_qr': (group_clipboard_qr_settings, 'clipboard_qr_enabled'),
        'camera_roll_qr': (group_camera_roll_qr_settings, 'camera_roll_qr_enabled'),
        'location_qr': (group_location_qr_settings, 'location_qr_enabled'),
    }
    if feature not in _map:
        return True
    d, global_key = _map[feature]
    if group_id and group_id in d:
        return d[group_id]
    return globals().get(global_key, True)


def get_group_welcome(group_id):
    """Get per-group welcome settings with global fallback."""
    if group_id and group_id in group_welcome_settings:
        gw = group_welcome_settings[group_id]
        # Merge with global defaults for missing keys
        return {
            "enabled": gw.get("enabled", welcome_settings.get("enabled", True)),
            "text_zh": gw.get("text_zh", welcome_settings.get("text_zh", "")),
            "text_id": gw.get("text_id", welcome_settings.get("text_id", "")),
        }
    return welcome_settings

# Translation cache: key = (text, src, tgt), value = (result, timestamp)
translation_cache = {}
_cache_lock = _threading.Lock()
CACHE_MAX_SIZE = 500
CACHE_TTL = 3600  # 1 hour

# Message cache for quoted message context: {message_id: {"text": str, "ts": float}}
message_cache = {}
MESSAGE_CACHE_MAX = 200

LANG_FLAGS = {
    "zh": "\U0001f1f9\U0001f1fc",
    "id": "\U0001f1ee\U0001f1e9",
}

LANG_NAMES = {
    "zh": "Traditional Chinese",
    "id": "Indonesian",
}

LANG_NAMES_ZH = {
    "id": "\u5370\u5c3c\u6587",
}

# Valid target languages
VALID_TARGETS = ["id"]


def extract_mentions(text):
    """Extract @mentions from text. Skip @Indonesian_word (not real mentions)."""
    _id_skip = {
        'tolong','semua','untuk','yang','dan','ini','itu','ada','tidak','akan',
        'sudah','bisa','juga','saya','kami','kita','mereka','dia','apa','belum',
        'sedang','harus','boleh','mau','bukan','jangan','terima','kasih','baik',
        'bagus','benar','salah','kerja','pulang','pergi','karena','tapi','atau',
        'kalau','masih','lagi','nanti','sekarang','siap','izin','minta','cepat',
        'capek','sakit','gak','udah','gimana','dong','banget','kipas','mesin',
        'rusak','bocor','macet','stok','habis','ganti','pasang','gudang','masuk',
        'keluar','tutup','buka','material','selesai','beres','datang','besok',
        'kemarin','libur','lembur','cuti','proses','produksi','diperhatikan',
        'selalu','mohon','pakai','pake','cek','lihat','bilang','ambil','kirim',
        'tunggu','bantu','butuh','perlu','panggil','suruh','hati','awas',
        'bahaya','lantai','mesin','pompa','pipa','oli','besi','baja','batang',
    }
    mentions = []
    # English @mentions: grab @word + up to 2 more, trim Indonesian words from end
    for m in re.finditer(r'@([A-Za-z0-9][A-Za-z0-9_.-]*)(?:\s+([A-Za-z0-9_.-]+))?(?:\s+([A-Za-z0-9_.-]+))?', text):
        first = m.group(1)
        if first.lower() in _id_skip:
            continue
        parts = [first]
        for g in [m.group(2), m.group(3)]:
            if g and g.lower() not in _id_skip:
                parts.append(g)
            else:
                break
        mention = '@' + ' '.join(parts)
        if mention not in mentions:
            mentions.append(mention)
    # Chinese @mentions
    for m in re.findall(r'@[\u4e00-\u9fff\u3040-\u30ff]+(?:\s*[\uff08(][^\uff09)]*[\uff09)])?', text):
        m = m.rstrip()
        if m and len(m) > 1 and m not in mentions:
            mentions.append(m)
    # @All
    for m in re.findall(r'@[Aa][Ll][Ll]', text):
        if m not in mentions:
            mentions.append(m)
    return list(dict.fromkeys(mentions))


def extract_line_mentions(text, message):
    """Extract @mention strings using LINE's actual mention data (the blue text).
    Returns list of exact mention strings from the message."""
    mentions = []
    try:
        mention_data = getattr(message, 'mention', None)
        if mention_data and hasattr(mention_data, 'mentionees'):
            for m in mention_data.mentionees:
                idx = m.index
                length = m.length
                mention_text = text[idx:idx+length]
                if mention_text and mention_text not in mentions:
                    mentions.append(mention_text)
    except Exception:
        pass
    return mentions


def protect_mentions(text, line_mentions=None):
    # Use LINE's actual mention data if available (100% accurate)
    # Fall back to regex extraction if not
    if line_mentions:
        mentions = line_mentions
    else:
        mentions = extract_mentions(text)
    protected = text
    placeholders = {}
    for i, m in enumerate(mentions):
        ph = f"__MENTION_{i}__"
        if m in protected:
            placeholders[ph] = m
            protected = protected.replace(m, ph, 1)
    return protected, placeholders


def restore_mentions(text, placeholders):
    restored = text or ""
    for ph, original in placeholders.items():
        idx = ph.replace("__MENTION_", "").replace("__", "")
        variants = [
            ph,
            ph.replace("_", " "),
            ph.replace("__", ""),
            f"MENTION_{idx}",
            f"MENTION {idx}",
            f"__MENTION {idx}__",
            f"[[MENTION_{idx}]]",
        ]
        for v in variants:
            restored = restored.replace(v, original)

    # Final safety net: if any original @mention disappeared during translation,
    # prepend it back so the tagged person is not lost.
    missing = [original for original in placeholders.values() if original not in restored]
    if missing:
        prefix = " ".join(missing)
        restored = (prefix + " " + restored).strip()
    return restored


def strip_mentions_for_detect(text, line_mentions=None):
    """Strip @mentions for language detection."""
    if line_mentions:
        # Use LINE's actual mention data - most accurate
        clean = text
        for m in line_mentions:
            clean = clean.replace(m, ' ')
        return clean
    _id_skip = {
        'tolong','semua','untuk','yang','dan','ini','itu','ada','tidak','akan',
        'sudah','bisa','juga','saya','kami','kita','mereka','dia','apa','belum',
        'sedang','harus','boleh','mau','bukan','jangan','terima','kasih','baik',
        'kerja','pulang','pergi','karena','tapi','atau','kalau','masih','lagi',
        'siap','izin','minta','capek','sakit','gak','udah','gimana','dong',
        'kipas','mesin','rusak','bocor','macet','stok','habis','ganti','pasang',
        'gudang','masuk','keluar','tutup','buka','material','selesai','beres',
        'datang','besok','kemarin','libur','lembur','cuti','proses','produksi',
        'selalu','mohon','pakai','pake','cek','lihat','bilang','ambil','kirim',
        'tunggu','bantu','butuh','perlu','panggil','suruh','hati','awas',
    }
    def _replace_en(m):
        first_word = re.match(r'@([A-Za-z0-9]+)', m.group(0))
        if first_word and first_word.group(1).lower() in _id_skip:
            return m.group(0)  # Keep: not a real @mention
        return ' '
    clean = re.sub(r'@[A-Za-z0-9][A-Za-z0-9 _.-]*(?:\s+[\u4e00-\u9fff]{1,4})?(?=(?:\s|[\n,\uff0c\u3002!\uff01?\uff1f:\uff1a;\uff1b()\uff08\uff09\[\]{}<>\u201c\u201d]|$))', _replace_en, text)
    # Strip Chinese @mentions
    clean = re.sub(r'@[\u4e00-\u9fff]+(?:\s*[\uff08(][^\uff09)]*[\uff09)])?', ' ', clean)
    return clean


def has_chinese(text):
    return len(re.findall(r'[\u4e00-\u9fff]', text)) >= 2


def has_japanese(text):
    hira = len(re.findall(r'[\u3040-\u309f]', text))
    kata = len(re.findall(r'[\u30a0-\u30ff]', text))
    return (hira + kata) >= 2


def has_korean(text):
    return len(re.findall(r'[\uac00-\ud7af]', text)) >= 2


def has_thai(text):
    return len(re.findall(r'[\u0e00-\u0e7f]', text)) >= 2


def has_vietnamese(text):
    vi_special = re.findall(r'[\u01a0\u01a1\u01af\u01b0\u0110\u0111]', text)
    vi_chars = re.findall(r'[\u00e0-\u00ff\u1ea0-\u1ef9]', text.lower())
    vi_marks = re.findall(r'[\u0300-\u036f]', text)
    words = text.lower().split()
    vi_markers = set([
        'cua', 'nhung', 'trong', 'duoc', 'khong', 'nhu', 'mot',
        'toi', 'ban', 'anh', 'chi', 'em', 'ong', 'ba',
        'la', 'va', 'cac', 'cho', 'voi', 'tai', 'nay', 'khi',
        'con', 'roi', 'lam', 'biet', 'muon', 'den', 'di',
        'xin', 'cam', 'chao', 'dep', 'ngon', 'tot', 'xau',
    ])
    marker_count = sum(1 for w in words if w in vi_markers)
    if len(vi_special) >= 1:
        return True
    if len(vi_chars) >= 3 and marker_count >= 1:
        return True
    if len(vi_marks) >= 2 and marker_count >= 1:
        return True
    return False


def has_indonesian(text):
    if has_chinese(text) or has_thai(text) or has_korean(text) or has_japanese(text):
        return False
    words = re.findall(r'[a-zA-Z]+', text.lower())
    if len(words) < 2:
        return False
    id_words = set([
        # ── Pronouns / titles ──
        'saya', 'aku', 'gue', 'gw', 'kamu', 'lu', 'elo', 'dia', 'mereka',
        'kami', 'kita', 'kalian', 'bapak', 'ibu', 'pak', 'bu', 'mas', 'mbak',
        'bang', 'kak', 'om', 'tante', 'bos', 'boss', 'gan',
        # ── Particles / fillers ──
        'ya', 'lah', 'loh', 'dong', 'sih', 'nih', 'kok', 'deh', 'kan',
        'tuh', 'nah', 'wah', 'aduh', 'masa', 'emang', 'kayak', 'kayaknya',
        'soalnya', 'makanya', 'jadinya', 'aja', 'doang', 'cuma', 'gitu',
        'gini', 'sini', 'sana', 'situ', 'mana', 'iya', 'oke',
        # ── Prepositions / conjunctions ──
        'di', 'ke', 'dari', 'pada', 'oleh', 'untuk', 'dengan', 'supaya',
        'agar', 'karena', 'tetapi', 'tapi', 'namun', 'sehingga', 'meskipun',
        'walaupun', 'sebelum', 'sesudah', 'setelah', 'selama', 'ketika',
        'sambil', 'tanpa', 'antara', 'tentang', 'terhadap', 'atau', 'dan',
        'jika', 'kalau', 'biar', 'sampai',
        # ── Question words ──
        'apa', 'siapa', 'dimana', 'kapan', 'kenapa', 'bagaimana', 'berapa',
        'gimana', 'mana', 'mengapa',
        # ── Common verbs ──
        'ada', 'adalah', 'ambil', 'angkat', 'antar', 'atur', 'bangun',
        'bantu', 'bawa', 'bayar', 'beli', 'berangkat', 'berhenti', 'bicara',
        'bilang', 'bisa', 'bikin', 'boleh', 'buat', 'buang', 'buka', 'butuh',
        'cari', 'catat', 'cek', 'coba', 'cuci', 'dapat', 'datang', 'duduk',
        'ganti', 'hapus', 'hitung', 'hubungi', 'ikut', 'ingat', 'isi',
        'jaga', 'jalan', 'jawab', 'jemput', 'jual', 'kasih', 'kejar',
        'keluar', 'kembali', 'kirim', 'kurang', 'lari', 'lepas', 'lewat',
        'lihat', 'lupa', 'makan', 'masak', 'masuk', 'mau', 'minum', 'minta',
        'naik', 'paham', 'pakai', 'pake', 'panggil', 'pasang', 'perbaiki',
        'pergi', 'periksa', 'pindah', 'potong', 'pulang', 'selesai',
        'sembuh', 'simpan', 'suruh', 'tahu', 'tau', 'tambah', 'tanya',
        'taruh', 'tiba', 'tidur', 'tinggal', 'tolong', 'tukar', 'tulis',
        'tunggu', 'turun', 'tutup', 'ngerti', 'paham', 'ngomong', 'ngobrol',
        'nyari', 'nyoba', 'nunggu', 'ngitung', 'ngirim', 'ngecek', 'ngangkat',
        'ingin', 'harus', 'boleh', 'perlu', 'wajib',
        # ── Common adjectives ──
        'bagus', 'baik', 'baru', 'benar', 'berat', 'besar', 'bersih',
        'buruk', 'cepat', 'dingin', 'gampang', 'gelap', 'jelek', 'kecil',
        'keras', 'kotor', 'kuat', 'lambat', 'lama', 'lebar', 'lemah',
        'lurus', 'mahal', 'miring', 'murah', 'panas', 'panjang', 'pendek',
        'penuh', 'rata', 'ringan', 'salah', 'sehat', 'sempit', 'susah',
        'tajam', 'tebal', 'terang', 'tipis', 'tua', 'muda', 'lembut',
        'kasar', 'kosong', 'basah', 'kering',
        # ── Nouns (general) ──
        'air', 'api', 'asap', 'barang', 'batu', 'biaya', 'botol', 'cat',
        'dinding', 'ember', 'gelas', 'helm', 'kabel', 'kaca', 'kain',
        'kamar', 'kayu', 'kertas', 'kotak', 'kursi', 'lampu', 'listrik',
        'meja', 'mobil', 'motor', 'obat', 'paku', 'papan', 'pintu',
        'plastik', 'rak', 'roda', 'sabun', 'sapu', 'selang', 'sepatu',
        'surat', 'tangga', 'tali', 'tas', 'tiang', 'topi', 'truk',
        # ── Nouns (work / factory) ──
        'absen', 'alat', 'atasan', 'bahan', 'bengkel', 'bor', 'crane',
        'debu', 'forklift', 'gaji', 'gerinda', 'gudang', 'jadwal',
        'kartu', 'kecelakaan', 'kerusakan', 'kualitas', 'laporan',
        'las', 'limbah', 'lini', 'logam', 'lubang', 'mandor', 'masalah',
        'meter', 'mutu', 'pabrik', 'pekerja', 'pelindung', 'peralatan',
        'perbaikan', 'peraturan', 'permukaan', 'produksi', 'produk',
        'rapat', 'sabuk', 'sampel', 'shift', 'sisa', 'sparepart',
        'supervisor', 'tabung', 'tekanan', 'timbangan', 'toleransi',
        'tungku', 'upah', 'wadah',
        # ── Factory equipment / materials ──
        'kipas', 'angin', 'mesin', 'pompa', 'kunci', 'baut', 'mur',
        'pipa', 'oli', 'besi', 'baja', 'batang', 'stok', 'material',
        'lantai', 'atas', 'bawah', 'ukuran', 'nomor',
        'bocor', 'macet', 'mati', 'hidup', 'nyala', 'jalan',
        # ── Factory actions ──
        'ukur', 'timbang', 'sortir', 'pisah', 'gabung', 'campur', 'cetak',
        'press', 'poles', 'tekuk', 'lipat', 'gulung', 'tarik', 'dorong',
        'geser', 'putar', 'balik', 'susun', 'tumpuk', 'bungkus', 'ikat',
        'segel', 'proses', 'bagian', 'tempat',
        # ── Safety / quality ──
        'bahaya', 'aman', 'keselamatan', 'cidera', 'luka', 'awas', 'hati',
        'peringatan', 'darurat', 'masker', 'kacamata', 'rompi', 'cacat',
        'retak', 'gores', 'bengkok', 'penyok', 'standar', 'inspeksi',
        'audit', 'lapor',
        # ── Time ──
        'detik', 'menit', 'jam', 'hari', 'minggu', 'bulan', 'tahun',
        'pagi', 'siang', 'sore', 'malam', 'subuh', 'kemarin', 'sekarang',
        'besok', 'lusa', 'nanti', 'dulu', 'tadi', 'segera', 'selalu',
        'sering', 'kadang', 'jarang',
        # ── Numbers / quantity ──
        'satu', 'dua', 'tiga', 'empat', 'lima', 'enam', 'tujuh',
        'delapan', 'sembilan', 'sepuluh', 'puluh', 'ratus', 'ribu',
        'juta', 'setengah', 'cukup', 'terlalu', 'sekitar', 'kira',
        'banyak', 'sedikit', 'semua', 'beberapa',
        # ── States / emotions ──
        'senang', 'sedih', 'marah', 'takut', 'capek', 'cape', 'males',
        'lapar', 'haus', 'sakit', 'sehat', 'ngantuk', 'bosan', 'bingung',
        'kaget', 'malu', 'bangga', 'puas', 'kecewa', 'khawatir', 'tenang',
        'sibuk', 'santai', 'mantap', 'keren', 'asik',
        # ── Daily / HR ──
        'izin', 'cuti', 'libur', 'lembur', 'istirahat', 'kerja', 'masuk',
        'pulang', 'absen', 'telat', 'terlambat', 'ijin', 'sakit', 'mangkir',
        'resign', 'kontrak', 'tetap', 'harian', 'bulanan', 'THR',
        # ── Negation / affirmation ──
        'tidak', 'bukan', 'belum', 'jangan', 'sudah', 'akan', 'sedang',
        'masih', 'lagi', 'saja', 'juga', 'pernah', 'tidak', 'tanpa',
        'hanya', 'bahkan', 'sangat', 'amat', 'sekali', 'paling',
        # ── Slang abbreviations ──
        'gak', 'nggak', 'ga', 'gk', 'udah', 'udh', 'uda',
        'gmn', 'bgt', 'org', 'yg', 'tdk', 'dg', 'dgn', 'krn',
        'blm', 'bs', 'sy', 'trs', 'tp', 'tpi', 'sm', 'lg',
        'dl', 'skrg', 'hr', 'msh', 'brp', 'dpt', 'hrs', 'kmrn',
        'bsk', 'wkwk', 'otw', 'gpp', 'jgn', 'tlg', 'cb', 'emg',
        'stlh', 'sblm', 'tgl', 'mksd', 'kl', 'krj', 'plg', 'msk',
        'klr', 'btw', 'fyi', 'cmn', 'drpd', 'blg', 'klo', 'knp',
        'dmn', 'gmna', 'bkn', 'sbg', 'ttg', 'scr', 'utk',
        # ── Common responses ──
        'siap', 'beres', 'selesai', 'oke', 'sip', 'mantap', 'lanjut',
        'betul', 'benar', 'setuju', 'mengerti', 'paham', 'jelas',
        'terima', 'kasih', 'makasih', 'maaf', 'permisi', 'selamat',
        'halo', 'hai', 'assalamualaikum', 'waalaikumsalam',
        # ── Misc common ──
        'orang', 'baru', 'lain', 'beda', 'sama', 'sendiri', 'bersama',
        'bareng', 'duluan', 'belakangan', 'awal', 'akhir', 'mulai',
        'setiap', 'tiap', 'per', 'masing', 'soal', 'hal', 'cara',
        'jenis', 'tipe', 'macam', 'warna', 'bentuk', 'sisi', 'ujung',
        'tengah', 'tepi', 'pinggir', 'depan', 'belakang', 'kiri', 'kanan',
        'dalam', 'luar', 'atas', 'bawah', 'samping', 'sebelah',
        'dekat', 'jauh', 'sini', 'sana', 'situ',
        'yang', 'ini', 'itu', 'rumah', 'kantor', 'uang', 'harga',
        'yuk', 'ayo', 'banget', 'ruang', 'baca', 'ujian', 'terakhir',
        'punya', 'jadi', 'mohon', 'saat', 'secara', 'harap', 'rusak',
        'habis', 'bulat', 'kamu',
        # ── Food / break ──
        'nasi', 'ayam', 'ikan', 'sayur', 'teh', 'kopi', 'susu', 'roti',
        'mie', 'goreng', 'rebus', 'pedas', 'manis', 'asin', 'pahit',
        'warung', 'kantin',
        # ── Missing common words (comprehensive) ──
        'kata', 'sandi', 'nama', 'alamat', 'telepon', 'email', 'buku',
        'dengar', 'pikir', 'rasa', 'cinta', 'suka', 'benci',
        'teman', 'musuh', 'keluarga', 'anak', 'istri', 'suami',
        'adik', 'kakak', 'ayah', 'nenek', 'kakek', 'paman', 'bibi',
        # ── Places ──
        'negara', 'kota', 'desa', 'gedung', 'toko', 'sekolah',
        'masjid', 'gereja', 'pasar', 'bandara', 'stasiun', 'terminal',
        'rumah', 'hotel', 'restoran', 'warnet', 'bengkel',
        # ── Transport ──
        'sepeda', 'pesawat', 'kapal', 'kereta', 'bis', 'taksi', 'ojek',
        # ── Food detail ──
        'makanan', 'minuman', 'buah', 'sayuran', 'daging', 'beras',
        'garam', 'gula', 'minyak', 'tepung', 'bumbu', 'sambal',
        'telur', 'tempe', 'tahu', 'soto', 'bakso', 'sate',
        # ── Body ──
        'tangan', 'kaki', 'kepala', 'mata', 'telinga', 'mulut',
        'hidung', 'perut', 'punggung', 'dada', 'bahu', 'jari',
        'lutut', 'siku', 'leher', 'pinggang', 'tumit', 'bibir',
        # ── Abstract / reasoning ──
        'milik', 'hak', 'kewajiban', 'tugas', 'tanggung', 'jawab',
        'solusi', 'metode', 'alasan', 'tujuan', 'maksud', 'arti',
        'makna', 'contoh', 'info', 'informasi', 'berita', 'pesan',
        # ── Ability / certainty ──
        'mampu', 'sanggup', 'berani', 'gembira', 'sulit', 'mudah',
        'lebih', 'hampir', 'nyaris', 'mungkin', 'pasti', 'tentu',
        'yakin', 'ragu', 'percaya',
        # ── ber- prefix verbs ──
        'bekerja', 'belajar', 'bermain', 'berlari', 'berjalan',
        'berbicara', 'berpikir', 'berharap', 'berdoa', 'bernyanyi',
        'beristirahat', 'berbelanja', 'bertemu', 'bercerita',
        'berdiri', 'berbaring', 'berputar', 'bergerak', 'berhenti',
        'bergabung', 'berpisah', 'bertugas', 'bertanya', 'berubah',
        # ── me- prefix verbs ──
        'membuat', 'membeli', 'menjual', 'membawa', 'mengambil',
        'memberikan', 'menerima', 'mengirim', 'menyimpan', 'membuang',
        'mencari', 'menemukan', 'menunggu', 'melihat', 'mendengar',
        'menulis', 'membaca', 'menghitung', 'mengukur', 'memotong',
        'membuka', 'menutup', 'menyalakan', 'mematikan', 'menghubungi',
        'menelepon', 'mengecek', 'memeriksa', 'memperbaiki', 'mengganti',
        'memasang', 'melepas', 'mengisi', 'mengosongkan', 'membersihkan',
        'mencuci', 'membantu', 'meminta', 'memakai', 'memasak',
        'memilih', 'memiliki', 'mengerti', 'mengetahui', 'memulai',
        'menyelesaikan', 'mengerjakan', 'melapor', 'melaporkan',
        'mengatur', 'mengantar', 'menjaga', 'menjemput', 'menaruh',
        'memindahkan', 'mengangkat', 'menurunkan', 'mendorong',
        'menarik', 'memutar', 'menekan', 'mengunci', 'merasa',
        # ── di- prefix (passive) ──
        'dibuat', 'dibeli', 'dijual', 'dibawa', 'diambil', 'dikirim',
        'disimpan', 'dibuang', 'dicari', 'ditemukan', 'dilihat',
        'ditulis', 'dibaca', 'dihitung', 'diukur', 'dipotong',
        'dibuka', 'ditutup', 'dinyalakan', 'dimatikan', 'dicek',
        'diperiksa', 'diperbaiki', 'diganti', 'dipasang', 'dilepas',
        'diisi', 'dibersihkan', 'dicuci', 'diminta', 'dipakai',
        'dipilih', 'diketahui', 'dikerjakan', 'dilaporkan', 'diatur',
        'dijaga', 'ditaruh', 'dipindahkan', 'diangkat', 'diturunkan',
        'diperhatikan', 'disampaikan', 'dilakukan', 'diberikan',
        'diterima', 'digunakan', 'disediakan', 'dibutuhkan',
        # ── Documents / admin ──
        'formulir', 'dokumen', 'berkas', 'file', 'data', 'rekening',
        'tabungan', 'pinjaman', 'hutang', 'bayaran', 'diskon', 'gratis',
        'untung', 'rugi', 'modal', 'surat', 'izin', 'tanda', 'tangan',
        'stempel', 'cap', 'kuitansi', 'faktur', 'invoice', 'nota',
        # ── Dimensions / colors ──
        'tinggi', 'rendah', 'dangkal', 'halus', 'cair', 'padat', 'lunak',
        'merah', 'kuning', 'hijau', 'biru', 'putih', 'hitam',
        'coklat', 'abu', 'emas', 'perak', 'ungu', 'oranye', 'pink',
        # ── Weather / nature ──
        'hujan', 'awan', 'mendung', 'cerah', 'badai', 'banjir',
        'gempa', 'petir', 'kabut', 'embun',
        # ── Speed / manner ──
        'bentar', 'sebentar', 'langsung', 'pelan', 'keras', 'kencang',
        'terburu', 'santai', 'harian', 'mingguan', 'bulanan', 'tahunan',
        # ── ke-...-an nouns ──
        'keselamatan', 'kecelakaan', 'kerusakan', 'kebersihan',
        'keamanan', 'kesehatan', 'kecepatan', 'keterlambatan',
        'kekurangan', 'kelebihan', 'kesalahan', 'keberhasilan',
        'kemampuan', 'kebutuhan', 'keperluan', 'keterangan',
        # ── per-...-an nouns ──
        'perbaikan', 'perubahan', 'perbedaan', 'perhatian',
        'perkembangan', 'pertemuan', 'perjanjian', 'perusahaan',
        'pekerjaan', 'peralatan', 'peraturan', 'perlengkapan',
        'permintaan', 'pengiriman', 'penggantian', 'pemasangan',
        'pemeriksaan', 'pembersihan', 'pengisian', 'pengecekan',
        'penggunaan', 'pemakaian', 'pelaksanaan', 'pelaporan',
        # ── Common endings -kan / -i ──
        'pastikan', 'perhatikan', 'sampaikan', 'lakukan', 'berikan',
        'gunakan', 'sediakan', 'siapkan', 'selesaikan', 'kerjakan',
        'beritahukan', 'hubungi', 'temui', 'cari', 'ambilkan',
        'tolong', 'mohon', 'harap', 'silakan', 'silahkan',
    ])
    count = sum(1 for w in words if w in id_words)
    if count >= 2:
        return True
    if count >= 1 and len(words) >= 2 and count / len(words) >= 0.4:
        return True
    return False


def has_english(text):
    if has_chinese(text) or has_thai(text) or has_korean(text) or has_japanese(text):
        return False
    if has_vietnamese(text) or has_indonesian(text):
        return False
    words = re.findall(r'[a-zA-Z]+', text.lower())
    if len(words) < 3:
        return False
    en_words = set([
        'the', 'is', 'are', 'was', 'were', 'have', 'has', 'had',
        'will', 'would', 'could', 'should', 'can', 'may', 'might',
        'this', 'that', 'these', 'those', 'what', 'which', 'who',
        'where', 'when', 'how', 'why', 'not', 'but', 'and', 'or',
        'for', 'with', 'from', 'about', 'into', 'your', 'you',
        'we', 'they', 'she', 'him', 'her', 'its', 'our', 'their',
        'just', 'also', 'very', 'much', 'more', 'most', 'some',
        'any', 'all', 'each', 'every', 'been', 'being', 'does',
        'did', 'doing', 'going', 'want', 'need', 'know', 'think',
        'come', 'make', 'like', 'time', 'good', 'new', 'first',
        'please', 'thank', 'thanks', 'sorry', 'hello', 'okay',
        'yes', 'yeah', 'already', 'still', 'here', 'there',
    ])
    count = sum(1 for w in words if w in en_words)
    if count >= 2:
        return True
    if len(words) > 0 and count / len(words) > 0.25:
        return True
    return False


def detect_language(text):
    """Detect language: Chinese → 'zh', Latin text → 'id'.
    For mixed messages (factory codes + Chinese), Chinese dominates."""
    clean = strip_mentions_for_detect(text).strip()
    if not clean or len(clean) < 2:
        return None
    zh_count = len(re.findall(r'[\u4e00-\u9fff]', clean))
    latin_words = re.findall(r'[a-zA-Z]{2,}', clean.lower())
    # Has Chinese characters — if Chinese dominates or Latin is minimal, it's Chinese
    if zh_count >= 2 and zh_count >= len(latin_words):
        return "zh"
    # No Chinese but has Latin → Indonesian
    if zh_count == 0 and latin_words:
        return "id"
    # Both exist but Latin dominates → Indonesian
    if latin_words and len(latin_words) > zh_count:
        return "id"
    # Only Chinese (1+ chars)
    if zh_count >= 1:
        return "zh"
    # Has some Latin words
    if latin_words:
        return "id"
    return None


def contains_source_script_outside_placeholders(text, src):
    cleaned = re.sub(r'__MENTION_\d+__', ' ', text or '')
    cleaned = re.sub(r'__CUST_\d+__', ' ', cleaned)
    # Also strip known customer names (they are kept in original language intentionally)
    for name in CUSTOMER_NAMES:
        if name in cleaned:
            cleaned = cleaned.replace(name, ' ')
    patterns = {
        "zh": r'[\u4e00-\u9fff]',
        "ja": r'[\u3040-\u30ff\u4e00-\u9fff]',
        "ko": r'[\uac00-\ud7af]',
        "th": r'[\u0e00-\u0e7f]',
    }
    pattern = patterns.get(src)
    if not pattern:
        return False
    return len(re.findall(pattern, cleaned)) >= 2


def is_translation_valid(result, src, tgt):
    if not result or not result.strip():
        return False
    if src != tgt and contains_source_script_outside_placeholders(result, src):
        return False
    return True


# === Hard replacement tables ===
# These bypass GPT entirely - applied BEFORE sending to GPT (zh->id)
# and AFTER receiving from GPT (id->zh result post-processing)

ZH_TO_ID_HARD = {
    # 製程/站別
    "爐號標籤": "label heat number",
    "爐號": "heat number",
    "無心研磨": "centerless grinding",
    "光輝退火爐": "furnace bright annealing",
    "光輝退火": "bright annealing",
    "退火爐": "tungku annealing",
    "過帳": "input data ke sistem",
    "放行": "release data",
    # 品質/缺陷
    "殺光痕": "bekas grinding mark",
    "車刀痕": "bekas pisau bubut",
    "砂光痕": "bekas sanding mark",
    "軋輥印痕": "bekas roll mark",
    "環狀擦傷": "goresan melingkar",
    "表粗": "surface roughness",
    "偏小": "under size",
    "偏大": "over size",
    "風險批": "lot berisiko",
    "走ET檢測": "jalankan pengujian ET",
    "開立重工": "buat work order rework",
    "不允收": "pelanggan tidak terima",
    # 設備
    "矯直機": "mesin straightening",
    "壓光機": "mesin press polish",
    "砂光機": "mesin sanding",
    "拋光機": "mesin polishing",
    "眼模": "die/cetakan",
    "引拔座": "drawing bench",
    "皮膜槽": "coating tank",
    "氣壓缸": "silinder pneumatik",
    "安全圍籬": "safety fence",
    "集塵設備": "dust collector",
    "計長器": "length counter",
    "冷水機": "chiller",
    "馬蹄環": "shackle",
    "吊掛物": "beban gantung",
    "護罩": "pelindung mesin",
    "interlock": "pengunci keamanan",
    "標籤機": "mesin label",
    # 管理
    "品保": "QC",
    "儲運": "bagian gudang",
    "生計": "production planning",
    "業務": "bagian sales",
    "營業": "bagian sales",
    "人事": "HRD",
    "處長": "kepala divisi",
    "稼動率": "utilization rate",
    "線速": "kecepatan lini",
    "速差": "selisih kecepatan",
    "主機手": "operator utama",
    "印勞": "pekerja Indonesia",
    "在製品管制表": "tabel kontrol WIP",
    # 包裝/入庫
    "套紙管": "pasang tabung kertas",
    "太空包": "jumbo bag",
    "噴漆罐": "kaleng spray",
    "木箱": "kotak kayu",
    "櫃子": "kontainer",
    # 訂單
    "允收": "toleransi terima",
    "訂尺": "panjang pesanan",
    "短尺": "ukuran pendek",
    "異型棒": "batang bentuk khusus",
    "遞延單": "order ditunda",
    "急單": "order urgent",
    "不擋非本月": "order bukan bulan ini boleh masuk gudang",
    "不擋": "tidak dibatasi",
    "溢量": "kelebihan produksi",
    "併包": "gabung packing",
    "出貨差": "kekurangan pengiriman",
    # HR/紀律
    "忘卡補": "input lewat sistem lupa kartu",
    "造冊": "buat daftar absensi",
    "班股": "rapat shift",
    "堆高機複訓": "pelatihan ulang forklift",
    "天車複訓": "pelatihan ulang crane",
    "扣績效": "potong penilaian kinerja",
    "劣項": "pelanggaran",
    "納入劣項": "dicatat pelanggaran",
    "提報懲處": "laporkan untuk sanksi",
    "三定": "3 tetap",
    "不要物": "barang tidak terpakai",
    "被釘": "kena tegur",
    "綠卡": "kartu hijau",
    # 環境
    "煙蒂": "puntung rokok",
    "檳榔渣": "sisa pinang",
    "廚餘": "sisa makanan",
    "漏油": "bocor oli",
    "積水": "genangan air",
    "粉塵": "debu",
    # 口語
    "感溫": "terima kasih",
    "有夠": "sangat",
    "母湯": "jangan",
}

# Post-replacement: fix common GPT mistakes in output
ID_POST_FIX = {
    # 爐號 corrections
    "nomor panas": "heat number",
    "label nomor panas": "label heat number",
    "nomor tungku": "heat number",
    "label nomor tungku": "label heat number",
    "nomor oven": "heat number",
    "label nomor oven": "label heat number",
    # 有包到 corrections
    "paket datang ke": "kalau ada packing untuk",
    "saat paket datang ke": "kalau ada packing untuk",
    "Mohon diperhatikan saat paket datang ke": "Nanti kalau ada packing untuk",
    "Mohon diperhatikan saat kalau ada packing untuk": "Nanti kalau ada packing untuk",
    # 三米六米 corrections
    "tiga meter di atas enam meter": "batang 3 meter ditaruh di atas batang 6 meter",
    "Tiga meter di atas enam meter": "Batang 3 meter ditaruh di atas batang 6 meter",
    "3 meter di atas 6 meter": "batang 3 meter ditaruh di atas batang 6 meter",
    # 品保 corrections
    "jaminan kualitas": "QC",
    "penjaminan mutu": "QC",
    # 點名 corrections (NOT roll call)
    "panggilan nama": "inspeksi pengawas",
    "absen nama": "inspeksi pengawas",
    "roll call": "inspeksi pengawas",
    # 感溫 - should not be translated literally
    "suhu perasaan": "terima kasih",
    "merasakan suhu": "terima kasih",
    # Common GPT errors
    "Polymetal": "寶麗金屬",
    "Bao Li Metal": "寶麗金屬",
    "Bao Li Logam": "寶麗金屬",
    "Changzhou Zhongshan": "常州眾山",
    "Da Shun": "大順",
    "Da Cheng": "大成",
    "Bei Ze": "北澤",
    "Hong Yun": "鴻運",
    "Tian Hua Rong": "田華榕",
    "Jia Dong": "佳東",
    # 營業 common mistranslation
    "bagian operasional": "bagian sales",
    "operasional perlu": "sales perlu",
}

# Customer names - protect from translation by wrapping
# Storage area lookup data (from 儲區查詢.xlsx)
_STORAGE_JSON = '{"6C422209": [["<=3200", "EH28"], [">4200", "EG38"], [">3200<=4200", "EH26"]], "ABE": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "AIK": [[">3200<=4200", "EG14"], [">4200", "EH33"], ["<=3200", "EH28"]], "ALCONIX JP": [["<=3200", "EG14"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "AMERICAN STAINLESS": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "AMS": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "ANCHOR": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "ANIL METALS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "APEX METAL": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "AWACS": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "B&B": [[">4200", "EH33"], ["<=3200", "EH22"], [">3200<=4200", "EG14"]], "B&J": [["<=3200", "EC40"], [">4200", "EC40"], [">3200<=4200", "EC45"]], "BOBCO": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "BOLLINGHAUS": [[">3200<=4200", "EC43"], ["<=3200", "EC43"], [">4200", "EC43"]], "CA-ASD": [[">4200", "EH11"], ["<=3200", "EH12"], [">3200<=4200", "EH12"]], "CA-AUSTRAL": [[">3200<=4200", "EH12"], ["<=3200", "EH12"], [">4200", "EH11"]], "CA-DALSTEEL": [[">4200", "EH11"], ["<=3200", "EH12"], [">3200<=4200", "EH12"]], "CA-FLETCHER": [[">3200<=4200", "EH12"], [">4200", "EH11"], ["<=3200", "EH28"]], "CA-M&S": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-MICO": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-MIDWAY": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-S&T": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-VAN LEEUWEN": [["<=3200", "EH12"], [">4200", "EH11"], [">3200<=4200", "EH12"]], "CA-VES": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-VULCAN": [[">4200", "EH11"], ["<=3200", "EH12"], [">3200<=4200", "EH12"]], "CA-VULCAN NZ": [["<=3200", "EH12"], [">3200<=4200", "EH12"], [">4200", "EH11"]], "CA-WAKEFIELD": [[">4200", "EH11"], [">3200<=4200", "EH12"], ["<=3200", "EH12"]], "CAMELLIA": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "CASTLE": [[">3200<=4200", "EH12"], ["<=3200", "EH28"], [">4200", "EH11"]], "CHANDAN": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "CHANG HSIN": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "CHANGSU": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "COGNE AOSTA": [[">3200<=4200", "EG34"], ["<=3200", "EH28"], [">4200", "EG14"]], "COGNE CELIK": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "COGNE DE": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH28"]], "COGNE DG": [[">3200<=4200", "EC47"], ["<=3200", "EC47"], [">4200", "EC41"]], "COGNE FR": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "COGNE KR": [["<=3200", "EH26"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "COGNE UK": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH28"]], "COMINOX": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "COMPRINOX": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "CSMU": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "DACAPO": [["<=3200", "EH25"], [">3200<=4200", "EG14"], [">4200", "EH31"]], "DACAPO-K STOCK": [["<=3200", "EH25"], [">3200<=4200", "EG14"], [">4200", "EH31"]], "DAECHANG": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "DAMSTAHL": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "DAVER": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "DK METAL": [[">4200", "EG35"], [">3200<=4200", "EG14"], ["<=3200", "EC47"]], "DUFU": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "EGMO": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "EIAM": [[">3200<=4200", "EG14"], ["<=3200", "EG14"], [">4200", "EH33"]], "ESP": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "EURO STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "FASTENAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "FINE METAL TRADE": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "FSS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "G HWA": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "GIC": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "GLH": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "GS METAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "HADCO": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "HAKUDO": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "HAMATECH": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "HANWA": [["<=3200", "EH21"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "HEAP SING HUAT": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "HH": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "HRMETAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "HUA GUAN METAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "HWA GUAN METAL": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "IM": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "INTEGRITY STAINLESS": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "IPE": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "ISE": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "IWATANI": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "JANG ANN": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "JFE SHOJI": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "KANGRUI": [[">3200<=4200", "EG14"], [">4200", "EC45"], ["<=3200", "EH28"]], "KANSAI": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "KDK": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "KIAN": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "KIM ANN": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "KJ": [[">4200", "EG32"], ["<=3200", "EC47"], [">3200<=4200", "EG14"]], "KJ PRECISION": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "KOMINOX AB": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "LAI KING": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "LAURIE": [["<=3200", "EG14"], [">3200<=4200", "EH28"], [">4200", "EH33"]], "LE": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "LEE & STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "LIM MENG SENG": [[">3200<=4200", "EG14"], [">4200", "EG34"], ["<=3200", "EH28"]], "LINSTER": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "LOTUS METAL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "LTM": [["<=3200", "EG15"], [">3200<=4200", "EG15"], [">4200", "EG34"]], "M.R. STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "MAINCHAIN": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "MAN TAK": [[">4200", "EG34"], ["<=3200", "EG15"], [">3200<=4200", "EG15"]], "MARINE": [["<=3200", "EG14"], [">3200<=4200", "EH28"], [">4200", "EH33"]], "MCB": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "MENAM": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "METAL ESTABLISH": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "METALINOX": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "METALLSERVIS": [[">3200<=4200", "EH14"], ["<=3200", "EH28"], [">4200", "EG35"]], "NAKAYAMA": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "NDE": [["<=3200", "EH28"], [">4000", "EG34"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "NM": [[">3200<=4200", "EG14"], ["<=3200", "EH22"], [">4200", "EG34"]], "NMSK": [[">4200", "EG34"], ["<=3200", "EH28"]], "NOVA TRADING": [["<=3200", "EH27"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "NOXFAP": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "NS METAL": [[">3200<=4200", "EG14"], ["<=3200", "EG14"], [">4200", "EH18"]], "NSC": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "OKAYA": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EH14"]], "OLYMPIC STEEL": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "OME": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "PACKER(ISRAEL)": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EH14"]], "PASCAL": [[">3200<=4200", "EH14"], ["<=3200", "EH28"], [">4200", "EG34"]], "PF": [["<=3200", "EH28"], [">3200<=4200", "EH14"], [">4200", "EG34"]], "PLUTUS": [[">3200<=4200", "EI30"], ["<=3200", "EI25"], [">4200", "EI40"]], "PRECISION": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EH14"]], "PRECISION METAL": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EH14"]], "PRECISION METALS": [[">3200<=4200", "EH14"], ["<=3200", "EH28"], [">4200", "EH33"]], "QPLUS": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "RAAJRATNA": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "RHS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "RINO": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "RISEBM": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SAGAMI": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SAMWON": [["<=3200", "EC47"], [">3200<=4200", "EG14"], [">4200", "EG32"]], "SCM": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SCOT": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EI40"]], "SD-BK": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "SD-BKL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "SD-KHS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH33"]], "SD-LIM METAL": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SD-METALPHILE": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EG34"]], "SD-METHA": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SD-TPS": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SENG HUAT": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "SENG HUAT METALPLEX": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "SGH": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "SHIMIZU MATERIAL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SHINKO": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SHINKO TH": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SING LEONG-雙馬": [["<=3200", "EH28"], [">4200", "EH34"], [">3200<=4200", "EG14"]], "SLA": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "SMG": [["<=3200", "EH28"], [">4200", "EG33"], [">3200<=4200", "EG14"]], "SPECTROMATRIX": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "STEELINC": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "STEWART": [[">3200<=4200", "EG14"], [">4200", "EH33"], ["<=3200", "EH28"]], "STIRLINGS": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "STIRLINGS(5%)": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH28"]], "STKSTAINLESS": [["<=3200", "EH28"], [">4200", "EH33"], [">3200<=4200", "EG14"]], "STRONG STEEL": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH34"]], "SUNGEUN": [[">4200", "EG33"], ["<=3200", "EG37"], [">3200<=4200", "EG14"]], "SUNGSIL METAL": [[">4200", "EG35"], ["<=3200", "EC47"], [">3200<=4200", "EG14"]], "SUPERFIX": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "SUPREME": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "TAN VIET": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "TCI": [["<=3200", "EH32"], [">3200<=4200", "EH32"], [">4200", "EH32"]], "TEKPOINT": [[">3200<=4200", "EG14"], ["<=3200", "EG14"], [">4200", "EG34"]], "TITAN METALS": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "TK-SCHULTE": [[">4200", "EH33"], [">3200<=4200", "EG14"], ["<=3200", "EH22"]], "TKMP": [[">3200<=4200", "EG14"], [">4200", "EH34"], ["<=3200", "EH26"]], "TMC": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "TOP SUNNY": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "TOZZHIN THAILAND": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "TSA": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "TSM": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EG34"]], "TUBE SUPPLY": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "TUSCO": [[">3200<=4200", "EG15"], [">4200", "EG34"], ["<=3200", "EH28"]], "WESCO": [[">4200", "EG34"], [">3200<=4200", "EG15"], ["<=3200", "EH28"]], "WEST COAST": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "WING KEUNG": [[">3200<=4200", "EG14"], [">4200", "EH33"], ["<=3200", "EH29"]], "WPS": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "YGS": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "YIEH CORP LTD(HK)": [["<=3200", "EH28"], [">4200", "EG34"], [">3200<=4200", "EG14"]], "YONGTA": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "YOSHU": [[">4200", "EH33"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "YOUCHANG": [[">4200", "EG34"], [">3200<=4200", "EG14"], ["<=3200", "EH28"]], "YOUNG DONG": [[">3200<=4200", "EG15"], ["<=3200", "EG15"], [">4200", "EH33"]], "？頂": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "？暉": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "力常(觀音)": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "三大興": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "三卯鍛壓": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "三利": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "上晉": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "上海凡斯": [["<=3200", "EC47"], [">4200", "EC40"], [">3200<=4200", "EC45"]], "上海坤成": [["<=3200", "EC47"], [">3200<=4200", "EC40"], [">4200", "EC40"]], "上海億科": [[">3200<=4200", "EC40"], [">4200", "EC40"], ["<=3200", "EC47"]], "上海町芃": [["<=3200", "EH10"], [">4200", "EH10"], [">3200<=4200", "EH10"]], "上銀": [["<=3200", "EH99"], [">4200", "EC40"], [">3200<=4200", "EH99"]], "凡立": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "千里眼": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "大甲永和": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "大成": [[">4200", "EH32"], [">3200<=4200", "EH32"], ["<=3200", "EH32"]], "大連德邁仕": [["<=3200", "EC47"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "大順": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "中國防蝕": [[">4200", "EH35"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "元盈": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "元偉勝": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "升暘": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EG38"]], "天津隆德": [[">4200", "EC40"], ["<=3200", "EC47"], [">3200<=4200", "EC40"]], "方鉦": [[">3200<=4200", "EH72"], [">4200", "EH72"], ["<=3200", "EH79"]], "世廷": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "世華": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "功億": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "北澤": [[">4200", "EG38"], [">3200<=4200", "EG39"], ["<=3200", "EG39"]], "北澤一廠": [["<=3200", "EG39"], [">3200<=4200", "EG39"], [">4200", "EG38"]], "北澤二廠": [[">4200", "EG38"], ["<=3200", "EG39"], [">3200<=4200", "EG38"]], "北澤三廠": [["<=3200", "EG39"], [">3200<=4200", "EG38"], [">4200", "EG38"]], "右勝鋼鐵": [[">3200<=4200", "EH78"], ["<=3200", "EG39"], [">4200", "EH71"]], "台芝": [[">4200", "EH10"], ["<=3200", "EH10"], [">3200<=4200", "EH10"]], "台灣亞錁": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "台灣林吉": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "台灣矽微": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "巨昌": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "巨頻": [[">3200<=4200", "EG14"], [">4200", "EG38"], ["<=3200", "EH79"]], "永川泰": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "永村": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "生計直棒": [[">3200<=4200", "EH99"], ["<=3200", "EH99"], [">4200", "EH99"]], "生計庫存": [[">3200<=4200", "EH99"], [">4200", "EH99"], ["<=3200", "EH99"]], "禾桀": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH38"]], "光翔": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "全利金屬": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "全敏尖端": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "向春": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "名威": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "合順": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "宇隆": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "宇慶": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "有光": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "江陰外庫": [["<=3200", "EC47"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "江陰華新": [[">4200", "EC40"], [">3200<=4200", "EC40"], ["<=3200", "EC47"]], "江蘇迪威": [[">4200", "EC40"], [">3200<=4200", "EC47"], ["<=3200", "EC47"]], "汎新": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "百呈": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "百堅": [[">3200<=4200", "EG37"], [">4200", "EH33"], ["<=3200", "EG14"]], "西邁金屬": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH33"]], "君立": [["<=3200", "EH79"], [">4200", "EH36"], [">3200<=4200", "EH78"]], "壯安": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "宏盈": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "宏荃": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "志典": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH33"]], "志聯": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "甫剛": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "貝加": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "貝克休斯": [[">3200<=4200", "EG38"], [">4200", "EG38"], ["<=3200", "EG38"]], "京碼": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "京鋼": [[">4200", "EC41"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "佳東": [[">3200<=4200", "EH76"], ["<=3200", "EH76"], [">4200", "EH70"]], "佳東-台中": [[">4200", "EH70"], ["<=3200", "EH76"], [">3200<=4200", "EH78"]], "佳東-台北": [[">4200", "EH70"], [">3200<=4200", "EH78"], ["<=3200", "EH76"]], "佳東-高雄": [[">3200<=4200", "EH78"], [">4200", "EH70"], ["<=3200", "EH76"]], "協崎": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "坤泰": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "奇賓": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EG38"]], "孟駿": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "尚智": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "岡山東穎": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "承總": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "易隆": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "昆山金富盈": [[">3200<=4200", "EC40"], ["<=3200", "EC47"], [">4200", "EC40"]], "明石": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "東栗": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "東莞峰作": [["<=3200", "EC47"], [">3200<=4200", "EC40"], [">4200", "EC40"]], "東萊": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "東徽": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH33"]], "武漢機械": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "金大": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "金利山": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "金亞洲": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "金城": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "金耘": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH71"]], "金耘-南營所": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH71"]], "金煜": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH71"]], "長盈": [[">4200", "EG38"], [">3200<=4200", "EG14"], ["<=3200", "EH79"]], "長圓": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "俊來(蘆洲)": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "俊益": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "厚群": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "威孚高科技": [["<=3200", "EC47"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "建新": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "柏緯": [[">3200<=4200", "EC47"], [">4200", "EC42"], ["<=3200", "EC47"]], "津展": [["<=3200", "EH75"], [">4200", "EH71"], [">3200<=4200", "EH71"]], "津展-台中": [[">4200", "EH72"], ["<=3200", "EH75"], [">3200<=4200", "EH72"]], "津展-台北": [[">3200<=4200", "EH72"], [">4200", "EH72"], ["<=3200", "EH75"]], "津展-台南": [[">3200<=4200", "EH72"], ["<=3200", "EH75"], [">4200", "EH72"]], "皇銘": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "研發部": [["<=3200", "EH99"], [">3200<=4200", "EH99"], [">4200", "EH99"]], "研發測試": [["<=3200", "EH99"], [">3200<=4200", "EH99"], [">4200", "EH99"]], "科威聯": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "英鈿": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "重慶九勝": [[">4200", "EC40"], ["<=3200", "EC47"], [">3200<=4200", "EC47"]], "重慶九環": [[">4200", "EC40"], ["<=3200", "EC47"], [">3200<=4200", "EC40"]], "展舵": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "峰作金屬": [["<=3200", "EH74"], [">3200<=4200", "EH78"], [">4200", "EH71"]], "峰勝": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "振家": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "振華興": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "時哲": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "晉易": [[">4200", "EH38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "晉椿": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH38"]], "晉椿(鹿港)": [[">4200", "EH38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "浙江三花": [[">4200", "EG34"], ["<=3200", "EH28"], [">3200<=4200", "EG14"]], "益陽": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "退庫重工": [[">4200", "EH99"], [">3200<=4200", "EH99"], ["<=3200", "EH99"]], "高立熱處理": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "高銪": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "商旺": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "域鑫科技": [[">4200", "EC40"]], "常州眾山": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EC43"]], "強淞": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "強實": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "捷流": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "淳康": [[">3200<=4200", "EH10"], ["<=3200", "EH10"], [">4200", "EH10"]], "眾山": [[">3200<=4200", "EH78"], [">4200", "EG35"], ["<=3200", "EH79"]], "祥日達": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "祥英": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "笠源": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "頂翔勝": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "麥億": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "備料庫存": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "凱記": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "勝初": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "勝新": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "勝盟": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "富億鑫": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "尊茂": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "復盛應用": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH78"]], "敦壹": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "朝盟": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "無錫永雋": [["<=3200", "EH28"], [">3200<=4200", "EG14"], [">4200", "EH33"]], "舜欽": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "華友(外)": [[">3200<=4200", "EG14"], ["<=3200", "EH28"], [">4200", "EH34"]], "華纜": [[">4200", "EH10"], [">3200<=4200", "EH10"], ["<=3200", "EH10"]], "詠勗": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "詠晟": [["<=3200", "EH79"], [">3200<=4200", "EC47"], [">4200", "EC40"]], "進達": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "開滋": [["<=3200", "EG39"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "隆明": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "隆門": [[">4200", "EG38"], [">3200<=4200", "EH28"], ["<=3200", "EH28"]], "隆順發": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "雅信億": [[">3200<=4200", "EH14"], ["<=3200", "EH79"], [">4200", "EH33"]], "廉喬": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "廉錩": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH77"]], "廉錩-台北": [[">3200<=4200", "EH78"], ["<=3200", "EH77"], [">4200", "EG38"]], "廉錩-台南": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH77"]], "慈溪龍華": [[">4200", "EC40"], [">3200<=4200", "EC40"], ["<=3200", "EC47"]], "新創捷": [["<=3200", "EH79"], [">3200<=4200", "EH14"], [">4200", "EH33"]], "新華特聯": [["<=3200", "EH79"], [">3200<=4200", "EH14"], [">4200", "EH35"]], "新萊應材": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "瑞鋼": [[">4200", "EC40"], [">3200<=4200", "EC40"], ["<=3200", "EC47"]], "盟鉦": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "萬揚": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EH33"]], "經捷": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "經貿": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "群鎰": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "聖泰": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "路竹新益": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "鉅泰昇": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EH36"]], "鉅銅": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "鉅豐": [["<=3200", "EH79"]], "鼎崴": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "嘉冠": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "嘉碁": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "寧波東葛": [[">3200<=4200", "EC40"], ["<=3200", "EC47"], [">4200", "EC40"]], "慷倫": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "睿緻佳": [["<=3200", "EH78"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "福泉": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "聚祥": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "銓宥": [[">4200", "EH10"], [">3200<=4200", "EH10"], ["<=3200", "EH10"]], "廣泰": [[">4200", "EH10"], [">3200<=4200", "EH10"], ["<=3200", "EH10"]], "慶鋐": [[">3200<=4200", "EH78"], [">4200", "EG38"], ["<=3200", "EH79"]], "歐承": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "毅鋼": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "磐石": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "誼山": [["<=3200", "EH29"], [">3200<=4200", "EH78"], [">4200", "EH38"]], "頭份": [[">4200", "EG38"], [">3200<=4200", "EH78"], ["<=3200", "EH79"]], "優普洛": [["<=3200", "EH79"], [">3200<=4200", "EH78"], [">4200", "EG38"]], "營三備庫(內)": [[">3200<=4200", "EC40"], ["<=3200", "EC47"], [">4200", "EC40"]], "營三備庫(外)": [[">3200<=4200", "EC40"], [">4200", "EC40"], ["<=3200", "EC47"]], "營業庫存": [[">4200", "EH99"], ["<=3200", "EH99"], [">3200<=4200", "EH99"]], "環友": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "聯岱": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "聯祥": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "邁達斯": [[">4200", "EG38"], ["<=3200", "EH79"], [">3200<=4200", "EH78"]], "鴻運": [[">3200<=4200", "EH27"], ["<=3200", "EH27"], [">4200", "EG38"]], "雙和": [[">3200<=4200", "EG14"], ["<=3200", "EH26"], [">4200", "EG34"]], "麒譯": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "町洋": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "晟田": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "畯圓": [["<=3200", "EH19"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "鐿順發": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "鑫誠鐵材": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "恒耀": [["<=3200", "EH79"], [">4200", "EG38"], [">3200<=4200", "EH78"]], "暉": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]], "頂": [[">3200<=4200", "EH78"], ["<=3200", "EH79"], [">4200", "EG38"]]}'
STORAGE_LOOKUP = json.loads(_STORAGE_JSON)
# Try loading from storage_data.json (auto-updated via admin panel)
_storage_json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "storage_data.json")
if os.path.exists(_storage_json_path):
    try:
        with open(_storage_json_path, "r", encoding="utf-8") as _f:
            STORAGE_LOOKUP = json.load(_f)
            logger.info("Loaded storage data from storage_data.json: %d customers", len(STORAGE_LOOKUP))
    except Exception as _e:
        logger.warning("Failed to load storage_data.json, using embedded: %s", _e)
# Try loading packaging_data.json
_packaging_json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "packaging_data.json")
if os.path.exists(_packaging_json_path):
    try:
        with open(_packaging_json_path, "r", encoding="utf-8") as _f:
            PACKAGING_LOOKUP = json.load(_f)
            logger.info("Loaded packaging data: %d codes", len(PACKAGING_LOOKUP))
    except Exception as _e:
        logger.warning("Failed to load packaging_data.json: %s", _e)
# Extra customers not in storage Excel but appear in factory chat
# Per-group protected names: {"__all__": [...], "group_id": [...]}
_DEFAULT_NAMES = [
    "寶麗金屬", "田華榕", "蘋果", "賽利金屬", "盛昌遠", "曜麟",
    "LOTUS", "LOTUS METAL", "shinko", "wing keung",
    "高侑", "十元", "小麥", "啊堂", "秋情", "政軒", "碩凱", "汶錡",
    "武駿", "凱銘", "小趙", "阿澤", "法比恩", "山多", "EggEgg", "fang", "Dato潘",
    "阿添", "小叮噹", "多啦A夢", "潘柏良", "大彭",
]
extra_names_by_group = {"__all__": list(_DEFAULT_NAMES)}
EXTRA_CUSTOMERS = []

def rebuild_customer_names():
    """Rebuild EXTRA_CUSTOMERS and CUSTOMER_NAMES from all groups."""
    global EXTRA_CUSTOMERS, CUSTOMER_NAMES
    merged = set()
    for names in extra_names_by_group.values():
        merged.update(names)
    EXTRA_CUSTOMERS = sorted(list(merged), key=lambda x: -len(x))
    CUSTOMER_NAMES = sorted(list(set(list(STORAGE_LOOKUP.keys()) + EXTRA_CUSTOMERS)), key=lambda x: -len(x))

CUSTOMER_NAMES = []
rebuild_customer_names()


def pre_replace_zh(text):
    """Apply hard replacements to Chinese text before GPT translation.
    Returns (modified_text, customer_placeholders_dict)."""
    result = text
    # Protect customer names with placeholders (these survive GPT translation)
    cust_ph = {}
    for i, name in enumerate(CUSTOMER_NAMES):
        if name in result:
            ph = f"__CUST_{i}__"
            cust_ph[ph] = name
            result = result.replace(name, ph)
    # Apply hard replacements (longest first to avoid partial matches)
    for zh, replacement in sorted(ZH_TO_ID_HARD.items(), key=lambda x: -len(x[0])):
        if zh in result:
            result = result.replace(zh, f"[{replacement}]")
    return result, cust_ph


def restore_customers(text, cust_ph):
    """Restore customer name placeholders back to original names."""
    if not text or not cust_ph:
        return text
    result = text
    for ph, name in cust_ph.items():
        # GPT might mangle placeholders, try variants
        idx = ph.replace("__CUST_", "").replace("__", "")
        variants = [
            ph, ph.replace("_", " "), f"CUST_{idx}", f"CUST {idx}",
            f"__CUST {idx}__", f"[CUST_{idx}]",
        ]
        for v in variants:
            if v in result:
                result = result.replace(v, name)
    # Safety: if any customer name placeholder pattern remains, try regex
    result = re.sub(r'__CUST_(\d+)__', lambda m: cust_ph.get(f"__CUST_{m.group(1)}__", m.group(0)), result)
    return result


def post_fix_translation(text):
    """Fix known GPT translation mistakes in output."""
    if not text:
        return text
    result = text
    # Fix specific wrong translations (longest match first)
    for wrong, correct in sorted(ID_POST_FIX.items(), key=lambda x: -len(x[0])):
        result = result.replace(wrong, correct)
    # Remove bracketed hints that leaked through from pre_replace
    result = re.sub(r'\[([a-zA-Z /&]+)\]', r'\1', result)
    # Clean up double spaces (preserve newlines)
    result = re.sub(r'[^\S\n]+', ' ', result)
    result = re.sub(r'\n{3,}', '\n\n', result)
    result = result.strip()
    return result


def translate_openai(text, src, tgt, strict_no_source_script=False, repair_mode=False, bad_result=None):
    if not oai:
        return None
    try:
        src_name = LANG_NAMES.get(src, src)
        tgt_name = LANG_NAMES.get(tgt, tgt)

        # Apply hard replacements before GPT for zh->other
        input_text = text
        cust_placeholders = {}
        if src == "zh":
            input_text, cust_placeholders = pre_replace_zh(text)

        protected, placeholders = protect_mentions(input_text)

        extra_rule = ""
        if strict_no_source_script and src != tgt:
            if src == "zh":
                extra_rule = (
                    " 10. IMPORTANT: Do not leave any Chinese words untranslated unless they are a person's name or __MENTION__ placeholder."
                    " Terms such as 印籍, 印尼籍, 早班, 夜班, 考試, 讀書, 下班後 must be translated into the target language."
                )
            elif src == "ja":
                extra_rule = " 10. IMPORTANT: Do not leave Japanese text untranslated unless it is a person's name or __MENTION__ placeholder."
            elif src == "ko":
                extra_rule = " 10. IMPORTANT: Do not leave Korean text untranslated unless it is a person's name or __MENTION__ placeholder."
            elif src == "th":
                extra_rule = " 10. IMPORTANT: Do not leave Thai text untranslated unless it is a person's name or __MENTION__ placeholder."

        # Get tone from thread-local (set by handler before calling translate)
        _tone = getattr(_tl, 'tone', 'casual')
        _tone_custom = getattr(_tl, 'tone_custom', '')
        tone_instruction = _tone_custom if _tone_custom else TONE_PRESETS.get(_tone, TONE_PRESETS['casual'])

        sys_prompt = (
            "You are a professional translator for a stainless steel factory (Walsin Lihwa/華新麗華, Yanshui plant) work group chat. "
            "This factory produces stainless steel bars, wire rods, peeled bars, cold-drawn bars using processes like rolling, annealing, pickling, peeling, cold drawing, and centerless grinding. "
            "This is a group with Taiwanese managers and Indonesian migrant workers operating centerless grinding (無心研磨) equipment. "
            "CRITICAL RULES: "
            "1. NEVER translate @mentions and NEVER translate or romanize person names. Keep all Chinese names in ORIGINAL CHINESE CHARACTERS. "
            "For example: 徐嘉騰 stays as 徐嘉騰, NOT Xu Jiateng. 陳弘林 stays as 陳弘林, NOT Chen Honglin. "
            "Chinese nicknames for people must stay unchanged. Do NOT translate them literally. "
            "2. Any text like __MENTION_0__, __MENTION_1__ etc are placeholders - keep them exactly as is. "
            "3. TRANSLATION TONE/STYLE: " + tone_instruction + " "
            "4. Indonesian slang: gak=tidak, udah=sudah, gimana=bagaimana, bgt=banget, org=orang, yg=yang, tdk=tidak, dg=dengan, krn=karena, blm=belum, hrs=harus, bs=bisa, lg=lagi, gw=saya, lu=kamu. "
            "5. TAIWANESE MANDARIN COLLOQUIAL (very important): "
            "乾/干=aduh/astaga, 靠=astaga/waduh, 幹=sial/buset, 傻眼=gak percaya, 扯/誇張=keterlaluan, 笑死=ngakak, 氣死=kesel banget, 累死=capek banget, "
            "啦=lah/dong, 喔/哦=ya/lho, 耶=dong/nih, 嘛=dong/kan, 蛤=hah?/apa?, 厚=ya kan, "
            "醬/降=begitu/gitu(=這樣), 母湯=jangan/gak boleh(=不要), 超/有夠=banget(=非常), 感溫=terima kasih(台語感恩), "
            "CRITICAL: Taiwanese rhetorical questions SUGGEST doing something: 需不需要X=perlu X gak nih(suggesting X should be done), 要不要X=gimana kalau X, 還在X=masih X(often implies criticism). "
            "搞什麼=ngapain sih, 搞定=beres, 人咧=orangnya mana, 怎麼搞的=kenapa bisa begini, 出包=ada masalah, 先這樣=segitu dulu ya, 再說=nanti aja, "
            "X到不行/X得要死/X到爆=X banget, 怎麼這麼X=kok X banget, 有夠X=X banget, "
            "ㄏㄏ=haha, QQ=sedih, 3Q=terima kasih, GG=tamat, XD=haha, @@=bingung. "
            "6. Target Traditional Chinese = Taiwan style, not mainland. "
            "7. Target Indonesian = simple clear daily language for factory workers. "
            "8. Context: factory work - shifts, overtime, orders, tasks, meals, breaks, meetings, exams. "
            "9. FACTORY VOCABULARY: "
            "【製程/Process】"
            "無心研磨=centerless grinding, 研磨=grinding, 砂輪=batu gerinda, 調整輪=roda pengatur, 刀板=work rest blade, 冷卻液=cairan pendingin, "
            "不鏽鋼=stainless steel, 棒鋼=steel bar, 盤元=wire rod, 削皮棒=peeled bar, 冷精棒=cold-drawn bar, "
            "熱軋=hot rolling, 退火=annealing, 酸洗=pickling, 削皮=peeling, 冷抽=cold drawing, "
            "鋼種=jenis baja, PMI=uji material, 來料=material masuk, 棒材=batang baja, 混料=tercampur material(SERIOUS), 料號=nomor material, "
            "拋光=polishing, 粗拋=rough polishing, 噴漆=spray paint, 洗料=cuci material, "
            "倒角=chamfer, 修磨=repair grinding, 盤元修磨=repair grinding wire rod, 線外修磨=offline repair grinding, "
            "壓光=press polish, 矯直=straightening, 重矯=straightening ulang, 精整=finishing, AP=mesin finishing, "
            "光輝退火=bright annealing, 回爐=kirim kembali ke furnace, "
            "側磨=side grinding(DILARANG/prohibited), 不可側磨=dilarang side grinding, "
            "【站別/Stations - numbers are STATION NUMBERS】"
            "400站=station 400, 401站=station 401, 420站=station 420, "
            "470站=station 470(UT station), UT=mesin UT(di station 470), 480站=station 480, "
            "490站=station 490(秤重站/timbang), 801站=station 801, "
            "OL=sedang produksi/online, 回400=kembalikan ke station 400, "
            "無主=tanpa pemilik/unassigned, 入無主=masukkan ke status tanpa pemilik, "
            "掛單/工單=work order, 重掛單=pasang ulang work order, 無工單資訊=tidak ada info work order, "
            "改制=ubah proses, 去化=ada order baru mau terima, 有單去化=ada order baru untuk serap material, 改制去化=ubah proses produksi, "
            "帳/帳務=data administrasi(ERP), 帳已回400=data sudah dikembalikan ke station 400, "
            "過帳=input data produksi(jumlah&berat)ke sistem tanpa release ke stasiun berikutnya, "
            "放行=release data ke stasiun berikutnya(setelah QC lulus), "
            "退庫=kembalikan ke gudang, 退庫拆包=keluarkan dari gudang bongkar packing untuk dibagi ulang, "
            "發料=issue material, 存檔=simpan data, 暫存=simpan sementara, 短尺=ukuran pendek, "
            "溢量=kelebihan produksi melebihi permintaan, 併包=gabung packing dari lot berbeda dalam order sama, "
            "出貨差=kekurangan pengiriman hari ini, 轉用=dialihkan untuk order lain, 跳無主轉用=pindah ke tanpa pemilik lalu dialihkan, "
            "【班次/出勤】"
            "點名=ada pengawas yang datang(inspection, NOT roll call), 早班=shift pagi, 夜班=shift malam, 中班=shift siang, "
            "加班=lembur, 排班=jadwal shift, 調班=tukar shift, 上班=masuk kerja, 下班=pulang kerja, 打卡=absen, "
            "請假=izin, 病假=izin sakit, 事假=izin pribadi, 特休=cuti tahunan, 代班=gantikan shift, "
            "忘卡補=lupa kartu ID, pakai sistem input waktu, 造冊=buat daftar absensi, "
            "班股=rapat shift, 堆高機複訓=pelatihan ulang forklift, 天車複訓=pelatihan ulang crane, "
            "紅包=angpao, 年終獎金=bonus akhir tahun, 過年不停機=Imlek tidak berhenti produksi, "
            "【產線/設備】"
            "產線=lini produksi, 機台=mesin, 開機=nyalakan mesin, 停機=mesin berhenti, 調機=setting mesin, "
            "上料=isi material, 備料=siapkan material, 產量=jumlah produksi, 目標=target, 達標=capai target, 超產=over production, "
            "訂單=order, 出貨=kirim barang, 交期=deadline, 趕貨=kejar order, 急單=order urgent, 急單備註=catatan order urgent, "
            "下製程=proses selanjutnya, 異常=abnormal/ada masalah, 維修中=sedang diperbaiki, "
            "天車=overhead crane, 台車=trolley, 吊秤=timbangan gantung, 馬蹄環=shackle, 鋼索=sling baja, 吊掛物=beban gantung, "
            "稼動率=utilization rate, 線速=line speed(m/min), 限速=batas kecepatan, 降速=turunkan kecepatan, 提速=naikkan kecepatan, 速差=selisih kecepatan, "
            "撥料=feed material, 過機=lewatkan mesin, 線外=offline, 印勞=pekerja Indonesia, "
            "砂光機=sanding machine, 眼模=die/cetakan drawing, 引拔座=drawing bench, 皮膜槽=coating tank, "
            "查修=investigasi&perbaiki, 修護=maintenance, 儀電=instrumen listrik, 備品=spare part, "
            "跳異常=error muncul, 復歸=reset, 復歸無效=reset gagal, 跳機=mesin trip, 恢復生產=kembali produksi, "
            "叫修=panggil teknisi, 進廠查修=teknisi masuk pabrik cek, 電聯儀電=hubungi instrumen listrik, "
            "斷料=material putus, 卡料=material macet, 擠料=material terjepit keluar, "
            "主機手=operator utama, 上料人員=petugas pengisian material, 點檢=cek rutin, 護罩=pelindung mesin/safety guard, "
            "interlock=pengunci keamanan(jangan ditahan pakai benda), "
            "【印尼文機械/設備詞彙 Indonesian Mechanical Terms】"
            "as=軸/軸心(axle/shaft), as roda=輪軸(wheel axle), roda=輪(wheel), roda penarik=拉料輪(pulling wheel), "
            "penarik barang=拉料車/拖料車(material puller/cart), "
            "kopel=萬向接頭(universal joint/coupling), cross joint=十字接頭(cross joint), "
            "as roda penarik barang patah=拉料輪的萬向接頭斷裂(pulling wheel universal joint broken), "
            "patah=斷了/斷裂(snapped/broken off), bengkok=彎了(bent), "
            "retak=裂了(cracked), aus=磨損(worn out), bocor=漏(leak), macet=卡住(jammed), "
            "bearing=軸承(bearing), rantai=鏈條(chain), sabuk=皮帶(belt), engsel=鉸鏈(hinge), "
            "kawat=鋼線/線材(wire), selang=軟管(hose), katup=閥門(valve), baut=螺栓(bolt), mur=螺帽(nut), "
            "tekanan=壓力(pressure), getaran=震動(vibration), gesekan=摩擦(friction), pelumas=潤滑油(lubricant), "
            "gigi/gear=齒輪(gear), kipas=風扇(fan), kipas angin=電風扇(electric fan), motor=馬達(motor), "
            "pompa=泵浦(pump), kompresor=壓縮機(compressor), pipa=管(pipe), tabung=鋼瓶/桶(tank/cylinder), "
            "dongkrak=千斤頂(jack), kunci=扳手/鑰匙(wrench/key), obeng=螺絲起子(screwdriver), tang=鉗子(pliers), "
            "las=焊接(welding), gerinda=砂輪機(grinder), bor=鑽孔機(drill), gergaji=鋸子(saw), "
            "forklift=堆高機(forklift), crane=吊車(crane), conveyor=輸送帶(conveyor), "
            "NOTE: 'As' in Indonesian mechanical context ALWAYS means axle/shaft(軸), never translate as 'as/像'. "
            "NOTE: 'patah' means snapped/broken off(斷了), different from 'rusak'(壞了/故障). "
            "【包裝/入庫】"
            "套紙管=pasang tabung kertas, 入庫=masuk gudang, 優先包裝入庫=prioritas packing masuk gudang, "
            "需求單=formulir permintaan, 可以全收=bisa diterima semua, "
            "櫃子=kontainer(shipping container), 櫃子在路上=kontainer sedang di jalan, "
            "木箱=kotak kayu, 裝箱=masukkan ke kotak kayu, 2700大的木箱=kotak kayu ukuran besar 2700mm, "
            "NOTE: 木箱 context: 3200/2400=box LENGTH mm, 500/1000=weight CAPACITY kg. "
            "把=bundel(bundle), 捆=bundel/ikat, 支/根=batang(piece/rod), 批=lot/batch, "
            "NOTE: X米(三米,六米)=batang X meter(bar LENGTH not distance). 三米上面放六米=batang 3m ditaruh di atas batang 6m. "
            "包(verb)=packing/kemas(NOT wrapping). 秤重=timbang, 貼標=tempel label, 綁鐵=ikat besi, "
            "【訂單管理】"
            "允收=jumlah yang boleh diterima pelanggan, 允收0支=zero tolerance, 不收短尺=tidak terima ukuran pendek, "
            "訂尺=panjang sesuai pesanan, 爐號=heat number/nomor furnace(NEVER translate as 'panas'), 爐號標籤=label heat number, "
            "分捆=pisah bundel, 遞延單=delayed order, 非本月=bukan order bulan ini, "
            "非本月不入庫=order bukan bulan ini jangan masuk gudang, 檔非本月=tahan order bukan bulan ini, "
            "異型棒=batang bentuk khusus, 異型棒不擋=batang khusus tidak dibatasi, "
            "不擋=tidak dibatasi/boleh masuk(exemption), 不擋非本月=order bukan bulan ini BOLEH masuk gudang(exception/exemption, NOT blocked), "
            "入庫目標=target masuk gudang, 壓日期=ada deadline ketat, "
            "管控=kontrol, 不管控=tidak dikontrol(bebas), "
            "【品質/缺陷】"
            "品保=QC, 會驗=joint inspection, 暫留=hold sementara, HOLD=tahan, "
            "客訴=komplain pelanggan, 夾帶樣品=sertakan sampel, 掛檔=simpan ke arsip, 稽核=audit, "
            "螺紋=thread mark, 車刀痕=turning tool mark, 砂光痕=sanding mark, 殺光痕=grinding mark, "
            "剝片=flaking, 軋輥印痕=roll mark, 碰傷=luka benturan, 黑皮=unfinished surface, "
            "偏小=under size, 偏大=over size, 表粗=surface roughness, 目視=visual inspection, "
            "開立重工=buat WO rework, 重工研磨至尺寸下限=rework grinding sampai batas bawah ukuran, "
            "不允收=pelanggan tidak terima, 風險批=lot berisiko, 走ET檢測=jalankan pengujian ET, "
            "卡料需關閉電源後再取料=material macet HARUS matikan listrik dulu baru ambil, "
            "【部門/人員】"
            "業務=sales, 營業=sales(=業務), 生計=production planning, 資訊=IT department, 品保=QC, 儲運=gudang&logistik, 人事=HRD, 工安=safety officer, "
            "處長=kepala divisi, 抓資料=ambil data, "
            "【標籤/系統】"
            "TAG=label, 儲區=area penyimpanan di sistem, 轉檔=konversi data, "
            "MES=MES(sistem produksi), 報表=laporan produksi, 條碼=barcode, "
            "標籤機=mesin label, 包裝電腦=komputer packing, "
            "在製品管制表=WIP control sheet, "
            "【安全/環境/紀律】"
            "太空包=jumbo bag/FIBC, 噴漆罐一定要打洞才能丟棄在太空包=kaleng spray HARUS dilubangi sebelum buang ke jumbo bag, "
            "扣績效=potong kinerja(sanksi), 劣項=pelanggaran, 納入劣項=dicatat pelanggaran, "
            "三定=3 tetap(tempat/barang/jumlah tetap), 不要物=barang tidak terpakai, "
            "漏油=bocor oli, 生鏽=berkarat, 掉漆=cat mengelupas, 積水=genangan air, 粉塵=debu, "
            "煙蒂=puntung rokok, 檳榔渣=sisa pinang, 被釘=kena tegur atasan, "
            "提報懲處=laporkan untuk sanksi, 會嚴罰=dihukum berat, "
            "綠卡=kartu hijau(catatan safety), KYT=pelatihan prediksi bahaya, 防火演練=latihan pemadam kebakaran, "
            "調班單=formulir tukar shift, 簽核=tanda tangan persetujuan, "
            "【生活/薪資】"
            "宿舍=asrama, 便當=bekal makan, 餵狗=kasih makan anjing, "
            "薪水=gaji, 加班費=uang lembur, 績效=penilaian kinerja, 匯款=transfer, "
            "尾牙=pesta akhir tahun, 春酒=pesta tahun baru, 伴手禮=oleh-oleh, 便當費=biaya makan siang, "
            "量測=mengukur, 尺寸=diameter, 公差=toleransi, 校正=kalibrasi, "
            "【客戶 - NEVER translate】"
            "DACAPO, CASTLE, LOTUS, METALINOX, KANGRUI, SUNGEUN, STEELINC, GLH, shinko, wing keung, "
            "田華榕, 佳東, 蘋果, 常州眾山, 大順, 大成, 巨昌, 北澤, 鴻運, 畯圓, 名威, 右勝, 貝克休斯, 皇銘, "
            "台芝, 百堅, 津展, 曜麟, 廉錩, 盛昌遠, 永吉, 光輝, 寶麗金屬. "
            "NOTE: 蘋果=customer NOT fruit. 光輝=customer OR 光輝退火(bright annealing), context determines. "
            "10. CRITICAL CONTEXT RULES: "
            "a) X米(三米,六米)=bar LENGTH. 三米上面放六米=batang 3m ditaruh di atas batang 6m. "
            "b) 把/捆=BUNDLE counters. 包2把=packing 2 bundel. "
            "c) 包(verb)=packing NOT wrapping. 高侑的今天包2把都這樣=Yang di-packing 高侑 hari ini 2 bundel semuanya kayak gini. "
            "d) Names(" + ",".join(EXTRA_CUSTOMERS) + ")=keep as-is. "
            "e) Customer names=keep as-is, do NOT translate. "
            "f) R+number=round bar diameter(R28.57=bulat 28.57mm). Non-R=hex/special(H26=hex 26mm). "
            "g) S/B=straight bar. E1~E11=cold drawing lines. I1~I21=grinding machines. BF2/3/5=polishing machines. "
            "h) 5F/5L/6S/6T/6U/6W/7E/7F/7G+numbers=work order ID, keep as-is. "
            "i) 課料=section chief designated material. G包=packing method code. AP=finishing equipment. "
            "j) 爐號=heat number(NEVER 'nomor panas'). 有包到X=kalau ada packing untuk X(NOT 'paket datang ke X'). "
            "k) 放=POLYSEMY(multiple meanings, judge by context): "
            "放+把/單/批/工單號/這把/這單/這批=RELEASE data(放行). e.g. 先放這把=release bundel ini dulu, 放了=sudah di-release, 幫放一下=tolong bantu release. "
            "放+地點/方位(地上/旁邊/上面/那邊/架上)=PUT/PLACE(taruh). e.g. 放地上=taruh di lantai, 三米上面放六米=batang 3m ditaruh di atas 6m. "
            "放+料/材料(without location)=FEED material. e.g. 放料=isi material. "
            "放假=libur/holiday. "
            "When ambiguous and context is about work orders or production flow, default to RELEASE(放行). "
            "l) 再=POLYSEMY: "
            "X再Y(condition+action)=hanya X yang Y / X baru Y(=才). e.g. 急單再幫忙安排入庫=hanya order urgent yang tolong bantu atur masuk gudang. "
            "再+verb(without preceding condition)=lagi/sekali lagi(=again). e.g. 再確認一下=confirm sekali lagi. "
            "m) 非本月=bukan order bulan ini(order that is NOT for the current month). 非本月包裝不入庫=yang bukan order bulan ini jangan packing masuk gudang. "
            "n) 不擋=tidak dibatasi/boleh masuk(EXEMPTION, means ALLOWED). 不擋非本月=order bukan bulan ini BOLEH masuk gudang. "
            "CRITICAL: 不擋 means NOT blocked = ALLOWED. Do NOT translate as tidak boleh(=blocked). "
            "e.g. DACAPO不擋非本月=DACAPO order bukan bulan ini boleh masuk gudang. "
            "o) When H、S appear in a list with 異型棒 or customer names, they are SEPARATE product categories(H=hex bar, S=straight bar). "
            "Keep them as individual items with commas. e.g. H、S異型棒=H, S, batang bentuk khusus(three separate types). "
            "11. TRANSLATION EXAMPLES (follow strictly): "
            "【中→印尼】"
            "乾 需不需要提報一下 → Aduh, perlu dilaporkan gak nih? "
            "UT囤一堆料了 → UT udah numpuk banyak material. "
            "品保還在下班 誇張 → QC udah pulang, keterlaluan. "
            "三米上面放六米 → Batang 3 meter ditaruh di atas batang 6 meter. "
            "麻煩他們不要這樣放料 → Tolong bilang ke mereka jangan taruh material kayak gini. "
            "高侑的今天包2把都這樣 → Yang di-packing 高侑 hari ini 2 bundel semuanya kayak gini. "
            "來料都短少4-5公斤 → Material masuk semuanya kurang 4-5 kilogram. "
            "已轉達 → Sudah disampaikan. "
            "這批料有問題 → Lot material ini ada masalah. "
            "幫我盯一下 → Tolong awasin ya. "
            "怎麼搞的啦 → Kok bisa kayak gini sih. "
            "人咧 → Orangnya mana? "
            "辛苦了 → Makasih kerja kerasnya. "
            "靠 又壞了 → Astaga, rusak lagi. "
            "先這樣 → Segitu dulu ya. "
            "叫他快點 → Suruh dia cepatan. "
            "砂輪要換了 → Batu gerinda harus diganti. "
            "公差超過了 → Toleransinya udah lewat. "
            "這6把再麻煩今晚入庫 → 6 bundel ini tolong masukin gudang malam ini. "
            "明早業務要抓資料 謝謝 → Besok pagi sales perlu ambil data, makasih. "
            "BF2拋光機維修中 → Mesin polishing BF2 sedang diperbaiki. "
            "44.45前天有跟妳說超產，業務回覆了嗎 → Diameter 44.45 kemarin sudah bilang over produksi, sales udah balas belum? "
            "噴漆後照訂單量拆包 → Setelah spray paint, bagi packing sesuai jumlah order. "
            "品保點錯製程，麻煩退回400-無主 → QC salah pilih proses, tolong kembalikan ke station 400 tanpa pemilik. "
            "帳已回400、料要回去那一個單位？ → Data sudah dikembalikan ke 400, materialnya mau ke unit mana? "
            "去削皮退火 感溫 → Ke proses peeling dan annealing, makasih. "
            "7F414020 請幫放至480轉用收回400，要改制去化，謝謝 → 7F414020 tolong pindahkan ke station 480, lalu kembalikan ke 400, mau ubah proses, makasih. "
            "業務說收～ 請包～ → Sales bilang terima, tolong di-packing. "
            "班長～ 7F656502A 這把溢量請再入無主～ 謝謝! → Kepala shift, 7F656502A bundel ini kelebihan, tolong masukkan ke tanpa pemilik, makasih! "
            "客需求支數7支、不收短 來料只有6支、其中一支短、剔除掉剩5支、能包嘛？ → Pelanggan minta 7 batang, gak terima pendek. Masuk cuma 6, 1 pendek dibuang sisa 5, bisa packing gak? "
            "因為櫃子在路上 9點到 這樣可能可以等一下入庫 → Karena kontainer sedang di jalan, sampai jam 9, mungkin bisa tunggu sebentar baru masuk gudang. "
            "DACAPO都入完了 → DACAPO semuanya sudah masuk gudang. "
            "班長～ 請用2700大的木箱裝，再麻煩幫我抓一下幾點會好，業務下午要出，謝謝 → Kepala shift, tolong pakai kotak kayu 2700, cek jam berapa selesai, sales sore mau kirim, makasih. "
            "那就是帳沒入到 → Berarti datanya belum masuk ke sistem. "
            "資料異常，凱銘在處理了 → Data ada masalah, 凱銘 sedang urus. "
            "研磨排程已更新，急單再麻煩安排洗料拋光 謝謝 → Jadwal grinding diupdate, order urgent tolong atur cuci material dan polishing, makasih. "
            "粗拋完已放行 → Rough polishing selesai, sudah di-release. "
            "麻煩先放這把 → Tolong release bundel ini dulu. "
            "放了 → Sudah di-release. "
            "先放這單 → Release order ini dulu. "
            "幫放一下 → Tolong bantu release. "
            "這批先不要放 → Lot ini jangan di-release dulu. "
            "料放旁邊 → Material taruh di samping. "
            "放地上 → Taruh di lantai. "
            "今日出貨差 DACAPO 7G63837在490 7G687108A在420 OL → Hari ini pengiriman kurang: DACAPO 7G63837 di 490, 7G687108A di 420 sedang produksi. "
            "METALINOX 差2噸等等K4會在出料 可以的在幫包裝 感謝 → METALINOX kurang 2 ton, nanti K4 keluarkan material, kalau bisa tolong packing, makasih. "
            "7G108519D 請幫收回400，有單去化 謝謝 → 7G108519D tolong kembalikan ke 400, ada order baru untuk serap material, makasih. "
            "洗給E7拋了 → Sudah dicuci dan dikasih ke E7 untuk polishing. "
            "包裝遇到常州眾山再注意這個料號，剛接單後續才會投料生產，此訂單不收短尺需將短尺分捆 → Kalau packing ketemu 常州眾山 perhatikan nomor material ini, baru terima order nanti baru produksi, order ini gak terima pendek harus pisah bundel. "
            "剛剛開會決議過年不停機，如果A班D班出勤人數不夠12人，想賺紅包可以代班 → Rapat keputusan Imlek tidak stop, shift A D kurang 12 orang, mau angpao bisa gantikan shift. "
            "人事有通知堆高機複訓課程，1/29 1700-2000三樓會議室。當天來上課就好，加班時數改天用忘卡補 → HRD info pelatihan forklift, 29/1 jam 17-20 ruang rapat lt.3. Datang ikut aja, jam lembur diinput lewat sistem lupa kartu di hari lain. "
            "處長走了 → Kepala divisi sudah pergi. "
            "有壓日期的急單再幫忙處理一下，很多未到站，拋光會一邊產出 → Order urgent deadline tolong diproses, banyak belum sampai, polishing produksi sambil jalan. "
            "噴漆罐一定要打洞才能丟棄在太空包，本週被查核兩次缺失 → Kaleng spray HARUS dilubangi baru buang ke jumbo bag, minggu ini kena audit 2 kali. "
            "本月入庫目標2950，異型棒不擋，其餘非本月不入庫 → Target gudang 2950, batang khusus bebas, sisanya bukan bulan ini jangan masuk. "
            "本月入庫目標量已達標，目前只入急單、異型棒跟二月以前的遞延單 → Target tercapai, sekarang hanya urgent, batang khusus, dan order ditunda sebelum Feb. "
            "今天沒點名，昨天來過了 → Hari ini gak ada inspeksi, kemarin sudah datang. "
            "應該是上週四D班，傍晚要注意一下小趙跟處長行蹤，免得凱銘被釘 → Harusnya shift D Kamis kemarin, sore perhatikan 小趙 dan kepala divisi, supaya 凱銘 gak kena tegur. "
            "自己稍微看一下設備的料源，有料就是要生產。月底我們不可能是停機的單位 → Cek material di mesin masing-masing, ada material ya produksi. Akhir bulan kita gak boleh mesin berhenti. "
            "之後有包到寶麗金屬注意一下，有一批訂單會備註客戶不要爐號標籤 → Nanti kalau ada packing untuk 寶麗金屬 perhatikan, ada order dicatat pelanggan tidak mau label heat number. "
            "非本月只有異型棒不管控，其他麻煩不要入了，昨天早班沒管控被檢討 → Bukan bulan ini cuma batang khusus bebas, sisanya jangan masuk, shift pagi kemarin gak kontrol kena tegur. "
            "非本月包裝不入庫 → Yang bukan order bulan ini jangan packing masuk gudang. "
            "急單再幫忙安排入庫 → Hanya order urgent yang tolong bantu atur masuk gudang. "
            "大成、SUNGEUN/佳東/麒譯/津展/DACAPO不擋非本月，各班在注意一下 → 大成, SUNGEUN/佳東/麒譯/津展/DACAPO order bukan bulan ini boleh masuk gudang, semua shift tolong perhatikan ya. "
            "H、S異型棒、大成、SUNGEUN、佳東……以上不擋非本月 → H, S, batang bentuk khusus, 大成, SUNGEUN, 佳東... yang di atas order bukan bulan ini boleh masuk gudang. "
            "開天車務必遵守規定目視吊掛物 → Operasi crane WAJIB lihat beban gantung sesuai aturan. "
            "護罩跟外勞宣導一下要蓋好 → Sosialisasi ke pekerja Indonesia pelindung mesin harus ditutup rapat. "
            "印勞打錯系統有提示 可是他們看不懂把他按掉了 → Pekerja Indonesia salah input, sistem ada peringatan tapi mereka gak ngerti jadi ditutup. "
            "拋光機interlock都不要拿東西擋著，上面會查 → Pengunci keamanan polishing jangan ditahan pakai benda, atasan akan periksa. "
            "來料自由端偏小 → Material masuk ujung bebasnya under size. "
            "殺光痕嚴重但表粗有過 → Bekas grinding parah tapi surface roughness lulus. "
            "表粗有過目視沒過 → Surface roughness lulus tapi visual tidak lulus. "
            "涉及軋輥印痕的批次，請協助開立重工研磨至尺寸下限 → Lot kena roll mark, tolong buat WO rework grinding sampai batas bawah ukuran. "
            "護罩要隨時關閉，卡料需關閉電源後再取料 → Pelindung mesin harus ditutup, material macet HARUS matikan listrik dulu baru ambil. "
            "嚴禁運轉中設備直接以手搬動棒材 → DILARANG pindahkan batang baja dengan tangan saat mesin jalan. "
            "矯直機前壓輪故障，卡死無法上昇，已請修護協助處理 → Roda tekan straightening rusak macet, sudah minta maintenance bantu. "
            "氣壓缸更換備品回裝完成，測試OK正常生產 → Silinder pneumatik ganti spare part selesai, tes OK produksi normal. "
            "來料盤元不佳退回線外修磨 → Wire rod masuk kualitas buruk, dikembalikan offline repair grinding. "
            "不可側磨已宣導多次，納入劣項 → Larangan side grinding sudah disosialisasi berkali-kali, dicatat pelanggaran. "
            "E5線速是否過慢，僅2.4～3.6m/min → Kecepatan lini E5 terlalu lambat, cuma 2.4-3.6 m/min? "
            "眼模刮傷整修一次，無法改善，更換眼模 → Die tergores, perbaiki sekali tidak membaik, ganti die. "
            "E11已抽完，要回精整，請放行過帳 → E11 selesai drawing, harus kembali ke finishing, tolong release dan input data. "
            "更換備品後已恢復生產 → Setelah ganti spare part sudah kembali produksi. "
            "報表要記得確實填寫，尤其是雷射校正部分 → Laporan produksi ingat diisi benar, terutama bagian kalibrasi laser. "
            "幫追料 → Tolong kejar materialnya. "
            "幫追帳 → Tolong kejar data administrasinya. "
            "已2900別入帳了噢 → Sudah 2900 jangan masukkan data lagi ya. "
            "【印尼→中文】"
            "Saya mau izin besok → 我明天要請假 "
            "Mesinnya rusak → 機台壞了 "
            "Materialnya udah habis → 料用完了 "
            "Kapan gajinya keluar? → 薪水什麼時候發？ "
            "Saya gak ngerti → 我聽不懂 "
            "Boleh pulang duluan? → 可以先下班嗎？ "
            "Lembur sampai jam berapa? → 加班到幾點？ "
            "Bos, ini udah selesai → 老闆，這個好了 "
            "Ukurannya gak pas → 尺寸不對 "
            "Stoknya masih ada? → 庫存還有嗎？ "
            "Tolong ajarin saya → 請教我一下"
            + extra_rule +
            " IMPORTANT: Preserve the original line breaks and blank lines exactly. If the source has a blank line between paragraphs, keep a blank line in the same position in the translation."
            " Only output the translation. No quotes, no explanation, no prefix."
        )

        if repair_mode and bad_result:
            msg = (
                "Original text (source language): " + protected + "\n\n"
                "Bad translation that leaked source-language words: " + bad_result + "\n\n"
                "Rewrite the bad translation into pure " + tgt_name +
                ". Preserve names and __MENTION__ placeholders exactly. Translate every remaining source-language word."
            )
        else:
            msg = "Translate from " + src_name + " to " + tgt_name + ": " + protected

        _model = pick_model(text)
        r = oai.chat.completions.create(
            model=_model,
            messages=[
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": msg}
            ],
            temperature=0.1 if strict_no_source_script or repair_mode else 0.2,
            max_tokens=2000,
        )
        track_tokens(r)
        result = r.choices[0].message.content.strip()
        result = restore_mentions(result, placeholders)
        # Fix known GPT translation mistakes and restore customer names
        if src == "zh":
            result = post_fix_translation(result)
            result = restore_customers(result, cust_placeholders)
        return result
    except Exception as e:
        logger.error("OpenAI error: %s", e)
        return None


def translate_google(text, src, tgt):
    try:
        protected, placeholders = protect_mentions(text)
        lang_map = {
            "zh": "zh-TW", "id": "id", "en": "en",
            "vi": "vi", "th": "th", "ja": "ja",
            "ko": "ko", "ms": "ms", "tl": "tl",
        }
        sl = lang_map.get(src, src)
        tl = lang_map.get(tgt, tgt)
        q = urllib.parse.quote(protected)
        url = "https://translate.googleapis.com/translate_a/single?client=gtx&sl=" + sl + "&tl=" + tl + "&dt=t&q=" + q
        req = urllib.request.Request(url)
        req.add_header("User-Agent", "Mozilla/5.0")
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            parts = []
            for item in data[0]:
                if item[0]:
                    parts.append(item[0])
            result = "".join(parts)
            result = restore_mentions(result, placeholders)
            # Fix known translation mistakes
            result = post_fix_translation(result)
            return result
    except Exception as e:
        logger.error("Google translate error: %s", e)
        return None


def cache_get(text, src, tgt):
    """Get translation from cache if exists and not expired."""
    key = (text.strip(), src, tgt)
    with _cache_lock:
        if key in translation_cache:
            result, ts = translation_cache[key]
            if time.time() - ts < CACHE_TTL:
                logger.info("Cache hit: %s -> %s", src, tgt)
                return result
            else:
                del translation_cache[key]
    return None


def cache_set(text, src, tgt, result):
    """Store translation in cache, evict oldest if full."""
    key = (text.strip(), src, tgt)
    with _cache_lock:
        if len(translation_cache) >= CACHE_MAX_SIZE:
            oldest_key = min(translation_cache, key=lambda k: translation_cache[k][1])
            del translation_cache[oldest_key]
        translation_cache[key] = (result, time.time())


def translate_with_retry(func, text, src, tgt, max_retries=2):
    """Call a translation function with retry on failure."""
    for attempt in range(max_retries + 1):
        result = func(text, src, tgt)
        if result:
            return result
        if attempt < max_retries:
            wait = 1 * (attempt + 1)
            logger.warning("Retry %d/%d after %ds for %s", attempt + 1, max_retries, wait, func.__name__)
            time.sleep(wait)
    return None


def translate(text, src, tgt):
    # Check cache first
    cached = cache_get(text, src, tgt)
    if cached:
        return cached

    result = translate_with_retry(translate_openai, text, src, tgt, max_retries=2)

    # If source-language leakage is detected, retry with strict mode.
    if result and not is_translation_valid(result, src, tgt):
        logger.warning("Source-language leakage detected in translation, retrying with stricter prompt")
        strict_result = translate_openai(text, src, tgt, strict_no_source_script=True)
        if strict_result and is_translation_valid(strict_result, src, tgt):
            result = strict_result
        else:
            repaired = translate_openai(
                text,
                src,
                tgt,
                strict_no_source_script=True,
                repair_mode=True,
                bad_result=(strict_result or result)
            )
            if repaired and is_translation_valid(repaired, src, tgt):
                result = repaired

    if result and is_translation_valid(result, src, tgt):
        cache_set(text, src, tgt, result)
        return result

    # Fallback to Google with retry.
    result = translate_with_retry(translate_google, text, src, tgt, max_retries=1)
    if result and is_translation_valid(result, src, tgt):
        cache_set(text, src, tgt, result)
        return result

    # Last chance: ask OpenAI to repair the latest output instead of returning a leaked translation.
    if result:
        repaired = translate_openai(
            text,
            src,
            tgt,
            strict_no_source_script=True,
            repair_mode=True,
            bad_result=result
        )
        if repaired and is_translation_valid(repaired, src, tgt):
            cache_set(text, src, tgt, repaired)
            return repaired

    return None


def detect_work_order(ocr_text):
    """Detect if OCR text is from a factory work order (製造指示書).
    Returns customer name if detected, None otherwise."""
    if not ocr_text:
        return None
    wo_keywords = ["冷精棒製造指示書", "製造指示書", "訂單編號", "客戶名稱", "成品尺寸",
                   "FINAL流程", "FINAL", "MIC_NO", "ID_NO", "HRITABPDIL", "退火代碼",
                   "冷精棒", "收貨人", "短尺", "品保", "特殊", "削皮", "訂單資訊",
                   "成品尺寸MIN", "成品尺寸MAX", "製造指示"]
    keyword_count = sum(1 for kw in wo_keywords if kw in ocr_text)
    logger.info("Work order detection: %d keywords matched in OCR text (%d chars)", keyword_count, len(ocr_text))
    if keyword_count < 2:
        return None
    # Try multiple patterns to extract customer name
    patterns = [
        r'客戶名稱[:\s：]*([^\s\n|,，]+)',
        r'客戶[:\s：]*([^\s\n|,，]+)',
        r'客[户戶]名[称稱][:\s：]*([^\s\n|,，]+)',
    ]
    for pat in patterns:
        m = re.search(pat, ocr_text)
        if m:
            customer = m.group(1).strip()
            if customer and len(customer) >= 2:
                logger.info("Work order customer detected: %s", customer)
                return customer
    # Fallback: try to match any known customer name in the text
    for name in CUSTOMER_NAMES:
        if len(name) >= 2 and name in ocr_text:
            logger.info("Work order customer matched from list: %s", name)
            return name
    logger.info("Work order detected but no customer name found")
    return None


def format_length_zh(code):
    """Convert length code to Chinese."""
    if code == "<=3200":
        return "未滿3200"
    elif code == ">4200":
        return "超過4200"
    elif code == ">3200<=4200":
        return "3200～4200"
    elif code == ">4000":
        return "超過4000"
    else:
        c = code.replace("<=", "未滿").replace(">=", "超過").replace(">", "超過").replace("<", "未滿")
        return c


def format_storage_for_work_order(customer_name):
    """Format storage lookup for work order image detection."""
    entries = STORAGE_LOOKUP.get(customer_name)
    if not entries:
        for key in STORAGE_LOOKUP:
            if key.lower() == customer_name.lower() or customer_name in key or key in customer_name:
                entries = STORAGE_LOOKUP[key]
                customer_name = key
                break
    if not entries:
        return None
    lines = []
    lines.append("\U0001f4cb \u5de5\u55ae\u5075\u6e2c")
    lines.append("\u5ba2\u6236\uff1a" + customer_name)
    lines.append("")
    lines.append("\U0001f4e6 \u5132\u5340\u67e5\u8a62")
    lines.append("=" * 18)
    for length, area in entries:
        zh = format_length_zh(length)
        lines.append(zh + " \u2192 " + area)
    lines.append("=" * 18)
    return "\n".join(lines)


    """Use OpenAI Vision to extract text from image."""
    if not oai:
        return None
    try:
        r = oai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are an OCR assistant. Extract ALL text visible in the image. "
                        "Output ONLY the extracted text, preserving line breaks. "
                        "If there is no text in the image, output exactly: NO_TEXT_FOUND"
                    )
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": "data:image/jpeg;base64," + image_base64,
                                "detail": "high"
                            }
                        },
                        {
                            "type": "text",
                            "text": "Extract all text from this image."
                        }
                    ]
                }
            ],
            temperature=0.1,
            max_tokens=2000,
        )
        track_tokens(r)
        result = r.choices[0].message.content.strip()
        if result == "NO_TEXT_FOUND" or not result:
            return None
        return result
    except Exception as e:
        logger.error("OpenAI Vision OCR error: %s", e)
        return None


def ocr_image_openai(image_base64):
    """Use OpenAI Vision to extract text from image."""
    if not oai:
        return None
    try:
        r = oai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are an OCR assistant. Extract ALL text visible in the image. "
                        "Output ONLY the extracted text, preserving line breaks. "
                        "If there is no text in the image, output exactly: NO_TEXT_FOUND"
                    )
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": "data:image/jpeg;base64," + image_base64,
                                "detail": "high"
                            }
                        },
                        {
                            "type": "text",
                            "text": "Extract all text from this image."
                        }
                    ]
                }
            ],
            temperature=0.1,
            max_tokens=2000,
        )
        track_tokens(r)
        result = r.choices[0].message.content.strip()
        if result == "NO_TEXT_FOUND" or not result:
            return None
        return result
    except Exception as e:
        logger.error("OpenAI Vision OCR error: %s", e)
        return None


def ocr_and_translate_image(image_base64, tgt_lang):
    """OCR + translate image text in one API call, preserving layout."""
    if not oai:
        return None, None
    tgt_name = LANG_NAMES.get(tgt_lang, tgt_lang)
    tgt_flag = LANG_FLAGS.get(tgt_lang, "")
    try:
        r = oai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are an OCR + translation assistant for a factory work group chat.\n"
                        "Task: Extract ALL text from the image, then translate each section.\n\n"
                        "OUTPUT FORMAT:\n"
                        "For each distinct section/paragraph in the image, output:\n"
                        "original text...\n"
                        + tgt_flag + " translated text...\n"
                        "(blank line before next section)\n\n"
                        "EXAMPLE:\n"
                        "1.研磨來料前需紀錄來料三點式尺寸\n"
                        + tgt_flag + " 1.Sebelum grinding material masuk, catat dimensi 3 titik\n\n"
                        "2.拋光棒需清洗\n"
                        + tgt_flag + " 2.Batang polishing harus dicuci\n\n"
                        "RULES:\n"
                        "1. Keep the SAME structure, numbering, and line breaks as the original.\n"
                        "2. Each section: original text first, then translation with " + tgt_flag + " flag. Do NOT add section titles or brackets.\n"
                        "3. If there are numbered items (1. 2. 3.), keep the same numbering.\n"
                        "4. Do NOT repeat the original text. Show it only ONCE then show the translation.\n"
                        "5. Translate naturally, casual daily language for factory workers.\n"
                        "6. Target Traditional Chinese = Taiwan style.\n"
                        "7. NEVER translate or romanize person names. Keep Chinese names in original Chinese characters (e.g. 陳弘林 stays as 陳弘林, NOT Chen Honglin). Do NOT convert to pinyin.\n"
                        "7b. NEVER translate customer/company names. Keep them EXACTLY as-is: "
                        "賽利金屬, 寶麗金屬, 田華榕, 佳東, 蘋果, 常州眾山, 大順, 大成, 巨昌, 北澤, 鴻運, 畯圓, 名威, 右勝, "
                        "貝克休斯, 皇銘, 台芝, 百堅, 津展, 曜麟, 廉錩, 盛昌遠, 永吉, 光輝, "
                        "DACAPO, CASTLE, LOTUS, METALINOX, KANGRUI, SUNGEUN, STEELINC, GLH, SHINKO, WING KEUNG, "
                        "BOLLINGHAUS, COGNE, TCI, PLUTUS, SAMWON, DK METAL, KJ. "
                        "If you see ANY company name in the image, keep it unchanged. Do NOT translate 金屬=metal, 鋼鐵=steel etc. when part of a company name.\n"
                        "8. If no text found, output exactly: NO_TEXT_FOUND\n"
                        "9. TABLES/SPREADSHEETS: If the image is a table or spreadsheet, output it as a COMPACT table. "
                        "Only translate column headers and labels. Keep person names as-is in original characters. "
                        "Keep numbers as-is. Use a simple format like:\n"
                        "姓名/Nama | 3/17止/Hingga 3/17\n"
                        "陳弘林 | -600\n"
                        "蔡佳佳 | 200\n"
                        "Do NOT output each cell as a separate translated section. Keep it compact.\n"
                        "10. Factory vocabulary: "
                        "交辦事項=hal yang harus dikerjakan, "
                        "研磨=grinding, 無心研磨=centerless grinding, 拋光=polishing, 來料=material masuk, "
                        "量測=mengukur, 尺寸=diameter/dimensi, 三點式=3 titik, "
                        "雷射=laser, 設備=peralatan, 故障=rusak, "
                        "紀錄=catat, 拋光棒=batang polishing, "
                        "清洗=cuci, 輕調輕放=handle dengan hati-hati, "
                        "環狀擦傷=goresan melingkar, "
                        "重工=rework, 料回削皮=material kembali kupas/peeling, "
                        "補上=lengkapi, C行套環=C-ring, "
                        "廠內=di dalam pabrik, 禁止=dilarang, 餵狗=kasih makan anjing, "
                        "宣導=sosialisasi, "
                        "包裝站=stasiun packing, 啟動=mulai, "
                        "PMI全檢=inspeksi penuh PMI, 抽查機制=sistem sampling, "
                        "每捆=setiap bundel, 鋼種=jenis baja, "
                        "棒材=batang baja, 混料=tercampur material, "
                        "出貨=pengiriman, 依情節=sesuai tingkat pelanggaran, "
                        "增加績效=tambah penilaian kinerja, "
                        "確實=pastikan, 防止=mencegah, "
                        "精整=finishing, AP=mesin finishing, 矯直=straightening, 壓光=press polish, "
                        "退火=annealing, 光輝退火=bright annealing, 酸洗=pickling, 削皮=peeling, 冷抽=cold drawing, "
                        "熱軋=hot rolling, 煉鋼=steelmaking/peleburan baja, 碳廠=pabrik karbon, "
                        "職安署=Dinas K3(inspeksi keselamatan kerja), 查核=audit/inspeksi, "
                        "品保=QC, 儲運=gudang&logistik, 生計=production planning, 業務=sales, 營業=sales, 人事=HRD, "
                        "處長=kepala divisi, 點名=inspeksi pengawas(NOT roll call), "
                        "加班=lembur, 排班=jadwal shift, 早班=shift pagi, 夜班=shift malam, "
                        "砂輪=batu gerinda, 天車=overhead crane, 堆高機=forklift, "
                        "油桶=drum oli, 太空包=jumbo bag, 噴漆罐=kaleng spray, "
                        "入庫=masuk gudang, 退庫=kembalikan ke gudang, 出貨差=kekurangan pengiriman, "
                        "掛單/工單=work order, 重掛單=pasang ulang work order, 取樣=ambil sampel, "
                        "二道門=pintu kedua(gate 2), 捐血=donor darah, "
                        "爐號=heat number(NEVER nomor panas), 過帳=input data ke sistem, 放行=release data\n"
                        "11. Only output the result. No extra explanation."
                    )
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": "data:image/jpeg;base64," + image_base64,
                                "detail": "high"
                            }
                        },
                        {
                            "type": "text",
                            "text": "Extract and translate all text from this image to " + tgt_name + ". Keep the same layout structure."
                        }
                    ]
                }
            ],
            temperature=0.2,
            max_tokens=3000,
        )
        track_tokens(r)
        result = r.choices[0].message.content.strip()
        if result == "NO_TEXT_FOUND" or not result:
            return None, None
        return result, None
    except Exception as e:
        logger.error("OpenAI Vision OCR+translate error: %s", e)
        return None, str(e)



def download_line_image(message_id):
    """Download image from LINE and return (base64_string, raw_bytes)."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_message_content(message_id)
            img_base64 = base64.b64encode(content).decode("utf-8")
            return img_base64, content
    except Exception as e:
        logger.error("LINE image download error: %s", e)
        return None, None


def download_line_audio(message_id):
    """Download audio from LINE and return bytes."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_message_content(message_id)
            return content
    except Exception as e:
        logger.error("LINE audio download error: %s", e)
        return None


def transcribe_audio_openai(audio_bytes):
    """Use OpenAI Whisper to transcribe audio to text."""
    if not oai:
        return None
    try:
        with tempfile.NamedTemporaryFile(suffix=".m4a", delete=True) as tmp:
            tmp.write(audio_bytes)
            tmp.flush()
            tmp.seek(0)
            r = oai.audio.transcriptions.create(
                model="whisper-1",
                file=tmp,
            )
            text = r.text.strip() if r.text else None
            return text
    except Exception as e:
        logger.error("OpenAI Whisper error: %s", e)
        return None


def make_notice(content, target="id"):
    tgt_text = translate(content, "zh", target)
    if not tgt_text:
        tgt_text = "(translation failed)"
    sep = "=" * 18
    lines = []
    lines.append("\U0001f4e2 \u516c\u544a / Pengumuman")
    lines.append(sep)
    lines.append("\U0001f1f9\U0001f1fc " + content)
    lines.append(LANG_FLAGS.get(target, "") + " " + tgt_text)
    lines.append(sep)
    return "\n".join(lines)


def make_notice_from_other(content, src, target="zh"):
    zh_text = translate(content, src, "zh")
    if not zh_text:
        zh_text = "(translation failed)"
    sep = "=" * 18
    lines = []
    lines.append("\U0001f4e2 \u516c\u544a / Pengumuman")
    lines.append(sep)
    lines.append("\U0001f1f9\U0001f1fc " + zh_text)
    lines.append(LANG_FLAGS.get(src, "") + " " + content)
    lines.append(sep)
    return "\n".join(lines)


def get_help_text(group_id):
    sep = "=" * 18
    lines = []
    lines.append("\U0001f310 翻譯機器人")
    lines.append(sep)
    lines.append("【開關】")
    lines.append("/on ・ /off 翻譯")
    lines.append("/img on・off 圖片")
    lines.append("/voice on・off 語音")
    lines.append("/wo on・off 拍工單查儲區")
    lines.append("【個人】")
    lines.append("/skip 不翻譯我")
    lines.append("/unskip 恢復翻譯")
    lines.append("【管理】")
    lines.append("/skipadd 名字 加入白名單")
    lines.append("/skipdel 名字 移出白名單")
    lines.append("/skiplist 查看白名單")
    lines.append("【功能】")
    lines.append("/notice 內容 雙語公告")
    lines.append("/qry 客戶 查儲區")
    lines.append("/pkg 代碼 查包裝碼")
    lines.append("/pw1 班長密碼")
    lines.append("/pw2 儲運密碼")
    lines.append("/scrap 廢料顏色")
    lines.append("/status 查看狀態")
    lines.append("\U0001f4f7 拍工單→自動查儲區")
    lines.append(sep)
    lines.append("中文 ⇄ 🇮🇩 印尼文 即時互譯")
    return "\n".join(lines)


def handle_lang_command(text, group_id):
    return "ℹ️ 本機器人僅支援 中文 ⇄ 🇮🇩 印尼文 互譯"


def handle_qry_command(text):
    """Handle /qry <customer_name> command to lookup storage area."""
    parts = text.strip().split(None, 1)
    if len(parts) < 2:
        return "\u26a0\ufe0f \u8acb\u8f38\u5165\u5ba2\u6236\u540d\u7a31\n\u7bc4\u4f8b: /qry ABE\n\u7bc4\u4f8b: /qry \u4f73\u6771"
    query = parts[1].strip()
    # Try exact match first
    entries = STORAGE_LOOKUP.get(query)
    # Try case-insensitive match
    if not entries:
        for key in STORAGE_LOOKUP:
            if key.lower() == query.lower():
                entries = STORAGE_LOOKUP[key]
                query = key
                break
    # Try partial match
    if not entries:
        matches = [k for k in STORAGE_LOOKUP if query.lower() in k.lower() or query in k]
        if len(matches) == 1:
            query = matches[0]
            entries = STORAGE_LOOKUP[query]
        elif len(matches) > 1:
            result = "\U0001f50d \u627e\u5230\u591a\u7b46\u7b26\u5408:\n"
            for m in matches[:10]:
                result += "  \u2022 " + m + "\n"
            if len(matches) > 10:
                result += "  ...(\u5171" + str(len(matches)) + "\u7b46)\n"
            result += "\n\u8acb\u8f38\u5165\u5b8c\u6574\u5ba2\u6236\u540d\u7a31"
            return result
    if not entries:
        return "\u274c \u627e\u4e0d\u5230\u5ba2\u6236: " + query + "\n\u8acb\u78ba\u8a8d\u540d\u7a31\u662f\u5426\u6b63\u78ba"
    # Build response
    lines = []
    lines.append("\U0001f4e6 " + query + " \u5132\u5340\u67e5\u8a62")
    lines.append("=" * 18)
    for length, area in entries:
        zh = format_length_zh(length)
        lines.append(zh + " \u2192 " + area)
    lines.append("=" * 18)
    return "\n".join(lines)


def handle_pkg_command(text):
    """Handle /pkg <code> command to lookup packaging info."""
    parts = text.strip().split(None, 1)
    if len(parts) < 2:
        return (
            "⚠️ 請輸入包裝碼 / Masukkan kode kemasan\n"
            "範例 / Contoh: /pkg U\n"
            "範例 / Contoh: /pkg G"
        )
    query = parts[1].strip()
    query_upper = query.upper()
    if not PACKAGING_LOOKUP:
        return "⚠️ 包裝碼資料尚未上傳\nData kode kemasan belum diupload"
    # Try exact match (case-insensitive)
    entry = PACKAGING_LOOKUP.get(query) or PACKAGING_LOOKUP.get(query_upper)
    matched_key = query if PACKAGING_LOOKUP.get(query) else query_upper
    if not entry:
        for k in PACKAGING_LOOKUP:
            if k.upper() == query_upper:
                entry = PACKAGING_LOOKUP[k]
                matched_key = k
                break
    # Try partial match
    if not entry:
        matches = [k for k in PACKAGING_LOOKUP if query_upper in k.upper()]
        if len(matches) == 1:
            matched_key = matches[0]
            entry = PACKAGING_LOOKUP[matched_key]
        elif len(matches) > 1:
            result = "🔍 找到多筆符合:\n"
            for m in matches[:15]:
                result += "  • " + m + "\n"
            return result
    if not entry:
        return "❌ 找不到包裝碼: " + query
    # Build response - show specific fields in order
    # Match Excel headers by keyword → display label
    PKG_DISPLAY = [
        ("簡稱",       ["簡稱"]),
        ("詳細包裝方式", ["詳細包裝", "包裝方式說明", "包裝方式"]),
        ("內包裝",     ["內包裝"]),
        ("外包裝",     ["外包裝"]),
        ("固定繩",     ["固定繩", "固定"]),
    ]
    lines = []
    lines.append("📦 包裝碼: " + matched_key)
    lines.append("=" * 20)
    if isinstance(entry, dict):
        for display_label, keywords in PKG_DISPLAY:
            # Find matching field in entry
            for field_name, field_val in entry.items():
                if field_val and any(kw in field_name for kw in keywords):
                    lines.append(display_label + ": " + str(field_val))
                    break
    elif isinstance(entry, str):
        lines.append(entry)
    lines.append("=" * 20)
    return "\n".join(lines)


def get_display_name(group_id, user_id):
    """Get user display name from cache or LINE API. Also caches user language."""
    if group_id in group_user_names and user_id in group_user_names[group_id]:
        return group_user_names[group_id][user_id]
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            profile = api.get_group_member_profile(group_id, user_id)
            name = profile.display_name
            if name:
                if group_id not in group_user_names:
                    group_user_names[group_id] = {}
                group_user_names[group_id][user_id] = name
            # Cache user language and picture from LINE profile
            lang = getattr(profile, 'language', None)
            if lang and user_id not in user_languages:
                user_languages[user_id] = lang
                logger.info("User %s language: %s", name or user_id, lang)
            pic = getattr(profile, 'picture_url', None)
            if pic:
                user_pictures[user_id] = pic
            return name
    except Exception as e:
        logger.warning("Failed to get display name for %s: %s", user_id, e)
    return None


def record_user_name(group_id, user_id):
    """Record user display name in background (best effort)."""
    if not group_id or not user_id:
        return
    if group_id in group_user_names and user_id in group_user_names[group_id]:
        return
    get_display_name(group_id, user_id)


def find_user_by_name(group_id, name_query):
    """Find user_id by display name (partial match). Returns list of (user_id, display_name)."""
    if group_id not in group_user_names:
        return []
    matches = []
    query_lower = name_query.lower().strip()
    for uid, dname in group_user_names[group_id].items():
        if query_lower == dname.lower() or query_lower in dname.lower() or dname.lower() in query_lower:
            matches.append((uid, dname))
    return matches


def handle_command(text, group_id, user_id=None):
    bot_stats["commands"] += 1
    cmd = text.strip().lower()
    if cmd == "/help":
        return get_help_text(group_id)
    elif cmd == "/on":
        group_settings[group_id] = True
        save_settings()
        return "\u2705 \u7ffb\u8b6f\u5df2\u958b\u555f / Penerjemah aktif"
    elif cmd == "/off":
        group_settings[group_id] = False
        save_settings()
        return "\u274c \u7ffb\u8b6f\u5df2\u95dc\u9589 / Penerjemah nonaktif"
    elif cmd == "/img on":
        group_img_settings[group_id] = True
        save_settings()
        return "\u2705 \u5716\u7247\u7ffb\u8b6f\u5df2\u958b\u555f / Terjemahan gambar aktif"
    elif cmd == "/img off":
        group_img_settings[group_id] = False
        save_settings()
        return "\u274c \u5716\u7247\u7ffb\u8b6f\u5df2\u95dc\u9589 / Terjemahan gambar nonaktif"
    elif cmd == "/voice on":
        group_audio_settings[group_id] = True
        save_settings()
        return "\u2705 \u8a9e\u97f3\u7ffb\u8b6f\u5df2\u958b\u555f / Terjemahan suara aktif"
    elif cmd == "/voice off":
        group_audio_settings[group_id] = False
        save_settings()
        return "\u274c \u8a9e\u97f3\u7ffb\u8b6f\u5df2\u95dc\u9589 / Terjemahan suara nonaktif"
    elif cmd == "/wo on":
        group_wo_settings[group_id] = True
        save_settings()
        return "\u2705 \u62cd\u5de5\u55ae\u67e5\u5132\u5340\u5df2\u958b\u555f"
    elif cmd == "/wo off":
        group_wo_settings[group_id] = False
        save_settings()
        return "\u274c \u62cd\u5de5\u55ae\u67e5\u5132\u5340\u5df2\u95dc\u9589"
    elif cmd == "/skip":
        if not user_id:
            return "\u26a0\ufe0f \u7121\u6cd5\u8b58\u5225\u4f60\u7684\u8eab\u4efd"
        if group_id not in group_skip_users:
            group_skip_users[group_id] = set()
        group_skip_users[group_id].add(user_id)
        save_settings()
        return "\u2705 \u5df2\u5c07\u4f60\u52a0\u5165\u767d\u540d\u55ae\uff0c\u4f60\u7684\u8a0a\u606f\u4e0d\u6703\u88ab\u7ffb\u8b6f\nAnda ditambahkan ke daftar skip"
    elif cmd == "/unskip":
        if not user_id:
            return "\u26a0\ufe0f \u7121\u6cd5\u8b58\u5225\u4f60\u7684\u8eab\u4efd"
        if group_id in group_skip_users:
            group_skip_users[group_id].discard(user_id)
        save_settings()
        return "\u2705 \u5df2\u5c07\u4f60\u79fb\u51fa\u767d\u540d\u55ae\uff0c\u4f60\u7684\u8a0a\u606f\u6703\u88ab\u7ffb\u8b6f\nAnda dihapus dari daftar skip"
    elif text.strip().lower().startswith("/skipadd"):
        name_query = text.strip()[8:].strip()
        if not name_query:
            return "\u26a0\ufe0f \u8acb\u8f38\u5165\u540d\u5b57\n\u7bc4\u4f8b: /skipadd \u79cb\u60c5"
        matches = find_user_by_name(group_id, name_query)
        if len(matches) == 0:
            return "\u274c \u627e\u4e0d\u5230\u300c" + name_query + "\u300d\n\u8a72\u7528\u6236\u9700\u5148\u5728\u7fa4\u7d44\u767c\u904e\u8a0a\u606f\u624d\u80fd\u88ab\u8a8d\u5230"
        if len(matches) > 1:
            names = "\n".join(["  \u2022 " + m[1] for m in matches])
            return "\U0001f50d \u627e\u5230\u591a\u4eba\u7b26\u5408\uff1a\n" + names + "\n\u8acb\u8f38\u5165\u66f4\u5b8c\u6574\u7684\u540d\u5b57"
        uid, dname = matches[0]
        if group_id not in group_skip_users:
            group_skip_users[group_id] = set()
        group_skip_users[group_id].add(uid)
        save_settings()
        return "\u2705 \u5df2\u5c07\u300c" + dname + "\u300d\u52a0\u5165\u767d\u540d\u55ae\uff0c\u8a0a\u606f\u4e0d\u6703\u88ab\u7ffb\u8b6f"
    elif text.strip().lower().startswith("/skipdel"):
        name_query = text.strip()[8:].strip()
        if not name_query:
            return "\u26a0\ufe0f \u8acb\u8f38\u5165\u540d\u5b57\n\u7bc4\u4f8b: /skipdel \u79cb\u60c5"
        matches = find_user_by_name(group_id, name_query)
        if len(matches) == 0:
            return "\u274c \u627e\u4e0d\u5230\u300c" + name_query + "\u300d"
        if len(matches) > 1:
            names = "\n".join(["  \u2022 " + m[1] for m in matches])
            return "\U0001f50d \u627e\u5230\u591a\u4eba\u7b26\u5408\uff1a\n" + names + "\n\u8acb\u8f38\u5165\u66f4\u5b8c\u6574\u7684\u540d\u5b57"
        uid, dname = matches[0]
        if group_id in group_skip_users:
            group_skip_users[group_id].discard(uid)
        save_settings()
        return "\u2705 \u5df2\u5c07\u300c" + dname + "\u300d\u79fb\u51fa\u767d\u540d\u55ae\uff0c\u8a0a\u606f\u6703\u88ab\u7ffb\u8b6f"
    elif cmd == "/skiplist":
        skipped = group_skip_users.get(group_id, set())
        if not skipped:
            return "\u76ee\u524d\u767d\u540d\u55ae\u662f\u7a7a\u7684 / Daftar skip kosong"
        names_cache = group_user_names.get(group_id, {})
        lines = ["\u23ed\ufe0f \u767d\u540d\u55ae / Daftar skip:"]
        for uid in skipped:
            dname = names_cache.get(uid)
            if dname:
                lines.append("  \u2022 " + dname)
            else:
                lines.append("  \u2022 (\u672a\u77e5\u7528\u6236)")
        return "\n".join(lines)
    elif cmd == "/status":
        is_on = group_settings.get(group_id, True)
        if is_on:
            img_on = group_img_settings.get(group_id, True)
            img_status = "\u2705 開啟" if img_on else "\u274c 關閉"
            audio_on = group_audio_settings.get(group_id, True)
            audio_status = "\u2705 開啟" if audio_on else "\u274c 關閉"
            wo_on = group_wo_settings.get(group_id, True)
            wo_status = "\u2705 開啟" if wo_on else "\u274c 關閉"
            return "\u2705 翻譯：開啟中 / Aktif\n中文 ⇄ 🇮🇩 印尼文\n\U0001f5bc\ufe0f 圖片翻譯：" + img_status + "\n\U0001f3a4 語音翻譯：" + audio_status + "\n\U0001f4cb 拍工單查儲區：" + wo_status
        else:
            return "\u274c 翻譯：已關閉 / Nonaktif"
    elif cmd.startswith("/lang"):
        return handle_lang_command(text, group_id)
    elif text.strip().startswith("/notice ") or text.strip().startswith("/notice\u3000"):
        if not is_cmd_enabled(group_id, "notice"):
            return None
        content = text.strip()[8:].strip()
        if not content:
            return "\u26a0\ufe0f \u8acb\u8f38\u5165\u516c\u544a\u5167\u5bb9\n\u4f8b\u5982 / Contoh: /notice \u660e\u5929\u653e\u5047\u4e00\u5929"
        tgt = group_target_lang.get(group_id, "id")
        if has_chinese(content):
            return make_notice(content, tgt)
        else:
            src = detect_language(content)
            if src and src != "zh":
                return make_notice_from_other(content, src)
            return make_notice(content, tgt)
    elif text.strip().lower().startswith("/qry"):
        if not is_cmd_enabled(group_id, "qry"):
            return None
        return handle_qry_command(text)
    elif cmd == "/pw1":
        if not is_cmd_enabled(group_id, "pw1"):
            return None
        return "🔑 班長密碼 / PW Shift Leader\n" + "=" * 18 + "\n" + pw1_text + "\n" + "=" * 18
    elif cmd == "/pw2":
        if not is_cmd_enabled(group_id, "pw2"):
            return None
        return "🏭 儲運密碼 / PW Gudang\n" + "=" * 18 + "\n" + pw2_text + "\n" + "=" * 18
    elif cmd == "/scrap":
        if not is_cmd_enabled(group_id, "scrap"):
            return None
        return scrap_text
    elif text.strip().lower().startswith("/pkg"):
        if not is_cmd_enabled(group_id, "pkg"):
            return None
        return handle_pkg_command(text)
    return None


@app.route("/callback", methods=["POST"])
def callback():
    sig = request.headers.get("X-Line-Signature", "")
    body = request.get_data(as_text=True)
    try:
        handler.handle(body, sig)
    except InvalidSignatureError:
        abort(400)
    return "OK"


@handler.add(MessageEvent, message=TextMessageContent)
def handle_message(event):
    text = event.message.text.strip()
    if len(text) < 2:
        return

    source = event.source
    is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
    group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
    user_id = getattr(source, 'user_id', None)

    # --- DM (private message) mode ---
    if is_dm and user_id:
        # Record DM user for admin panel
        if user_id not in dm_known_users:
            try:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    profile = api.get_profile(user_id)
                    dm_known_users[user_id] = profile.display_name or user_id
            except Exception:
                dm_known_users[user_id] = user_id

        # DM commands
        cmd = text.strip().lower()
        if cmd == "/help":
            sep = "=" * 18
            lines = []
            lines.append("\U0001f310 私訊翻譯模式")
            lines.append(sep)
            lines.append("傳訊息給我就會翻譯！")
            lines.append("中文 → 🇮🇩 印尼文")
            lines.append("印尼文 → 🇹🇼 中文")
            lines.append("")
            lines.append("/qry 客戶 查儲區")
            lines.append("/pkg 代碼 查包裝碼")
            lines.append("/pw1 班長密碼")
            lines.append("/pw2 儲運密碼")
            lines.append("/scrap 廢料顏色")
            lines.append("\U0001f4f7 拍工單→自動查儲區")
            lines.append(sep)
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text="\n".join(lines))]
                ))
            return
        if cmd.startswith("/to"):
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text="ℹ️ 本機器人僅支援 中文 ⇄ 🇮🇩 印尼文 互譯")]
                ))
            return
        # DM: handle /qry command
        if text.strip().lower().startswith("/qry"):
            qry_result = handle_qry_command(text)
            if qry_result:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=qry_result)]
                    ))
            return
        # DM: handle /pw1, /pw2, /scrap, /pkg commands
        dm_cmd_result = None
        if cmd == "/pw1":
            dm_cmd_result = "🔑 班長密碼 / PW Shift Leader\n" + "=" * 18 + "\n" + pw1_text + "\n" + "=" * 18
        elif cmd == "/pw2":
            dm_cmd_result = "🏭 儲運密碼 / PW Gudang\n" + "=" * 18 + "\n" + pw2_text + "\n" + "=" * 18
        elif cmd == "/scrap":
            dm_cmd_result = scrap_text
        elif text.strip().lower().startswith("/pkg"):
            dm_cmd_result = handle_pkg_command(text)
        if dm_cmd_result:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=dm_cmd_result)]
                ))
            return
        # DM: skip other / commands
        if text.startswith("/"):
            return

        # DM master toggle check
        if not dm_master_enabled and user_id not in dm_whitelist:
            return

        # DM translation: strip mentions, detect language, translate
        text_clean = strip_mentions_for_detect(text).strip()
        if not text_clean or len(text_clean) < 2:
            return

        lang = detect_language(text_clean)
        tgt = dm_target_lang.get(user_id, "id")
        if lang is None:
            return
        if lang == tgt:
            return

        # Set translation tone for DM (use global default)
        _tl.tone = translation_tone
        _tl.tone_custom = translation_tone_custom

        _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
        result = translate(text_clean, lang, tgt)
        track_group_usage("__dm__", _bp, _bc)
        if not result:
            return
        reply = LANG_FLAGS.get(tgt, "") + " " + result


        bot_stats["text_translations"] += 1
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.reply_message(ReplyMessageRequest(
                reply_token=event.reply_token,
                messages=[TextMessage(text=reply)]
            ))
        return

    # --- Group mode (original logic) ---
    # Record user display name for /skipadd lookup
    if group_id and user_id:
        record_user_name(group_id, user_id)
    # Track group for admin panel
    if group_id and not is_dm and group_id not in group_tracking:
        gname = ""
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                summary = api.get_group_summary(group_id)
                gname = summary.group_name or ""
        except Exception:
            pass
        group_tracking[group_id] = {"name": gname, "joined_at": time.time()}
        save_settings()

    if text.startswith("/"):
        # Set tone before commands that may translate (e.g. /notice)
        _tone, _tone_custom = get_group_tone(group_id)
        _tl.tone = _tone
        _tl.tone_custom = _tone_custom
        cmd_result = handle_command(text, group_id, user_id)
        if cmd_result:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.reply_message(ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=cmd_result)]
                ))
        return

    is_on = group_settings.get(group_id, True)
    if not is_on:
        return

    # Check skip list
    sender_id = getattr(source, 'user_id', None)
    if sender_id and sender_id in group_skip_users.get(group_id, set()):
        return

    if text.startswith("!"):
        return

    # Extract LINE's actual @mention data (blue text = 100% accurate)
    line_mentions = extract_line_mentions(text, event.message)

    # Cache this message for future quote references
    msg_id = getattr(event.message, 'id', None)
    if msg_id:
        message_cache[msg_id] = {"text": text, "ts": time.time()}
        # Trim cache
        if len(message_cache) > MESSAGE_CACHE_MAX:
            oldest = sorted(message_cache.items(), key=lambda x: x[1]["ts"])[:50]
            for k, _ in oldest:
                message_cache.pop(k, None)

    # Check if this is a reply to another message (quoted message)
    quoted_text = None
    quoted_id = getattr(event.message, 'quoted_message_id', None)
    if quoted_id and quoted_id in message_cache:
        quoted_text = message_cache[quoted_id].get("text", "")

    # Strip @mentions for language detection only
    text_for_detect = strip_mentions_for_detect(text, line_mentions).strip()
    if not text_for_detect or len(text_for_detect) < 2:
        return

    lang = detect_language(text_for_detect)
    if lang is None:
        return

    tgt = group_target_lang.get(group_id, "id")

    # Show typing indicator while translating
    show_loading(group_id)
    if get_group_feature(group_id, 'mark_read'):
        mark_as_read(group_id)

    # Protect LINE mentions before translation
    text_to_translate = text
    mention_placeholders = {}
    if line_mentions:
        text_to_translate, mention_placeholders = protect_mentions(text, line_mentions)

    reply = None
    _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
    # Set translation tone for this group
    _tone, _tone_custom = get_group_tone(group_id)
    _tl.tone = _tone
    _tl.tone_custom = _tone_custom
    if lang == "zh":
        result = translate(text_to_translate, "zh", tgt)
        if result and mention_placeholders:
            result = restore_mentions(result, mention_placeholders)
        if result:
            reply = LANG_FLAGS.get(tgt, "") + " " + result
    else:
        result = translate(text_to_translate, lang, "zh")
        if result and mention_placeholders:
            result = restore_mentions(result, mention_placeholders)
        if result:
            reply = LANG_FLAGS.get("zh", "") + " " + result
    track_group_usage(group_id, _bp, _bc)

    if reply is None:
        return

    bot_stats["text_translations"] += 1

    # Build reply message based on settings
    sender_display = None
    if sender_id:
        sender_display = (group_user_names.get(group_id, {}).get(sender_id) or
                       get_display_name(group_id, sender_id))

    src_flag = LANG_FLAGS.get(lang, "")
    tgt_flag = LANG_FLAGS.get("zh" if lang != "zh" else "id", "")
    translated_text = reply.split(" ", 1)[1] if " " in reply else reply

    # Flex or plain text based on setting
    flex_msg = None
    if get_group_feature(group_id, 'flex'):
        flex_msg = build_translation_flex(text, translated_text, src_flag, tgt_flag, sender_display, quoted_text)
    qr = build_quick_reply(group_id) if get_group_feature(group_id, 'quick_reply') else None
    custom_sender = get_sender_object()
    # Get quoteToken from original message for reply linking
    qt = getattr(event.message, 'quote_token', None)

    _use_retry = get_group_feature(group_id, 'retry_key')
    _retry_key = generate_retry_key() if _use_retry else None

    with ApiClient(configuration) as api_client:
        api_line = MessagingApi(api_client)
        if flex_msg:
            if qr:
                flex_msg.quick_reply = qr
            if custom_sender:
                flex_msg.sender = custom_sender
            if qt:
                try: flex_msg.quote_token = qt
                except Exception: pass
            req = ReplyMessageRequest(reply_token=event.reply_token, messages=[flex_msg])
            if get_group_feature(group_id, 'silent'):
                req.notification_disabled = True
            try:
                if _retry_key:
                    api_line.reply_message(req, x_line_retry_key=_retry_key)
                else:
                    api_line.reply_message(req)
            except TypeError:
                api_line.reply_message(req)
        else:
            msg = TextMessage(text=reply)
            if qr:
                msg.quick_reply = qr
            if custom_sender:
                msg.sender = custom_sender
            if qt:
                try: msg.quote_token = qt
                except Exception: pass
            req = ReplyMessageRequest(reply_token=event.reply_token, messages=[msg])
            if get_group_feature(group_id, 'silent'):
                req.notification_disabled = True
            try:
                if _retry_key:
                    api_line.reply_message(req, x_line_retry_key=_retry_key)
                else:
                    api_line.reply_message(req)
            except TypeError:
                api_line.reply_message(req)


@handler.add(MessageEvent, message=ImageMessageContent)
def handle_image(event):
    """Handle image messages: OCR + translate with layout-preserving text."""
    source = event.source
    group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
    user_id = getattr(source, 'user_id', None)
    is_dm_img = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
    logger.info("Image received from %s", group_id)

    # Record user for whitelist (even if translation is off)
    if group_id and user_id and not is_dm_img:
        record_user_name(group_id, user_id)

    # Check if translation is on
    is_on = group_settings.get(group_id, True)
    if not is_on:
        return

    # Check skip list
    sender_id = user_id
    if sender_id and sender_id in group_skip_users.get(group_id, set()):
        return

    # DM master toggle check for image
    if is_dm_img and sender_id:
        if not dm_master_enabled and sender_id not in dm_whitelist:
            return

    # Check if image translation is on
    img_on = group_img_settings.get(group_id, True)
    if not img_on:
        return

    # Need OpenAI for image OCR
    if not oai:
        logger.warning("No OpenAI key, cannot do image OCR")
        return

    show_loading(group_id)
    if get_group_feature(group_id, 'mark_read'):
        mark_as_read(group_id)

    # Download image from LINE
    message_id = event.message.id
    img_base64, img_raw = download_line_image(message_id)
    if not img_base64:
        logger.warning("Failed to download image %s", message_id)
        return
    logger.info("Image downloaded: %d bytes", len(img_raw) if img_raw else 0)

    # Determine target language
    tgt = group_target_lang.get(group_id, "id")

    # Quick OCR to check if there's text and detect language
    _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
    extracted = ocr_image_openai(img_base64)
    logger.info("Image OCR result: %s chars, text: %s", len(extracted) if extracted else 0, (extracted[:100] + "...") if extracted and len(extracted) > 100 else extracted)
    if not extracted or len(extracted.strip()) < 2:
        return

    # === Check if this is a work order (製造指示書) ===
    try:
        wo_customer = detect_work_order(extracted)
        if wo_customer:
            # It's a work order — never translate work order content
            wo_on = group_wo_settings.get(group_id, True)
            if wo_on:
                reply = format_storage_for_work_order(wo_customer)
                if reply:
                    bot_stats["work_order_detections"] += 1
                    with ApiClient(configuration) as api_client:
                        api = MessagingApi(api_client)
                        api.reply_message(ReplyMessageRequest(
                            reply_token=event.reply_token,
                            messages=[TextMessage(text=reply)]
                        ))
            # Whether storage found or not, skip translation for work orders
            track_group_usage(group_id, _bp, _bc)
            return
    except Exception as e:
        logger.error("Work order detection error: %s", e)
    # === End work order check ===

    lang = detect_language(extracted)
    if lang is None:
        return

    # Determine actual translation target
    if lang == "zh":
        actual_tgt = tgt
    else:
        actual_tgt = "zh"

    # Translate OCR text using the same translation engine as text messages
    # Set translation tone for this group
    _tone, _tone_custom = get_group_tone(group_id)
    _tl.tone = _tone
    _tl.tone_custom = _tone_custom
    if lang == "zh":
        result = translate(extracted, "zh", tgt)
    else:
        result = translate(extracted, lang, "zh")

    if not result:
        track_group_usage(group_id, _bp, _bc)
        return

    reply = "\U0001f5bc\ufe0f " + LANG_FLAGS.get(actual_tgt, "") + "\n" + result

    # LINE message limit is 5000 chars
    if len(reply) > 5000:
        reply = reply[:4990] + "\n..."

    track_group_usage(group_id, _bp, _bc)
    bot_stats["image_translations"] += 1
    with ApiClient(configuration) as api_client:
        api = MessagingApi(api_client)
        api.reply_message(ReplyMessageRequest(
            reply_token=event.reply_token,
            messages=[TextMessage(text=reply)]
        ))


@handler.add(MessageEvent, message=AudioMessageContent)
def handle_audio(event):
    """Handle audio/voice messages: Whisper STT + detect language + translate."""
    source = event.source
    group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
    user_id = getattr(source, 'user_id', None)
    is_dm_aud = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)

    # Record user for whitelist
    if group_id and user_id and not is_dm_aud:
        record_user_name(group_id, user_id)

    # Check if translation is on
    is_on = group_settings.get(group_id, True)
    if not is_on:
        return

    # Check skip list
    sender_id = user_id
    if sender_id and sender_id in group_skip_users.get(group_id, set()):
        return

    # DM master toggle check for audio
    if is_dm_aud and sender_id:
        if not dm_master_enabled and sender_id not in dm_whitelist:
            return

    # Check if audio translation is on
    audio_on = group_audio_settings.get(group_id, True)
    if not audio_on:
        return

    # Need OpenAI for Whisper
    if not oai:
        logger.warning("No OpenAI key, cannot do audio transcription")
        return

    show_loading(group_id)
    if get_group_feature(group_id, 'mark_read'):
        mark_as_read(group_id)

    # Download audio from LINE
    message_id = event.message.id
    audio_bytes = download_line_audio(message_id)
    if not audio_bytes:
        return

    # Transcribe with Whisper
    transcribed = transcribe_audio_openai(audio_bytes)
    if not transcribed or len(transcribed.strip()) < 2:
        return

    # Detect language
    lang = detect_language(transcribed)
    if lang is None:
        return

    tgt = group_target_lang.get(group_id, "id")

    reply = None
    _bp, _bc = bot_stats.get("tokens_prompt", 0), bot_stats.get("tokens_completion", 0)
    # Set translation tone for this group
    _tone, _tone_custom = get_group_tone(group_id)
    _tl.tone = _tone
    _tl.tone_custom = _tone_custom
    if lang == "zh":
        result = translate(transcribed, "zh", tgt)
        if result:
            reply = "\U0001f3a4 " + LANG_FLAGS.get(tgt, "") + "\n\U0001f4ac " + transcribed + "\n\U0001f4dd " + result
    else:
        result = translate(transcribed, lang, "zh")
        if result:
            reply = "\U0001f3a4 " + LANG_FLAGS.get("zh", "") + "\n\U0001f4ac " + transcribed + "\n\U0001f4dd " + result
    track_group_usage(group_id, _bp, _bc)

    if reply is None:
        return

    bot_stats["voice_translations"] += 1
    with ApiClient(configuration) as api_client:
        api = MessagingApi(api_client)
        api.reply_message(ReplyMessageRequest(
            reply_token=event.reply_token,
            messages=[TextMessage(text=reply)]
        ))



if StickerMessageContent:
    @handler.add(MessageEvent, message=StickerMessageContent)
    def handle_sticker(event):
        """Record user name when they send a sticker (for whitelist tracking)."""
        source = event.source
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None)
        user_id = getattr(source, 'user_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)


if VideoMessageContent:
    @handler.add(MessageEvent, message=VideoMessageContent)
    def handle_video(event):
        """Handle video messages: download thumbnail, OCR, translate."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
        user_id = getattr(source, 'user_id', None)
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)
        # Video OCR: try to get preview image and OCR it
        if not get_group_feature(group_id, 'video_ocr'):
            return
        if not group_settings.get(group_id, True):
            return
        if not group_img_settings.get(group_id, True) and not is_dm:
            return
        try:
            msg_id = event.message.id
            with ApiClient(configuration) as api_client:
                blob_api = MessagingApiBlob(api_client)
                # Get video content (preview/thumbnail)
                content = blob_api.get_message_content_preview(msg_id)
                if content and len(content) > 100:
                    b64 = base64.b64encode(content).decode()
                    # OCR the preview frame
                    ocr_result = ocr_image_openai(b64)
                    if ocr_result and len(ocr_result.strip()) > 2:
                        lang = detect_language(ocr_result)
                        if lang:
                            # Set translation tone for this group
                            _tone, _tone_custom = get_group_tone(group_id)
                            _tl.tone = _tone
                            _tl.tone_custom = _tone_custom
                            if lang == "zh":
                                tgt = group_target_lang.get(group_id, "id")
                                result = translate(ocr_result, "zh", tgt)
                                actual_tgt = tgt
                            else:
                                result = translate(ocr_result, lang, "zh")
                                actual_tgt = "zh"
                            if result:
                                reply = "🎬 " + LANG_FLAGS.get(actual_tgt, "") + " " + result
                                with ApiClient(configuration) as ac2:
                                    api2 = MessagingApi(ac2)
                                    api2.reply_message(ReplyMessageRequest(
                                        reply_token=event.reply_token,
                                        messages=[TextMessage(text=reply)]
                                    ))
                                bot_stats["image_translations"] += 1
        except Exception as e:
            logger.warning("Video OCR failed: %s", e)


if FileMessageContent:
    @handler.add(MessageEvent, message=FileMessageContent)
    def handle_file(event):
        """Handle file messages: record user, log file info."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None)
        user_id = getattr(source, 'user_id', None)
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)
        fname = getattr(event.message, 'file_name', '未知檔案')
        fsize = getattr(event.message, 'file_size', 0)
        logger.info("File received: %s (%d bytes) from %s", fname, fsize, group_id)


if LocationMessageContent:
    @handler.add(MessageEvent, message=LocationMessageContent)
    def handle_location(event):
        """Handle location messages: translate location info."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', None)
        user_id = getattr(source, 'user_id', None)
        is_dm = not getattr(source, 'group_id', None) and not getattr(source, 'room_id', None)
        if group_id and user_id and not is_dm:
            record_user_name(group_id, user_id)
        if not group_settings.get(group_id, True):
            return
        if not get_group_feature(group_id, 'location'):
            return
        # Translate location title/address if available
        title = getattr(event.message, 'title', '') or ''
        address = getattr(event.message, 'address', '') or ''
        if title or address:
            loc_text = (title + " " + address).strip()
            lang = detect_language(loc_text)
            if lang:
                # Set translation tone
                _tone, _tone_custom = get_group_tone(group_id)
                _tl.tone = _tone
                _tl.tone_custom = _tone_custom
                if lang == "zh":
                    tgt = group_target_lang.get(group_id, "id")
                    result = translate(loc_text, "zh", tgt)
                    actual_tgt = tgt
                else:
                    result = translate(loc_text, lang, "zh")
                    actual_tgt = "zh"
                if result:
                    try:
                        with ApiClient(configuration) as api_client:
                            api = MessagingApi(api_client)
                            api.reply_message(ReplyMessageRequest(
                                reply_token=event.reply_token,
                                messages=[TextMessage(text="📍 " + LANG_FLAGS.get(actual_tgt, "") + " " + result)]
                            ))
                    except Exception:
                        pass


if JoinEvent:
    @handler.add(JoinEvent)
    def handle_join(event):
        """Track when bot joins a group."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None)
        if not group_id:
            return
        gname = ""
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                summary = api.get_group_summary(group_id)
                gname = summary.group_name or ""
        except Exception:
            pass
        group_tracking[group_id] = {"name": gname, "joined_at": time.time()}
        save_settings()

if MemberJoinedEvent:
    @handler.add(MemberJoinedEvent)
    def handle_member_joined(event):
        """Send bilingual welcome when a new member joins the group."""
        source = event.source
        group_id = getattr(source, 'group_id', None)
        if not group_id:
            return
        # Record new members
        members = getattr(event, 'joined', None)
        if members and hasattr(members, 'members'):
            for member in members.members:
                uid = getattr(member, 'user_id', None)
                if uid:
                    record_user_name(group_id, uid)
        # Send welcome if enabled
        ws = get_group_welcome(group_id)
        if not ws.get("enabled", True):
            return
        if not group_settings.get(group_id, True):
            return
        try:
            zh = ws.get("text_zh", "")
            id_text = ws.get("text_id", "")
            welcome = zh + "\n\n" + id_text if zh and id_text else (zh or id_text)
            if welcome:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    api.reply_message(ReplyMessageRequest(
                        reply_token=event.reply_token,
                        messages=[TextMessage(text=welcome)]
                    ))
        except Exception as e:
            logger.warning("Failed to send welcome: %s", e)


if MemberLeftEvent:
    @handler.add(MemberLeftEvent)
    def handle_member_left(event):
        """Track when a member leaves the group."""
        source = event.source
        group_id = getattr(source, 'group_id', None)
        if not group_id:
            return
        left = getattr(event, 'left', None)
        if left and hasattr(left, 'members'):
            for member in left.members:
                uid = getattr(member, 'user_id', None)
                if uid:
                    # Remove from skip list
                    if group_id in group_skip_users:
                        group_skip_users[group_id].discard(uid)
                    logger.info("Member %s left group %s", uid, group_id)


if FollowEvent:
    @handler.add(FollowEvent)
    def handle_follow(event):
        """Track when a user adds the bot as friend."""
        user_id = getattr(event.source, 'user_id', None)
        if not user_id:
            return
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                profile = api.get_profile(user_id)
                dm_known_users[user_id] = profile.display_name or user_id
                lang = getattr(profile, 'language', None)
                if lang:
                    user_languages[user_id] = lang
        except Exception:
            dm_known_users[user_id] = user_id
        bot_stats["followers"] = bot_stats.get("followers", 0) + 1
        save_settings()
        logger.info("New follower: %s", dm_known_users.get(user_id, user_id))


if UnfollowEvent:
    @handler.add(UnfollowEvent)
    def handle_unfollow(event):
        """Track when a user blocks/removes the bot."""
        user_id = getattr(event.source, 'user_id', None)
        if user_id:
            bot_stats["unfollowers"] = bot_stats.get("unfollowers", 0) + 1
            logger.info("Unfollowed by: %s", user_id)


if BotLeaveEvent:
    @handler.add(BotLeaveEvent)
    def handle_bot_leave(event):
        """Clean up when bot is removed from a group."""
        group_id = getattr(event.source, 'group_id', None) or getattr(event.source, 'room_id', None)
        if group_id:
            group_tracking.pop(group_id, None)
            group_settings.pop(group_id, None)
            group_target_lang.pop(group_id, None)
            group_img_settings.pop(group_id, None)
            group_audio_settings.pop(group_id, None)
            group_wo_settings.pop(group_id, None)
            group_skip_users.pop(group_id, None)
            group_user_names.pop(group_id, None)
            save_settings()
            logger.info("Bot removed from group %s", group_id)


if PostbackEvent:
    @handler.add(PostbackEvent)
    def handle_postback(event):
        """Handle postback actions from Quick Reply etc."""
        data = event.postback.data if hasattr(event.postback, 'data') else ""
        logger.info("Postback: %s", data)


if UnsendEvent:
    @handler.add(UnsendEvent)
    def handle_unsend(event):
        """Clean up cached message when user unsends."""
        msg_id = getattr(event.unsend, 'message_id', None) if hasattr(event, 'unsend') else None
        if msg_id and msg_id in message_cache:
            del message_cache[msg_id]
            logger.info("Unsend: removed message %s from cache", msg_id)


if VideoPlayCompleteEvent:
    @handler.add(VideoPlayCompleteEvent)
    def handle_video_play_complete(event):
        """Handle video viewing complete event — log for analytics."""
        source = event.source
        group_id = getattr(source, 'group_id', None) or getattr(source, 'room_id', None) or getattr(source, 'user_id', '')
        user_id = getattr(source, 'user_id', '')
        tracking_id = getattr(event.video_play_complete, 'tracking_id', '') if hasattr(event, 'video_play_complete') else ''
        logger.info("VideoPlayComplete: group=%s user=%s tracking=%s", group_id, user_id, tracking_id)
        bot_stats["video_play_complete"] = bot_stats.get("video_play_complete", 0) + 1


def show_loading(chat_id):
    """Show typing indicator before translation."""
    if not ShowLoadingAnimationRequest:
        return
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.show_loading_animation(ShowLoadingAnimationRequest(chat_id=chat_id))
    except Exception:
        pass


def mark_as_read(chat_id):
    """Mark messages as read in the chat (shows 'read' indicator)."""
    if not MarkMessagesAsReadRequest:
        return
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.mark_messages_as_read(MarkMessagesAsReadRequest(chat_id=chat_id))
    except Exception as e:
        logger.debug("mark_as_read failed: %s", e)


def generate_retry_key():
    """Generate a UUID v4 for X-Line-Retry-Key header to prevent duplicate sends."""
    return str(uuid.uuid4())


def safe_reply(reply_token, messages, retry=True):
    """Reply with optional X-Line-Retry-Key for idempotency."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            req = ReplyMessageRequest(reply_token=reply_token, messages=messages)
            if retry:
                # SDK v3 supports x_line_retry_key param
                try:
                    api.reply_message(req, x_line_retry_key=generate_retry_key())
                except TypeError:
                    api.reply_message(req)
            else:
                api.reply_message(req)
    except Exception as e:
        logger.warning("safe_reply failed: %s", e)


def safe_push(to, messages, retry=True):
    """Push with optional X-Line-Retry-Key for idempotency."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            req = PushMessageRequest(to=to, messages=messages)
            if retry:
                try:
                    api.push_message(req, x_line_retry_key=generate_retry_key())
                except TypeError:
                    api.push_message(req)
            else:
                api.push_message(req)
    except Exception as e:
        logger.warning("safe_push failed: %s", e)


# ---- Webhook Management API ----
def get_webhook_info():
    """Get current webhook endpoint info."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            info = api.get_webhook_endpoint()
            return {
                "endpoint": getattr(info, 'endpoint', ''),
                "active": getattr(info, 'active', None),
            }
    except Exception as e:
        logger.warning("get_webhook_info failed: %s", e)
        return None


def set_webhook_url(url):
    """Set webhook endpoint URL via API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            from linebot.v3.messaging import SetWebhookEndpointRequest
            api.set_webhook_endpoint(SetWebhookEndpointRequest(endpoint=url))
            return True
    except Exception as e:
        logger.warning("set_webhook_url failed: %s", e)
        return False


def test_webhook(endpoint=None):
    """Test webhook endpoint."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            from linebot.v3.messaging import TestWebhookEndpointRequest
            if endpoint:
                resp = api.test_webhook_endpoint(TestWebhookEndpointRequest(endpoint=endpoint))
            else:
                resp = api.test_webhook_endpoint(TestWebhookEndpointRequest())
            return {
                "success": getattr(resp, 'success', None),
                "timestamp": getattr(resp, 'timestamp', ''),
                "status_code": getattr(resp, 'status_code', None),
                "reason": getattr(resp, 'reason', ''),
                "detail": getattr(resp, 'detail', ''),
            }
    except Exception as e:
        logger.warning("test_webhook failed: %s", e)
        return {"success": False, "reason": str(e)}


# ---- Content Preview & Preparation Status ----
def get_content_preview(message_id):
    """Get a preview image of an image or video message."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_message_content_preview(message_id)
            return content
    except Exception as e:
        logger.warning("get_content_preview failed: %s", e)
        return None


def check_content_preparation(message_id):
    """Check if video/audio content is ready for download."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            resp = blob_api.get_message_content_transcoding_by_message_id(message_id)
            status = getattr(resp, 'status', 'unknown')
            return status  # "processing", "succeeded", "failed"
    except Exception as e:
        logger.debug("check_content_preparation failed: %s", e)
        return "unknown"


# ---- Validate Message Objects ----
def validate_message_objects(messages, msg_type="reply"):
    """Validate message objects before sending. Returns True/error dict."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            if msg_type == "push":
                api.validate_push({"messages": messages})
            elif msg_type == "broadcast":
                api.validate_broadcast({"messages": messages})
            elif msg_type == "multicast":
                api.validate_multicast({"messages": messages})
            else:
                api.validate_reply({"messages": messages})
            return {"valid": True}
    except Exception as e:
        return {"valid": False, "error": str(e)}


def get_line_quota():
    """Get LINE monthly message quota info."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            quota = api.get_message_quota()
            consumption = api.get_message_quota_consumption()
            return {
                "quota": getattr(quota, 'value', None),
                "type": getattr(quota, 'type', None),
                "used": getattr(consumption, 'total_usage', None),
            }
    except Exception:
        return None


def get_follower_count():
    """Get bot follower count."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_number_of_followers(var_date=time.strftime("%Y%m%d"))
            return getattr(resp, 'followers', None)
    except Exception:
        return None


def get_bot_info():
    """Get bot's own profile info."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            info = api.get_bot_info()
            return {
                "name": getattr(info, 'display_name', ''),
                "picture": getattr(info, 'picture_url', ''),
                "status": getattr(info, 'chat_mode', ''),
            }
    except Exception:
        return None


def get_group_member_count(group_id):
    """Get group member count from LINE API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            count = api.get_group_members_count(group_id)
            return count
    except Exception:
        return None


def fetch_all_group_members(group_id):
    """Fetch all member IDs in a group using LINE API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            member_ids = []
            token = None
            while True:
                try:
                    if token:
                        resp = api.get_group_members_ids(group_id, start=token)
                    else:
                        resp = api.get_group_members_ids(group_id)
                except AttributeError:
                    # Try alternative method name
                    if token:
                        resp = api.get_group_member_ids(group_id, start=token)
                    else:
                        resp = api.get_group_member_ids(group_id)
                # Extract member IDs from response
                ids = getattr(resp, 'member_user_ids', None) or getattr(resp, 'member_ids', None) or []
                member_ids.extend(ids)
                token = getattr(resp, 'next', None) or getattr(resp, 'next_token', None)
                if not token:
                    break
            # Record names for all members
            for uid in member_ids:
                record_user_name(group_id, uid)
            logger.info("Fetched %d members from group %s", len(member_ids), group_id)
            return member_ids
    except Exception as e:
        logger.warning("Failed to fetch group members: %s (type: %s)", e, type(e).__name__)
        return []


def push_message_to_group(group_id, text):
    """Push a message to a group (not a reply)."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.push_message(PushMessageRequest(
                to=group_id,
                messages=[TextMessage(text=text)]
            ))
            return True
    except Exception as e:
        logger.warning("Push message failed: %s", e)
        return False


def setup_rich_menu():
    """Create a rich menu with common bot actions."""
    if not RichMenuRequest:
        logger.warning("RichMenuRequest not available, skipping rich menu setup")
        return None
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            # Delete ALL existing rich menus first
            try:
                api.cancel_default_rich_menu()
            except Exception:
                pass
            try:
                existing = api.get_rich_menu_list()
                for rm_old in (existing.richmenus or []):
                    try:
                        api.delete_rich_menu(rm_old.rich_menu_id)
                    except Exception:
                        pass
            except Exception:
                pass
            # Create new rich menu (2 rows x 3 columns)
            rm = RichMenuRequest(
                size=RichMenuSize(width=2500, height=1686),
                selected=True,
                name="翻譯Bot選單",
                chat_bar_text="📋 選單",
                areas=[
                    RichMenuArea(
                        bounds=RichMenuBounds(x=0, y=0, width=833, height=843),
                        action=MessageAction(label="說明", text="/help")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=833, y=0, width=834, height=843),
                        action=MessageAction(label="狀態", text="/status")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=1667, y=0, width=833, height=843),
                        action=MessageAction(label="查儲區", text="/qry ")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=0, y=843, width=833, height=843),
                        action=MessageAction(label="翻譯開", text="/on")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=833, y=843, width=834, height=843),
                        action=MessageAction(label="不翻我", text="/skip")
                    ),
                    RichMenuArea(
                        bounds=RichMenuBounds(x=1667, y=843, width=833, height=843),
                        action=MessageAction(label="公告", text="/notice ")
                    ),
                ]
            )
            result = api.create_rich_menu(rm)
            rid = result.rich_menu_id
            # Generate simple image for rich menu
            _upload_rich_menu_image(api_client, rid)
            # Set as default
            api.set_default_rich_menu(rid)
            logger.info("Rich menu created: %s", rid)
            return rid
    except Exception as e:
        logger.warning("Rich menu setup failed: %s", e)
        return None


# Rich Menu image (2500x843 JPEG, base64 encoded)
_RICH_MENU_IMG_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAA0JCgsKCA0LCgsODg0PEyAVExISEyccHhcgLikxMC4pLSwzOko+MzZGNywtQFdBRkxOUlNSMj5aYVpQYEpRUk//2wBDAQ4ODhMREyYVFSZPNS01T09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT09PT0//wAARCAaWCcQDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwDgKKKWthCUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUABooopgFFFFAC0lAooAWiiikAUUUtACUUUUwCiiigAooooAKKKKAFooooAKKKKACiiigAooooAKKKKACiiigBaKKKBBSUtFABRRRQAUUUUAFFFFMApKWkoAWiiikMO1FFFMAooooAKKKKACiiigApaSigAooopgFFFFAC0UlLSAKKKKACjFFJTAWiiigAooooAKKKKACiiigApKWigAooooAWikpaACiiigBKKWigBKKKKACiiloASlopKACiiigApaKSgApaSlzQAUUUUAFFFFABRRRQAUUUUwCiiikAUUUUAFFFFAAKWiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFACiigUUwCkpTRQAUUUUgCiiimAUdaKSgBaKKKQBRRRTAKKKKACiiigYUUtFAhKWkpaACiiigAooooAKKKKBhRRRQIKWkooAWikpaACiiigAooooAKKKKACiiigAooooAKKKKAA0lLSU0AtFFFABRRRQAUUUUAFFFFABRRS0AJRRRTAKU0lFABRRRQAUUUUgFpKKKAFooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLRRQAUUUUAFFFFAwooooAKKKKBBRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKAFooooAKKKKACiiigAopKWgAooooAKKKKACiiigAooooAKKKWgYlFFFAgooooGFFFFABRRRQAUvaiimAlFLRQAUUUUCDiiiimMp0UUVykhSUUUDCiiigAooooAKKKKAClpKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUd6AFoopaQCUUUUAFFFBpgFFFJQAtFFFABS0lLQAUUUUAFFFFABRRRQAUUUUAFFFFABS0lLQAUUUUCCiiigAooooAKKKKACiiimAUlFLQMKKKKACiiigAooooAKKKKAFpKWkoAKO9FFABRRRTAKKKWgBKWkpaQBRRRQAUlFFMBaKKTvQAtFFFABRRSUALRRRQAUUUUAFLSd6KAFoopKACiiigAooooAKKKKACiiigApaSigAooooAKKKKAFoo7UUAFFJS0wCkopaACiiigAooooAKKKKQBRRRQAUopKUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoo70UAFLSUooAKKKSmAtFJS0AFFFFABRRRQAlFLRQAUUUUAFFFFABRRRQAUUUtABRSUUALRSUtAwooooEFFFFACUtFFABRRQKAFpKWkoAWiiigAooooAKKKKACiiigAooooAKKKKACiiigBKWkNFMBaKKKACiiigAooooAKKKKACiiigApaSlpgJRS0lAAaKKKQB1ooooAKKWkpgLRRRQMO1FFFAgooooAKKKKACiiigAooooAKKKKAFooooAKKKKBhRRRQAUUUUAFFFFAg7UUUUAFFFFAAKKBS0AJRRRTAKKKKACiiigAooooAKKKKAFooooGFFFFAgooooAKKKKACiiigAooooAKKKKACiiigYUtJRQAtJS0lABRRRQAUUUUAFFFLTAKKKKACikpaBBRRRQAUUUUDKdFJ3ormJCiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABS0lFABRRRQAUUUUAFFFFABRRRQAUUUUAFFLSUAFFFFMAooooAKUUlKKAFpKKKQC0neiimAUUUUAFFFFABSUtFABS0gpaACiiigAooooAKKKKACiiigAooooAKWkpaACiiigQUUUUAFFFFACUtFFACUUUvamAgpaKKQwooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUtMBKKKKACloopAFFFFCASiilpgFFFFABRRSUALRSUUAFGaKKAFpaQUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUCgBaKSloASloopgJS0UUAFFJS0AFFFFABRRRQAUUUUgFooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAd6WkpaYBQaKKACiiigAooooAKSiloAKKSloAKKKKACiiigAooooAKWkpaACiiigAooooAKKKKACiiigAooooAKKKKAF5pKKKAClpKWmAUUUUgCiiigAooooAKKKKACiiigApKU0lMApaSloAKKKKACiiigAooooAKKKKACiiigAooopgLSUtJQAUUUUAFFFFAwpRSUtAgooooAKKKKACiiigAooooAKKKKACiiigAopaSgApaSigBaKKKBhRRRQAUUUUAFFFFABRRRQIKKKKACiiimAUUUUAFFFFABRRRQAUUUtACUUUtABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUDCiiigAooooAKKKKACjNFFABS0gzmlpgFFFFABRRRQIKKKKQBRRRVAU6KSiuUAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAClpKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKWgQUUUUAFFJS0AJRS0lMYtFAopAFFFFABRRRTAKKM0UAFFFFABRRR3oAWiiigAooooAKKKKACiiigAooooAKKKKACloFFAgooooAKKKKACiiigAooooASiilpjCiiikAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKYBRRRSAKWkpRQAUlLSUwClpKKAFopKKAFpKKKAFpKKWgAFFFJQAvaiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACl7UCigAoopKAFopKKAFopKWgAopKWmAUUUUAFFFFABRRRSAWiiigAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFAoAKWiimAUUUUAFFFFABSUtJQAUUUUAFFFGaAClopKAFopKWgAooooAKWkpaACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFLSUtABR3pKKAFooopAFFFFABRRRQAlFFFMApaSloAKKKKACiiigAooooAKKKKACiiimAtFFJQAUUUUAFFFFABRRS0AJS0CikAUUUUwCiiigAooooGFFFFABRRRQIKKKKACiiigAoFFFAxaKKSgBaBSUtABRRRQAUUUUAFFFFAgooooAKKKKYBRRRQMKKKKBBRRRQAUUUUAHelpKWgAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKBhRRRQAUUUUAFFFFABRRRQAoopBS0wCiiigQUUUUAFFFFIAoooqgKVFFFcoBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLRRQAlFFFABRRRQAtFFFAgooooAKKKKACiiimMBS0UUgCkoooAKKKKYBRRRQAUUUUAFLSUCgBaKKKACiiigAooooAKKKKACiiigAooooAWiiigQUUUUAFFFFACUtFFABRRSUwCloooGFFFFABRRRQAUUUUAFFFFAAKKKKACiiigAooo7UwCiiigAooopALRQKKACkoopgFFFFABS0lFAC0lFFAC0UlGaAClFJ1paACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAFooopgJRS0UgEooooAKKKKAFooFFMAooooAKKKKQBS0lLQAUUUUAFFFFMBKWiigAooooAKKKKACiikoAWkpaKACiiigAooooAKKKKACiiigApaKKYBRRRQAUUUUAFFFJQAtFFJQAUUUUAFFFFABRRRQAUUUtABRRRQAtFJS0AJS0UlAC0UUUAFFFFABRRRQAUUUUAFFFFAwooooEFFFFMYUUUUAFFFFAgooooAKKKKAFoooNIBKKKWmAUlLRQAUUUUAFFFFABRSUtABRRRQAUUUUwCiiigAooooAKKKKAClpKKBi0UCigQUUUUAFFFFABRRRQMKKKKACiiigQUUUUAFFLSUAFFFLQMKSlooAKKKKACiiigAooooEFFFFAwooooEFFFFMAooooGFFFFAgooooAKKKKAFFFHaikAUUUUxhRRRQIKKKKACiiigAooooAKKKKACiiigAooooGFFFFABRRRQAUUUUAFFFFAC0UUlMBaKKKBBRRRSAKKKKBhRRRTApUUUVzCCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigApaSloAKKKKBBRRRQAUUUlABS0UCgYtJRRQAUUUUAFFFFMAooooAKKKKAClpKWgAooooAKKKKACiiigAooooAKKKKACiiloEFFFFABRRRQAUUUUAFJS0UAFJS0lMYtFFFIAooopgFFFFABRRRQAUUUUAFFFFABRRRmgANFGaKYBRRRQAUtJRSAWigUUAJRRRTAKWkooAKKKKACiiigApaSigApaO1FABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUGgBaSiigBaKSigAooooAWkopaACiiimAUUUUAFFFFIBaKKKACiiigAooopgFFFFABRRRQAUUUUAFFJS0AFFFJQAtFFFABRRRQAUUUUAFFFHegBaKKKYBRRRQAUUUlAC0lFFABRRRQAUUUUAFFFFABRRRQAUtJS0AFFFFABRRRQAUtJRQAtFFFABRRRQMKKKKACiiigQUUUUAFFFFMYUUUUAFFFFABRRRQIKKKKACiigUALSUtJQAUUUtACUtFFABRRRQAUUUUAFFFFABRRSUALRRRTAKKKKACiiigAooo7UALSUUUALRRRQMKKKKBBRRRQMKKKKACiiigQUUUUAFFFFAwooooELRRRQMKKKKACiiigAooooAKKKKACiiigQUUUUAFFFFABRS0lMAooooAKKKKACjvRS9qACiiikAUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKBhRRRQAUUUUAFFFFABS0lFABS0lLTAKKKKBBSUtFABRRRSAKKKKBhRRRTApUUUVzCCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoopaBBRRRQAUUUUAFFJS0AFApBS0DCiiloASiiimAUUUUAFFFFABRRRQAUtFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUtABRRRQIKKKKACiiigAopKWmAUUUUDCiiigBaSiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUtIBKKWigApKWkoAKKKKYBRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAKKSlpKACiiigBaKSigBaKSloASloooAKKWkoAKKKKACiiloAKKKKACiiimAUUUUAFFFFACUtFFABRRRQAUUUUAFFFFABRRRQAfSiiigAooooAKWkooAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFFABS0UUAFGKKKACiiigAooooAWiiigAooooAKKKKACiiigYUtJRQIKKKKACiiimAUUUUDCiiigQUUUUAFFFFABRRRQAtJRRQAtJS0UAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKKKKACiiikAtFFFABRRRTAKKKKACiiikAUUlLTAKKKKACiiigAooooAWiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUCCiiigAooopgFFFFACiiiikAUUUUAFFFFAwpKWkpiFopKWgAooooGFFFFAgooooGFFFFABRRRQAUUUUAFFFFABRRRQAUtFFMApKWigQUUUUAFFFFABRRRQAUUUUDKVFFFcwgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKWkoAKKWigBKKKKACiiloASiiigAoopaAEpaSigBaKSigBaKSimAUUUUALRRR2pALSUUUAFFFFMAooooAKKKKACilpKAFooooAKKKKACiiigAooooAKKKKACiiigApaSloEFFFJQAtFFFABRSUUALSUUUwFooopDCiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACilooASlFJS0gCkpaKYCUtJRQAUtFJQAUUUUAFFLRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRQAtFJRQAUUUtACUtFFABRRRQAUUUUAGaKKWgBKKKKAClpKWgAooopgFFFFABRRRQAlLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAAooooAKKKKACiiigBaSiigAooopgFFFFABRRS0AFJS0UAJS0UUAFFFFABRRRQAUUUUAFLSUUAFLRRQAUUlLQAUUUUAFFFFABRS0UAFJS0UwEooooAKKKKACiiigAooooAKKKKBhRRRQIWikpaACiiigApKWigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFAooFAC0UUUAFFFFACUtFFABRRRSAKKKKYBRRRQAUUtJQAUtFFABRRRQMKKKKACiiigAooooAKKKKBBRRRQAUUUUAFFFFABRRRQAUUUUwCjvRQKAFopKWkAUUlFAC0UlLQMKSiimIKWkpaACiiigAooooAKKKDQAUUUUDCiiigAooooAKKKKACiiigBaKKKYBRSUtAgooooGFFFFAgooopAFFFFUMp0lGeaK5RBRS0lABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFBIHWrtnpN7en93HsU/wAT0AUqNwHUiuptvCkSgG6mYn/Z6VfTRdKiGGCH/epXA4bcv94Ubl9RXeDT9IH8EFH2HSP7kFFwOE3L/eFJuX+8K7z7DpH9yCj7BpH9yCi4HB7l/vCjcv8AeFd59h0j+5BR9h0j+5BRcDg9y/3hRuX+8K7z7BpH9yCk+w6R/cgouBwm5fUUu5fUV3X2HSP7kFH2LSP7kFFwOE3L6il3L6iu6+xaR/cgpPsOkf3IKLgcKWX1FG5fUV3f2HSP7kFJ9h0n+5BRcDhty+opNy+oruvsOkf3IKPsOkf3IKLgcNuX1FJuX1Fd19h0j+5BR9i0n+5BRcDhty/3hRuX+8K7n7DpP9yCj7DpP9yCi4HDbh6ijcvqK7j7FpP9yCl+xaT/AHIKLgcNuX+8KNy+oruPsWk/3IKPsWk/3IKdwOHyvqKMr6iu4+xaT/cgo+xaT/cgouBw+4eoo3L6iu3+xaV/cgo+xaT/AHIKLgcRuX1FG5fUV232LSv7kNH2LSv7kNFwOK3L6ijcvqK7X7Hpf9yGk+x6V/choA4vcvqKNy+ortDZ6V/chpPsel/3IaLgcZuX1FG5fUV2f2PS/wC5DSfY9L/uQ0XA43K+oo3D1Fdl9k0z+5DR9k0z+5DRcDjdy+oo3D1Fdl9k0z+5DSfZNM/uQ0XA47cvqKNw9RXY/ZNM/uQ0fZNM/uQ0XA47cPUUbh6iuw+yaZ/cho+yab/chouByG5fWk3D1Fdf9l03+5DR9l03+5DRcRyG5fUUbh6iuu+yab/chpPsmm/3IqLgcluHqKNw9RXW/ZdO/uRUn2XTv7kVO4HJ7h6il3D1FdX9l07+7FR9l07+7FSuM5TcPUUZHqK6r7Np392Kk+zaf/dip3A5bcPUUbh6iup+zaf/AHYqPs2n/wB2Ki4HLbh6ijI9RXUG20/+5FSfZtP7JFRcDmMj1FG4etdP9n0/+5FSfZ9P/uxUXA5nI9aNw9a6b7PYf3YqPs9h/diouBzO4etGR610v2ew/uxUfZ7H+7HRcDmtw9RRuHqK6U29j/djpPs9h/djouBzeR60ZHrXSeRY/wB2Ok8iy/ux0XA5zI9aMj1ro/Isj/DHTGsrN+gUfSi4GBRWxLo6MMxOc+9UJ7G4g+8u4eq07gV6KQUtACUUUUwClpKKAFopKWgBKKKKAFooooAKKBRQAUUUUAFFFFABRS0lABRRToopJ22Qxs59hQA0UEgdTitu08M3UwDXDqinsOtasXhywhH712Y/7VK4HHbl/vCjcvrXcDTNJT+GL8aX7BpX9yGlcDhty+oo3L6iu6+w6T/cgpPsOk/3IKLgcNuX1FG5fUV3P2DSf7kNH2DSf7kFO4HDbl9RRuX1Fdz9g0n+5DR9h0n+5DRcDh9y+oo3L6iu4+w6T/cgo+w6T/chouBw+5fUUm5fUV3H2HSf7kNH2HSf7kNK4HD7l9RS7l9RXb/YtJ/uQ0fYdJ/uQUXA4fcvqKXcvqK7f7DpP9yGj7DpP9yGi4HEbl9RRuX1Fdv9i0r+5DR9h0r+5DTuBxG5fUUbl9RXb/YtK/uQ0n2LSv7kNFwOJ3L6ijcvqK7b7FpX9yGj7DpX92Gi4HEhl9RS7h/eFdr9i0r+7DSfYtK/uQ0XA4vcvqKTcvrXa/Y9L/uQ0n2LS/7kNFwOM3L6ijcvqK7P7Fpf9yGj7Fpf9yGi4HF7l9RS7l/vCuy+x6X/AHIaPsel/wByGi4HGbl9RRuX1Fdl9j0v+5DR9j0v+5DTuBxu4eoo3D1Fdl9j0z+5DR9j0z+5DSuBxu4eopdw9RXY/Y9M/uQ0n2PTP7kNFwOP3D1FG5fUV2H2TTP7kNH2PTP7kNO4HH7l9RRuHqK6/wCx6Z/choNppn9yGi4HIZHrRuHqK6/7Jpv9yGj7Hpv92Gi4HIbh6ijcvqK677Jpv92Gj7Jpv92Gi4HI7h6ijcPUV132XTf7kVIbXTf7sVFwOS3L60bh6iut+y6b/dipPsunf3YqLgcpuHqKMj1FdX9l07+7FSfZdO/uxUXA5XcPUUZHrXVfZdO/uxUn2bT/AO7FRcDldw9RS5HrXU/ZtP8A7sVH2XT/AO7FRcDlsj1FJuHrXU/ZtP8A7sVH2bT/AO7FRcDltw9RS7h6iun+zaf/AHYqT7NYf3YqLgczuHqKMj1rpvs9h/djo+z2H92Ki4HM7h6ijI9a6X7PYf3Y6Ps9j/djp3A5rI9aMj1rpfs9h/djpPs9h/djouBzeR60ZHrXSfZ7D+7HSfZ7H+7HRcDnNw9aMj1ro/s9j/djpPs9j/djouBzuR60ZHrXRfZ7L+7HR5Fl/djouBz2R60ZHrXQeTZf3Y6PIsz/AAx0XA57I9aWt42do/AC/hUEukRsMxOw+tFwMiirM9jPByV3D/Zqt/OmAtFA6UUAFFFFAC0lFLQAUlLRQAlFLSUwCiiigAooooAKKKKACiiigAooooAWik6naAST0ArSs9EvbvB2iNP9rg0mxmceOtJuX+8K6qDwvbIAZ5XJ/Sri6RpcY5EZ+tLmEcTuX1FG5fUV3H2DSh/BDS/YdJ/uQ0cwHDbl9RRuX1Fdz9g0n+5DSfYdJ/uQ0cwHD7l9RRuX1Fdx9h0n+5DR9h0n+5DRzAcPuX1FG5fUV3H2HSf7kNJ9h0n+5DRzAcRuX1FG5f7wrt/sOlf3IaPsOk/3IaOYDiNy+oo3D1Fdv9h0n+5DR9h0r+5DRzAcRuHqKNy+ort/sOlf3YaT7DpX92GjmA4jcvqKXcvqK7Y2Ok/3IaPsOk/3YaOYDidy+ooDL612v2HSv7sNL9h0r+7DT5gOJ3L6il3L6iu0+xaV/dho+w6V/dhpcwHFbl9RRuX1Fdr9i0v+5DR9i0r+5DRzAcVuX1o3L6iu1+xaX/dho+xaX/dho5gOL3L6ik3L6iu0+xaX/dho+xaX/dhouBxm4etG4eors/sel/3IaT7Hpf8Adhp8wHG7h60bl9a7L7Hpn92Gk+x6Z/dho5gOO3L60bl9a7H7Hpn9yGk+x6Z/cho5gOP3D1FG4etdh9k0z+7DR9k03+7FRcDj9w9RS7hjqK677Jpn92Kj7Jpv92KjmGchuHqKXcPUV1v2TTf7kVH2TTf7kVFwOR3D1FLuHqK6z7Lpv9yKj7Lp39yKjmA5PcPUUbh611f2XTv7sVH2XTv7sVO4jlNy+tG4eorq/sunf3YqPsunf3IqVwOU3D1o3D1FdV9l07+7FR9l07+7FRzAcruHqKNw9RXU/ZdP/uxUfZdP/uxU7gctuHqKNw9a6k2un/3YqT7Np/8AdiouBy+4etGR611H2aw/uxUn2aw/uxUcwHMZHqKMj1rp/s1h/dipPs1h/djouBzOR60uR6iul+z2H92Ok+z2H92OjmA5vI9RRketdJ9nsf7sdJ9nsf7sdFwOcyPWjI9a6P7PY/3Y6Q29j/djouBzuR60ZHrXReRY/wB2Oj7PY/3Y6Lgc5ketLketdD5Fl/djo8iy/ux0XA57I9aMj1rofIsv7sdJ5Fl/djp3A5/I9aMj1roPIsv7sdJ5Fn/djouBgZHrRketb/2ezP8AClNawtXHGB9KLgYfaitSXSB1hc/8CqhNbzQnDoceo6UXAiooopjCiiigAooooAKKKKACiiigBaKKKYgooooAKKKKACiiigAooooAKKKKBlOkpTSVzCCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKkt4JbmURQIWc+naltbeW7uFggGWb9K7SytLbRbPLEb8fM/c0gKmmaBBZqJrzEknUg/dFT3WtQQfurddxHYdKyr/U5rtiqkpF2A71SGB0oAuTapdzfx7B6LVZpJW+/Ix+pplLQAc+po59TRRQAc+po59TRRQAfiaOfU0UUAHPqaOfU0UUAH4mj8TRRQAnfqaX8TRRQAn4ml/E0UUAH4mj8TRRQAnPqaOfU0tJQAc+po/E0UUAHPqaOfU0UUAGPc0n4mlooATn1NH4mlooATHuaPxNLSUwD8TR+JoooAPxNH4miigAx7mjHuaKKADn1NJ+JpaSgA59TR+JpaKAE/E0c+ppaKAE/E0fiaWkoAPxNH4miigA59TR+JopKAF/E0n4miigA/E0fiaKKAD8TSfiaWigBPxNH4miimAn40fjRRQAc+tH40UUAH40fjRRQAfjR+NFFAB+Jo59TRRQAfjR+dFFAB+JpPxNLRQAnPqaUMw6MRSUUATx3c8f8AFu+tXINQST5JRtJ/Ksyj60AaV1p0Vwu+HCt2x0NYssbwyFJFww/WtC2u3gOCdyenpV+eCK/gyMbv4WoA56inzRPDKY5Bhh+tMqgCiiigApaSigAoopaACkpaKACiiigAooooAKKKKACjoOaXtW/oGi+fi7u1/dj7invSbsBW0rQpr7Es+Y4P1NdHix0mLCqqkDt941FqWqrbjyLcAuOOOi1gSs8rl5GLMe5qQNK51uWQkQIAvqetZ8l1cyHLzMfbNR0lAC5Y9WNHPqaSimAvPqaOfU0UUAJz6mjn+8aWkoAOf7xo59TRRQAc+po59TRRQAfiaPxNFFAB+Jo/E0UUAHPqaOfU0UUAHPqaOfU0UUwDn1NJz6miikAc+po/E0UUAHPqaPxNFFMA59TR+JoooAOfU0fiaKKAD8TR+JoooAPxNJ+JpaKAD8TSfiaWigBPxNH4miigA59TRz6miimAc+ppOfU0tFABz6mjn1NFJQAvPqaPxNJRQAfiaPxNFFAB+Jo/E0UlAB+Jo/E0UUAH4mjHuaKKADn1NHPqaKKAD8aPxNFFAB+NFFFACfjR+JoooAKOfWiigA/E0fjRSUwF/Gk/GiigA/E0UUUAH4mk/GlFJQAv40c+tFFAwDMOjkVKl3PH/FuHvUNFAjSh1BH+WUbSfyoutPiuF3w4Vu2OhrOqe3uXgPXK+lIClJE8T7JFwf50yt+WKG+gyMZ7HuKxJonhlMcg5H600wI6WikpgFLSUUALRSUtABRRRQAlFLSUwCiiikMKKKKACiiimIKsWNlPfzeXbrx3Y9BTtN0+TUbgRpxGPvNXXFrXSLMIgAwOAOrGpbGQ2Wk2enR75Nrv3Z/6Uy61xVJS3XcR3bpWVd3k12+ZDhOyjpUFSBZlv7qYndKV9hUBd2PzOxptFMBcn1NAz6mkpaAEyf7xpcn1NFFACc+po59TS0lABz6mjn1NFFABz6mjn1NFFABz6mj8TRRQAfiaOfU0UUwD8TSfiaWkoAX8TSc+ppaKBB+Jo59TRRQAnPqaPxNFFAw59TRz6miloATn1NHPqaWigBPxNH4miigQfiaPxNHaigA/E0c+popKYBz6mj8TRRQMPxNH4miigA/E0fiaKKAD8TRz6miigA59aOfU0UUAH4mjn1NFFABz60n40UtACc+tHPrRRQAn4mjn1paKAD8aOfWiigBPxo/E0tJQAfjR+NFFMA/Oj8aKKAEpfxpKKAFopKKBB+NH40UUAH40UUUxhRRRSAOfU0odx0cikooAsR3kqdTuHvVuK6inG18fQ9KzKTr0osBcu9NDAyW/B/u1lsCGIYYI6itO1vGjIWQ5XsamvbRblPNiwJAO3ei4GLRQQQSGGCOooqgCiiigAooooAKWkpaACikpaYgooooAKKKKACiiigAooooAKKKKBlOkoormEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFLQAlGCSABkngCitjw1Yi6v8AzXGUh5+tIDc0TT49MsjNNjzXGWJ7CsnUb1r2c8kRKeBWlr95jFsh6/ex6Vh9qQCUUtFMAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKSigAooooAKKKKAFpKKKACiiigApKWkoAWikopgFFFFABRRRQAUUUUAFFFFABSUtFACUtJRQAtJRRQAUUUUAFGaKKAEpaO9JQAUUUUAFFJS0AFJRRQAUlLRTAKKKKACkoooAKKWigBKKWkoAKWkooAKKKKACiikoAWkoooAKKKKACprW4NvIO6HqKhooA1NQtluoPMj++oyD6isHnv1rb0yfOYW7ciqOqW/kXO4fdfmmgKVFFFMApaSigApaKKACgdaKKAA9aKKKACiiigAooowSQo6scCgDR0PTjqF4N4/cxnLH3rpdVvxawiCDAcjAA/hFFhAml6SNwAbGWPqawJpWnmaVzksajcBvXJJyT1NJQKKYBRRRQAUCiigApKKKACiiigAooooAKKKKACiiigAooooAOlFFFMApKKKAFpKKKACiiigAooooAKKKKACiiigAooooAKKKSgAooopgFFFFABRRRQAUUlBoAKKKKACiiigAoopKAClpKKACiiigAooooAKKKKACiiigAoopM0ALRSUZpgFBpKKAClpKWgApKWkoAKKKKACiiigAoopKAClpKWgAFFFGaAJracwSA/wAJ6ir19brdW+9PvgZU+tZdX9Nn6wsfpQwMfnkHqKKuapB5NxvUfK9U6YBRRRQAUUUUwFooopAFJS0lABRRQKYwooooAKdFE88yQxjLOcCm10Xhax3F7yQf7K5pNiNW3hg0jTgOMgZY+prAubh7qYyyH6D0q5rN2Z7jylPyJ1+tZ1QMKKSlFMApaKKACiiigQUUmaKACiiimMKKKKBBRRRQMKSlpKAFpKWigBKWiigAooooEFFFFABSdaWigAFFGaSgBaTNFFABRRRQAUUUUxhRRRQIKSjNFAw4ooooAKKKKACiiigAooooAKSiigBaSiigAooooAKKKSmAtJmiigAooooAKKKSgQvakoooAKKKKACilpM0AGaKKKYwooooAKKKKACikooAKKKKACrdjc+W/luflPQ+lVKKQFvVbXI+0Rj/AHgKy63bSUXFuUbkgYNY9zCYJ2j7Z4poCKiiimAUUUUAApaKKYBRRRQIKKKKACiiigAooooAKKKKACiiigClRRRXMAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAB4Ga7Xw/biz0cSMMFuT9K4xV3yIn94iu6viLfRtq8fuxSYHN3EpmuZJDzk4qOgdM0UAFFFFABSUtFABRRRQAUlLRQAUUUUAFFFFABRSZooAKKKKYC0lFFABRRRSAKKKKYBRRRSAKKSigAooopgFFFFABRRRQAUZpKKACiiigAooooAKKKKACiiigAooooAKKKSgAopaTNABmiiigApKKWgApM0UUAFJS0UwCijNJQAZpaSloAKKKSgAooooAKKKKBBRSUUDFpKKKACiiigAooooAKKKKACiikpgPhcxzIw9ea0tTjE1lvHUc1lHoa2YP3unkf7OKQHO9RmilI2sV9DikqgCiiigApe9JRQApooooAKKKKACiiigAq/oVr9q1WNWGVXk1Qro/B8QLTzEcg8UmBc8QT4RIFP3utYdXdYk8zUZB2XpVOpAKKKKYBRRRQAUUlFIAoo70UwCiiigAooooAKKKKACiikoAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUlABRRRTAKKKKACiiigAozRSUAGaKKKACiiigApKKKACiiigAooooAKKKKACiikoAWkpaKACiikzTAKM0fWk7UAFFFFABRRS0AFJRRQAUUUUAFFJRQAZooooAKKKKACiiimAUUUUAFPjcxyq47GmUh6UAauox+dZbx1UZFYfat+3Pm2GD6VgHhmHoaSAKKKKYBRRRTGGaWkpaQhKKKKBhRRRTEFFFFAAFLkIOrHAruEC2GiqBwdn61yOmR+bqUC+jZrqfEEmy1SMd2xUSAwCSxLHq3JpKX2ooGJS0UUAFFFJQAUUUUwCiiigAooooAKKKSgQuaSiloGFFJRQAtFFFABRSUtAgo7UZpKBi0UUUCCk+lFFABRRRQAUUUUAFFFFMYUUUlAB9KKKKACiiigAooooAKKKKACikooAKMUUUAFFFFMAooooAKKSigAzRRRQAUUUUCCkzRRQAtFJS0AFFHekoAKKKKBhRRRTAKKKSgBaSiigAoFFFABRRRQAYooozQAUUUlAFmxk8u4APQ0/WYuElHbg1UU4dT6GtW9XzbAn8aXUDBopAeKWqAKKKBQAtJS0UxBRRRQAUUUUAFFFFABRRRQAUUCigAooooApUUUVzAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAE1mM38A9Wrs9eONNQeorjbD/kI2/+9XY+If8AkHx0mBzY6CigdBRQAUUUUAFFFFABRRRQAUUUlAC0UlFABmiiigAooooAKKTNLQAZopKKAFopKWgAopKKACiiimAUUUUAFFFFABRRSUALRSUUAFFFFABRRRQAUUUUAFFFFABRSUtABRSUUAFFFFABRRRQAUUlFABmiiigAoopKYC5pKKKACiiigAooooAKKKKACiiigApKWkoAKKKKACiiigAooooAKKKKYB2pOaKKACiiikAdjWvpnNoR71kVraT/wAezfWgDEnGLmUf7VR1Lc/8fUv+9UVUgCiiigAoopaAAUUUUAFFFFABRRRQAdjXW+EVAsJm9a5Lsa67wn/yDZaTAy75t1/MfeoKlu/+Pyb61FSAKKKKAFooopAFJRRQAUUUZpgFFGaM0AFFFJQAUUUUAFFFFMAooooAKKKKACiiigAooooASloopgFFFFIAooooASiiigAooopgFFFFABRRRQAUlFFABRRRQAUZopKACiiloASiiigAooooAKKKKACiiimAUUUUAFJRRQAZNJS0lAC0lLSUALRSUUAFFFHagAoopKAFpKKKACiiigAooopgFFFFABRRQKAFpKWkNABQaKKANXTDm1x9axZxtuJB71s6V/x7n8ayLr/j7k+tJARUUUUwCiiigYUUUUAFLSUtAhKKKKYBRRRQBpeHl3auvsM1teI2/eovbrWP4b/5C4/3a1vEf/H0n0qHuBk96KO9FAwoozSUAFFFFMAooooAKKKSgAooooAKKKKACiiigAooooAWkoozQAtJRS5oASlpKKACiiigQUUUUAFFFFMAoopM0DFpKKKACiiigAooooAKKSigBaKSigAooooAKKKKACiiimAUUUlAC0lFFABRRRQAUlFFAgooooAKKKKAFpKKKBiUtFFABRRRmmIKSiigYUUUUAFFFFABRR2ooAKKSigBc0lFFABRRRQAVsJ82n8/3TWOa2IP+PD/AICaTA54UtJ3P1paoApaQUtABRRRTEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUDKVFFFcwgooooAKKKKACiiigAooooAKKKKACiiigAooooAnsP8AkI2/+9XZeIf+QfHXG2P/ACEbf/ersfEP/IPjpMDmuwpaOwooAKKKKACiiigApKKKACiiigAooopgFFJmikAUUUUAFFFFABRRRQAUZoopgFFFFABRRRQAc0UUUAJRS0UAJRRRQAUUUUAFFFFABRRRQAUUUUAFJRRQAUUUUAFHeiigAopM0UALmkoooAKKKKACkopaYCZopaKAEoopaAEoooNABRRRQAUUUUAJRRRQAUZoooAM0UYooAKKKKYBQaKSgAoopaAEoo70UAFFLSUAHY1raV/x6n/erJrW0n/j1P1pMDFuf+PqX/eqKpbn/j6l/wB6oqoAooooAKKKKAFoo7UUAFFFFABRRRQAdjXXeEv+QdLXInoa67wl/wAg6WkwMm7/AOPyb61FU13/AMfk31qGpAKKKKAClpKKACiiimAlLRSUAFLSUUAFFHFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKSgAooopgFFFFIAooopgFFJRQAUUUUAFGaDSUAGaKKWgBKKWkoAKKKKACiiigAopKKAFooooAKKSigAzRRRQAUUUUwDNJRRQAUUUUAFFFFABSZoooAKKKKACiiigAooopgFFFFABRSUUALRSUtAC0neiigAoNFJQBq6T/x7n8ayLr/j6k+ta2k/8e5/Gsm6/wCPqT60luBFRRRTAKKKKBhRRRQAtFHekoEFFFFMYUUUUCNTw5/yGB9K1vEn/H0n0rI8Of8AIYX6Vr+I/wDj6T6VD3AyaDRmkoGFFFFMAooooAKKKSgAooooAKKKKACiiigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUGiimMSjNFJQAuaSiigApaSlzQAUUUUAFFFJQAUUUtAhKKWigYlLxSUUAFFFJQAtJRRTAKKKKACkoooEFFFFABRRRQAUUUUAFFFFABRRRQAUlFFABRRRTGFFFFABRRRQAUUUlABmiiigAooooEFFFFAwooooADWxB/x4f8ANYxrYg/48P+AmkwMDufrRSdz9aWqAKKWimAUUUUCCiiigAooooAKKKKACiiigAooooASilooApUUUVzAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAE9h/yEbf/ersvEP/ACDo642w/wCQjb/71dj4h/5B8dJgc32FFIOlLQAUUhNGaADNFFFABRRRQAUUmaKAFpKKWgBKWkooAOKKKKACiiimAlLRRQAUUUlAhaKSkz3NAx3Wius8MeDf7Vt1vNQZkgb7qDgmuk/4QHRD/DN/31U8yCx5filxXpzfD/RT084f8CqCT4daaf8AVyyD6mjmQWPNyKK72T4cr/yxu9v1FU5vh5qK5MN7E3tijmQHHUlb9z4O1yDOIDN/u1k3Gnajak/aLGWMDuRTugK1FM3rnBJB9MU7tTAKWkooAWikpaACkpaSgAooooAKSlooASiikNABRRR70ALRTQ4LYXLH0Aq7b6Vqd1/x76fK4PcCgCpS4roLbwZrlx96LyR/tCtOH4c37f6++iHsBS5kBxmDRg16BF8N4P8AltdFvpVpPhzpQ+/JKfxpcyA81xRjNeoD4e6IOvnH/gVKfh/ohHSYf8Co50B5dikrq/E3gyXSYGu7FjLbqfmT+Ie9clkEZHQ1SdwFzRSUUwFpKKKAFpKKKACiiigAooopgFFFFABRRRQAlFFFIAooopgFFGaKACiiigArW0n/AI9W+tZNa2k/8ezfWkwMW5/4+pf96oqluf8Aj6l/3qiqgCiiloASiilAoAKKKKACiiigAooooAOxrr/Cf/INlrkPWuu8J/8AINlpMDKu/wDj8m+tQ1Ld/wDH5N9aiqQClpKM0AFFJRTAWikpaACiikoAKKKKYBRRRQAUUUUAFFFFABRRRQAUUlFAC1e0zRtQ1Yn7HCfLBwZCOBWe3QD1OK9l8N28dtodrHGoUbMnHepk7AeY6p4b1TS4vOuI/MiHV1HArJHrXuN7Ek1nNHIoZWQ8H6V4jNGIbmaIHhGIFKMrgNoopKsBaKSigBaSiigApKKKYBS0UlABRRRQAZooooAKKKSgBc0lFFABRRRQAUUUlAAaKM0UAFFFFABRRRQAUUUUAFFFJTAM0UUUAFFFFABRSUUAFFFFABRRRQAUUUUAFGaM0UwCiikoAWiiikAlLRRQAUUUlMAooooAKKKKANXSf+Pc1kXX/H3JWvpP/HuayLr/AI+5PrSQEVFFFMAopaSgAooooAWkpaSgAooopgFFFFAGn4c/5DA/3a1/Ef8Ax9J9KyPDv/IXH0rX8R/8fSfSoe4GRRR3opjCiiigQZpKWigBKKKWgYlFFFABRRRQAUUUUhBRRRTAKKKM0AFFFFABRRRQAUUUmaAFpM0UUxhRRRQAlFOpp6ZPSgAzR9KsWNhe6i+2xtnl9WA4FdVp3gCWTa+p3I2nqicEVLkkBxRkUHBJJ9MVat7DULr/AI9rGWQHuBXqmn+GtKsB+5tlc+sgzWrHFHEMRxqg/wBkYqHU7AeUweE9cmxmDyv94Vfj8B6uwy1zCv4V6UCfWlzUe0YHnP8Awr/Uf+f2H8qjk8A6qvK3ULfhXpdFHOwPJ5/CGtw5xF5v+6Kzp9L1K2/4+LCVAO5Fe00140kGHRWHuM01UYHhm5QcEkN6EUpFewX2gaZfIVmtUGe6DBrl9S8AY3PpdxtHXY/OatVEBw1FW7/Tr3TXK3tu8Y7ORwaqdsitE0wCkozRTAKKKSgBaSiigQUUUUAFFFFABRRRQAUUUUAFFJRQAUUUUDCiijNMAooooASlopKAFpKKKACiiigQUUUUwCiiigYUUlFIAooooADWxB/x4f8AATWP2rYg/wCPD/gJpMDA7n60Udz9aBVALRRRTEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAJRRRQBTooormAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAJ7D/AJCNv/vV2PiH/kHx1x1j/wAhG3/3q7HxD/yDo6TA5rsKKOwooAKKKKACikozQAtJRRQAUUtJQAUUUUAFFFFABRRRTAKKSloAKKKKAEooooAShRuljTsXH86U0Rf8fEX++P50Ae52UaxWUCIAAI14H0qwKitv+PaH/rmv8qj1KR4tPnkjOGVCQaxGWqK8di8W6+oY/akOGI5FWE8ba+nWWJvqtVysLnrXHpRXmEXxA1VP9bDG/wBFq9B8SAMC40+Q+60uVgeg0ySKOQfvI0f/AHhmuVtfH2kzECVXh/3q2LXxBpN3jyL6Nie2aLAJe+H9KvQRNaIM90GK56++HdhJlrCV4W/2zkV2auHGUII9QadwaLsDyXUfBes2JJjT7Uo/uCufmSSB/LuImicdiK97xVO90mxv4ylzbRsD3C8/nVKQrHiGO9Fd7q/w9UbpdIm2d/LfnNcVfWN3p03lX1u8LdiR1qk0wK9JSnim5pgLRSZozQAtFAPGe1W9O0y+1STZYW7yDu+OBRcCpjjNLDHLcSeXbRPK/oBXoOkfD2FMS6tMZm67U4xXX2emWVlGEtraNQO+0Z/OpcgPLtO8E6zekGVRaKf74rptP+HenxYa+leZ/wDZOBXbYo4qeZjM2z0LTLNQsNpHx3ZQTWhHGkYwiKv+6MUkkiRqWkZVUdyay7rxJo9rnzb+IEdhU6sDYorjLr4h6XASIYZJvdazJ/iU7f8AHvp7r/vU+Vgei8UteWP8QdYf/VxxJ9Vqu3jfxA3SaFfotPlYrnrdFePHxfr7yRA3KYLgHA7V6zZSNJYwu5yzKCaTVhkepqH0y7UjIMTfyrwoDaWUdFJAr3bUf+Qddf8AXJv5V4Sf9ZJ/vGrgDFooFFWIKKKKACiiigAooopgFFJRQAtFJS0gCikopgLRSUUABooooEFFFFABRRRQAVraT/x6t/vVk1raT/x7N9aTGYtz/wAfUv8AvVFUtz/x9S/71RVQBRS0UAJS0UUAFFFFABRRRQAUUtJQAdq67wl/yDZa5Guv8J/8g2WpYGRd/wDH5N9aizUt3/x+y/WoaQC0lFFMApetJRQAUUUUAFFFFMAooooAKKKKACiiigAooooAKSlpKACiiiiwCH+D/fFe2aP/AMgm2/3BXiZ/g/3xXtejf8gm2/3BWcwLU/8AqJP9w/yrxG8/5CN1/v17fP8A6iT/AHT/ACrxC+/5CV1/v0QGRUlFJWghaKKSgBaKKKYBRSUUAFFFFABRRSUALRRSUALSUUUAFFFFABRRSUAFFFJQAUZAGTSMdqk+ld34S8H29xZx6hqYLmQbo0HYe9JuwHB+bH/eP5UCWP8AvH8q9oXw/pQ/5c4/++ad/YGlf8+cf/fNR7QdjxYSR/3v0pd8f94/lXtP9gaV/wA+cf8A3zR/YGlf8+cf/fNHtBWPFd8f94/lR5if3v0r2r+wNK/584/++aR/D2lOpU2ceD6Cj2gWPF+1Fdj4y8KRabCL7TgRFnEiHnHvXHZHUdDVp3GFFFJVCFpKKKACiiigAooooAKKKKACikpaYBRSUUgClFFFABRRSUALRSUtACUUUUwCiiigAoxRRQBq6T/x7msi6/4+5K1tK/49zWTdf8fUn1pICKiiimAtJRRQAtFFJQAtFHWigBKKKKYBRRRQM0/Dv/IXH0rW8R/8fSfSsnw7/wAhcfStbxH/AMfSfSoe4jJPWig9aKYBRSUUDCiiigAooooAKKKKQBRRRQIKKKM0wCiikoAWikpaACiikoAWkoopjCiiigAooooAKWgdau6TpV1rF19ns1OB9+QjhRSbsBViilnlWGCNpJW6KK7TQ/A24Lcau2e4iHGPrXSaF4es9GhAjQPMfvSHr+FbFYynfYCC3tILWMR28SRqP7oxUuKJJEiQvIwVV5JPauU1nxxY2ZaKyU3MnTcvQGoSbA6rNUrvV9OtM/abyOMjsTXl2peJdX1Jj5tx5cfYR8GshtznMjvIf9s5rRUwPULnxvo8JISQzf7tUX+Idip+SymP4155tUdFX8qWq9mgO/HxFtv+fCb86lT4h2J+/ZTD8a89BNGaPZoD1C28b6RMQJHMP+9Wxa6xp15j7NeROT2BrxfCnqoP4Uq5Q5jd4z/sHFL2YHuuaK8f0/xJq2nEeTceYndZOa7DSPHVndFYr9DbydN56GocGgOqubeC5jKTxJIpGPmGa4vXfAqPun0h9j9TG3OfpXaxSxzRiSJw6HkMD1p9Sm0B4dcQTWs7QXUTRSr1Vqjr1/XNBs9ZtzHcIFk/hkHUH3ry7WdIu9FujBdKShPySAcGt4zuBQpaTHNLWggooooAKKKKACiiigAopKM0ALRSUUwClpKKQBRSUtMYUUUUAGaSiigAooooAKKKKBBRRRQMKTNFFAhaKSigAopaSgBaKSigAPStiD/jw/4Caxz0rYg/48P+AmkxmB3P1paQ9T9aWqAKKKKYgooooASloooAKKKKACiiigAooooAKKKKACiiigClRS0lcwBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBPY/8hC3/AN6ux8Q/8g+OuOsf+Qjb/wC9XY+If+QfHSYHNdhRR2pKAFopKKACjFFFABRS0lABRRSUALRRSUwFpKKWgAopKKQBRS0UwCiilxQAlGKuabpt5qlx5FjCXYfebstdVB8OrlkBn1BAx7BelK6QHEkURj/SIv8AfH867z/hXC974flQPhwFkVxfD5SD0pcyCx3Nt/x7Q/8AXNf5VDqv/ILuf9w1YjTy40T+6oX8qrar/wAgy4/3DWYzxBTw3++aXNMX+L/fNLmthDs0mTSUZoACAeqg/WkAwcoSn+7xS0UAW7XVNRs2DW95Jx0DNkV0Wn+P9SgIF7Gs6j+4MGuSxS5pWQHrGl+NNJv8LJJ9mkP8L10UciSoHjYMp5BBrwUgHnAz6960NN1rUtLkDWly2O6yHIqXEZ7ZVW+sLW/haK6hWRWGORyK5XRPHtrclYdTX7PJ08w/dNdhHLHNGJInDowyCD1qbWA828Q+BZ7MPc6QTLD1MP8AEK4xgyuUdSrqcFT2r301zHinwlb6xEZ7VVhvFHBAwGqlIR5TSqGeRY40LyMcKo71bh0jUZtTOmpbN9pBw3HAHrXpvhnwnaaLGJZQs14w+ZyOB9KpsDn/AA54DaYJda0SFPKwDgj616Ba2tvaRCK2iSNAMfKMVJ3pk9xFbwtLPIqRqMlielZt3GTUyWWOFC8rhFHUk1xGt/ECCEtDpUfnP083+GuK1HWNR1OQveXTnP8AChwKai2B6TqnjbSbHKxP9qcdVjrk9Q8fapclhZIluh/vDmuTwByAM+tKST1qlFCJ7rUL+7Ytc3kpJ7K2BVQqD975j6tzTqDVWAQADooH0pdxHekopgLmjNJRQA9T+9h/66Cvc9O/5B1v/uCvCV/10P8A10Fe66b/AMg63/3BWcxoNR/5B11/1yb+VeFf8tJP94171cxefbSw5x5iFc+ma4QfDZdzE3w5OelKLsBwOKMV6B/wrZMf8fv6VWufhxcpGWtr9Cw/hK9avnQrHEUVa1DT7vTLk299C0T9s/xCqvSqAKQ0tJQAUUlFABS0UUAFJRRmgQUUUUxhRRRQAUUUUCCiiigYUUUUAHatbSf+PVvrWTWtpP8Ax6n60mBi3X/H1L/vVFUt1/x9S/71RVQBS0lFAC0UCigAooooAKKKWgAoopKADtXX+Ev+QbLXIGuv8J/8g2WkwMi8/wCP2b61DUt3/wAfs31qKkAUUUUAFFFFABRRRQAUYpKKYC0daTNFAC0UlFAC0UlFAC0lFFABS0lFABS0lFAA/wDD/vivatF/5BFr/uCvFG/h/wB8V7Von/IItf8AcFZzAuTf6mT/AHTXiF9/yE7v/fr2+b/Uyf7prxG//wCQnd/79EAIKKKStQFpKKKACjNFFABRRSUAFFFFABS0lFABRRRQAUUUUAFBopKBC0lFFAwoopKAGv8Awf7wr23RABotmB0EYrxNhyn+8K9u0Uf8Se0/65is6g0XaZ50Q/5aJ/30KS4yLeUj+4a8Ouprj+0Lv/Spx+9P8ZqFG4Hufnxf89E/76FHnR/89E/76FeEedcf8/U//fdL59x/z93H/fZqvZge8CRCcB1/OnV4poM9wdfsAbmcgyjIL8Gvau5qGrAY/i0A+Gr3P/POvG0/1a/SvZPFv/ItXv8A1zrxuP8A1a/StaYhaSlpK0AKKKKACiiigAopKKAClpKKACilpKACj6UUUwCiiigAoopKQC0UUUAFFFFMAooooAKKKKANTSv+Pc1k3X/H1JWtpX/Huaybr/j7k+tJARUtJRTAWkoooAWikpaACiiigBKKWkpjCiiigDT8O/8AIXH0rW8R/wDH0n0rJ8O/8hcf7ta3iP8A4+k+lQ9xGTSUtFAxKKKKACiiimAUUUUAFFFJQIWjNJRQAtFJRQAtJRRQAUUUUAFFFFMAooopDCkoozTAM0ZzSVc0jTJ9Yv1s7cEDP7x+yik3YCxoOjXGuXnlQgrAp/eSdsV6tpmm22mWi29rGFVRye5NN0vTrfS7JLW1QKqjk92NXQeK55SbAWsfXfENlo0R85t82PliU8mszxV4sj0wNaWJEl2RyeoT615xNNLcTNNcSGSVjkknNVGFwNLWNf1DWJCZ5THD/CiHH51lgAdBikzRWySQARTcVYtLW5vpRHZwPM/oBXTaf4CvrhQ95cLCv9zHNJySA5A4HU03zEHc/lXqFr4E0eEAyo8j+u7itODw7pUH3LRf+BDNT7RAeOh1/wBr8qfXs02kafNEYntItpGDhcGvMfFWkLouq+VFnyJRujB7CnGdwMiim0tWAGmnnqM0tFAGlo+vaho8u63lLxfxRuc5HtXpOgeI7LWogI2EdwB80THmvI/pToZJIJlmt5DHKhyGHH51EoJge6daqalpttqdo1tdxhlYcHuPpXP+FPFcepqtpfER3ajAJ6PXW1g7pgeOa/oVxoV35coLW7H93J2+lZle1anp1vqdk9rcoGRhwe4PrXkesaXPo9+1pODt6xv6it4TvowKNJS0VoISijNJQAUUUUDCiiigQUUlFABS0lFMYUUUUAFFFFACUtFJQIWikooAWikooGGaKKKBBRRRQAUUUUDCg0UUAFFFFMANbEH/AB4f8BNY56Vrwf8AHj/wE0mBg9z9aWk7n60tMAoopKYhaKKKAEpaKKAEopaKACiiigAooooAKWkooAKKKKAKdJS0lcwBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBPY/8hG3/wB6uw8Qn/iXxVx9h/yEbf8A3q7DxD/yD46TA5r0ooooAKKKKACiiigAoopKYC0UUUAJS0lFABRRRQAUUUUAFLSUUALS4ZiFTlmOBSVNZf8AIRtc9N4pAeu+GNJh0nSYo41G91DO3ck1sVHb/wDHvF/uD+VQarcvaaZcXEYy8aZFZDLfFHFeWjx/qxz+6Trj7tKvj7VjIi+WnzMB92nysD1Cqeq/8gy5/wBw/wAqswuZII3PVlBP5VW1X/kGXP8AuGkB4an8f++adTU6P/vmlrYQUtFFABRRRQAUUUUALSUUlAA2D1Ga1tD8R6hocoMMjSwE/NGxz+VZJpKGgPZ9C16y1u38y1cCQD54ieVrVrwmzuriwukubKQxyoc8dD9a9a8KeIE1/T/NK7J4/lkX1PrWbVhmyIYxMZREgkIwXA5NP6UtYXirxBHoNh5mN1xJ8sS+/rSAl8QeIbLQrfdcMHlI+SIHk15ZrWv6hrc5e5kKRfwxKcYHvVG7uri+unuryQySuc89B9Kiq1EQDgYAx9KcDTaWqAWikozQAUUUUwA0lFFACUUtJQAL/rof+ugr3XTf+Qdb/wC4K8LX/Ww/9dBXuum/8g63/wBwVEhosilqK7lMFpNKoyY0LD8BXmv/AAsHVCzfuVwDjpUJXA9P4pcV5cfiDqmR+6Tk/wB2vSNPna5sLedxhpEDGhqwGX4t0mHVNFnDoDLEpeNu+RXjozjDfeXhvrXvF7zZzf7hrwmf/j6ucf8APU1cBMZSUUVYC0lLSZoAWkoopiCiiigYUUUUAFFFFABRRRQAUUUUAFFFFABWtpP/AB7N9aya1tJ/49W+tIDFuv8Aj7l/3qiqW6/4+5f96oqoBaKSloABRRRQAUUUCgAopaSkAtFFFMBD3rr/AAn/AMg2WuQ7Guu8J/8AINlpMDIu/wDj9m+tRVLd/wDH7N9aipAFFFJQAtJS0lABRRRTAKKKKQBRRRTAKM0UUAFFFFMAooooAKWkopAFFFFACN/D/vivatE/5A9r/uCvFW/h/wB4V7Von/IHtf8AcFZ1ARcm/wBTJ/umvEL/AP5Cd3/v17fN/qZP9014hf8A/ITu/wDfogBBRRScVqAUUUUAFLSUUAFFFFABRRRmgAozRRQAUUUUAFFJRQAUZopKAFzRRSUALRSUtMAPVP8AeFe26L/yB7T/AK5ivEW/g/3hXt2if8ga0/65isagyzc/8e8v+6a8LvRjUbv/AK6mvdZ/9RJ/umvDL3/kI3n/AF1NFMCGilpK1EX9A/5GGw/66ivbe5rxPQP+RhsP+uor2z1rKpuMxvF3/ItXv+5XjUf+rX6V7N4u/wCRavf+udeMp/q1+lOmIdRRRWoBRmkooAKKKKACiiigApaSimAtJRRmgAopKKAFopKWkAlLRRQAUUUUwCiiigAooooGFFFFAjU0r/UGsm6/4+5K1tK/49zWTdf8fclJAR0ZpKKYC0lFFABRRRTAKWkpaAEooooGFFFFAGn4d/5C4/3a1vEf/H0n0rJ8O/8AIXH+7Wt4i/4+k+lQ9xGTSUppKBhRRRTAKKKKBBRRSUALRSUUAFFFFABRRRQAUUUUDCiiimAUUUlAC0lLRQISg0tLjNAxqq8jpHEpaSQ4UDvXq/hbQ49G01VwDcSDdI39K5bwBowurptUuEzHGcRZ/vV6NisKkugDelcr4x8SjS4TZ2bA3cg5I/gHrWt4j1iPRtMe4JBlIxEvqa8hmnluZ3uJ2LSyHJJ7e1EI3ARmZnZ3Ys7HLMeppM02nRpJLKkUSF5HOFUd622AcDlgACWPAA7112geCp7sLcarmKHqIu5ra8KeE4tORbu/USXbDIB6JXW4rKU+wFWysLWwhWK1hVFUdcc/nVqkZgqksQAOpNcxq3jXTrFjHbA3Uo6hO1ZpNgdRTWYDqQPqa8uvvGms3JIhZIIz2xzWJPf307FpbuUk+jYq1TYHs01zBEheWZFQdTury7xnq8Orasn2Y7ooBtDf3qwmkmYYe4mYehemYx0rSMLagLRRRVgLSUtFACUtFLQAqlldXRirqcqw6ivS/B3iUanCLO7IW7jHf+IV5nUsE8trcR3NuxWaM5UjvUzjdAe41ieKNFTWdNZAv+kR/NE3vU3h3WYtZ01LhDiQfLIvoa1a59UwPCpEaOR4pAQ6Haw96aTXYfEHR/s10mp26/JL8sgA6e9cb9K6Yu6ELRSUZqhi0ZpKKACiiimAUUUUAGaM0lFAC0UlLQAUlFFABRRRQIKKKKBhRRRQIKKKKACiiigYUUlFAhaSiigBaKSlpjEPStiD/jx/4CayK14P+PD/AICaTAwe5+tFHc0tUAUUlLQAUUUUCCiiigAooooAKKKSgBaKKKACiiigAooopgU6SlpK5QCiiigAooooAKKKKACiiigAooooAKKKKACiiigCew/5CNv/AL1dh4i/5B8VcfYf8hG3/wB6uw8Rf8g+KkwOaoo5wKKACiiigAooooAKKKTNAC0lFFABRRRTAKKSl7UAFFFFABRRRQAVLaf8hC1/66CoqltP+P8Atf8AfFJge5W3/HtF/uD+VU9f/wCQHef9czVy2/49ov8AcH8qqa9/yBLz/rmayGeJqeG/3qch/fxf74/nUa/xf7xp6f6+L/fH8616CPdrX/j2h/65r/KodV/5Btx/uH+VTWv/AB6w/wDXNf5VDqn/ACDbj/cP8qyGeGD+P/falpF/j/32pa2EFFFFABRRRQAUUnWigBaSiigAoopM0AKOorvfhd/q7r61wQ6iu9+F33Lr61MthnoHavOPih/x9Wlej9q85+J//HzaVEdwOHI5NFKeppK1EFGaKSgBaSiimAUUUUAFFFFABRRRQAq/62H/AK6CvddN/wCQdb/7grwpf9bD/wBdBXuumf8AINt/9wVnMaDUv+Qbdf8AXJv5GvC8nL8/xGvdNS/5Bt1/1yb+Rrwr+J/940QAVifl+te4aL/yBrP/AK5CvDT2+te46J/yBrP/AK5CnMRYvP8Ajzm/3DXhE/8Ax93X/XU17vef8ec3+4a8In/4/Lr/AK6mlAGMoopK0AWikopgLRRRmgAopKWgAopKKAF7UlLSUAFLSUUAFLSUUAFLSUtACVr6T/x7N9aye1aulf8AHs31pMDGuv8Aj6l/3qiqW5/4+pf96oqoApRSUtABRRRQAUUUUALRSUtABRSUUABrr/Cf/INlrkDXXeEv+QdLSYGRdj/TZvrUVS3n/H7N9ahpAFFFFABRRRQAUUUUwCiiigAopKWgAooooAKKKKACikpaACiiigApKWkoARui/wC8K9r0P/kD2v8AuCvFG6L/ALwr2vQ/+QNa/wC4KzqAXJv9S/8AumvENQ/5Cl3/AL9e3zf6mT/dNeH6h/yFLz/fopgQUUZpK1AWigUUAFFFFABRRRQAUUUUAFFFJQAUUUUAFFJRQAUtJRTAKKKWgApKKKQA38H+8K9u0X/kDWn/AFzFeIt/B/vCvbtF/wCQPaf9cxWVQZan/wBRJ/umvDL7/kJXn/XU17nP/qJP9014Xff8hK8/66mimBH2oo7UVsI0NA/5GGw/66ivbPWvE9A/5GKw/wCuor2zuaxqbgjF8Xf8ize/9c68aj/1a/SvZfF3/Is3v/XOvGU/1a/Sqp7AOooorQAooooAKKKKYBxRRSZoAWkoooAKKKKACiiikAUtJS0AFFFFABRSUUwFo7UlFAC0UUUAFGaKKBmppX+oNZN1/wAfcla2lf6g1k3X/H3JSQiKiiimAUUUUDCiiimAUUUUAFFFFABRRRQBp+Hf+QuPpWt4i/4+k+lZPh3/AJC4+laviI/6UlQ9xGSaKKKYwoopKBC0lLSUAFFFFABRRRQAUUUUDCiikzQAtFJRQIWkzRRTGFFFFABS0lGaAHU+GF7meK2j5eZtoqMcnFdP4C08Xesvcuv7uBcr9amTsgPQdJsU0/TobaNdu1Ru+verbEAEk4A70tc9401T+ztEcIcSz/Ih9K592Bwfi7V21bWGCE/Z4DtQf7Q71h+9OAOOep5P1pCK6UrIBCcDOMk8AV6R4K8NCwgGoXyA3Uoyqn+AVzvgbRBqWom9uFzb2/Kg9GNeoAYHt2rKcugC1T1PVLXS7Rri7kCKOg7mk1TUYNLsZLu5YKiDgepryTWdXudavGubkkJn93H2AqYxuBoa94nvdYkKKzQWuflQHBP41iDA6D8e9NzS1ukkApNJRRTAKSlpKACkpaSmAUUUtABRRmkoAWjNJRQBteFdYbR9XQlsW85CyDsPevW0cOiupyrDIPtXhDDcpX1r1HwJqx1HRVilbM0B2kf7PasKkeoG5q1jHqWnTWkmMSLjPpXi08L2tzLbSAho2KgH0r3TtXmXxD077Lq0d8i4S4G049aKctbAcpRRS10AFFFGaACikooAKKWigBKKWkoAKKM0UAFFFFABRRRQAUUUUCCiiigAoopKACiiigAopaKBiUtFJQAUtJS0xCGtiD/jx/4CaxzWxB/x4/8AATSYzBP3jRR/EfrS0xCUtJRTAWiiigYUlLRQIKKKSgBaKSloASloooAKKKKACijFFAyl3opaSuYQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAT2H/IRt/wDerr/EX/IPirkLH/kI2/8AvV1/iL/kHxUmBzQpaBRQAUUUmaACiiimAUUUUAFFFGaACikooAKKKKAClpKUUAFFFFABUtp/x/23++KhqW0/4/7X/fFJge523/HtF/uD+VVNe/5Al3/1zNW7b/j2i/3B/Kqmu/8AIFu/+uZrIZ4iv8X+8acn+ui/3x/Omr/F/vGnR/6+L/fH8616CPdrX/j1h/65r/KotU/5Blx/uGprX/j1h/65r/KodU/5Blx/uGshnhi/xf77UtIv8X++1LWwgopKKACiiigAooozQAUlFFABS0UUAKOoru/hd0uvrXCDqK7v4XfduvrUy2A9C7V5x8UP+Pm0+tejdq85+KH/AB8Wn1qY7jOHPU0UHqaStRBRRRQAUUUUAFFFJQAtFJS0AGaKSloAVf8AWw/9dBXuumf8g23/ANwV4Sv+th/66CvdtM/5Btv/ALgrOYw1L/kG3X/XJv5GvCf43/3jXu+pf8g66/65N/I14R/G/wDvGiACHt9a9y0P/kC2X/XIV4a3b617lof/ACBbL/rkKcxFi9/485v9w14Rcf8AH3df9dTXu95/x5zf7hrwi4/4/Ln/AK6mlAGR0UUVoAUUUUAFLSUUwCiiigAooooAKKKKACiiigAooooAKKKKADtWtpP/AB7N9ayq1dK/492+tJgY1z/x9S/71RVLc/8AH1L/AL1RVQBS0lLQAUUUUAFFFLQAlFFFABRS0lAAa63wn/yDpa5LtXW+Ez/xLpaTAyLv/j8m+tRVLd/8fkv1qKkAUUUZoAKKKKYBRRRQAUUUUAFFFFABRSUtABSUtJQAtFJRQAtJRmigAzRSUUwB+i/7wr2rQv8AkDWv+4K8Ubov+8K9r0L/AJA1r/uCsqg0XZv9S/8AumvD9Q/5Cl3/AL9e4Tf6mT/dNeH6h/yFLv8A36KYEFFFJWohaKSigAopaKACikooAWkoooAKKKMUAJRS4oI9x+dACUUhI9R+dJkeo/OgB1Gabkf3h+dGR/eH50AOzRTdw/vD86AR6j86AHUUmaMigAb+D/eFe36J/wAge0/65ivD2P3P98V7hon/ACB7T/rmKyqDLU/+ok/3TXhd9/yErz/rqa90n/1En+6a8Lv/APkJ3n/XU0UwI6KO1JWwjR8P/wDIxWH/AF1Fe2eteJeH/wDkYrD/AK6ivbe5rGpuMxfF/wDyLN7/ANc68Zj/ANWv0r2bxf8A8ize/wDXOvGY/wDVr9KqmIdRRRWgBRRRQAUlFFMAooooAKKKKACiiigAooooAKWkpaQBSUGimAUUUUALSUUUAFLSUtABRSUtAGppX+oNZN1/x9yVq6X/AKg1lXX/AB9SUkBFRRRTAKKKKBhRRRTAKKKKACiiigAooooA0/Dv/IXH0rV8Rf8AH0n0rK8Pf8hcfStTxF/x9L9Kh7iMo0lKaSmAtJRRQAUUUUAFFFFAwooozQIKKSigYtJRRQIKKWigYlFFFMAopKKACikpaAAttUse1eoeA7H7JoKuw+eVt2favL1QyyRwj/lowFe16bEINOtogMbYwKyqMC3Xl/j+/N1rYtlbMcCg/jXps0ixRPI3RBk14jfzm51G5mY53SHH0qaa1AhzzSqjyukUYy7kBRTa6DwTY/bfEKOVytt8xraTsgPRtB02PTNJgtYx0Xcfqa0GwASeg5p1YPjDUzpuhytG2JpBtSubdgcJ4y1ptW1Q28TZtbc4XHRjWBQBge55P1pa6YqyASloopgFLSZooAWkoopgFFFJQAUUUUAFFFFABSUtIaAAmug8DX5svEQiJwl0NmK56n285tryC5XgxODSkroD3Wud8cWIvPD0zgZeAblrdtJPOtIZeu9Af0ou4VuLSWFhw6kVzLRgeGKcqPpzS06VDFczxHqsjCm11LYAooozTAKKSigAooNFABRRRQAUUUUAFFFFABRRRQAUZopKACiiigAooooEFFFFABS0lFAwpaKSgAooozTEB6VsQf8AHj/wE1j9q2IP+PH/AICaTGYJ+8aKO5opiCiilpjCiiigQUUUUAFFFFABRSUtABRRRQMKKKKBBRSUUDKnekpaSuYQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAT2P/IQt/wDerr/EP/IPirkLH/kIW/8AvV13iH/kHxUmBzfaij0ooAKKKKACijNFMAopM0tABSUUUAFFFFABRRRQAtJRRQAUUUUAFS2n/H/a/wC+KiqW0/4/7X/roKTA9ztv+PaL/cH8qqa7/wAgW7/65mrdt/x7Rf7g/lVTXf8AkC3f/XM1kM8RH8X+8adH/r4v98fzpi9X/wB406P/AF8X++P51r0Ee8Wv/HtD/wBc1/lUOqf8gy4/3DUtr/x6w/8AXNf5VFqv/IMuP9w1kM8MB+//AL7UUi9H/wB80tbCCiijNABQTSUUAFFFFABS0lLQAUUlFACg8iu8+F33br61wa/eFd58Lul19aUtgPQe1ec/FD/j4tPrXo3avOfij/x8Wn1qI7gcOeppKD1NFagFFFFABRRmkoAPpRRRQAUtJRQAtJRRQIVf9bD/ANdBXu2mf8g23/3BXhK/62H/AK6CvdtM/wCQbbf7gqJjQal/yDbr/rk38jXhP8T/AO8a921L/kG3X/XJv5GvCf4n/wB40oDA9vrXuOif8gaz/wCuQrw49vrXuOif8gaz/wCuQpzEWLz/AI85v9w14Pcf8flz/wBdTXvF5/x5zf7hrwi4/wCPu6/66mlACOiiitACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALWrpP/AB7N9aya1tK/49m+tAGNc/8AH1L/AL1RVLc/8fUv+9UVMBaKKKACiiigAooooAKKKKACg0UUAHY11nhT/kHS1yeetdZ4U/5B8tJgZN3/AMfkv1qKpbv/AI/JfrUVIApKWkpgFLRRQAUUUd6ACiiigApKWigBKKWkoAKKTNFAC0lFFAC0lFFABRRRQAjdF/3hXtehf8ga1/3BXijdF/3hXtehf8ga1/3BWdQaLs3+pf8A3TXh+of8hW8/369wm/1L/wC6a8P1H/kK3n+/RTAr0UtJWogooooAM0UUUAFFFJQAuaKSigBaWkooAlt4Jbm4jt4F3SyHCivQtN+H9gkCnUWaaUjJwcY9q5jwQgfxPCSPudK9brKcncDmP+ED0D/n2b/vqkPgPQP+fZv++q6iio5mM5b/AIQLQf8An3b/AL6pP+EC0H/n3b/vquqoo5mByv8AwgWgj/l3b/vqqt/8PdMkhIsWaCXsxOa7PFJjmnzMDwm/s59PvpbO5XEsZwfeoK6v4kRqmvW7AYLqc+9cnW0XdAKf4P8AfFe46L/yB7T/AK5ivDT1T/fFe46J/wAge0/65is6gFuf/USf7prwu/8A+Qnef9dTXuk/+ok/3TXhd/8A8hO8/wCupopgRdqKKWthF/w//wAjFYf9dRXtvrXiWgf8jFp//XUV7b61jU3AxfF//Is3v/XOvGU/1a/SvZvF/wDyLN7/ANc68ZT/AFa/Sqp7AOoopK0AKKKKACiiimAUUUUAFFFFABRRRQAUUUCgBaKKKQBSUtJTAKKKKACiiigAoopaAEooooA1NL/1BrKuv+PqT61q6V/qDWVc/wDH1J9aSAiooopjCiiigAooopgFFFFABRRRQAUUUUCNPw9/yFh9K1PEX/H0v0rL8Pf8hYfStTxF/wAfSfSoe4GUaSiimMKKKKACiiigAopKKBBRRRQAUUUtACUUUUAFFJS0wCkooxQMKMUUUAJRS0lAi5osXn69YR4480Zr2gLjgdBXkXhJN/ia3/2SDXrx6msKj1GZfiaf7P4evZM4Pl8V40pJXJ7816p48lKeHJF/v8V5Uv3FHtVU9gHg16B8NLbFrcXZH+sO38q88zgE16t4Cj8rwxD6sxNOpsB01ecfEa887UbezB4h+Yj616NmvIfFk/n+JrpvQACs6a1AyDRSZoroAKSlpKACiiimAUUUUgCiiigAooopgFFFFACUUUUAJTJRmJhT6Q9DQwPYvClx9p8P2zk5wNv5VsVy3w8k3+F4wf4ZDXU1yS3A8Y8RwfZvEd3Fjj71Z1b3jpdniyYj+JBWBmumOwC0lGaCasAopKKAFopKWgAooooAKKKKACg0lFAC0lFFABRRRQIKKKKACiiigAooooGFLSUtABSUUUALSUUUxB2rYg/48f8AgJrHNa8H/Hj/AMBNJgYXc0Udz9aBTGLRRRTEFFFFABRRRQAUUUUAFFFFAwooooAKKKKBCUUtFAylRRRXMIKKKKACiiigAooooAKKKKACiiigAooooAKKKKBE9j/yELf/AHq67xD/AMg+KuRsf+Qhb/71dd4g/wCQfFSGc32ooooAKKKSmAppKKKACiiigBaSiigAzRRRQAUUUUAFFFFABSUUUALUtp/x/wBr/wBdBUVS2f8AyELX/rpSA9ztv+PaL/cH8qqa7/yBbv8A65mrdt/x7xf7g/lVTXf+QLd/9czWQzxBf4v9405P9fF/vj+dNX+P/eNOj/18X++P51r0Ee8Wv/HrD/1zX+VQ6p/yDLj/AHDUtr/x6w/9c1/lUWqf8gy4/wBw/wAqyGeFr/H/AL7UuaaP4/8AfalrYQUUUUAFFFKKAEpaSigAooooAKKDSUAKOorvPhb0uq4Qda7v4W/8vVTLYEehdq85+KP/AB8Wn1r0btXnHxR/4+LT61MdxnDnqaKD940laiFozSUtACUUUdqACiiigAopaSgQUUUUDFX/AFsX/XQV7rpf/INtv9wV4Uv+si/66CvddL/5Blt/uCs5ghdR/wCQddf9cm/ka8J/if8A3jXu+pf8g66/65N/I14R/E/+8aIDEPb617lof/IFsv8ArkK8Nbt9a9y0P/kC2X/XIU5iLF5/x6Tf7hrwi4/4/Lr/AK6mvd7z/jzm/wBw14Rcf8fd1/11NKAMjooorQAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAB2rW0r/AI9m+tZPatXSv+PZvrSYGPc/8fUv+9UVS3P/AB9S/wC9UdUAUUUUAFFLSUAFFFFABS0lFABRQaKADtXWeFP+QfLXJ+tdZ4U/5B8tJgZN3/x+S/Woqlu/+PyX61DQAtJS0UAJS0lFAC0lFFABRRR2oAKKKKACiiigAopKKACiiigAooooAKDSUUwEbov+8K9s0L/kDWv+4K8Tb+H/AHhXtmhf8ga1/wBwVlUAuzf6l/8AdNeH6h/yFbz/AH69wl/1L/7prw/UP+Qref79FMGV6KKK1ABRRRQAUlLSUAFFFFABRRRQAUoptKKAOi8C/wDIzx161XkvgX/kZ469aFYz3Gc345v7vT9Caeym8qQH71ee/wDCT69tGdQ6gdq7v4i/8i4/1ry7+EfQVUErAa3/AAk2vf8AQQ/Sj/hJte/6CH6Vk5oFXyoD0n4e6lfajDetf3HnFGAU46V2Zrgvhf8A8e9//viu8NYS3A8v+Jn/ACG7T/cNchXYfEz/AJDVp/uGuPraOwCHqn++K9y0T/kDWf8A1zFeGt/B/vivctE/5A1n/wBcxUVALc/+ok/3TXhd/wD8hS8/66mvdJ/9RJ/umvCr/wD5Cl7/ANdTRTAjooorURf0D/kYtP8A+uor271rxHQP+Ri0/wD66ivbh1NZVNwRi+L/APkWb3/crxmP/Vr9K9m8X/8AIs3v/XOvGU/1a/SqpgLRRRWgBRRRQAUUUUwCiiigAooooAKKKKAClopKAFooooAKKKKAEpaKSgAooooAKKKKACiiloA09K/1BrKuf+PqStTS/wDUGsu5/wCPqSkgIqKKKYBRRRQAUUUUxhRRRQAUUUUCCiiigDT8Pf8AIXH0rU8Rf8fSVl+Hv+QuPpWp4i/4+k+lS9wMk9aKKKBhRRRQAUlFFAgooooAKKKKACjNFFABRSA0UALSUUUwDNJS0lAC0UUUALSUZooA2/Bo/wCKmi+lesnqa8i8KPs8S2uf4mAr1/1rCpuM5T4gg/8ACP8A/Aq8uA+UfSvWvHMPmeG5yB9wZrydR8i/Srp7ANIypFeqeBLqK48ORIjDfGx3L3FeW4q3pupXuk3BmsJShb7y9jVSjdAe1swVSW4ABJJrxbV51uNYupkOVLEZq9qHizWdQgaB5fKiYYYDqaxAMcVMI23Admlpopa0AKKKKACiiigAooooAKSilpgJRRRmgQtFJmigYGiiimIMUmKWkPAJpDPTPhwMeGgf+mhrrK5vwBH5fhiL3cmukrlluB5R8QD/AMVU/wD1zFc7mt3xy4fxbPj+FBWFXRDYApaSiqAWikpaACiiimIKKSigYppKKKBBRRRQAUUUUAFFFFABRRRQMKKKKACiiigQUUUtAxKKWimAUUUUCENa8H/Hj/wE1kmtaD/jx/4CaTGYXc/WijuaKYhaKKKYCUtFFABRRRQAUUUUAFFFFAwooooAKKKKACiiigClRRRXMIKKKKACiiigAooooAKKKKBBRRRQAUUUUAFFFFAE9j/yELf/AHq67xD/AMg+KuRsf+Qhb/71db4h/wCQfFSYznOwpKPSimAUUUUAFFLSUAFFFFABRSUtABRSUUAFFFFAgooooGFFFFABUtn/AMhC1/66CoqltP8Aj/tf+ugpAe6W/wDx7Rf7g/lVTXf+QJd/9czVu2/49ov9wfyqpr3/ACBLz/rmayGeID+L/eNOj/18X++P50xf4v8AeNPj/wBfF/vj+da9BHvFr/x6w/8AXNf5VDqn/IMuP9w1Na/8esP/AFzX+VQ6p/yDbj/cP8qyGeFr/H/vt/Oihf4/99v50VsIKKKKACiiigAooNFABSUUUAFFFLQADqK7z4W/8vVcGOorvPhb0uqmWwHoXavOPij/AMfNp9a9H7V5x8Uf+Pm0+tTHcZw7dTSUHqaK1EFFFBoAKKKKBBiiiigYUUUUAFFFFACr/rIv+ugr3XSv+QZbf7grwpf9ZF/10Fe66V/yDLb/AHBWcwQ7Uv8AkHXX/XJv5GvCP4n/AN417vqX/IOuv+uTfyNeEH7z/wC8aIAxG6D617lof/IFsv8ArkK8NboPrXuWhf8AIEsv+uIpzAsXn/HnN/uGvB5/+Py6/wCupr3i8/485v8AcNeD3H/H5df9dTSgDGUUUVoAUUlLTAKKKKACiiigAooooEFFFFAwooooAKKKKACiiigArV0r/j2b61lVq6V/x7H60mBj3P8Ax9S/71R1Jc/8fMv+9UdUAUUUUAFFFFAC0UUlIApaSigAooopgHY11nhX/kHS1yddX4V/5B8tJgZV3/x+S/Woqku/+Pyb61FSQBS5pKKYgooooAKKKKBhRRSUALRSUUAFFHeigAooooAKOKM0lABRRRTAKKKKAEbov+8K9s0L/kDWv+4K8Tbov+8K9s0P/kDWv+4KyqAi7L/qn/3TXh2o/wDIWvP+ule4y/6p/wDdNeHaj/yFrz/rpRACvmiiitQAUtJRQAtJS0UAGKSlpKADNFFFACYpaKSgDofAp/4qiKvXBXkPgY/8VTDXrhPNYz3Gct8Rf+Rbf615aD8o+gr1D4in/imn+ory4fdX6CrhsAtFJS1Yj0D4X/8AHvf/AO+K7w1wfwu/497/AP3xXe1hLcZ5h8Tf+Q1af7hrjq7H4mf8hq0/3DXH1rHYBp/g/wB8V7lon/IGtP8ArmK8OPVP98V7lon/ACBrT/rmKioBan/1En+6a8Kv/wDkKXn/AF1Ne6z/AOok/wB014Xf/wDIUvP+upop7iIu1FFFbAX9A/5GLT/+uor271rxHQP+Rh0//rqK9t7msam4IxvF/wDyLN7/ANc68Yj/ANWv0r2fxf8A8ize/wDXOvGE/wBWv0qqewDqKKK0AKKKKACiiigBaSlopgJRS0UAFJS0UAJS0UUAFFFJSAKKWimAlFFFABRRRQAUUUUAFLSUtAGlpf8AqDWXc/8AH1JWppf+oNZdz/x9SUkBFRRRTAKKWigBKKKKYBRRRQAUUUUAFFFFAGn4e/5C4/3a0/EX/H0lZnh7/kLD6Vp+Iv8Aj5Soe4zKJpKWkpgFFFFAgoo70UwCiiikAUUUlAC0maKKAEopaKYBRRSUAFFFFABRRRQAUtJRQMu6RL5GtWMvYSjNe0IwdAw6MM14WWKASDqhyK9p0adbnSbWVTnMYz9cVjVQEfiCH7RoV5EBktGcV4wBtBU9VOK92dQ6FWHDDBrxXVbZrTVrqBhg+YWH0opvoBTxSYpaK2ASilooAKKKKACiijNMAoopKAFpKKKBBRRSUALmkoooGLSUUUCFopKKBi02TIiP4U6p7G3N3qNtaqMmV8UPYD13wzbfZtBtY8dVDfnWoe/0plvH5VvFH/cQL+QqO+nW2spp26IpNcnUDxvxDN9p8RXcuc87aoU6RjJPNIf45CabXUlZAFLSUtUAUUUlAgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAFLSUGmMWigUUgCiiimAHpWtB/x4/8AATWSela0H/Hj/wABNJiMLufrS0dz9aKYCUtFFMAooooAKKKKACiiigYUUUUAFFFFABRRRQAUUYopgUqKKK5RBRRRQAUUUUAFFFFAgooooAKKKKACiiigAooooAnsf+Qhb/71db4h/wCQfFXJWP8AyELf/errfEP/ACD4qTGc32FFFFMAooooAM0UUlIApaKSmAUUUZoAKKKKACikooAWikopALSUZopgFTWn/H/a/wDXQVDUtp/x/wBr/wBdBQwPdbb/AI9ov9wfyqpr3/IEvP8Armat23/HtF/uD+VVNe/5Al5/1zNYjPD1/i/3jTo/9fF/vj+dNH8X+8acn+vi/wB8fzrXoI95tf8Aj1h/65r/ACqHVP8AkG3H+4f5VLa/8esP/XNf5VFqn/INuP8AcP8AKsuozwpf4/8Afb+dLSD+P/fb+dLWogoopKYBRmlpKACiiimAUUUUALSUlFIBw6iu7+Fv/L1XBjqK7z4W/wDL1SlsB6H2rzj4o/8AHxafWvR+1ecfFH/j5tPrUR3GcMepopT1NJWogopKKAFpKKKAFooFFAgooooGFFFFACr/AK2L/roK910v/kGW3+4K8KX/AFsX/XQV7rpf/IMtv9wVEwHal/yDrr/rk38jXhB++/8AvGvd9S/5B11/1yb+Rrwg/ff/AHjSgDEboPrXuOhf8gSy/wCuIrw1jxXuWhf8gSy/64inMCzef8ec3+4a8Hn/AOPy6/66mveL3/jzm/3DXg8//H5df9dTSgMZRRRWggooopgFFFFABRRRQAUUUUAFFFFAgooooAKKKKBhRRmjNABWrpf/AB7n61lVqaX/AMezfWkwMi5/4+Zf96o6kuP+PmX/AHqjqgCiiigAopaKAEopaKQBSUtJTAKKKKACur8K/wDIPlrlPWur8K/8g+WkwMm7/wCPyX61FUt3/wAfkv1qKkIKKKSmMWikooAWkoooAKKKKACiiigAopKKACiiigAooopgHeiikoAKWkzS0ANb+H/eFe26F/yBrX/cFeJN0X/eFe26F/yBbT/cFZVARdm/1Mn+6a8N1H/kK3n+/XuU3+pf/dNeGaj/AMha8/36KYENApKUVqAoFBAHcVo6HpFxrd99mt/lUcvJ/dFdtF8PdPVAJLl3PripckgPOOP7w/Ojj+8Pzr0v/hX2l/8APR6X/hX+l/32pe0QHmfH94fnRx/eH516X/wr/Sv77Uv/AAr/AEv++1HtEFjzPj+8PzpMV6Yfh/pf/PR6wvEfgmTTbVrzT5WmjTl0PYe1CmmBx9FGcjNJVgdB4G/5GmGvWz1NeR+Bv+Rqhr1w9TWM9wOU+Iv/ACLT/UV5cPur9BXqPxF/5Fl/qK8uH3V+gq4bDClFJS1Yj0D4Xf8AHvf/AO+K76uB+Fv/AB73/wDviu+rCW4zzD4mf8hq0/3DXH12PxM/5DVp/uGuPrWOwhrdU/3xXuOif8ga0/65ivDm/g/3xXuOif8AIGtP+uYqKgy3P/qJP9014Vf/APITvf8Arqa91n/1En+6a8Kvv+Qnef8AXU0UwI6KSithGhoH/Ixaf/11Fe3V4joH/Iw6f/11Fe3d6xqbgjF8X/8AIs3v/XOvGE/1afSvZ/F//Is3v/XOvGE/1afSqp7AOooorQAooooAKKKKAFpKWimIKKKKACkpaKACkpaM0DEooooAWkopaACkpaSgAooooAWkoooAKWkpaANLTP8AUGsu5/4+nrU0z/UGsu5/4+npICKilpKYBRRRQAUUUUwCiiigAooooAKKKKANPw9/yFh9K0/EX/HylZnh/wD5Co+laXiL/j5Soe4GUaKDRTAKKKKACiikoAKKKKACikpaACikopiCiiigYUUUUAFFJRQAUtJRQAtFJRQMdgHg9DXo/wAO77z9Ha2dv3kTn8q83BxW/wCCtS+weIVRjiO5AT8aiaugPWK81+I1gYNSi1BV+WYbDj1r0j+VZPifSxqmizQBcyqN0f1rGLswPHu9LQEKko4+ZDtP1FLXSgCiikoADRRSUxBRRRQAUUUnvQAtJS0UDEopaKACkoopgFFJQaAClzSUUCHd8V1Hw+0/7Xrb3TD5LcZU+9coWIHyjLHgD1r13wbpX9laHEjD97L87H69qyqSshm/XNeO74Wnh+WMHD3HyrXS15j8Q9R+1avHZI2Utxu49axgrsDkhwqj2paD1orrAKKKSgQUUUUAFFLSUAFFFFABRRRQAUUUUAFFFFABRRRQMKKKKBBRRRQAUUUUwClpKUUDCiiikAUUUUwA1rQf8eP/AAGsntWtB/x5f8BNJiMPufrRR3P1opgFFFFMAooooAKKKKBhRRRQAUUUUAFFFFABRRRQIKKKKYylRRRXKIKKKKACiiigBaSiigQUUUUAFFFFABRRRQAUUUUAT2P/ACELf/errPEB/wCJfFXJ2P8AyELf/erq/EH/AB4RUmM53sKKOwooAKKKSmAUUtJQAUUUUAFFJS0AFJRRQAUUUUAFFFFABRSUUALUtp/yELX/AHxUVS2n/H/a/wDXQUmB7rbf8e0X+4P5VU17/kCXn/XM1btv+PaL/cH8qqa9/wAgS8/65mshnh6/xf7xpyf6+L/fH86av8X+8aVP9fF/vj+da9BHvVr/AMesP/XNf5VDqn/IMuP9w/yqW1/49Yf+ua/yqLVP+Qbcf7h/lWQzwpf4/wDfaihf4/8AfalrZCEooooAKKKKACiiigBKKDRTAKKSigBw6iu9+FvS6+tcCvUV33ws6XX1qZbAehdq84+KX/HxafWvR+1ecfFL/X2n1qI7jOGPU0lKeppK1EFFFFABRRRQAoooooEFFFFAwooooAVf9bF/10Fe66X/AMgy2/3BXhS/62H/AK6CvddL/wCQZbf7gqJgh2pf8g66/wCuTfyNeEH77/7xr3fUv+Qdc/8AXJv5GvCD/rJP940oDGt0/GvctC/5All/1xFeGt0/GvctC/5All/1xFOYizef8ec3+4a8HuP+Py6/66mveL3/AI85v9w14Pcf8fd1/wBdTSgDGUUUVoAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiimAVqaX/x7N9ay61NM/492+tJgZFx/wAfMn+9UdSXH/HzJ/vVHTAKWkooAKKKKAFooooASiiigAooooAPWur8LH/iXyVynauq8Lf8g+WkwMq7/wCPyX61DUt3/wAfkv1qKkIKKKKYwpKKKAFopKKAFopKKACiiigAoopKAFo5oopgFJRmigAooooAKKKWkA1ui/7wr23Qv+QLaf7grxJ+i/7wr23Qv+QNa/7grOYIuzf6l/8AdNeGal/yFrz/AH69zl/1T/7prwzUv+Qtef8AXSimMr0ucc0lIehrUR6V8MolXR7iXHzNL1rtM1x3w1/5AMv/AF1rscVzy3GGfalrzbxd4k1jT/EBtrO4KRBM4rIHjHX/APn6P5U1BgewfhSZryH/AITHX/8An6P5Uj+MdfVCRdHj2p8jA9gpkqLJGyOMqwIIqvpM0lxpVrNKcu8YZj71ZJqAPCLxBHqF2g4AmbFQ1Y1H/kKXn/XZqr10LYRv+Bf+Rqhr171ryLwN/wAjVBXruOaynuByfxF/5Fp/qK8uH3V+gr1L4jf8i0/1FeWj7q/QVcNhhQKKUVYjv/hb/qNQ/wB8V39cD8Lv9RqH++K76sJbjPMfiZ/yGrT/AHDXHV2HxM/5DVp/uGuPrWOwCHqn+8K9x0T/AJA1p/1zFeHH+D/fFe46J/yBrP8A65ioqAW5/wDUSf7prwq+/wCQnef9dTXutx/x7yf7prwm9/5Cd5/11NFMBlJRR2rYRoaB/wAjDp//AF1Fe3eteIaB/wAjDp//AF1Fe3d6xqbgjG8X/wDIs3v/AFzrxhP9Wn0r2fxf/wAize/9c68YT/Vp9KqmA6ijvRWgBRRRQAUUUUALSUUtMQUlLRQAnNHailoGJRS0lABRRRQAUUUUALSUUUAFFFFABRS0UAJS0UUAaWmf6g1l3P8Ax9SVqaZ/qTWXc/8AH09JARUUtJTAKKKKYBRRRQAUUUUAFFFFABS0lFAGl4f/AOQqPpWn4i/4+lrM0D/kKj6VpeIf+Plal7gZZooNJQAtJmiigAzRRRQAlLRSZoAKM0UUwEpaKKQBRRRmgAopKKYC0lFFABRSUUDCiiloAKAzoyyRnEiHKmiigD2Lw3qaaro8Nwp+YDaw9xWtXlfgjWf7M1QWszYt7g4Gein1r1MHI4rmkrMDzLx1ohsL8X9umLec4YD+FvWuWIr23UrGHUbGW0uBlJBj6V49q+mz6RfvZ3KkYOY27MO1a05X0ApUUE03NagLSE0ZooEGaKKWgYnelopM0ALRSUUAFFFFAgooopjEopfpRikAlJS4qexsp9RvY7K1UmSQ4JH8I9aG7CNrwTop1XVluZF/0a2O7PZj6V6uo2gADAHQVQ0TTIdI02O0hUDaMufVu9aBrlk7sZT1bUI9N02e7kIxGuQPWvFp53uriW4lJLSMWB9q6n4ga2Ly8XTbZ/3UPLkdz6VyNbU421AXNLTRS1qIWikooAKKWkoAKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFAgooooAKKKKYBS0lLSGFFLSUAFFFFMQHpWrB/wAeX/AayjWrB/x5f8BpMZh9z9aWk7n60tUIKKKKACiiigYUUUUAFFFFABRRRQAUUUUAFFHeigQUUlFMZTooorlEFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigCex/wCQhb/71dXr/wDx4R1yll/yEIP96uq8Qf8AHhFSYznR2oo7UUwCiikoAWikooAWkopaAEooooAKKKKACikooAM0UUUAFFFFIBaltP8AkIWv/XQVFUtr/wAf9r/10FAHutt/x7Rf7g/lVTXv+QLd/wDXM1btv+PaL/cH8qqa7/yBbv8A65mshnh6/wAf+8acn+vi/wB8fzpq/wAf+8acn+vi/wB8fzrXoI95tf8Aj1h/65r/ACqHVP8AkGXH+4f5VNa/8e0P/XNf5VDqn/IMuP8AcNZDPCl/j/32paRf4/8AfalNbCCkopKAFpKKDQAUGkpaYCUUUtABRRRQADqK734W9LquDHWu8+Fv/L1Uy2A9E7V5v8Uv9fafWvR+1ecfFL/X2n1qI7jOGPU0lB6mitRBRRRQAtJRRQAUtFJQAtFJRQAtFFJQAq/62H/roK920v8A5Blt/uCvCV/1sP8A10Fe7aX/AMgy2/3BUTBDtR/5B11/1yb+Rrwg/ff/AHjXu2pf8g26/wCuTfyNeE/xv/vGlAY1ug+te5aF/wAgSy/64ivDX6V7loP/ACBLL/riKcxFm8/485v9w14Pcf8AH5df9dTXvF5/x5zf7hrwe4/4/Lr/AK6mlAGR0tFFagFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAK09M/492+tZlaWm/wCob60mBlXH/HzJ/vVHT7j/AI+ZP96mUwCiiloASlopKAFpKWigApKKKACiiigArqvC/wDyD5K5Wuq8L/8AIPkpMDIu/wDj8l+tRVLd/wDH5L9aioAKKKKACiiigAooooAKKKSmAueaSiigApaSigAooooAKKKKACiiigAoopKABjwv+8K9t0L/AJAtp/uCvEW6L/vCvbtC/wCQLa/7grKoCLsv+qf/AHTXhmpf8he8/wCule5y/wCqf/dNeG6l/wAhe8/66UQGVqD0NLSHoa0Eem/DX/kAy/8AXWuxrj/ht/yAZf8ArrXYd6wluM8i8d/8jS3+5XPV0Pjz/kaT/uVztbR2EOBpsv8Aqm/CjNNlP7o02B7joR/4kdn/ANclq6ev4VR0D/kB2f8A1yFXz/SsOozwrUP+Qpef9dmqvVjUP+Qpef8AXZqgrdCN/wADf8jVBXrwryLwMP8AiqYK9eFZT3Gcn8Rv+Raf6ivLR90fQV6n8Rf+Raf615aOi/QVUNgEoFFFaCO/+Fv+o1D/AHxXf1wHwt/1Gof74rvqwluM8x+Jv/Ias/8AcNcdXY/E3/kNWf8AuGuOrWGwAeqf74r3HRP+QNaf9cxXhp6p/vivctE/5A1n/wBcxUVALc/+ok/3TXhN9/yE7z/rqa91uP8Aj3k/3TXhV9/yErz/AK6mimBH2ooorYRoaD/yMOn/APXUV7d614joH/Iw6f8A9dhXt3rWNTcEYvi//kWb3/rnXjCf6tfpXs/i/wD5Fm9/6514wn+rX6VVPYB1FFFaAFFFFABRRRQAUUtJTAKKKKAClFJS0AJRQaKACiiigAooooAKKKKACiiigBaSiigBaKSigDS03/UGsy5/4+nrT03/AFJrMuf+Pp6SAiooopgFFFFMYUUUUAFFFFAgooooAKKKKANLQP8AkKj6VpeIf+PlKzdA/wCQqPpWl4h/4+VqXuBld6KDRQAlLRSUAFFFFMAopKWkAUUUlABRRRTAKKKM0AFFJRQMWkoooAKKKKACiiigApelFJQAvUcHB7H0r03wT4hGpWf2O6YC7hGOf4l7V5jmpbW6nsruO6tWKzRnI96mcboD3PNY/iPQ4NbsTFIAsy8xv3BpvhvXoNcsVkQhZ1GJE759a2M1z7MDw69tLiwu3tbtCkqHv0YVBXr3iPw9ba5bYcbLhfuSDrXlepadd6VdNbX0ZRh0f+FvxreE7gVqMUClrQQlFLSUxhRRRSEFFJRQMWikooAXNFJS0wCloFT2Nnc6hcrbWURkkY4z2H1NJuwEcEEtzOlvbIXmc4AH9a9V8J+G4tEtN8gD3cg+dz29qXwx4Yt9EhEj4lu2HzyHt7CugrnnO4CEVzvjDXl0bTSsbD7VMNsYHUe9bl7M8FpLLFH5jouQnrXi2r391qWpS3F7uEuduw/wilCN2BSJZmZ3OXc7mPvRRS10gFLTadTEFLSUUhhRS0lMAooooEFFFFAwooooAKKKKACiiigAooooEFFFFABRRRTAKKKWgA+tFFFIYUUUUAFFFFMQGtWH/jy/4DWVWpD/AMef/AaTGYnc/WlpO5+tLVCCiiigYUUUZoAKKKKACiiigAooooAKKKKACijvRQIM0UlFMZTooorlEFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigCex/4/7f8A3q6rX/8AjwjrlbL/AI/4P96up1//AI8IqQznu1FJniimAZooooAKKKSgBaKKKACikooEGaKKKBhRRRQAUUUUAFFFFABUtp/yELX/AK6CoqltP+Qha/8AXQUgPdrb/j2i/wBwfyqprv8AyBbv/rmat23/AB7Rf7g/lVXXP+QLd/8AXM1kM8OXq/8AvGnx/wCvh/3x/Omr/H/vGnR/6+H/AHx/Otegj3i1/wCPaH/rmv8AKodU/wCQZcf7hqa1/wCPaL/rmv8AKodU/wCQZcf7hrIZ4UvR/wDfb+dFNU/f/wB9v50E1sIdSZpKKYBmiiigAooooAKWkxS4oAKKBS0AA6iu8+Fv/L1XBjqK7z4W/wDL1Uy2A9D7V5x8Uv8AX2n1r0ftXnPxS/19p9aiO4zhD1NFKeppK1EFFFJQAtFFGaAClpKWgApKWigBDRRiigAU/vYf+ugr3fS/+Qbbf7grwgf62H/roK920z/kGW3+4KiY0O1L/kG3X/XJv5GvCf4n/wB417rqR/4ll1/1yb+Rrwkfef8A3jSgDEf7te5aD/yA7L/riK8NbpXuWg/8gSy/64inMRZvP+POb/cNeD3H/H5df9dTXvF5/wAek3+4a8IuP+Py6/66mlAGR0UUVqAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMArS03/UN9aza0tN/wBQ31pMDKuP+PiT/eqOpLj/AI+JP96o6YBS0lLQAUlLSUALSUUUAFFFFABRRRQAHoa6nwv/AMg+SuWrqfDH/HhLSYGTd/8AH5L9aizUl1/x9y/WoqAFzSUUUAFFFFAC0lFFMAooooAKKKKQBRRRQAUUUUAFFJRQAtJRRTAKKKKAGv0X/eFe36D/AMgW0/3BXiL9F/3hXtug/wDIFtP9wVlUGi9L/qn/AN014bqX/IXvP9+vcpf9U/8AumvDdT/5C97/ANdKKYFfNIehopD0NaiPTvhr/wAgGX/rrXYd64/4a/8AIBl/6612HeueW4zyLx4f+KpP+5XO10Pj3/kaT/uVzwraOwgpsp/dH8KdTJf9U34UwPcfD/8AyA7P/rkK0D/Ss7w//wAgKz/65CtE/wBKwe4zwrUB/wATO8/67NUIqfUP+Qnef9dmqCt1sI3/AAP/AMjTBXrwryHwP/yNMFevCsp7jOU+Iv8AyLUn1FeWD7o+gr1P4i/8i2/1ryz+EfQVdPYAoooqxHf/AAt/1Oof74rvq4D4W/6nUP8AfFd/WEtxnmHxN/5DVp/uGuOrsfib/wAhq0/3DXHVrHYAP8H++K9y0T/kDWf/AFzFeGnqn++K9y0T/kDWf/XMVFQC1cf8e8v+6a8Kvf8AkJXn/XU17rcf8e8n+6a8Kvf+Qjef9dTRTAjope1JWwjQ0D/kYdP/AOuor27ua8R0D/kYdP8A+uor271rGpuCMXxf/wAize/9c68YT/Vr9K9n8X/8ize/9c68ZT/Vp9KqmAtFFFaAFFFFABRRRTAKKKKACiiigApaSigAooooAKKKKACiiigAopaSgAooooAKKKKAFopKKANHTv8AUmsy5/4+XrT07/Ums25/4+XpICKiiimAUUUUwCiiigYUUUUCCiiigYUUUUCNHQP+Qqv0rT8Q/wDHytZmg/8AIVX6VpeID/pK1L3GZfeig9aSgQUUUUwCikpaAEooooAKKKKACikpaBiUUUUAFFFLQAUUlFABRRS0wEooopABooooASkpaKYFjTr+60y8W7s2IcH5l7MK9X8Pa/a65ah4mCzqPnjPXNeQVNZ3VxY3a3NpIY5V7jvWco3A9yAqpqel2mqWxgvIldexI5WsTw34vtdUVYLsrBdjqCflP411APFYO6YHleu+DL7TC01lm5tuv+0PwrmT8rFXBRh/C3Br3qsXVvDGmaoGaWBY5T/y0Uc1pGpbcDx/pQa6zU/AWo2pZ9PkW4jH9881zNxZXtqxW4s5gR3C8VqppgQZpaZuUfeO3/epQyH+NfzqrgLS0ZX+8PzpCyd3X86ACjijIP3ct/u81Yt7C+umC29nMSe5XildAVqActtUF2P8K8mur07wDqVyVa/kW3jP9w812WkeFdL0sKyQiWUf8tHHNQ6iQHEaF4Ov9U2zXWba29f4j+Fej6TpFlpNuIrOFV4+ZsctV0cUuaxlJsBaTIzjPPpWB4g8VWOjoUDCa4I+VFORn3rgB4q1UauNRaUnB5g/h2+lCg2B6/XGeM/Ci36Nf6eoW6QZdR/GK6LRtXttYsUubZgcj5l7qavn2pJuLA8Ew4cxmNvMBwUA5zV230fVLkZhs5Mf7S4r18aNpovGu/scfnt1fFXgMDAxj6Vr7UDxWbQ9WgUtLZuR/sjNUSGVtrqyN/dYYNe9AH2/KsPXvC9hrETExiK47SqOaFV7geRUlXNT0650q8a1vE2sPut2YVUNbJ3AKKSimAUUUUAFFFFABRRRQAUUUUABooooAKKKKACiiigAooFLTASloopAFFFFABRRRQAUUUUwA1qQ/wDHn/wGss9K04f+PP8A4DSYGL3P1oo7n60VQC0UUUCCiiigYUUlLQIKKKKACiiigAooooGFFFFAgooopgUqKKK5QFpKKKBBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAE1l/x/wAH+9XU6/8A8eMdctZf8f8AB/vV1Ovf8eMdJjOe7CijtSUALRRSUwFpM0UUAFFBooAKKKKACiiigAozRSUALRRRQAlLRRmgA6VLZ/8AIQtf+ugqLmpbP/kI2n/XQUgPdrb/AI9ov9wfyqprv/IFu/8ArmauW3/HtF/uD+VU9e/5Al3/ANczWQzxBT9//eNKn/HxD/vj+dMB+/8A7xpN+ySNz0Dj+da9BHvlsf8ARof+ua/ypl/G09jNEn3nUgUllKsljBIhBVo1wR9KnrMZ5AfBPiJC37iE5YnrUL+EfECdbVD9K9mApfxp8wjw6XQNYh+/ZOf90VUktbyI/vLK4H/AK9759aa8SSDEiKw9xT5x2PACdp+dWT/eGKAVPRgfoa9xn0LSrjPnWELZ9RWReeBNGuM+Whg/3KfOI8nwfSiu7vfhvImWsLwv6CQ4rnr/AMLa1YZ8618xR3iGaakgMWinSK0TbZo3iPo4xTeoyORTAKKM0lMBwPIru/hZ1u64IfeFd78LP+XqplsB6J2rzn4o/wCvtPrXooPFedfFE/vrP61EdxnDN940mKU9TSVqISiloIwMnge9ACUUqBpG2wxPKfRBmtew8La3qGDFbeWp/wCegxSuBj80hZR1ZR9TXd2Xw1dsHULxk9RHzW7Z+A9GtsGRDPj+/S5kB5Sp3fcVnP8As81PHaXkpxFZXDf8Ar2i30LSrfHk2EK474q+kaRjEaKo9hU84WPFIvD+sTfcsmH+8Ksr4Q19+lsg+tey0UudjPHk8E+IDJGfIiAVwTz2r1eyiaGyhif7yKAas0hobuBV1P8A5Bl3/wBcm/ka8IB5f/eNe66u6ppN4zEACFv5V4Qh3At6k1UAY5jxXueg/wDIDsv+uIrwtj8te56B/wAgOx/64iiYi3d/8ek3+4a8Huf+Py6/66mveLv/AI9Jv9w14Nc/8ft1/wBdTSgAyiiitQCiiigAooooAKKKKAEpaKKACiiigAooooAKKKKACiiimAVpab/qG+tZtaOnf6g/WkwMu4/4+JP96o6kn/4+JP8AeqOgAooopgLSUUtABSGlpO9ABRRRQAUUUUAHauo8MH/QJa5ftXT+GP8AjwlpMDKuv+PyX61FUt1/x9y/WoqACjNFFABRRRTAKKKKQBRRRQAUUUUAFGaSigAzRRRTAKKKKACiikoAKWkpaAGt0X/eFe3aD/yBLT/cFeIt0X/eFe3aD/yBLT/cFZVAL0v+qf8A3TXhmp/8he9/66V7nL/qn/3TXhmp/wDIXvf+ulEBlYmg9DRQe9aiPTvhr/yAZf8ArrXYd64/4af8gGX/AK612Nc8txnkHjz/AJGk/wC5XPV0Xj0f8VSf9yudxW0dhBTZf9Ufwp9Nl/1Z/CmwPb9A/wCQFZ/9chWgaoaB/wAgOz/65CtA9Kwe4zwrUP8AkKXn/XZqgqfUP+Qnef8AXZqgrdbCN/wP/wAjTBXrwryHwP8A8jVBXrwrKe4zlPiL/wAi2/1ryz+FfoK9S+I3/Itv9a8szwv0FXT2AKKKK0A7/wCFv+p1D/fFd/Xn/wALP9TqH++K9ArnluB5f8Tf+Q1af7hrj67D4m/8hq0/3DXH1rHYAP8AB/vivctE/wCQNZ/9cxXhp/g/3hXuWh/8ga0/65ioqAW7j/j3k/3TXhV8P+Jlef8AXU17rP8A6iT/AHTXhV9/yErz/rqaKYEdFFFbCL+g/wDIw2H/AF1Fe3eteI6B/wAjFYf9dRXt3rWNTcEYvi//AJFi9/6514wn+rT6V7P4v/5Fm9/6514wn+rX6VVMB1FFFaAFFFFABRRRTAKKKKACiiigAFFFFIAooopgFFFLQAlFFLQAlFLRQAlFLRQAUlFFABRRRQBo6d/qTWZc/wDHy9aWnf6k1nXP/Hy9JARUUUUwCiiimAUUUUAFFFFABRRRQAUtJS0DNDQv+Qov0rR8Qf8AHytZuhf8hRfpWl4g/wCPlfpUvcRl96KDSUwCiiigAooooAKKKKACkopaBhRRRQAUUUUAFFJRQAUUUUAFFFFABRRRQAUlLSUALSUUUwFxRSZooAXPIIJDDoR2rqNC8a3mnBYdQzcW/wDfP3lFctRk0nFMR7Tpmt6fqkQe1uFOf4WOG/KtDOOteDRvJDJ5tvI0Ug/iU10ml+N9VscJcgXUf95zzWLp9hnqtRyRJKu2VFdfQiub03xvpN7tV3aGTvuGBW/De2k4BiuYnz6MKizQFOfw5o9wcy6fESe+KoyeCtEc/LbhPpXRj2pfwpXYHLf8ILo/o1TReC9FjOWtw/1ro6KOZgZkHh/SLfBisIQR3xWgkaRrtjVVA7AU+oJb21hB825iTH95gKNWBLijr0rm9S8baRZblWRpZB0CDIrkNU8c6pegpaIttH2dTzVKDYHoepazYaZE0l3cIMfwg/N+VcDrvjm7vg0OmAwQn/lr/Ea5SZ5LiUy3MrTSH+JjSda1jTS3AUszOXdizt1Y96M0lJitANXw7qd3pmqxNZZbzW2tF2avZIizxIzDaWUEj0rynwJYfbfEQkYfLbDeK9ZrnqbgGKz9Q1zTtOO25uUDf3QckVR8X64dF0rdF/x8TfLH7GvJJWeeVprljJK5yzE0oQ5gPa9P1iw1Ef6Lco7f3c8/lV+vB7aWW0nWe0kaKVTkMDXqvhPxLHrVr5UxCXkYw6f3vcUShYC9r+h22t2RhnGJB/q5B1U15Jqdhc6XetaXabXX7rdmHtXuFY/iPQbfXLIxSALMozHIOoNOE+UDxyipr+yudNvHtLxCsing9mHtUFdCdwFooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUwCloooAKKKKQBRRRQAUUUUAFFFFMAPStOH/j0/wCA1mGtKH/jz/4DSYGP3P1opO5+tLVAFJS0lAC0UUUCCiiigAooooAKKKKAEpaKKACiiigAopD1opjKdFFFcogooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUATWX/AB/Qf71dRr3/AB4x1y9n/wAf0H+9XT69/wAeEdIZz/pRmj0opgGaKKKACiiigAooooAKM0lLQAlLSUUALSUUUAFFFFABRRRQAVNZ/wDIRtP+ugqGpbM/8TG0/wCugpMD3e3/AOPeL/cH8qp6+f8AiSXn/XM1bt/+PaL/AHB/KqWv/wDIDvP+uZrMZ4cD97/eNB5GDSL/ABf7xpa1EdN4d8aXmjW4tZ4xPbj7pJ5FdAvxLsv4reT8q85pcmlygelL8StLP3oZh+FWI/iLoj/eE6/8Bry3J9aKXKgPX4PG2hzfdnZf94YrTg1zS5wDHfQ5PYuK8M2Keq5oCIOQuD9aXIFz3+O4hl/1U0b/AO6wNS/hXgsN7eW5zb3ksf0Natp4t1y1IP2ppwOzmlyjPZaK84sviRMpA1CzUDuY+a6bTvGWi35AW48ontJxSswNG+0bTtQBF3aRyE/xEc1yGrfDqNgZNLuGV+0b/dru4Z4p13QyJIPVTmnE9qLsDw3UtG1PS5Cl5auQP40GVqgrBh8pB+le+zRRzxmOZFkQ9VI4rkNd8BWV7um04/ZpuoQfdNWpCPMh1Fd58LP+XquP1HTL3Srkw38JQg8OB8p/Guw+FnW6py2A9Dxha85+KZ/f2n1r0XPFecfFQgT2n1qI7jOIzzQWC9Tj0FXNK0fUNZuBFYQkqTzIw+X869G0HwJYaeFlvv8ASp+6t90H2q3KwjgNM0LVNVcLaWzKp/jkGBXZaV8OoECyancM7941+7XdRRpFGEjUIg6KB0p9Q5MZn2Ojadp4H2SzijI7gc1fpk08UC7ppEjHqxxWDqXjPRtPJVpjKw/5580tWB0HApcn0rze++I9y5IsLRNvYv1rn7vxXrt0Tm7aEHshp8rYHscl1BF/rJ40/wB44qnPr2l24/eX0P0DV4pNd3k5/wBIvJZM+pqv5aH7wyfrVcgj2GbxvocJwZ2b/dGaqP8AETQ1+757f8BryoKo6Ling46UciC56W/xJ0sfdhmP4VE3xJsv4beT8q863GjJp8qA6XxH4yutatzawR+Rbn7zA8n2rl8ADA6CnE5pKaVgGN0r3TQf+QHY/wDXEV4Y44/Gvc9B/wCQHY/9cRUzAtXf/HpN/uGvB7j/AI/br/rqa94u/wDj0m/3DXg9x/x+3X/XU0QGMopKK0ELRRRQAUUUUAFFFFABRRRQAUUUUAFJS0UAFFFFABRRRQAVo6b/AKk/Ws6tHTv9SfrQwMu4/wCPiT/eqOpLj/j4k/3qjoAKKKWmAlFLSUAFFFFABRRRQAUUUUAFdN4Z/wCPCSuZNdN4a/48ZKTAyrr/AI+5frUdSXP/AB9y/Wo6ACiikoAWiiigAooooAKKKSgAooopgFFFFABRmikoAWkoooAKKKKADNFFJQAN0X/eFe3aB/yBLT/cFeIN0X/eFe36B/yBLT/cFZVARel/1T/7prwvU/8AkL3v/XSvdJf9U/8AumvC9T/5C97/ANdKKYytSkcUUE1qI9N+Gn/ICm/6612Vcb8M/wDkBTf9da7KueW4zyPx7/yNJ/3K52uh8enHio5/551zu4e9bReghcU2X/VH8KUMPeklYeU3Wm2B7hoH/IDs/wDrkKvnpWf4f50Kz/65CtBuh+lYPcZ4Vf8A/ITvP+uzVXzU1+f+Jnef9dmqDNbrYR0Hgc/8VVb169nFeP8Agf8A5Gq3r1/GSaynuM5P4jn/AIpp/qK8tB+VfoK9S+JHHhlz6EV5WGG1eD0FXDYB2aM03cPejcM9DV3EegfCz/Vaj/vCvQcV598KzmLUf98V6FWEtxnl/wATP+Q1af7hrj8V2XxMH/E6tP8AcNcdWsNhAf4f94V7hof/ACBrP/rmK8PP8P8AvCvcND/5Atn/ANcxUVBlu4/1En+6a8Lv/wDkJXn/AF1Ne6XH/HvJ/umvCr4/8TK74P8ArTRT3AiooyPQ03cB61rcRo6B/wAjDYf9dRXt/rXh2gsP+Eh0/r/rRXuGeTWVTcDG8X/8ize/9c68ZT/Vp9K9l8Xn/imb3/rnXjMf+rT6VVPYB1FFFaAFFFFABRRRTAKKKKACiiigAooopAFFFFMAooooAKKWigBKKKKACiiigAooooAKKKWgC/p3+qNZtz/x8vWlp/8AqTWbcf8AHy9CAjooooAKKKKYBRRRQAUUUUAFFFFAwpaSloAv6H/yE1+laOv/APHyv0rO0P8A5CY+laGv/wDHytS9xGZQaDSGgBaKSimAUUUUDDNFFGaAClptLQAtFFJQAtJRRnigAooooAKKKKACiiigAopKKAFpKKKADNFFJTAKKKKYBRRRQISilopAIcN94Zp8UssBzBK8R9jTaKLIZpweINZg4TUZSB2Jq0njHXE/5bbvqawqSp5EB0f/AAm+t4/g/Oo38Za4/wDy0CfQ1gUUciA05vEOtT8PqMqg9gaoyzTTnNxO8p/2jUVLTUUAABfujFFFFMBKMUtFMAxQBS0ooEdp8Miovbxf4tg/nXovavHvCmqDStdjkc4imOyQ+gr19GV1DKcqwyD6iuWotRnCfE23mMVpcAFolfDf7NcCev1r3S9tIb61ktrlA8bjBBryTxL4fn0G6PBe0c/JJ/d9jV05dAMapLa5ntLlLm1kMc0ZyCKiNJmttwPXPC/iaDW7YI5Ed2g+eP19xW/mvCLa4ntLlLm1kMcyHIIr1bwt4lh1y22OQl5GPnT19xXPOFgJvEvh+31yzKsAlwgzHIOorya9s7jT7t7W7TZKp/BvpXufWsXxL4dt9ctCGAS4Qfu5B1ohO24HkFGanvLO40+7e1u0KSqcezfSocV0J3AKKKKYBRRRQAUUUUAFFFFABRRRTAKBRRQAtFFFABRRRSAKKKKACiiigAooopgB6VpQ/wDHp/wGs09K0ov+PT/gNJgY3c/Wlo7n60UwCiiimIKKKKACiiigAooooAKKKKACiiigAooopDCijFFMClRRRXMIKKKKACiiigAooooEFFFFABRRRQAUUUUAFFFFAE1n/wAf0H+9XTa7/wAeMdczZ/8AH9B/vV02u/8AHjHSGYHYUUdqKYBRRRSAKKSlpgFJS0lABRRQaACiikoAWkpaKACiikoEGaKKQmgYtS2f/IRtP+ugqHNS2Z/4mNof+mgoYHu9v/x7Rf7gqlr5/wCJFef9c6u2/wDx7Q/7gqtrEElxpF1DEMu6YUVkM8KX+L/eNLWp/wAI3rilh9ib7x7Uf8I3rf8Az5P/AN81pdCMuitX/hG9b/58n/75pR4a1v8A58n/AO+ad0BlUVrf8I1rf/Pk35Uf8I1rf/Pk/wCVHMgMqgVrf8I1rf8Az5P+VH/CNa3/AM+bflSugMqg1q/8I1rf/Pm/5UHw3rf/AD5P+VF0BlBsUxlRuWXJrWPhvW/+fJ/ypP8AhG9b/wCfJ/yougK1lqmo6c4azvZUA/gzwa6/SPiM6FYtYt+Om+Pk1zB8Oa3/AM+T/wDfNMbQtZXrYS/gtJ2YHsmnapZanCJLOdJAR90HkVd614faWWv6dMJrO1uoXBz8oOD9a9D8M+KL27ZLTWNOmhn6Bwvyn61DQHR6hp9rqNuYLyFZEPTI6Vi+GfDR8P31z5Um+2l5XPUe1dNijFK4xm3iub8ReFxr+p28ly+22hwSB1b2rpqCKdwK9nZ29jAILSJYoh/Co61P7UjNtUsegriPEfinVcva6Hps7Ho0rL0+lLcDqtS1iw0uIveXCIQPuZ+Y1w+rfEaSQtHpFuNvTfJwa5SfS9fvpvNu7W5mkPdweKVdA1o8Cxk/75q0kIZe6rqGoMTeXkjg/wAGeBVNQq/dGK0h4d1v/nxf/vmnDw5rf/Pk/wD3zVXQGXnNJWt/wjet/wDPk35Uf8I3rf8Az5P+VO6AyaStf/hG9b/58n/Kk/4RvW/+fJ/yougMmlrV/wCEa1v/AJ8n/Kj/AIRrW/8Anyb8qOZAZVLWp/wjet/8+T/lS/8ACN63/wA+T/lRdAZNLitX/hG9b/58n/KlHhvW/wDnyb8qLoDIYfKPrXuWhf8AIDsv+uIryQ+G9bI/482/KvXtHieHSbSKQYdIgGHoaibBE13/AMek3+4a8HuP+P26/wCupr3i84s5v9w14Pcf8fl0f+mppwGR0tJS1oIKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigArR07/UH61nVo6d/qT9aGBl3H/HxJ/vVHUlx/x8Sf71R0wCilooASilpKACiiigAooooAKKKKAEPSun8Nf8eElcyehrpfDX/HjJSYGXdf8AH3L9aiqW5/4+pfrUVIBaSl7UUAJS0maKYC0lFFABRRQaYBRRRQAlLSUUAFFFFABRmkooAWikooAKKKB0oARui/7wr2/QP+QJaf7grxBui/7wr2/QP+QJaf7grKoBel/1T/7prwvU/wDkL3n/AF0r3Sb/AFL/AO6a8K1I/wDE2vP+ulEAK9IelFFagem/DI/8SGb/AK612gryvwDr8Wl3L2F422CdtyuegavT0uIXUMkyMD3BrCS1GQXGl2FzL5txaRSP03MOai/sPSv+gfB/3zV7zY/76/nR5if31/OpAo/2HpX/AED4P++aDoWlEYOnwf8AfNXvNj/vr+dHnR/89F/OnqAsUaRRrHGoVFGAB2obofpSedF/z0X86xPEniG00vTpWWZHnYFY0U5OfehAeSX3OpXh/wCmzVBTiS7vI33pGLH6mm4roSEb3gY/8VXb17AOprw3SL86Vq9vfAbhE3zD1FeyWGrWWoWyT29xGQwyVzyDWM1qMsXVrBdxGK5iWWM/wt0qp/Yek4/5B0H5Vd8+L/nqn50nnR/89U/OpAp/2HpX/QOg/wC+aBoelf8AQOg/KrvnR/8APVfzpwlj/wCeifnRqBDZ2FpZbvslvHDv+9sHWrVM82P/AJ6L+dQ3V/a2sTST3EaKozyaQHnXxM/5DVp/uGuPrW8T6uNZ1l7mMHyU4jz1xWTW8VZCEP8AD/vV7foRzolmR/zzFeIEblI7kcV6d4I8R21xpkdhcyLFPbjaNxxuFTUQzsSAQQRkGqDaNpjMWaxgLMck7etWxPCRxKn50edF/wA9E/OsgKf9iaX/AM+EH/fNJ/YWlf8AQPg/75q750X/AD0T86Xzo/8Anon509QKSaLpkUiyR2MKupyrBeQavfWk86P/AJ6J+dNeeFVJaVAB3JoAyPGJx4Zvf9yvGY/9Wn0r0Lx74jt5LX+zLGQSux/eMpyMV58BgADoOK1gtBC0tJS1oAUUUUAFFFFMAooooAKKWkoAKKKKQBS0lFMBaKKKACiikoAWikpaAEooooAKKKKACiiigC/Yf6k1nXH/AB8PWjYf6o1nXH/Hw9JAR0UUUwCiiimAUUUUAFFFFABRRRQMKWiigC/of/ITH0rQ1/8A4+U+lZ+h/wDITH0rQ17/AI+V+lT1EZdFFFMAooooGFFJRQAUtJRQAUtFGaACikpaBBRRRQAYoopKAFpKKKBhRRRQAUUUUAFFFFMBKKWigAoopDQIKM0UUAGaKSigAoopaYBRRRSAKKKKACiiigAooopjCiiikAtFFJQIU4Iweldx4K8VGMppepScdIZWP6GuGppGe+McgjtUyjdDPfAwNV76zgv7V7a6jDxuMEHtXE+DPFu/ZpmqSYccRSk9fY13ma52mmB494l8PXGg3R4L2bn5JP7vsaxq9yvbOC+tXtrqMPG4wQRXkviXw9PoN4RgvZufkk/u+xraE+jAx6ltrmezuUubWQxzIcgjvUeKBWlrget+FvEkGuWoVyEu0GHT19xXQ14Ta3M9ncpc2shjmQ5BHf2r1jwv4jh1y0+bCXSDEkfr7iuecLAP8SeHrbXLQqwCXCjMcg6/SvJr20nsLt7W6TbLGcex+le6VyXj3RVvdNN9Cn+kweg6jvRCdnYDzKkpQQRkdKK6QEooooAKKKKACiiigAooopgFLSUUALRRRQAUUUUgCiiigAooooAKKKKYgPStGH/j0/4DWcelaMP/AB6/8BpMZj9z9aKM/MfrS0wCiiimIKSlooAKKKKACiiigAooooAKKSloGFFFFABRRRQBSooormEFFFFABRRRQAUUUUCCiiigAooooAKKKKACiiigCaz/AOP6D/erptd/48Y65mz/AOP6D/erpdc/48Y6QzB7Cik9KKAFpKKKYC0UmaKAFpKPpRQAUUUUAFFJRQIWkoooGFBopKBBRS0UANo3MhV0+8hBFBooA9u8O6lDqmj29xCwOECsPQitSvEtB1690G58y1O6JvvxN0rvrP4haRNGDMsscncEcVm0M7Dn1o59a5j/AITrRP8Anq/5Uv8AwnWif89X/KlZjOn/ABo/GuY/4TrRP+er/lSf8J1on/PR/wAqLMDqPxo/GuX/AOE70T/no/5Un/Cd6J/z0f8AKizA6n8aPxrlv+E80T/no/5Uf8J5of8Az0f8qLMDqfxpOfWuX/4TzQ/+er/lR/wnmh/89H/KizA6g59aTJ9a5f8A4TzQ/wDno/5Uf8J5of8AfeizA6nn1o59a5ceO9EP/LR/yrR0rxJpeqzGG1nHmddrdTRZiNcA9zQVU9QDS0UhhRRRQAUUUUAIRkc0mMDAAAp1FADefWl59aWsnV/EOm6Q6peTgO38APIosBqc560o+tcwfHOh/wDPVqT/AITrQ/8Anq1OzA6n8aPxrlv+E70T/nq9L/wneh/89X/KizA6j8aPxrl/+E70P/nq/wCVH/Cd6J/z1f8AKizA6j8aPxrl/wDhO9E/56v+VH/Cd6J/z0f8qLMDqPxo/GuW/wCE70T/AJ6P+VL/AMJ3on/PR/yoswOo/Gj8a5f/AITvRP8Ano/5Uf8ACd6J/wA9H/KizA6j8aK5j/hOtE/56P8AlVW7+IekxRnyElkkx8oxxmizA2vE+oxabolxLI2GZCqe5rxTJbLt95zuP1rV13Xb3XbkSXZ2xL9yJeg96yz1rSKsISig0VYBRRRQAUUUUAJS0UUAFFFFABRRRQAUUUUCCiiigAooooGFaOn/AOpP1rOrQ0//AFB+tDAzJ/8Aj4k+tR1JP/x8SfWo6YBRRRQAtJRRQAZooooAKKKKACiiigA9a6Xw3/x4yVzVdJ4b/wCPKSkwMy5/4+5frUdSXP8Ax9S/WoqQgooopjCiiigAopKKYC0UUUAJRRRmgAoopKACjNFFABRRRQAUUUUAApaSigAbov8AvCvbtBP/ABJLT/cFeIOeF/3hXtuhH/iSWn+5WcwRfmP7mT/dNeFajzq15/10r3OX/Uyf7prw3UB/xNrzj/lpSgBWpRS7T6UuPatQEwCMEVOl3eRrtjvZlUdADUIB9KXB9KALH2/UP+ghP+dIdQ1D/oIT/nVfn0pDn0pWQFg3+of9BCf86T7ff/8AQQn/ADqDB9KMH0osgJxf3/8Az/z/AJ1EztI++V2d/wC8x5puD6UYPpRZALSGl59KMH0qgEpY5JYf9RM8Wf7powfSjB9KTAl+2X3/AD/z/nS/bL7/AJ/5/wA6h2n0NGD6UrICb7bf/wDP/P8AnSi9v/8An/n/ADqDB9KMH0osgLP2+/8A+ghP+dRy3FxONtxcySj0Y1Fg+lHPpTsgFpKMH0owfSgAo6MGBIYdCO1GD6UnPoaAJ/tt6BgX0w/Gj7dff8/8/wCdQYPpRg+lFkBP9tvv+f8An/OnC+vv+f8An/Oq+D6Uc+lFkBZ+233/AD/z/nTXu7xgQ97Mw7gmocH0owfSiyAQADOO/WkxTsH0pMH0pgJRS4PpRg+lACUUc0UAFFFFABRRRTAWikopAFFFFAC0UlLTASlpKKACg0UUAFFFFABRRRQAUUUUAFFFFAF+w/1RrOuP+Ph60LD/AFRrPuP+Ph6SAjoooqgCiiigAooooAKKKKACilooGHaiiigRf0T/AJCY+lX9e/4+V+lUNE/5CQ+lXtd/4+V+lS9wM0nmkoNFMYUUUUAFFFFABRRRQAUUUUCClpKKAFozSUUAFFJRQAtJRRQMWiiigAopKKAFpKWkzTAKKKKACkpaSgQUtJS0AFFFFABRSUtMAooopAFFFFABSUtGKYwooopAFFFFMAooopAJRS0lAhCOnJBHII7V3/g3xb5mzTdVkxIOIpSevsa4GjHTnBHII7VMo3Ge9DmoL6xt7+0e2uow8bjGCOlcb4M8WeZs03VHxIOI5Sevsa7sGudppgeOeI/D8+g3ZVgXtXP7uT09jWPivcdQsbfUbR7a5QPG4xz2ryPxBodxoV4Y5AWt2P7uTt9DWsJ30YGVViwvZ9OvYru2Yq6MAfcd6gxQVLFEUEs7BQPrWj2A9x0+7S9sYbmP7si5qS6QSW0qEZDIR+lVNBtHstGtbeT7yJzV2UhYnJ7KTXL1A8MuYvs95PB/zzc1HU+pOJdXu5F6M/FV6647AFFFFMAooooAKKKKACiiimAUtFFABRRRQAUUUUAFFFFIAooopgFFFFAgrQi/49f+A1n1oRf8ev8AwGkxmQep+tLSdz9aKYC0UUUwCikpaBBRSUtABRRRQAUlLRQMKKKKACiiigAooooApUUUVzCCiiigAooooEFFFFABRRRQAUUUUAFFFFABRRRQBNaf8fsH+9XS65/x4x1zVn/x+w/71dLrn/HjHSGYFFHajPFABRSUUwFooooAKKKSgBaSg0UAFFFFAhKKWigApKWkoAWkoooAKKKSgYuaU4J5GaSigBdq/wB0UYX+6KKKBCbV/uijav8AdFFFABhf7oo2r/dFFFACbV/uijav90UtFMBu1f7gpdqf3BRRQMTav90Uu1f7oopaQAFX+6Kkgke2nSe2YxzIcqwNR0tFgPVfCnjCDVIltr5lhvFGOTgP9K63NfPwJDBgxVl6MOorqtE8c6hpqLDej7TAP4jy9Q49gPV6K5ay8d6NdAbneEntJxWmniLRpB8uowD6tU2Ga2aKzP7c0r/oI2//AH1UUviTRohzqEJ+jUWA2KK5S88e6NbA7DJMR02c1yWseO9S1BWislFvCf4hw1HK2B2Hinxda6NA0NsyzXjcKoOQp968pup5r24e5vHMsznJLHp7UxizOZJGLyN1ZuppK0UbCG7E/uCk2p/cFPNJVAJtT+4KNqf3BS0UgE2p/dFG1P7gpaKADav9wUbV/uiiimAbU/uijav90UUtACbV/uil2r/dFFFABhf7opRx0GKSigBc0UlFMAooooAKKKSgBaKKKACiiigAooooEFFFFABRRRQAUUUUDCiiimAVoaf/AKk/Ws+r9h/qT9aTAzZ/+PiT61HUk/8Ar5PrUdMBaSlpKACiiloASijtRQAUUUUAFFFFABXSeHP+PKSubrpPDn/HlJSYGXc/8fUn1qOpLn/j6k+tR0CCkoooGFFFFMAooooAWkopKACjNFFABRRRQAUUUdqACjtRRSAKKKKBBRSUUxjXPC/7wr2/QP8AkCWn+4K8QcfKPY5r2vwzNHP4ftHjYEbOcdqzmCNRl3Iy+oxXn118O5p76ecXYAlbIFeiUnFZptDPOf8AhW8//P4KP+FcT/8AP2K9G4o4p8zA85/4VxP/AM/go/4VvP8A8/gr0bijijnYWPOD8N5/+fwfnR/wraf/AJ/B+dej8UcUc7Cx5x/wraf/AJ/BR/wraf8A5/BXo/FHFHOwPOP+FbT/APP4KP8AhW03/P4K9H4o4o52B5z/AMK3n/5+xR/wreb/AJ/BXo3FHFHOwPOv+FcTf8/go/4VxN/z+CvReKOKOdgedf8ACuJ/+fwUf8K4m/5/BXovFHFHOwPOf+Fbzf8AP4KP+Fbz/wDP4K9G4o4o52B5z/wref8A5/BSf8K3n/5+xXo/FHFHOwPOP+FcXH/P2tH/AAri4/5/Fr0fijijnYHnH/CuLj/n8FH/AAre4/5/BXo/FHFHOwPOP+Fb3H/P4KP+Fb3H/P4K9H4o4o52B5x/wre4/wCfsUf8K3uP+fsV6PgUcUc7A84/4VvP/wA/gpf+FcT/APP4K9G4o4o52B5z/wAK4n/5/BR/wrif/n7FejcUcUc7A85/4VxP/wA/a0f8K3n/AOfwV6NxRxRzsDzSf4c3qRk291GzDs3euRurae0uXt7mMxyoeQa95OK8v+JIiGtRFMeaVG/6VcJNsRyNFB60VqAUUUUwFpKKKQBS0lFMBaKSigBaSiigAooooAKKMUUAFLSUUAFFFFABRRS0AXbHiI1n3H/Hw9aFj/qjWfcf8fD0kBHRRRVAFFFFABRRRQAUUUUALRRRQAUUUUAXtE/5CY+lX9d/4+V+lUNE/wCQkPpV/Xf+PlfpU9QM2koopjCiiigAooooEFFFFAC0lFBoAM0UlFAC0lFFAC0lFFABRRRQMWikopiCiiigAooooGFFFFAgpKWkoAKWikoAWkFLRTAKKSloAKKSigBaSlopAFFFFMYCigUUAFFFFABRRRQAUUUUAFGKKKBCjtgkEcgjtXoHg3xZ5uzTdTkxIOI5SfvfWvP6PQgkEcgjtUSjzDPegaq6hYW2o2rW93EHRh3HT6VxvhHxgrqmn6s4WQcJKeh+td0rhlDKQVPQ+tczTTA89vfh3Msx/s66Xyz2l6itTw54Ig024F3fOJ5x91f4RXXcUoIp8zaAWsXxVqiaVok0pI3sNijuc1e1LUrXTLZp7uVUUDgE8t9K8n8Ra5Nr1/5rZS3j4iT296cI3YGQoOPm5JJJNLRRXVYApKKKACiiigAooooAKKKKYC0UUUAFFFFABRRRQAUUUUgCiijNMAooooAO1aEX/Hr/AMBrP5q/F/x6/wDAaTAye5oo7n60tMAooopgFFFFAgooooAKKKKACiiigYUUUUAFFFFABRRRTApUUUVyiCiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBNZ/8AH7D/AL1dJrn/AB4x1zdp/wAfsP8AvV0et/8AHlHSGYVFJRTAWkoooAXNFJRQAUUUUAFFFFAgooooASjNFFABSUuKKBhSUtFABRRSUALRRRQAUUUUCCiiigAooopgFFFFAwooooAKKKKAClpKKAClyR0NJRQAjAN95QfrTfLT+4BT6KQDdi+lHlx/3BTqKLACgL91QPpS5z3pKKYC0UUUAFFFFABRSUtACUtFFABRSUUALRSUtABRSUtMAooooAKKKKACikpaAEoopaACiiigAooooAKKKKACiiigAooooAKKKKACiiigAq/Yf6k/Ws+r9j/qT9aAM6f/AF7/AFqOnz/6+T60ymAUUUUALSUUUAFFFFABRRRQAUUUUAHauk8O/wDHjJXN9jXR+Hf+PKSkwMu5P+lSfWo6kuf+PqX61HQAUUUUAFFJRQAGijNFMAooooAKKKKACiiikAZooopgFFFJQAtFJRQAUtFFABWxoPiO/wBCZhbkSQN1jboPpWPRSauB3I+JEnezpf8AhZD/APPn+lcLk+tJuPrS5EB3X/CyZP8Anz/Sk/4WTL/z5/pXDZPrRk+po5EB3H/CyZf+fP8ASj/hZMv/AD5/pXD5PqaMn1NHIgO4/wCFkzf8+f6Uf8LJm/58/wBK4fJ9TS5PqaORAdx/wsib/nz/AEo/4WRN/wA+f6Vw+T6mjJ9aORAdx/wsib/nz/Sj/hZE3/Pn+lcPk+tGT60ciA7j/hZM3/Pn+lH/AAsmb/nzrh8n1oyfU0ciA7j/AIWRN/z5/pR/wsmb/nzrh8n1oyfWjkQHcf8ACyZv+fP9KP8AhZM3/PnXEZPrSZPrRyIDuP8AhZE3/Pn+lH/CyZf+fP8ASuIyfWjJ9aORAdv/AMLJl/58/wBKP+Fky/8APn+lcPk+tGT60ciA7j/hZEv/AD5/pR/wsmX/AJ8/0rh8n1oyfWjkQHc/8LJl/wCfP9KP+FkS/wDPn+lcNk+poyfWjkQHc/8ACyJf+fOj/hZEv/Pn+lcNk+poyfWjkQHc/wDCyJf+fP8ASj/hZEv/AD5/pXDZPqaMn1o5EB3P/CyJf+fP9KP+FkS/8+f6Vw2T60ZPqaORAdz/AMLIk/58/wBKP+FkS/8APn+lcNk+poyfWjkQHc/8LIl/586P+FkSf8+dcNk+tGT6mjkQHazfEa8dCLe1QN2LVyN3dT3t09zdOXlfqT2qDJ9aKaikAGiiiqAKKKKYBRRRSAKKKKYC0UlFABRRRQAUUUUAFFFFABRRRQAUUUUALRRSUAXrH/VGs+f/AI+Hq/Y/6o1QuP8Aj4ekgI6KKKoAooooAKKKKACiiigApaKKACiiigZe0X/kJD6Ve1z/AI+Vqjo3/IRH0q9rf/Hwv0qXuIzqSjvRTGFFFJQIKKKKACiiigAooooAKKKKBhRRRQIKM0lLQAUUUlMAooooGFFFFABS0lLQAUUlLQAUUUUCCiiigAooopgFFJS0AFFJRQAtFJRQAUtJS0DCiiigAooooAKKKKBBRRRQAUUUUAFFFFAAQCOf/wBVbOleKNW0rCxTefH/AHZTnFY1FS4pjO4i+I7hf39oS3+yKr3fxDvpVK2duiZ7v2rj8n1oNT7OIE97fXeoS+ZezvKeyk8D6VBmkpatKwC0lFFMAooooAKKKKACiiimAUUUUALRSUtABRRRQAUUUUAFFFFABRRRQAUUUUABq/F/x6/hVA1fi/49v+A0mBknqaKO5+tLTAKKSlpgFJS0UCCiiigAooooAKKKSgBaKKKBhRRRQIWiiigZRooormEFFFFABRRRQAUUUUCCiiigAooooAKKKKACiiigCa0/4/Yf96uj1v8A48o65y0/4/Yf96ui1r/jyjpDMPsKSjtRTAKKSlzQAUUUUAFFFFABSZoooEFFFFABRRRQAUUUUAFBoooASlopKBi0UUUCCiiigYUUUUCCiiimMKKKKACiiigAooooAKKKKACiiigAFFFLQAlFFLQAlFFFAC0UUUAFFFJQAtFFFABRRRQAlFFLQAlLRSUAFFFLTAKKKKACiiigBKWikoAWikpaACiiigAooooAKKKKACiiigAooooAKKKKACkopaYCVfsf9SfrVCr9j/qT9aTAzZ/9e/1plPm/17/WmUwCiiigAooooAKKKKACiiigAooooAO1dF4d/wCPKSud7Gui8Pf8eT0mBmXH/H1J9ajp9z/x9SfWo6ACikzRTAKKKKACiiigBaKSigBaKKSkAtFJRmmAtFFJQAtJS0lAC0UUUAFFFFABSUUUAFFFFABRRRTAKKKKACiiikAtFFFMAooooAKKKKQwooooEFFFFABRRRTGFFFFAgooooASloooAKKKKAEpaKKACiiigAooooASilooAKKSloAKKKKACiiimAUUUUgCiiimAUUUUAFFFBoAKKKO9AC0lFFAC0UlLQAUUlLQAlLSUUAXbL/VmqE//Hw9X7PPlGqE/wDr3oQEdFFFMAooooAKKKKACiiloGFJS0UAJS0UUAXdG/5CI+lXtb/4+Vqjo3/IRH0q9rf/AB8LUvcDO70Ud6KYBSUtJQIKKKKACiiimAUUlLSGFJRRQAUCiigAoopaAEooopgFFFFABRRRQIKKKKBi0lFFAC0UlLQIKSlpKYwpaKKBCUUUUDCiiigAooooAKKKWgAooNFABRRRQAUUUUAFFFFAgooooAKKKKACiiigYUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKO9FFAC0UlFAC0lFFABRRRQAUtJRQAtFJS0AFFFFACdq0I/+PX/AIDWeavxf8ev/AaTAyu5paTuaKoBaKKKACikpaACiiigQUUUUAFFFFABRRRSAKKKWmAlFLRTGUaKKK5RBRRRQAUUUUCCiiigAooooAKKKKACiiigAooooAmtP+P2H/erotb/AOPKOuctP+P2H/erotb/AOPKOkxmH2opO1FMApaSigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUDEpaKSgBaKKKACiiigAooopiCiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALSUUUALRRRQAUUUlAC0UUUAFJRRQAUtFFACUtJS0wCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooASilopgJV6y/wBUfrVKrtl/qj9aQGdP/r3+tMp83+vf60ymAUUUUAFFFFABRRRQAUUUUAFFFFAAeldD4f8A+PKSuerofD//AB5PSYGZc/8AH1J9aiqS5/4+pPrUdABiiiimAUUUUAFFFFIAooopgFFJRQAUtFFABRRRQAUtFFACUtFFACUUUUAFFFFABRRRQAUUUUAFFFFABS0lLQAUUUUwCiiigAooopDCiiimAUUUUCCiiigYUUUUAFJRS0CCiiigAooooASlpKWgAopKWgAooooAKKKKACiikoAWiiigAoo7UUAFFFFMAooooAKKKKACiiigAooooAKKKKAFooooAKSiigAooooAuWZ/dGqE/wDr2q/Z/wCrNUJ/9e1JAMoooqgCiiigAooooAKWiigYUlLSUAFLRSUCL2j/APIQH0q7rf8Ax8LVLR/+QgPpVzW/+PhanqBQJpKKKYBRRRQAUUlFAxaKKSgBaKQ0tACUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKBC0lFFAxaSiimIWikooGLSUUUAFFFFABRS0lABRRRQAUtFFABRRRQAUUUUAFFFFAgooooAKKKKACiiigAooooGFFFFABRRRQAUUUUAFFFFABRRRTAKKKKBBRR2ooGFLSUUAFFFFAgooooAKKKKBhS0lLQAlFFLQAlX4/+Pb8KoGr8f/Ht+FJgZXc/Wlo7miqAKKKKACiiigAooooEFFFFABRRRQMKKKKAClpKWgQUUUUxlGiiiuUQUUUUCCiiigAooooAKKKKACiiigAooooAKKKKAJbT/j8h/wB6ui1r/jzjrnbT/j8h/wB6uh1r/jzSkxmHRR2pKYC0UUUAFFFFABRRRQAUUUUCCiiigAooooAKKKKACiiigYUUUUAFFFFABRRRQIKKKKYCUtFFAwooooAKKKKACgUUUAFFFFABRRRQAUUtJQAUUUUAFFFFABS0lFABS0lFAC0UUUAFFFFABRRRTASloooAKKKMUAFFFFABRRRQAUUUUAFFFFABRRRQAUUZooAKKKKACiiigAooooAKKKKACiiimAVcs/8AVH61Tq5Z/wCqP1pAZ83+vf60ynTf65/rTaYBRRRQAUUUUAFFFFABRRRQAUUUUAHauh8P/wDHm9c9XQaB/wAeb0mBmXP/AB8yfWo6kuf+PqT61HQAUUUUAFFJRQAUUtJTAKWijFABRRRQAlLRRQAUtJS0AFFFFABRRRQAUlFFABRRRQAUlLSUALRSUUALRRRQAtFFFABRRRTAKKKKACiiigAooooGFFFFAgooooAKKKKAEpaKKACiiigAopKWgAopKKAFopKKAClpKWgBKWiigAooooAKKKKYBRS0lIBaSiimAUUUUAFFFFABRRRQAUUUtACUUUtACUtFJQAUUtJQAUUtJmgC3af6s1Rn/wBe9XrT/VmqM/8Ar3oQDKKSlpgFFGaKACiiigBaKKKACiiigAooooAu6R/yEB9Kua1/x8LVPSP+QgPpVzWv9etT1Azz1pKDRTAKM0UUABoooFAxaSiigAopaKACkpaSgAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUtACUUtJTEFFLSUDCilooASilpKACiiigApaKKACiiloASiiigAoopKAFoopKAFooooAKKKKACiiigAooooEFFFFAwooooAKKKKACiiigAooopgFFFFABRRRQAUUUUCCiiigAooooGFFFFABRRRQAUtJS0AFXY/+Pf8ACqVXY/8Aj3/CkwMvufrS0dzRVCCiiigAooooASloooAKKKKBiUtJS0AFFFFAgpaSloAKKKKYyjRRRXKIKWkooEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUATWn/H7D/vV0Gtf8eaVz1p/x+Q/71dBrP8Ax5pSGYnaiiimAUUUUAFFFFABRRRQAUUdKO9AgooooAKKKKACiijFABRRRQMKKKKACiiigApKDT7e2ubuURWkDTOey0CGUBgTgHmu00X4eXd0Fl1aXyoz1iH3q2fEXgqyXQz/AGXEVngGV9WpcyGeZ0tHbBGCDg/WiqAKSig0AFFJySqqMszBQPeuvk+H2onT4ri3uF81ly0RHNJuwHI5pKtX+l6jpzEXtnJEB/ERnNU1YN0NFwHUopq/M4RQWduigda6vQvA+o6kVlvh9lt+pVhyw9qG7AcvRiuu8WeDDpEQvNMDSWyj94h5I965HIIyDkGhO4BRRRTASilpCQOScUAFFIXT+9+lJ5if3v0oAdRTPMT+9+lHmJ/e/SgB1FN8xP736UeYn979KAH0U0Mp6Gl70AOop3kXP/PtJ+VOEFz/AM+sn5UXAjxS1J9nuf8An1k/Kj7Pc/8APrJ+VFwIqKl+z3P/AD6yflR9nuf+fWT8qLgR0VJ9nuf+faT8qX7Pc/8APrJ+VFwIaKe8csf+tiZM+oplABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFXLP/AFR+tU6t2n+qP1pAUJf9c/1plOl/1z/Wm0wCiiigAooooAKKKKACiiigAooopgFdDoH/AB5PXPV0Gg/8eb1LAzLn/j6k+tR1Jc/8fMn1qOhAFJS0UwCiiigAooooAKKKKACiiigAooooAWikooAWikooAKKKKACiiigAooooAKKTegOCTn6UGSP+8fyoAKKbvQnAPP0pJCRGxHUCgBS6jgnml3r/AHv0r0zwz4Y0i80G1uZ4C0siZY1rDwfon/PsfzqHMDx8Mp6GnV6T4n8N6VZaHPcW8BWRRkGvNEOY1J6kU4yuAtFFITVABIFG4etb3gzSrTWdWmt75WaNE3AKcc123/CBaF/zyl/76qXKwHleR60o5r1T/hAdB/55S/8AfVcJ4t0220nXWtLJWWIIGwTk5oUrhYx8UUEhRknip2sb1bT7W1rILb/noRxVXAr0UtFMBKKWigBKSnVe0TS5dY1WKzi4XO529AKTdgM8MpOAeaU169ceDNEuIVje3I2jAKnBrnr/AOHGCW068EY7BxmoVRAcFSZrU1vw9e6GiveSqwY8Ad6ySQBzVp3Admik5BwQQe2RijNO4BRketXdEtor3Wre1uATFIwDAV6UfAeg5OIZf++qhzsB5RketGa9VPgLQf8AnlL/AN9Vw3i/SbbRtbW0sgREY92GPehTuBiUtAU0uDVgJRS4NGDQAlFFLQAneilpKYBRRRSAKKKKYBRRRQAUUUUAFFFFABS0UUAJRRRQAUtJS0AJS0lFABRRRQBbtP8AVmqM/wDr3q7af6s1Sn/17UICOlpKWmAUUUUDCiijFABSikpaBBRRRQAUUUUAXdI/5CA+lXNa/wBetUtI/wCP8fSrms/69al7gZ560UUUwCiiigAooooGFLSUUALRRRQAUUUUAFFFFABRRRQAlFLSUwEJpCwUZY4FBre8EWsF74jMN1GJI/KyFPrSbsBz/mx/3v0o81P736V7R/wjmkHrZp+VNbw1pBH/AB6L+VZ+0A8bDBhlTmnCuh8eWNtp2tRw2cfloY8kD1rnAauLugHUtJRmqAWkoooAWiiimAUUUUAJRmg0xunHUnApAOyCcA80ua9F0vwZpl7oNs13G63DLksDjmsy/wDh7dxZaxukZR0Qjmo9ogOOpat6jpOoaVzf27RoejnoaLDTNQ1FwtnavIP7w6Cq5luBUxmiu40z4eu+2TU7gFTyY14IrH8WeHv7CukaDJtJOhP8J9KlTTdgOfpKcRSVYDTSZpX4Qkda7nwn4W0vVNFW5vEdpS2CQcCplKwHC596WvUj4E0L/nlL/wB9VxnjTSLPRb+KGxVlRlyQxzSVRN2A5/NLmm55qeCzu7mJpbW3aVE+8V7VdwIqM00nDbWBDehFLQAtLTCau2ulaneQCa0spJYj0cdKG7AVc0Zq/wD2Brf/AEDJaX+wda/6BktLmQGfRWj/AGBrX/QMlo/sDW/+gZLRzIDPorR/sDWv+gZLR/YGtf8AQMlo5kBm0taP9ga3/wBAyWj+wNa/6BktHMgM6itH+wNb/wCgZLR/YGt/9AyWjmQGdSZrROga3/0DJaT+wdb/AOgZLRzoDPozWgdA1r/oGS0f2BrX/QMlp8yAz6K0P7A1r/oGy0j6FrEcbSSafIqIMsT2FHMgKFFJnNLVAFFFFABRRRQAUUUUAFFFLQAgpc0UUAIavR/8e/4VRq7H/wAe/wCFAGb3NFHc0UCCiiimAUUUlABS0UUAFFJS0DCiiigAooooAKWkpaBBRSUUxlKiiiuUQUUUUCCiiigAooooAKKKKACiiigAooooAKKKKBktr/x+Q/71dBrP/Hmlc/a/8fkP1rf1j/jzSkBiUtFJTAWiiigAooooAKKKSgBc0UlLQIKKKKACiiigAooooAKKKKBhRRRQAUUUtACV1fw81SOy1k2cyptueEYjlTXKUCR4pElibbIhBU+lJ6oD6CFBGQQeh4rK8NaqmsaLBdofmI2sPQitXNZDPH/G+kf2TrZkjXbb3PKex71zpNeo/EZ9OfRWiuJkW7HMIzzXlaklRnrWkXoIcTSUU+GGW6njtoFLSynAA9PWqEb/AIF0c6trYmkXNtbct/vdq9gA/CsrwzokWiaTHbKAZCMyN/eNa9ZN3GRTQxTLiWJJB6MM15X8RLfTLS+ig0+MJcuMuV6LXper6hDpenTXkzALGDjPc9hXht7eS6jfz305O+Zs4PanEDvfhqulT28g8lDfIed4zkeor0IYxx0rwTTNQn0nUYr62YhkPzgfxL6V7bo+qW+r6fHeWzAq4+Yf3T6USQy7LGksbRyKGRhgg9xXk3jLw0+i3bXVspaylbPH8B9K9bFQ3tpDfWkltcoHikGCDSTsB4LSVreJNDm0HUTA4JtnOYn9vSsnvWqdxBW94KsbfUPEDQXab4wmQKwK6f4dn/iqH/3KUtgO/wD+ES0T/n0FH/CI6J/z6Ct2isrjMA+END/59KT/AIQ/Q/8An0roKSncDA/4Q/Q/+fSkPg/Q/wDn1P510FIaLgeYfEHRdP0m1tXsYTGzkg1xsJ/0u3/66L/OvQfisf8ARLH/AHjXnsH/AB923/XRf51a2A96it4DFGTbxcoP4RUn2aD/AJ4Rf98ilh/1Mf8AuCpKzYEf2eD/AJ4Rf98ij7NB/wA8Iv8AvkVJnFGaAI/s0H/PCL/vkUfZoP8AnhF/3yKlooAi+zQf88Iv++RR9mg/54Rf98ipaKAPP/ijHHHYW3lxonz9VGK8+PWvQ/ip/wAg+2/3xXnh6mtYbCYlFFFUAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFW7X/Vn61Uq1a/6s/WkBQl/1z/Wm06X/XP9abTAKKKKACiiigAooooAKKKKACiiimAdq6DQf+PN65+t/Qv+PJ6lgZtx/wAfMn1qOpLj/j5k+tR0wCiiigBKWiigBKWkpaACiiigAooooAKBRRQAUUUUAFFFFABRRRQAUUUUALTW4RvoaWmt/q2/3TQB6h4M0jT7zw3by3FsjyEcsRW7/wAI7pP/AD5x/lWf4A/5Fa3+ldLWDeozjPGGh6ZaeH5p4LZUkU8ECvLGybds/wB017H48OPC8/8AvCvHm/492/3TVQ2Ee0eDh/xS9l/1zrbrE8GHPhez9krcqHuMwfGf/Is3X+6a8ZiP7lPpXs3jP/kWbr/dNeMRf6lPpWkBEtIaUUhrQDrPhn/yMNz/ANcq9TAryz4af8jDcf8AXKvVBWM9xhivLvFunXOreOXtLQgSeWvJ7V6lXlHjK+m0zx6L63J3xIpIH8QpRA6nQfA1jp5We9Iubkd/4fyrprizt7i0a1kiTyWGCoHFM0u/i1LT4buBgyyLzjse9W6TbA8T8QaRJomqyWr5MTHdG/Y+1ZteueM9EGr6QxjTNzCN0Z/nXkYzyCMFTtP1raEroAooxU1rZ3d7KIrK3aaQ9h2qmxESJJLKkMCl5ZDhVFeueEPD6aJpw8wA3U3zSt6e1UvCHhFNJUXl9iS8YcZ6J7V11YylcYVS1XUrbSrGS7u3Cog4z3PpVtyVRmCliBnA715j4n03xR4hvyWsnjtIziOPP6mpQGTI+qeNtbZocqoPyAniNfcV2uh+A7CwKzXzfabgc5/h/KuY0Xwt4m0zUoru3BjIIDjsV716om4opYYJHI9KpvsI5jxb4Xg1axMlpGsV3CPk2jAYeleUujxyvFMpSSM4dT2Ne/471wnj7wx58batp8f71B+9QfxD1pxkM43wv/yM1n/vivbe5rxHwqwbxLZkdN4r27+I0p7gGK5nxD4Ottd1Jb2aVlYJtwK6iiouBww+G+n/APPd/wA6d/wrjT/+ez/nXb0VXMwOI/4Vxp3/AD2f86bL8O9OjhdxK+VBI5ruKjuf+PaX/dP8qFJgeDuuyWRB0Ryv5UlPmP8ApNx/11amV0IQUlLSUAFFFFMBaSiigAooooAKKWigApKWigQUUUUDCkpaSgAopaSgAooooAKKKKALVr9w1Sn/ANc1XbX7hqlN/rmoQEdLSUtMAooooAWiiigAooooAKKKKACiiigC5pH/AB/j6Vc1n/XrVPSP+P8AH0q5rP8Ar1qeoGf3ooopgFFFFABS0lFAwoopaBCUUtJQMWikooEBZV+8aN6f3v0rsvh/pVjqVveNeQ+YyOAvsK68eFtG/wCfQVm6lnYZ49vT+9+lG9P736V7F/wi2j/8+gpf+EX0f/n0FL2oHje+P+9+lG+P+9+leyf8Ixo//PotL/wjGj/8+i0e1A8YMkf94/lXSfDtlbxWQp/5Zeleh/8ACMaP/wA+i1PZaJp1jcefa24STGMj0pOpdAX6Q9KfimkVkB5X8THVfEcQJ/5ZDtXKiSP+9+le2ajoenalOJryDzJAMA+1Vv8AhFNF/wCfStYzsgPHfMT+9+lHmR/3v0r2L/hFdF/59KP+EV0X/n0FV7UDx0SR/wB79Kd5kf8Ae/SvYf8AhFtG/wCfQU4eF9G/59BR7UDx3en979KAyk4B/SvY/wDhGNH/AOfRa5/xrounWGhPNa24SQHgihVLsDzw0UDkD6UGtQENWdKtGv8AV7a1UZLMG/Kqxrsfhrp/n6hPfuOIflQ+uaibsgPSI0CRqijAVQMU7FLimyuI4nkboilj+FcwHmvxMvvP1C3sVbKIp3j3rR+G2rLLbSabIFEsfKEdStcRrt7/AGhrd3dZyrv8vNN0fUX0rVoLyM42kK/P8PetuX3QPdazPEGmR6rpM1s65bG5P97tV20uY7u2juIWBjkXcpFOmmigjMk0ioi9STWWwHhckckE0kEoxJEdrU2tvxnLp8muefps6SrIMybf71YRPTnv610p6AI7qEbk9PSvVfh8QfDaEf3ql0vw7pU2l2sklqpZowSfWty0tILKEQ20YRBzgVjOdwJm6V5d8T5FXWoFJOfLB6V6lVC+0mxv5BJdQLIwGMkVEXZ3A8M8xcEg9K9U+HtiLfw6HkT5pWJOe4rM8b+HLUpYR2EPlyTzbGIHau3sbYWtlBboMCNAp+taSndAUdS8P6XqS7bm1Ue6DBrgvFnhW00K2W5t58K5wsbHk16l0ry74kX/ANp1qOyVspbgNx60oN3A5Mn5c+1esfD7/kU7f/eNeTk8GvWPh7n/AIRS3/3m71dTYDpxRRS1zgFFFFMAooooAKKKKACij8qTP0oAWikz9KM/SgBaMUn5UtACYqlrQ/4k93/1yNXqo61zo930/wBUaa3A8Q/if/ep1NB+Z/8Aep1diAKKKKYBRRRQAUUUUCFooooGFFFFAhO1XY/+Pf8ACqVXY/8AUfhSYGb3NFHc0UwCkpaKYBRRRQAUUlFAC0UlLQMKKKKACiiigQUtFFMBKKWigZRooorlEFFFFAgooooAKKKKACiiigAooooAKKKKACiiigZLa/8AH5D9a39Y/wCPNKwLX/j7h+tb2r/8eaUgMWlpKWmAUUUUAFFFFACUUUUCCiiigApaSigBc0lFFAC0UUUDCiiigApaSloAStLQdK/trVBY+aI8/wAVZtdF4C/5GuOk9gN7/hWa/wDP/wDpTT8Mh21Af9816LRWfMxnL+FfC9x4dllC3olgk6x46Vr6taXd7b+VaXQt88MSOTWhijFAHAT/AA3FxJ5k2pPI57uSai/4VjGOl/j8K9DwKMCi7Cx494o8Jnw/bRTLc+e0jbVjA5JrrPAfhQ6dGNS1FP8AS5BlEP8AyzFddPaW9xIjzxK7R/d3DOKnp3YDqaTxRXK+OvEi6Lpxt4GBvLgYUZ6DuakDkviFr/8AaWof2bbP/o8B/eEfxNXJCmA9SzZYnJJPWnAj1FarQQvSt/wd4gfQdTVJCTZTnDr/AHT61ggj1FKQrAg4IND1A+gYZEliWSJgyMMqw7in15x8PfEuxho9/KMf8sHJ/wDHa9HrNjM3XdHtta097W4UZI+Ru6n1rxnU9OuNJv5LK7Uh0+6395fWveK57xf4ej1zTzsUC6jGY29famnYDx2um+HX/I0P/uVzUiSQyvDMMSxttYe9dL8Ov+Rpf/cFXLYR67RRRWQzjdW8f2WmalLZSWcrvH1YHiqf/CzrEsq/YJfmIHWuO8Wj/iqrvjtWOR88fA++P51fLoB9BRSCWJJB0dQw/GnmobIf6Fb/APXJf5VP0qAPPfiv/wAeth/vGvPYP+Pu3/66L/OvQviv/wAelh/vGvPYP+Pq3/66L/OtFsB9AQ/6mP8A3B/Kn1HCf3Mf+4P5VJ2qAOc8a69caBpi3NsgZmYDkVyC/EvUP4rP9K2vip/yAY/98V5kM+vanFXA7YfEq+/58v0q1YfEO7ub6C3aywJXCk+ma4IZq7pBP9s2X/XVarlQHu9FHeiswOB+Kv8AyDrb/roK87PU16J8Vv8AkHW3/XQV52etaw2EFFFFUAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgAq1a/6s/WqtWbb/Vn60gKMv+uf602nS/61/rTaYBRRRQAUUUUAFFFFABRRRQAUUUUAHat7Qv8AjzesGt7Qv+PN6TAzbn/j5k+tMqS5/wCPmT61HTQBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQIKRvuN/umlPSkb/Vt/umkB654A/5Fa3+ldLXNeAP+RXt/pXS1g9yjmvH/wDyK83+8K8gb/j3b/dNeu/EE/8AFLTf74rx9j/o7cj7vrVw2A9p8Ff8ivZ/7lbprA8FH/imLP8A3K381m9wMLxl/wAizd/7hrxiH/UJ9K9o8Yj/AIpm7/3DXjEI/cp9K1gBIOlGKUCitBHVfDXjxFP/ANc69Vryr4b/APIxz/8AXOvVawnuMK8i+IP/ACNsn/XIV67XkXxC/wCRtk/65CiG4Gl8NtYNvdSaTM3yPzDn1716XXgUFxJZ3UV3ASJYWBFe12Gr2tzo8eoNKqxFMsxPQ96c1qBokZHNeReNtK/svXWeNQsFzyg9+9dHrnxCtod8OkJ58nQS/wAIrg9Q1K81S4M99L5j9sdBTgmBBXoPwxuVe2urYqm+JhtOOcV51muh8B34s/EsaO4WOYHcT0zVz1Qj2Gis2fXNLt2KzXsSkdRmqj+LdAj+9qUVYWGbtJ+JrmpPHXh2PpfKx9hVOX4jaGn3GL/SnYDsfxNFcHJ8TtOX7lnK/wBDVWX4pRAHy9MkOPU0WYHo1McKylWGVIwR61leH9ci1jSEviohz95SelZ+u+NtL0oNHC/2q5H/ACzTtRYDm7vQv7H8fWssI/0a4fcP9k+lemY5NeRWfiO+1nxXZPfMoh8wCNAMYr13uaGAvauA8b+KNW0fXEtbCSNYjEGO4Z5rv657xF4R0/X7hLi53LMg2hge1CA4MeOvEH/PaD/vkUv/AAnXiD/nrB/3yK6P/hW2nf8APZ/zo/4Vrp3/AD3f86u8QOc/4TrxB/z1g/75FI/jjxAyMpmgwwwflFdH/wAK10//AJ7v+dIfhrYdrh/zovEDztmLMzN1c7j9aAa3fF3h+Dw/LbpbyFxKec1gVqncQ6lptLTAKWkopgFLRRQAlLRRQAUUUlABRRRQAUtJRQAUUUUAFFFFABRRRQAUUtJQBZtfuGqc3+uarlt9w1Tm/wBc1AEdFLRTAKKKKAFooooAKKKKACiiigAooooAuaT/AMf4+lXNZ/161T0n/j/H0q3rH+vWp6gUKKKKYBRRRQAClpKWgBKWikoADRS0lAwptONNNAjv/hd/x73/APviu9Brx7w34mbw9HOi2xm84547Vt/8LKk/6BzfnWEou4HpFFebf8LMl/6BzfnSH4mTD/mHt+dTyMZ6VRXHeGfFOp+ILrEenNFar96ZumfSuxqQCiijNABRRRQAmKXFFFACUUtc/wCItR1fTAZ7GyN3CBllXqtAG9mlzXmQ+J0+4q2nMCOoPGKkHxMl/wCgc351XKwPSc1y/wAQ/wDkWpPrWB/wsuX/AKBzfnWbr3jKXXNOaza0MQbqxqlB3A51fuL9BSmmjgAegp1bgNIzhR1Y7R+Nex+E9MGl6FBCVxIRlz61514N0r+1deQuuYLfl/r2r14DAwO1Y1H0AWsLxffiy0YqGw87iIfQ9a3a8w+Il7I/iTTrUn91HIOPU5rNAaUfw20/y1zOxJ5zn1p3/CtdO6GZ+feu1hOYYz/sj+VSCq5mMytA0caJZm1jmaSLPyhjnbVbU/C1pq0xkvZ5z/so+BW/ilqbiOP/AOFdaGOiyD/gVR3PgDRoraSRRJuRcr83eu0qC8/485v9w0czAh0kbNLtl/uxgVdzXic/ijXLee4ihuwEjchRjoK9K8F3tzf6Ak94++UtyabiB0OaM009K4Lx94l1XRtTht9NkVQ6g4Izk1NrgdzLBHM6NIoJjOV9jUoNZ+hG8bSYJNQkDzyKHJA6Z7VfNAEV3cJa20txIcJGpJrwy+unvdRuLqQ5LuQD7V6R8RdV+x6OLKNv3l0drDuBXmCDaoHoMVtTQC9q2PDep31vrNjbRTkW5blM1kVc0P8A5GGx/wB6rlsB7kvQH2p1NT7q/QU6uYDiPiLqd9pyWRsZvLLk7veuN/4SjXv+fwV1HxT/ANXYfU1wXatoRTQGqfE+vf8AP6KkTW/FUq7oXZ17MB1rEb7jfQ1694IRG8LWZZEJ2nkiidkI89/tbxd6P+Rprat4u/2/yNex+VH/AM80/wC+aPKi/wCeaf8AfIqOYZ4ydX8W/wDTT8jTf7Y8Wesn5GvZ/Ji/55R/98ijyIv+eUf/AHyKOZAeM/2z4s9ZPyNOGs+LPV/yNeyeRF/zyj/75FHkxf8APKP/AL5FHMB44db8VKCXdwB1ODUY8Ta73vRXr2owxf2fcfuo/wDVn+EeleHOMTzAf3zVwtIDU/4STXP+f2mS6/rEsbJJd5Rhhh61nClrVRQAv6mlFApaoAooopiCiiigAooooAWiiigAooooADVuP/UfhVSraf6j8KQGd3NFHc0VQwooooEFFFFACUUtFAxKWkpaACiiigAoopaQgooopgFFFFMCjRRRXKAUUUUCCiiigAooooAKKKKACiiigAooooAKKKKBktr/AMfcX1re1f8A480rBtv+PuL61u6v/wAeiUgMeijtRTAKKKKACkpaSgAooooEFFFFABRRRQAUUtJQMWikpaACiiigApaSigBav6Dqq6LrCXzwmVV/hHWs80UWA9BPxQtAedNl/Oj/AIWlZf8AQOm/Osv4faNYatJeC/h8zywNvPSu0/4QrQD/AMuX61m7IZz/APwtKy/6B0350f8AC0rL/oHTfnXQ/wDCFaB/z5frR/whegf8+X60tAOePxSsv+gfL+dJ/wALRsv+gdL+ddD/AMIVoH/Pl+tNk8FaBsY/YugPejQCHw/4203Wrj7MAbec/dRz96uoFeAahH9j1i4W2JjML5jI6ivTfDXja0vLa2tb18XzDaR/eptAdliuf1Hwdo+p3r3d5HJJK3ffwPpXQ9qSpTA5X/hX/h7/AJ95P++qUfD/AMP/APPvJ/31XU0tF2Byw8AeH/8An3f/AL6pf+EA0D/ng/8A33XU0U7sDlV8A6EkiSJFIrodykP0NdPEnlxKmSdoxk0+kJ9KQATXB+NfGYtA+m6VIGuDxJKOie1UfF/jaZnn0zTFMTKSksv9BXBY5JJJJ5JPU1SiApZmYs7FnY5Zj3NdR8Oj/wAVS3+4K5fFdP8ADv8A5Gk/7lW9hHsFFFFZDPD/ABc4XxVd53dP7tY5kQsn3vvj+H3r3ubSbCeUyy2sbOepIqP+xNM/584v++arm0AtWP8Ax42//XJf5VMelCqFUKowAMClqQPPPiv/AMelj/vGvPIf+Pq3/wCui/zr0P4sf8elj/vGvPIf+Pq3/wCui/zq1sB9AQf6iL/cH8qkpkA/cR/7g/lUmKgDn/F/h+TxDpy20U4hIYHcRmuU/wCFZ3X/AEEk/wC+a9MxRTvYDzX/AIVpdf8AQTT/AL5qay+Hd3bX0FwdRRlicMRt64r0SijmYBQaKQ0gOB+K3/IPtv8AroK87PWvQ/iv/wAg62/66CvO+9aw2ELS0lFUAtFFFMAooooAKKKKACiiigAooooEFFFFAwooooAKKKKACiiimAYqzbf6s/Wq1Wbb7h+tICjL/rX+tNp0v+tb602mAUUUUAFFFFABRRRQAUUUUAFFFLQAnrW9of8Ax5tWD2Nbuh/8ebUmBnXH/HzJ9aZUlx/x8yfWo6YBRRRQAUUlFAC0lLSCgBaKKKACiiigAooooAKKKKACiiigAooooAKKKWgBDTXP7tvoadTX/wBW30NDA9b+H5/4peCumJrmPh//AMivBXTVzvcZQ1nS4NY057K5JEbHJxXLn4a6Vt2iR8fWu3pcUXsBS0nTotL0+KzgJKRjAzV0CjFLSAxPGP8AyLN57Ia810XwlqOq6XFdW0gCHjFeleMf+RYvf+uZqh8OufCsP1qk7IDk/wDhAtZx/rVpD4C1r/nopr1eijnYHCeDPC+paPrMt1eMDGybRiu7oopN3AK888X+EtU1XX3vLRgIygUfWvQ6Q/WhOwHkf/CBa7/eWnt4K8SG2W2M7CBTkIDxXrP40n4mnzMDyNfAGtqOCo+grN1rQb3Q/J+2kHzvu4r278687+Kv39N+ppxk7iOCzTWJAypIOeo606mSfc/GtWB6XofgzQ77Rra6uYp3llXLsZOprRHgHw5/z5ufq1aHhM/8U1Zf7lbFYMZzP/CBeG/+fE/nSjwJ4cHSx/WugluYIXVJpURm6BjjNOWeFhlZYz9GFAHPjwR4fHSyP/fVOHgnQP8AnzP/AH1W8Z4VHzSxj/gQp0cscq7o3DD1BzQBip4U0mNNkcUqL/dWQgVzfjvQNM0/QftNrBtmEgG8nJr0GuR+JX/IsH/rqKFuB5ZE7wyxzxf6yJty/WvZ/DGuQa3pkcqOPOUBZU7g141BHLPKkMCb5XOFX1NbNroPiqxuBPZ2UkUo9G4NaSSYHs1FedQav47hAE2miXHcCorzx3renTCC/wBN8qUjO32rOwHpVJXl5+JWof8APjSf8LK1D/nxp8rA9RJ9qT8K8u/4WVqH/PjR/wALJ1H/AJ8qOVgWPil/x9WX1ria1Nf8QT+IJInuIfK8roKy+tbQVkIUUtJS1QBRmlopgJS0UUAJS0UlABRRS0AJS0lFABRRRQAUUUUAFFFFABRS0UAJRS0UAWLb7hqnN/rmq5b/AHDVOb/XNQBHS0UUwAGlpKWgAopKWgAooooAKKKKACiiigC5pX/H+PpVvWP9etVNK/4/h9Kt6x/r1qeoFCiiimAUUUUDCloooEFJS0lABRS0UDEpMUtFAhMUwg09mVevPsOa1tI8MarrDKY4jBbn/lq3+FS2kBh8lwiAs7cBRzmuy8OeBLi8ZLnWMxQdRD3auv0Dwlp2iqHVBNckfNIwyK6EDFZSn2GQ2ltBZ26QW0YjjQYAAqbNBqveXcFlbPcXUgjiQZJNZrUCS4uIraB5p3CRoMsx7CuF0jxJJr/jopbsy2UMZCjP3z61zfi7xXNrsxtrYmOwQ9B1k96f8N8DxWQOP3VXy6AewUGkFKelQBw/ibxRdaB4qiQjzLJ4hvjHUe9dZpmp2mq2q3FlKsiHrjqPavNviYM+I4v+uIrn9K1O80i5FxYylT/Eh6MPpWnJdXA91pCM8Gud8N+LbLWkWJyILvvGx610dZtWA5jxD4M0/WN0sai3uuzrwPyrzfWfD+paI5F3Ezw54mUcGvb6jmhjmjKTIroeoYZqlJoDwJTkZBpwr0XXvAME5a40l/Ik6mM8hq4K/sLzTJjFfW7RN24yDW0ZJiIM0FsL7k4H1pMjFdJ4H0I6vqYu51/0S3OeejmnJ2Qzt/BGj/2VoqNKuLif5pP6V0lNHAwBwOBTq53qAGvI/iKxTxTC4/gIb8q9bJ6V5J8Rf+RmT6VUNwL8fxNeOJEOmsdoxnPWnf8AC0v+oW/51k+BNKstY1GaK/iMiKPlAOK7seCNA/59G/76puyYHNj4pj/oFP8AnTx8UlPXSpP++q6MeCdA/wCfRv8Avqnf8IToI/5dD/31U6Ac4Pigh/5hcn/fVKfiP9pR4U0mQs6kAA10X/CFaD/z6H/vquYl0y00v4h2lvZxbYWjBKtzzT0A42ew1GWSaQWUmZG3Yx0r1bwJBNB4cjSeMxvu6Gui8qP/AJ5p/wB804AAYAA+lJyuAhHFcjrnh+bVfGVncun+jW6hiT3PpXYUmKkBoAHAGAOlK2AMnoKXFLigDyHxPBrWsa5Lc/2fJ5K/Ig+nesa4sby0AN3btEG6E96922j+6PyriPiaANNs8AD94egrWE+gHnOKuaIP+Kgsf96qtXNE/wCRgsf96tZbAe3p91fpS0i/dH0pTXKB558UzhdP+prg88Cu8+Kn3dP+prhB0rop7AMc/I30Nev+Bj/xStn9DXkEn3G+hr1/wN/yKtp9DU1AOh7Vn6zqsGjae17c5ManBxV+ub8e20934YlitovNkLghayQFD/hZGkkcRuaP+Fj6V/zyevOk0PVQOdOb8qcdE1XB/wCJc35VpyxA9m0LWbbW7I3VqCEB2kGtKuT+HdpcWmhPHcwmJzJnaa6ys3uBW1H/AJB9x/1zP8q8LlH+kzf75r3TUP8AjwuP+uZ/lXhcv/HzP/vmtKQCCnUClroASloopgFFFFAgooooGFFFAoELRSUtACUtHeigAq1H/qPwqr2q0h/c/hQBn9zRQepopjCiikoAWikpaAEooooELRRRQMKKKKACiijNAC0UUUCCikopgUqKKK5QCiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQMltv+PuL61u6t/x6JWFbf8AH3F9a3NW/wCPRKQGPS0UUwCiiigApKKKACiiigQUUUUAFFFFABRRRQMWiigUAFFFFABRRRQIKBRSCgDvfhV/rb/6CvR68P0B9fWSb/hHwScfvK2PN8f/AN0/lWbWpR6xRXk/nfED+6fyo874gf3T+VKwHrFMk+43+6a8qM/j/wBD+VN+0ePhnhjxzxRYDnNY/wCQ7ef79SeHQP8AhJbM4HWqc5mN3Kbricn5/rV3w9/yMdn9a06CPdfT6UtAHT6UuORWQzktd8cWmi6mbKa2Z2H8QNUR8TNOP/LpJ+dcv8QVH/CVtnHTvXO4X/Zq1FAemf8ACy9P/wCfWT861/Dfi218QXEkMELRsgycmvHOP9muz+FhH9sXAGPuHpQ42A9TzTe4ozxR3qAPBtb/AORgv/8Arsapirmt/wDIfv8A/rsaqdq1Qgrpvh5/yNJ/3a5k10vw8P8AxVR/3RQ9gPYaKM0VkMKK8617x5f6brc9lFbBkj6GqI+JWogrmyGCQKdmB6nRUVtKZ7aKUjG9A2PqKlpAee/FcZtLH/eNefW65u7b/rov869D+Kv/AB6WP+8a8/tRm+th/wBNF/nVrYD3yAfuY/8AcFSUyL/Vp/uin1AEF3eW1nHvuZVjU9zVYa3ph/5fI/zrlPip/wAgeHkj5uxrzQoB/HJ/31VKNwPdf7b0z/n8j/OlTWdOdwiXcZYnAGa8HKf7cn/fVWdKXGs2Xzyf65f4qfKB79mkPSkoqQOC+K//ACDrb/roK87716J8V/8AkHW3/XQV53/Ea0hsIWiiiqAWiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAqzbfcP1qtVi3+4frSApS/wCtb602nS/65vrTaYBRRRQAUUUUAFFFFABRRRTAKKKKADsa3dE/482rC7Vu6J/x5tUsDOuP+PmT60yn3H/HzJ9ajpgLRSUtACUUUUALRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRQAE0xslWHqMU4000Ad54V8XaVpOiRWl05WVOvFbX/Cf6F/z2P5V5OQO6j8qNo/uL+VZuIz1n/hP9B/57/pS/wDCfaABk3HT2ryTav8AcX8qbIo8p/lXp6UcoHv+n3sGoWcd3bNuikGVNWawfBXHhaywMfJW9WTAxfGH/IsX3/XM1n/Dr/kVYPrWj4uGfDV6P+mZryfTPEes6fYR21ldLHCBnBFUldAe40V4yPGPiIf8vy/9804eMfEX/P8AL/3zRyMD2SiuO+H+ranq1tPNqM4lCnC4GK7GpasAUlBri/iHqeq6TDaT6bcCJXcq/GaAO0NJXi//AAmPiP8A5/l/75pf+Ex8Rf8AP8v/AHzVcjC57PmvO/ip/rNN+prnf+Ex8Rf8/wAv/fNUdT1fUNXaM6jMJTH9zAxiqjGzEUqZJ938afTZPuj6itGB7X4U/wCRasv92tgCsnwuu3w7Zj/ZrXFYMZw3jLw/ea94isIoJHigWM+ZKp6VrQeCtJhsxADcE45fzTkn1qHxN4qHh7VrSGSDzIJ1JcjqvvW3pmr2Oq24msrhJFPbOCKAPPvEHgPUbeOSfTL2e4Tk+TvOQK6P4dJJHoDRzbxIr4Ic5IrpL++t7C0kubmVEjjXJyetZPhPVI9Xtrq8hh8qJ5fl46j1ouBv1yHxL/5Fj/tqK6+uQ+Jf/Isf9tRRHcDz3wz/AMjHY/8AXQV7h614d4Z/5GSx/wCugr3L1qp7gFec+PPD+sapr6T2FsZIhEF3Z6GvRqb+NSnYDxseCfEZ62+Pxpf+EH8Rf88P1r2Pv1P50Y9zVc7A8d/4QjxD/wA8P1psngzxBEhc2/CjJ5r2P8TUN6D9jm5P3D3o5mB4NyCyt1U4P1pwok/4+J/+urUorZCCiloqgCkoooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACloooAKKKKALFt9w1Tm/1zVbt/uGqk3+uahAMooopgApaSloAKKKKACiiigAooooAKKKKALelf8AH8PpVzV/9etU9K/4/h9Kt6t/r1qeoFHvRQetHWmAUUUUALRRRQAUlLRQAUlFLQMSnxKkk8ccsvlRuQC/pTDQQCMHoaGI9V0HwhpFlHHcKouZSM+aeQfwrpVRVGFAUDsBgV5x4E8Sm1lGlX8n7pv9S5/h9q9JBrmle+oxMUUkkiRoXkYKo5JJxTYZo7iJZYXDI3Rh3qQM/W9csNFtTPfTBePlQdWPpXkfiPxRdeILg+YWitVPyRD+tesX/hrStQuDPeW5lf3bj8qrf8IfoI/5cf1qotIDxcOv+RW54O1W00rxB9qvHKRFNu7Fem/8IhoP/Pl+tL/wiOg/8+IP1qnMCt/wnnh0Hi8z/wABpf8AhPPDx/5fD/3zVj/hEdB/58Fpw8JaEOlglToB51441ax1XWY7ixlMsYj2k471z/mL7/lXtA8LaKOligpf+EY0b/nyWqU7AeLeaFcOjOjqchl4Ndr4a+IDW+y01rLR9BOB0+tdr/wjOj/8+SU1vC2isMNYoRSckwNS2uYbuBZreRZI2GQQamrBubODw7pN1daZGymNd2wtkGofDXi+w1yMISILr+KJjUWA6PFUdXjsPsEr6kqGBVyxYc1eZgqliQABkmvJ/GniJ9YvTaWzkWcJwSP4zVRTbA57UXtJdRkaxjMFmzgKpOePWvZfD1raWui28VgQYdoIYdz3rxUqCMY4xivQPhrqxaGXSpn5i5iz3rSadgO+xRRRWICHtXkvxF/5GaP/AHa9a7ivJ/iOP+Kki/3TVw3An+GH/IYuOR0NeoZ9x+deE6UdVFw39jZ83+LFau7xr6SU5RuwPYsj1H50bvcfnXjufG3pJQD429JKnlA9jyPUfnXCayQvxJstxAzGOprmAfG3o9ZepNqi3anViyXOPlOecVSiB7n50QPMqA/7woE8OcebHk/7QrwRriYDLXM3/fZrtvAvhu5uZV1TUXnESnMUbOeT60nGwHpNFFB4qAEZgoyxAHqTTfNT++n/AH1Xmvj7xMbq4/svT5sRxn97Ipxk+lciLi5AAF3LgdPnq1C4HvHmJ/fX/vquK+JjA6ZaYZT+8PQ158Lm6/5+5f8Avs0jyzSYEsryAdNzZxVxhZgJVzRP+Rgsf96qdXNE/wCRgsf96tJbAe3L90fSlNIv3R9KDXKB578U/u6f9TXCV3fxT+7p/wBTXCA+4rop7ANcfI30Nev+Bx/xStn9DXkTYKkZHI9a7vw74z0vS9Ft7Kc/vIxhqVRXA9DxSEAjkCuR/wCFhaL/AHzSH4h6L/fNY8rA67Yv91fyo2L/AHV/KuQ/4WJo3qaQ/EXRvU0crA7EAAcAD6Utcb/wsXRf7xpw+Imif3zRysDqL/8A48bj/rma8MlH+lT/APXQ16Rc+P8ARZbaWMOcspArzd2DzyuCNrtkc1rSVgAUUceopa3ASilpKYgooooAKKKKAFpKWkoAKWkpaBiUUUtACVaT/U/hVU1ZT/VfhQBQ7mijuaKYBRRRQAtFJRQAUUUUCFooooAKKKKBhRRRQAtFFFAgooopjKNFFFcogooooAKKKKBBRRRQAUUUUAFFFFABRRRQMKKKKAJbb/j7i+tbmq/8eqVh23/H3F9a29V/49VpAZHaiilpgJRRRQAUUUUCCiiigAooooGFFFFAgooooGLRRRQAUUUUAFFFFACUClNNzQB3nwpP+kagP9kV6RXlHw71Wx0q4vWv7gQiQDbnvXc/8JfoP/QQX8qza1Gb1FYP/CX6D/0EF/Kj/hL9A/6CC/lSsBvYFI4+VuB0NYY8YaD/ANBBfypG8YaBgj+0F6elIDyXXB/xUN9/vVJ4eH/FR2f1qLVpo7jWrueFt0TtlW9am8O/8jJZ/Wtegj3MdPwpaQdKWshnLa54KsdZ1E3k8jK57Cs//hWumf8APV67mindgcL/AMK10z/ns1avhzwjaeH7x7i2kLF12kGukxRii4CdqO9Lik7ikB4Nrf8AyMF//wBdjVQVc1r/AJGDUP8Arsap1qhAa6T4ef8AI1H/AHRXNmuj+Hv/ACNX/AaGB7D3NLSdzRWQzxLxj/yNt39Kxmz8n++v863fF1tdP4qunS3kZccEKax2tLzC/wCiy/eH8J9a0Wwj3rTv+Qfbf9cl/lVmqum5/s22yMHy14/CrVZjPP8A4q/8edl/vGuC09d+qWi+riu9+Kv/AB52P+8a4vw9F53iOxjx1OatbAe5R/cX/dFOpAMAfSlqAOF+KSu2kwiNGc7uijNebGKf/n3l/wC+DX0C6I4w6q31Gaj8iH/njF/3wKpSsB4B5U//AD7y/wDfBqzpUcv9s2WYJR++X+A17t9nh/54Rf8AfAoFvCDkQxAj/YFNyEPAp3agUdqgZwPxX/5Btt/10Fedd69F+K//ACDLb/roK87/AIjWsNhBS0lLVAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigAoFFFABRRRTAKsW/3D9ar1Yt/uUAUpf8AWt9abTpP9a31ptABRRRQAUUUUAFFFFMAooooAKKKKADsa3dE/wCPNqwq3dF/482qWBnXH/HzJ9ajqS4/4+ZPrUdMAooooAKKKKAFooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKQ0AIa1ND8P3evCY2bhPJOGrL713fwrP72/H+0KmTsgM7/hXusH/luKP+Feax/wA/Ar1fH1ox7msuZgeUf8K81j/n4FNf4d6wyMPtC8jFes49zRj60czGZnhywl0zRbeznbdJGuCa1KSlqQMbxb/yLd7/ANczXikX+pT6V7X4u/5Fu8/3DXikX+pT6VpAB+KOnNLTZAShVfvHgVoI9T+GtuYfDe4jl5Ca6+svw5aiz0K0iAwTGGP1IrUrB7jErmfH1kLvwzO4GWhG4V09QXlul1aS28gysikGhMDwBOUX6c07FT3ds9pfXFvIMFJDge1R4rdCG4p2KdjijFMBmKa4zsHq4FSYqzpdm1/q9raIMlnDflSewHs2iRGHR7WMjkIK0BTY0CRqgHCqBT652M8+8f6FqWr6zZ/YLfzEVDubPAqvpHw91CB1uJNTa3cfwL0r0KG7gnnmhikDSQnDjuKmp3A5mbwfbX+3+1riW5C9FDbQa37Gyt9PtUtrSIRxIMBRS3d3b2cDzXMqxxoMkk1BpWq2mrWxuLJ98QOA3rSAvVx/xM/5Fn/tqK7CuO+Jx/4pgf8AXUU1uB574a/5GSx/66Cvcu5rwzwz/wAjLY/9dBXufc1U9wA9K5/XfFmm6HeC1vGPmMu7AHaugPSvJvifx4liJjYjyRyFzUxQHSn4jaN/tflR/wALH0j+635V5XvX/nm3/fFLuX/nm3/fFacqA9RPxI0kdI2P0FRTfEbTHieP7PICykDcMVyfgfTm1DxHE32fdDAdz7kwMV6fqfhnSdUTbc2aZ/hZRjFS7JgeME75ZH7O5YfjTsV2mq+ARY2k11Ff/JENwQjt6VxYOc/WtotMQUUUVQBRRRQMSiiigQtFJS0AJS0lFABRRS0AJRRRQAUUUUAFFLRQAlFLRQAUUlLQBPbnCGqkv+uarcH3DVSb/XNQgGUUUUwFoopKAFopKWgAooooAKKKKACiiigC3pX/AB/fhVvVv9cv0qppX/H7+FW9V/1y1PUCjRRRTAKKKWgYUUUlAgpaSigYUtJS0AFJS0lAhOeCDhlOVPoa7rTvH62ujpHdQtNdxjHB+971wtJjmplFMDS13xJqetbhPMY4OcRLxivVPCS7fDGnj/pkK8Wk+4fpXtXhQ/8AFNWH/XIVlNWGa9NJAGSQB6mlNZuvwT3Gi3UNsCZnQhQDjmoQF3zov+esf/fQo86L/nrH/wB9CvF08KeLAgHkTD6yGpB4U8V/88Jv+/pqrAey+dH/AM9Y/wDvoUedH/z1j/76FeN/8Ip4r/54Tf8Af00f8Ip4q/54Tf8Af00coHsvnxf89Y/++hR50X/PSP8A76FeN/8ACK+Kv+eM3/f00j+FfFmw4hm/7+mlyge0g5GRyKKpaMksWk2sdwCJUjAYH1q7UgZfiUZ8PX3/AFzNeIwgr86MUcNwynBr2/xH/wAgC9/65mvEo/uH/erWmgN0+LNWfSX06WXcGG3zB1ArEAwMfnS4pcVqopAJV3Q746Zrlrdg/KG2t75qnSEZx7HIoaugPXPEviKPSre2SIg3NyyiNfY9a3kJZFJ6kA14YLi4u9UtpbqUyMrqFz25r3OP/Vr/ALornlGwC4ryr4jr/wAVHD/umvV68r+I3/Iwwf7tOG4DvhiP+JxcDAPBr1AKP7q/lXk3gPUbPTNTmlvZhErAgGu//wCEt0P/AJ/1/Kia1A2to9B+VGB6D8qxD4t0L/n/AF/KkPi7Qf8An/X8qkDbIH90flXlfxHV5fFNvFFGXdosKqjrXcf8JfoP/P8Aj8qm05tH1S6Op2ZSeZRs3nkgfSmnYDlPCngZgyX2tLyOUg9Pc16EiqiBUUKo4AHagUuaTbYC1W1G3kurKWCKYws4xvHarGaWkBwY+G1rks14zMxyxx1NOHw4tP8An7b8q7qiq5mBwrfDq1AOLwj8K4CVBFcSxA5EblQfXFe7yfcb6GvCrs/6fdf9dW/nWlNtgR1b0T/kYLH/AHqp5q1op/4qGw/3q0lsB7iv3R9KDSL91fpS1yged/FY4SwPpk1S8MeD7fW9HjvZbgoz9VHarnxX/wBXY/Rq2/hwP+KVgP1rS9ogUB8OLL/n5b8qP+Fb2B6zk/hXdUVPMwOF/wCFa6d/z2P5Uf8ACtdO/wCe36V3VIaOZgeV+KvBtpomkm9hlDsHC7cetUvCHhi28QpO07iMxNgDHWu0+JH/ACKzf9dRWR8K/wDV3f8Av1V3YCf/AIVnY/8APz+lH/CtLH/n5/8AHa7sU6p5mBwg+GtkOlz/AOO08fDmzH/L035V3NFHMwOGf4eWixswujkDPSuBmTyrmaIHIjcrn1r3Sb/Uv/umvDLw/wDExu/+uxrWnJsCI0UUVuIKKKKBhS0lLQAUhpaSgAooooAKWkooAKsp/qvwqtVhP9V+FAFH+I0UfxGimAUUUUAFFFFABS0UUAFFFFAgooooGFFFFAgpaSloAKKKKYFGiiiuUAooooEFFFFABRRRQAUUUUAFFFFABRRRQMKKKWgCS2/4+4vrW3qn/HqtYlt/x9xfWtvVP+PVaAMiiiigAooooAKKKWgQlFLRQAlFLRQAlFFFAxaKSloAKKKKACiiigAooooAQ00040lAGz4b8KS+JIpJY7ryfL7Y61s/8Kvuf+gh+laPwp/487oe9egAcVm3qM8v/wCFX3P/AEEf0pR8L7j/AKCP6V6fS0rgeYj4X3H/AEEf0p3/AArC4/6CP6V6YKDRdgeEazpbaNqklg8nmlP4/Wm6XdJY6rb3cgykZ5Fa3j7/AJG24+grnq0WqEeqwfELSJp44VDbpGCjjua7AHNeD6LF5+u2caqM7wele7is5Kwxaw/EniW18PpG1ypbf0xW5XmXxakzNaw/9Myf1pID0LS7+LU9PhvIP9XKuRVqsLwT/wAirZf7lb1AFDWNUg0iwa8uP9WvWuYX4k6QwB2sPwrQ8fjPheb6149Go8pOB09KqKuBYv51u9TurpBhJpCy/SoaXFGK0Qhpro/h7/yNX/Aa5w10fw9/5Gr/AIDSYHsPc0UdzS1kMYY0JyY0J91o8qP/AJ5p/wB80+loABxRRmigDz/4q/8AHpY/7xrmvAcHneLbZsZEakmuk+Kp/wBFsf8AeNVfhZZ77q8vGHCYCmrWwHpdFFIagAJpPwNeceIvHl9Ya5cWVpAGjhON3rWePiLq/wDz7CnYD1jPsaPzryf/AIWLq/8Az7ip7L4g6pPf28D24CySBSfrRygeo5o7Ugp3akBwHxX/AOQZbf8AXQV52eteifFf/kGW3/XQV52eprWGwgpaSlqgCiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMA7VPB9yoKnh+5+NAFOX/AFrfWm06T/Wt9abQAUUUUAFFFFABRRRTAKKKKACiiigArc0X/jzasPtW5ov/AB5tSYGfcf8AHzJ9ajqS4/4+ZPrUdABRRRQAUtFAoAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACkpaKAG4rtvhlPDBcXomkVNx43HGa4qjBzwzD/dOKUldAe9Q3dtM+yGdHb0BzU9eV/DKMtrk0m5yFUjls16oKwasxhmqrahZoxDXUQI6gsOKnmOIXPoprwHWSW1XUH3yffJ+9QlcD35JUkQNGwZT0Ip2awfBQx4Wsskn5O5zW9SegGL4vP/FNXn+4a8VhP7lPpXtHi/8A5Fq8/wBw14tD/qU+lawAkzWh4fs21DxBaW4GVVwz/Ss769K774YaWQs+rSry/wC7XNOT0EehxqqIqKMBRgU6mA04ViMWio4Zo5lJicMAcHHrUlAHl/xI0s2uqR6jGv7u4+RsdiK5Cva/EWlJq+kTWjAb2X5G9DXi0kTwSvBKpWSJipB9q1gxCUUUVoAnbNd58ONEfe+r3CYHSDPp3rgzzxXT+HPGd1pOy2vF86zHAwMeWKmd7aAesVHPIsMDyscBFLflUGn6jaalbLPZzLIjeh5/KuX+IWuCzsBp1u/+kT9SD90VildjOV0rxL9k8X3OpyufstwTvFamtfESZi0elQbUHSY85/CuIAAGMU1hlSPUYrXkQHoMfhW48TafFfX2tSOZl3YUYA9q6nw3oaaDpxtI5TKCc7sYrC+GuofaNGltmcboH2qCe1dkWGPvL+dZMBc1xnxQfHhpB6zCusmuYII2eaeNVUZJLCuB8TteeMpo7PRIy9nE2ZJm4G729aaA43Q5hb69YyscDzQCfSvd1YMNynIPINeTa74GvtM05bu1lNwyDMigcp71veDvGttPaR2OpyeVPGNqMf4h705a7Ad5UFxZ211j7RCkmO5FOjnikUMksbA+jVKCD0IqAKH9jab/AM+kf/fNL/Y2m/8APnH/AN81foouBDb2sFsu2CFIx/sjFTUhIHUgfjUFze21rC0txPGiKMk7qAMHx7ei18OSx5w0/wAq15Mv3R9K6Hxp4hTXL5IrUk2sByrf3jXPVvBWQC0UlFaAFFFFAhaSiigAooooAKKKKACiiigBaSlpKAClpKWgAooooAM0lLSUALRRRQMnt/u1Ul/1rVag+6aqy/61qBDKWkpaYBRRRQAUUUUAJS0UUAFFFFABRRRQBb0v/j9/Creq/wCuWqml/wDH6PpVrVf9ctLqBSooooGFLSUUCCilpKBhS0UUAFJS0lAC0lFFACUhpaQ0ARyH5D9K9o8KH/imtP8A+udeLyj5D9K9o8Kf8izYf9c6yqAbIoxQKqatdPZ6Vc3MYy8UZYD3rEC3+dH515InxH1cqCYDz/s07/hYurf88D/3zVcrA9Zoryf/AIWLqv8AzwP/AHzR/wALG1X/AJ9z/wB80crA9X/Ok5968oPxH1b/AJ9z+VMb4j6uFJ+zn8qOVgeuUtUdIuXvdLt7mUYeRAxFXSakDN8Rn/iQXv8A1zNeIxH5D/vV7Z4j/wCQBe/9czXiUX3D/vVtTAlp1NFLWwBRRSUAS2v/AB/W/wD10X+de6x/cX/dFeFWn/H9b/8AXRf517rH/q1/3RWFTcB9eV/Eb/kYIf8Adr1SvK/iN/yMMP8Au1MNwM7QPC9x4gillhn8sRNtPvWqfhren/l9H5Vq/C7/AI8b7/roK7unKTuB5cfhne/8/wAPypp+Gd9/z+j8q9ToqeZgeVf8Kzv+16Pyp1vpuoeA7yG/ln820mcRSj0z3r1OuL+KC7vDSg/89RQncDqft1otvHcNOixSAbCT1qwrB1DKcg8g14Gb28uGs7eed2hicbFBxivdbD/kH22f+eYokrAWa86v/iJd2uo3FqmnBxE5UNnrXoteX6l4B1S51W5uY5gElkLAUo26gaugeOrrVdYhspLARLIcbs13ledeG/BOo6ZrcN5PMGjQ5Ir0TNErdAEk/wBW30NeEXR/0+6/67N/OvdpD+7b/dNeDXJ/0+6/67N/OtKQDc1b0TnxDY/71U6uaH/yMNj/AL1ay2A9xX7q/QUuKRfuj6CnVygedfFf7lj9Grc+HP8AyKkH41h/Ff8A1dj9Grc+HX/Ip2/41b+EDqa4P4i61qWl3Vkun3Pkh1O/jrXeV558S9Pvr68sTZ25lCqd2O1KO4Evw61rUtUuLpdQufNCH5eOld5Xn3w102/sbm6a8tzEGPGe9ehAUS3A5L4kf8iu3/XUVkfCr/V3f+/Wz8SB/wAUw3/XVayPhWP3d5/v0/sgeggcVyHxF1S/0vToJNPn8l2fBOO1dj2rivibaXN1pcC2sRkYSAkAVMdwMPwl4i1m88RR293eeZER93Feo15F4Q0+/h8TxSTWzogHUivXKc9wGzf6l/8AdNeFXh/4mN5/12Ne5zf6l/8AdNeFXf8AyEbz/rsa0pANFLTRS10ALRRRQIKKKKACiiigYUUUUAFLTaWgAqwh/dfhVep1/wBX+FAFLuaKO5opgFFFFABQKKWgQUUUUAFFFFAwooooAKKKKBBS0lLTASilooGUaKKK5RBRRRQIKKKKACiiigAooooAKKKKBhRRRQIKWkooGS23/H1F9a2tU/49lrFtv+PuL61tap/x7LQBk0UUUAFFLSUAFFFFAgpaSigBaSiloASiiloASlooFAwooooAKKKKACiiigAptONNoA9C+FB/0W7r0EHivPvhQP8ARruvQQOKyluMyvEOtxaFYfa5oy656Vyx+J1hxi1Y5IFX/iaP+KYb615IQNkXT7y00rgfQlpcC5tY5wMCRQQKmz0+tU9IH/Eptv8ArmP5VbPUfWkB454/P/FWz/QVzwNdB8QP+Rtn/CudJwCfStFsI6j4e2v2rxSkpGUhQ5+tew1xPw10hrLSnvZlxJdHcoPUCu1rNvUYteRfEu6Fxr4jB/1MZU+1esyyiKF5G6IpY/hXg2uXv2/VdQus5WRiVpxA9d8DnPhWz/3K6Cud8CHPhW0/3a6KpYHN+Pf+RYnryXTrSe+litbVd0zDgetetePP+RYnrzjwVz4ktPoKuLsgM26tbiymaG7heORTg8cfnUPXoQfpXvF9ptnqEZju7dJF9xXF678P7cRyXOmTGHaN3ldqakKx50RXR/D4f8VV/wABrnmBV3Q9Ubaa3/A00Nt4k82dwiBepNN7AexY60nes/8At3S8/wDH5H+dH9u6XnH2yP8AOsrDOH13x3qen61PZw26skfSqI+I2rn/AJdV/OsHxJKlx4iupYm3I3QjvWcAa0UUB2P/AAsbVv8An1Wk/wCFi6tj/j1WuQwaWnyoRreIPEl5r0cSXUQQRn5cdya9L8CaadO8OQiRcSyjc1eaeFtIbWdcihwfIjO6U+npXtsahEVV6KMCol2GOqG6kENtLKTgIhP6VNXMeP8AUv7P8OTBGxLL8qj1HepQHkV5Obu/ubo9ZXNNHSmqu1QPxp/atkIbVnTP+QxZf9dlqvVjTP8AkMWX/XZaTA98XoPpTu1NXt9Kd2rIZwPxX/5Blt/10Fedd69E+K//ACDrb/roK88PWtYbCEpaKKsAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAVPD92oKnh+7SApyf6xvrTadJ/rG+tNpgFFFFABRRRQAUUUUwCiiigAooooAK29GP+iNWJ2rb0b/j0akwKFx/x8SfWmU+f/j4k+tR0AFLSUtACUtFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUtABSUtNY4Rj7UAd98LIPlvpyP48A16HXKfDuyNt4dWVhg3B3V1dc8txkN422znb0jJ/SvAb1vMnu5P7zE17l4guVtdDu5GOP3ZA/KvCSSbdiepBNVAD2rwWP+KXsv9ytysXwZ/wAitZf7lbZqWBieL/8AkWrz/cNeKw/6lPpXtXjD/kWbz/cNeKRMFt0J7D86uAFm0tJdQvYbKEEtMwU47D1r3DS7GPTtOhtIwAI1AOO59a5L4eeHWtYTqt4mJphiNT2X1rugKUncBMVXv7pLGxmu5D8sSljVquF+Jmr/AGbTU06Jv3k5xIB2WpWoFL4ca6Z9SvrKdziRzJFnvk9K9GHSvAdMu30rULa9i6wMM+4r3Wxu47yziuYmDJIoOR605KwFnNee/ELw8wb+2bNM44nUenrXoVNljSWNo5FDIwwQe9JOwHgYIIBHSlroPF3huTQ7wzwKWspTkEfwGufrdO4CUUuKRuFJ/KmBteDodQm11YdNuXtwVPmP1A/CtDxJ4R1iC6lvhIb5G5Zz1H4V0vw80c2Okm8mXE1182COgrsCM1i5ageAE4JVgykdQwwav6ZoOpaxFJLp0QkWM4bJxzXp/iLw1pmpW8s80axTKpbzV4xisH4XMPI1JEYsizfK3rVc+gHP2nhbxbYymSyiELnqQ3WprjS/HSwvJLL8iDLHd2r1fHuagvBus5x6xkfpUXA8g8L6c/iPV/J1O/cRR/MyF8b/AGr1+1tobS3SC2jEUaDCqBivB5S8GozNDI0ckcnysDjFel+C/F66kq6fqLBLxB8rH/loKckB2RUEEEZB6g968z8deEPsrvqmmxkwscyxr2PrXpuap6tf2Wn2Ek1+6rFjlT1b2xUpgeHW11d24BtbuRB2yxNX01/XYxhdSbH0qpfTQXGozTWkXlQyPlE9q0b/AMN6rp9tHcvAZIJFDbl5I/CtdAGf8JLr4/5iR/Knwa/4nupRFa3skjnphePzqjpz2i6pbrqaMLUviQMMYFe1aVp+nWdqn9nQRpE4DAgZyKmTSA5DSfDvia7HmazqzQr/AM8l5zXHeJLC+0/Vntb+aSQfejbccEV7fWB4t0FNc0wqqj7TF80Te9TF6geOgY6U6lZGjkaKQEOh2sD60ldCEFFLSUwCiiigAooooAKKKKACiiigAopaSgBaKKKAEpaKKACijrRQAZopKKAFFFFFAyaH7pqrL/rWq1D92qsv+tahCGUtFFMAooooAKKKKACiiigAooooAKKKKALWl/8AH6PpVzVf9ctU9L/4/R9Kt6p/rlqeoFKilpKYBRRRQMKWkooAKKKKACiiigAopaTrQAlBpcUYpiIpf9WfpXs/hMf8UzYf9c68akHyH6V7D4Xu7VfDlipuIgRHggsARWNQZvCmyxpLE0cihkYYIPeoft1p/wA/MP8A32KT7faf8/UP/fYrGwFT/hHtJxj7FH+VH/CP6T/z5R/lVv7faf8AP1D/AN9ij7faf8/MP/fYp6gVP+Ef0n/nyj/Kj/hH9J/58o/yq59utP8An5h/77FH220/5+Yf++xRqBT/AOEe0n/nyj/Kmnw7pB62SflV/wC22v8Az8w/99ik+22v/PzD/wB9ijUCSKJIo1jjXaijAHpT8VB9ttf+fmH/AL7FH260/wCfmH/vsUrAU/EQ/wCJBe/9czXiMf3D/vGvateuraTQ7xUuIiTGeA4rxiMfKf8AeNbUwHCloorYApDTqQjigB9r/wAf9v8A9dF/nXu0X+rT/dFeEW3/AB/23/XRf517vF/q0/3RWFXcB9eV/EY/8VDD/u16nXlHxHP/ABUcX+7Uw3A3Phcf9Dvv+ugrvDXAfCs5tb8f9NBXf9qUtwMHxL4mg8P+V58ZfzOmKj8OeK7bX7h4oIihQZOa5z4q4/0Q8f5NVfhcR/aVyBj7tFtLgen1x/xLGfDQ/wCuorsB0rkviR/yLX/bUUo7geUxr/pNt/10Fe9acP8AiX2//XMV4VEP39t/10Fe7WH/AB4W/wD1zH8quYE9J+dL3Feb6j8Q7201O4tVsVZYnKhs9ahK4Ho/50tef6B48u9U1mGxlsxGshxuzXf55NDVgGyf6t/9014Rc/8AH/df9dW/nXu8n+rf/dNeEXI/0+6/66t/OtKQDO1XNFO3xBYk9mqoKltJPJ1C2l/uuK1ewHuyfdH0FOqOE7oY29UBqSuUDzr4r/6ux+jVt/Dlh/wikA9M1T+Juny3WjpdRKW+znkAc81z3hDxjbaJpv2S8XKj7rDrV7oD1XNJwewP1rjP+Fj6R6N+VKPiNo/o35UuVgdoAB2Apa40fEXR/wDb/KlHxE0f/a/KlysCb4j/APIsN/11Wsj4Wf6u8/36r+K/Fum61ozWdsW8wuG5HpWf4N8QWegLOLwnMhyMVai+UD1qkIB6gH61x3/CxdIHXd+VIfiNo4/vflU8rA7HaAeg/Klri/8AhY+j+rflR/wsfRvVvypcrA7Cb/VP9DXhV2P+Jlef9djXorfETRmRly2SMdK85mkWW7uJlPyySFh9K2ppoBoFOoGPWlrcBKKWkoEFFFFAwooooAKKKKACiiigAqdf9V+FQVOv+q/CgRS7mijuaKYwooooEFLSUtABRRRQAUUUUDCiiigAooooAWiiigQUUlFMZSooorlEFFFFAgooooAKKKKACiiigAooooAKKKKBhRRRQBJbf8fcX1rb1P8A49lrEtv+PuL61tan/wAey0gMukoFFMApaSloASiiigQUUtJQAUUUUALSUUtAwopKWgAooooAKKKKACiiigAoxRRQB6D8Kf8Aj3u/rXoXavPfhT/x73f4V6FWT3GZfiHRYtc05rOaQop7iuW/4VlY4UfbZOCD09K72ii4ENrALa1jgU5EagA089R9afSHGaQHjPxAOPFs+eBgdaf4O8MS65eLc3ClLGE5JI++a6u+8KWWt+Kri7ubvcsZG63HU12Vtbw2sCQW8YjiQYVR2qr6APijWKNUjUKijAA7U40U2aVIomkkYKiAsxPpUgcz4+1f+y9AkRG/fT/IF9QeteOsNsDD0Brd8W622u600qsfs8OVi9xWG4zGw9qtIR7N4COfC1t/u10lcp8OphJ4ZRQfuHBrq6l7jOf8cQzT+GrhYELt1wOtcJ8NtPmn17z5IZEihj6suOa9bIBGO1IkaJ9xFXPoMUXAdVDXJPK0W8kzjbETV+ud8dXotPDVwCcGYbBSQHju7cWf+8c0EAjnP4GkQYjUdwKWtkIb5af7f/fVN8tfV/8AvqpMUUWAaBjgUtLRTASljjkmmSCBC8shwqilVHkkWOMAuxwAa9Q8F+El0lBfX4V71xlR1EY9qluwGn4Q0BNC0tUYA3MozK1dBSCjNZbjFNeRfEHVxqWuC1ibMVpxx0Ymu68Z6+miaQ5RgbmUbYl/rXjgLElnbc7HJJqooB1Xk0bVZI1kjs2ZG5Ugdao9VI/KvUvhzq5v9Ha0lIMlodo46irk7CPOTourj/lxf8qtaLomqya3ag2bKqSBmYjoK9t2j0H5UBR6D8qjmGN6GlzkUuKaeAakDgvisf8AQbYf7YrzzvXb/FS5Vrq1tActt3muHzzWsNhDqKQUtWAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAWpYfu1DU0P3aQFST/AFjfWm06T/WN9abTAKKKKACiiimAUUUUAFFFFABRRRQAdq29H/49GrE7VtaP/wAerUmBRn/4+JPrTKfP/wAfEn1plABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABS0lFABUtrbveXsFrECWdxke3eoq7L4b2FtLfSX80qedH8scZPX3qZOyA9GsLZLKyhto/uxqAKsU2g1gM4/wCJl8LfQBbqf3krjj2ryqXAicf7NdP8QdT/ALQ8Q+QhzHaDYfrXLSn90/0rWK0Ee2+DP+RWsv8ArnW5WF4N48L2P/XOt2s3uMw/GHHhm8J/uGuB8BeFX1NotQv0K2kfKKf4zXpurratp8n2/wD49gP3n0qWw+zGyh+x7fs+weXt6YoTAmRAqhVACgYAHanUUGkBDd3MVpbSXE7BY413Ma8P1zUpNY1ia9kOQTtT/d7V1XxC8SC5l/siyf5FOZnHf/ZriMAAAdBWkI9QExkYIru/htrvlltGun4BzCT3PpXDY4pYpJYJ47iBissZypFXJXEe/ZpwrD8K69DrmmpICBcIMSJ3+tborFjILy1gvbZ7e5jDxOMMDXlHibwvc6JO0sKtNZMflYDJX2r16uL8eeJo7G2bTbba9zMMN32CnFu4Hmta/hbRn1vWEjIP2eE7pG9+wrFBAABONx5PpmvYfB1lp9lo0a2MqTFxmRwfmY+9aSdkI3o0WNFRAAqjAAp1JmmSSLHGzuQFUZJPpWIzl/iFq/8AZ+hmCJsT3Bwo9u9YnwnbEV/H/t1zHivWDrWuSTKSbeE7YvpW58LZwmp3NuTy4LAVpayA9QpkyF4ZEHVlIp4pazA8F1a2uLbWLyOW3lyJDghDg1UVpo5FkjjnSRDlWCHg19APbwynMkSN9VFM/s+z/wCfaP8A75FXzgcTpvj0L4e8y7tZ3vo/kCbD8/vXE6vqmp63deffR3BAPyRhDhRXt62lsvS3jH/ARS/ZrcdII/8AvkUkwPE9C0q81LVreKK3kCBwZGZcACvbYoljgSIYIRQozTljRPuIq/QYp1Ju4GLq/hfSdWRhc2qh2HDrwRWb4Z1GGx1GTw3JP5skA3RyN/EPSupkkWKNpHOFUZJrwy+1KYeIptSt2IlimJQjuM00rge7UVmeH9Wi1jSYbuJgSww49G71pZqQPOPiHoPkTDVrVPkbiZQPu+9cR15Fe069faXBp8sWpzIIpBtKg5P5V4w+wTSCI5i3HYfbtW1N6ANpKdSVqIKKKKACilpKACilpKACilpKACilooAKSiigApaKKACkopaACiiigAooooAmh+6aqy/601ah+6aqy/600IBtFFFMAooooAKKKKACiiigAooooAKKKKALWmf8fv4Vb1T/AFq1U0z/AI/Pwq3qn+uWp6gU6SlopgFJRRQMWkoooAKKKKBC0lFFAwpaSgGgRp6RoV/rKSPYKrCI4bccc1fPgfX/APnlH/33W78L+be//wB8V3uKxlNpjPJD4H1//njF/wB90w+BfEHaID6S169ijFRzsDx4+BPEP/PL/wAi03/hBPEP/PL/AMi17FijH1/OjmA8d/4QTxF/zy/8i0f8IJ4i/wCeX/kWvYce5pce5p8wHjv/AAgniL/nn/5Fpw8C+Iv+ef8A5Fr2DHuaMUuYDyD/AIQTxD/zz/8AItH/AAgviH+5/wCRa9fxRijmYHkB8C+Iv7n/AJFo/wCEF8Rf88//ACLXr9GPrRzMDyEeBfEP9wf9/akXwPr4GPJj/wC+69bxRinzsDyb/hCNe/55R/8AfVU9T8Narpdqbm9iRYh1IbNey4rlviF/yLUv1FNTdwPLe2aDSD7q/QUvatwH2o/0+2/66L/OvdY/9Wv+6K8LtP8AkIW3/XQfzr3WP7i/7orCruAp6ivJviP/AMjJF/u16weoryb4jf8AIyxf7tKG4Gz8Kj/o9/8A9dBXoNeffCr/AI99Q/66CvQaUtwMLxL4ag8QiITzNH5fTFReG/CdtoFxJNDO0hcY5FdDSg1NwFrkviR/yLX/AG1FdbXJfEj/AJFr/tqKcdwPMIv+Pi3/AN8V7pYf8eNv/wBcx/KvC4v+Pi3/AN8V7pYf8eFv/wBcx/KrqAWO9eV6v4D1i51e5uYNhjlcsMtXqlGKhOwHmfhvwVq2m67b3lxs8uNsnDV6VjnNLilobuAyQfu2/wB014TdD/T7r/rq38693f7jfQ14Td/8hC6/66t/OtKQEdNcfLx1BzTqQ1swPY/CepLqeg2824F1G1h6Yrarxzwn4hbQdQImJNpMcOP7teuW1xFdQLNbyCSNhkEHNc0o2YEjoroVdQykYIPeuW1DwFo93IXhTyGJycc11dLilewHAt8M7Q/dvnH/AAEU3/hWdv8A9BF/++RXoFFPmYHAj4aWw/5iD/8AfIp3/CtbX/n/AH/75Fd5RRzMDg/+FbWv/P8Ayf8AfIpP+FbW3/QQk/75rvaKOZgcEfhra/8AP+//AHyKT/hWlr/z/v8A98iu9paOZgcD/wAK0tf+f9/++RSj4aWn/P8AP/3yK72ijmYHCL8NrMf8vr/98ipF+HViOt3Ifwrt6KOZgeVeLvDdtoNtBLBKXaR9pBrmicV3/wAUeLGz/wCulefZ5ropu6AdSUmaWtACiiigAooooAKKKKBBRRRQAVOv+r/CoKnX/V/hQMpHqaKO5ooAKKKKBAKWgUUwCiiigAooooAKKKKBhRRRQAUtIKWgQUUUUxlGiiiuUQUUUUCCiiigAooooAKKKKACiiigAooooGFFFFAElt/x9RfWtrU/+PZaxIOLiP61ual/x6rQBk9qKO1LQAUlLRQAUUUUCCkoooAKKKKAFopKWgAooooGFFFFABRRRQAUUUUAFAopKAPQvhV/qLz8K9CrwzR9dv8ARA62LYEn3q0v+E61/wDvis2ncZ6P4n19PD9itzJEZAxxiuUPxOT+HT8/jXKat4j1PWIBBfMDGDmsvHpTUe4jvj8TvTTf/Hqafie3/QMH/fVcHikxT5UFzWn8UXX/AAkja3aRmFmxviByGFet6HrFtrWnR3dsw5HzL/dNeGYrX8Na/ceHrxpYx5kEg/eRdiaHELntckixoXdgqKMlicACvL/G/i/+0S2maW5FuDiWUfxGsvxD4t1HXMxAm2tP+ean71YCqAMAcUlEBAoAwOBTgKXFLirA7D4ca7Fp91Jpl24SOdtyMegPpXqSkEAg5B7jpXz2VB+o6Edq2dP8V67p0Yjju2ljXordqhxC57bS5ryNfiJrQHMCH8aU/ETWSOLdB+NTysZ60zBVJJwB3PSvKviFr0ep3qafaPvggOWYdC1ZGo+LNc1KNopboxRN95F71jqABgVUYiHGgdKKWtAEopaSgBKKWigBuSCCpwR0PpXZeFfHMtkyWWssZICcLP3WuONNIzwalq4Hv8E8VxCssEivGwyGU5qDU9Qt9LsZLu7kCRoM+5PtXkfhjxNd6BcbGYy2TdYyfu1H4m8Q3XiG7y+Y7RP9XEO9TygVNc1efXNTe8nyEziJOyiqQoxSgVaQCit/wRqH9neJotxxFMNrfU1g06JzFPFMvBjcN+VDVwPoGivOl+Je1ADYAsByc0x/iZKR+705c+7VnysZ6RWfrGrWmk2b3N3IqhRwueWP0rzW8+IOs3GRBEsHuDmucvr281KbztQuWnftntTUWA/V9Sl1fVJb6bgMf3a/3RVQUuKUCtEhAKWiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQIKKKKBhRRRQAUUUUwFqWL7tQ1NF92kBUk/1jU2nSf6xqbTAKKKKYBRRRQAUUUUAFFFFIAooopgHatrR/8Aj1asXtWzo5/0ZqTEU7j/AI+H+tR1Jcf8fMn1qOgYUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRQAGkjklt5lmt5WjlXowPSlppFDA7zw74/wAFLXWxjsJx3+tdN4i8QWum6FJeJKjmRdsYU5JJ7142QDwelKS7KqNIzIv3VJ4FZuADWd5ZHlmO6SQ5Y+tRyj9030qXFI65Uj1qrAe0+EOPDFj/ANcxW5Xl+lePxp2m29mbEN5Kbd2etXP+Fmr/AM+H61k4sZ1nizB8NXwPIMZrgPAXiltMaLTr9ybWTiNz/CfSrOqeP11HTZ7T7EF81duc9K4gIPLCnsKqMNNQPoRWDKGUggjII71x3jfxWumwtp9i4a8kGCR/yzHr9a5TTvGmpWOitp+PMkAxHMTygrnHZ5ZGlmcySucsx7mhQ11AZhiSzsWdjlmPc0uKUClrWwgFGKWigC3pGqXOjagt5aE8H94nZxXsWh6zaa1Yrc2rg5HzJ3U14jVrStTu9IvRc2UhU/xp2aolG4HrPizxBFoWmNJkNcyDESeprxyaaa5ne4uHLyyHLMf5Vb1rVrnW9RN3c/LxhY+y1SFEY2ASrOmanfaTcCbT52Q/xLnIaoMU0iqauB6j4e8dWepFbe+Atro8Afwt+NUfiB4mWOD+ybCQGWTmV1Odo9K87Kg9f0oC88kknuTmp5NQEAAGBVvSdRn0jU4r63+8nDD1FVsUYqrAez6H4k07WoA9vMqSY+aNzjB9q2x6jmvn5MxyCSN2Rx0IOK1rfxPr9soWPUnKjoCKzcGM9szRmvHV8aa+o5uc/hTh4317/nrU8jA9gzRmvH/+E417/nrSf8Jxr3/PWnyMD2DPsaRmAGScD1NeOv428QEcXG38KpXXiTXbxCk+ovsPVQKORgdv478VQ21m+mWMgkuJRiQqchR9a8yUYGM5Pr60oXkkkknqSc07bWkY2A6TwR4ij0K6mivGYWkg4xztNaGtfEO6uN0OkxeUmcCbuR9K4vaPSlxRyK4h88811MZrqZpZD1JNIDikApwq0rAOpKKKYC0UlLQAUUUUCEpaKKACiiigAoopKBhRRS0AJSikpaACkoooAWiiigAopcUlAEsP3TVaX/WGrMP3TVaU/vDQgG0UUUwCiiigAooooAKKKKACiiigAooooAtaZ/x+fhVvU/8AWrVTTTi8H0q3qg/eKaXUCnSUtFACUUtJQAUUUUAFFLRQAlLSUtACUUUhoA774Xsq29+WZR846mu882P/AJ6J/wB9V4NHPcwZFvcNFu6470pvtRH/AC/yfnWMoXYz3jzY/wDnon/fVHmx/wDPRPzrwf7fqP8Az/yfnS/b9R/6CEn50vZsD3fzY/8Anon50ebH/wA9E/76rwj7fqP/AD/yfnR9v1H/AJ/5Pzo9mxHu/mx/89E/76o82P8A56J/31XhH2/Uf+f+T86Pt+o/8/8AJ+dHsxnu/mx/89E/76o82P8A56J/31XhH2/Uf+f6T86Pt+o/8/0n50ezYHu/mx/89E/76o82P/non/fVeEfb9R/5/pPzo+36j/z/AEn50ezA9382P/non/fVHmp/z0T868I/tDUf+f8Ak/Ol/tDUv+f+Sj2YHu3mJ/fT86PMT++n514V/aWpf8/8lL/aWpf8/wDJR7Nge6eYn99fzrmfiAynwzLhlPI6GvMhqepf9BCSkkvb2aMxz3byRnqppqm0wIh91foKWkFLW4Etn/yELb/roP517pH9xf8AdFeEQyeTcxTYzsYNj1ruB8SEUAfYe3rWNSLb0A9APavJviN/yMkf0rZPxKT/AJ8f1rk/EerjXNSW7EXl7RjbShFpgdZ8KwRb35x/y0Fd9+BrwzT9V1HSw40+6MIc5YDvVv8A4SrxD/0En/KhwbYHf+LvFcvh26t4o7MTiVS2ScYrAHxMuf8AoGJ/31XJajqN9qkiSahcGZkGFJHSqoWqUO4Hcj4l3HfTV/76rO8Q+MX13TfsbWYh+YNuBzXMhadiqUEA6Hm6t/8AfFe6WH/HjB/1zH8q8KQlJEcdUORXZQ/ES5igjjGnodigZ3damcW9gPS81yt5480ezvJbWYyeZGcNhawf+FkXWf8AkHIf+BVx99cG9vpbpkCmVixX0qFTfUD1TS/GukapfJZ2zP5r8AFa6SvCtMvZNM1CO9hQM8ZyF9a6j/hY2o5/48UH40Om+gHpb/cb6GvCbs/8TC6/67N/OuqPxF1Egj7CnIx1rkJHMs0kpGDIxYj0zV04tALSGgUuK1AjYZrS0bX9S0RsWkpaHPMRPBqhikIpNXA9DsfiRZuANQtzA3fbzWxB420Gb7t2R9VxXkRUU1o1PUVm6aA9qTxNo79LxPzqQeINKP8Ay+R/nXh3koex/OjyU/2v++qXsgPcv7f0r/n9j/76o/t7Sv8An8j/AO+q8N8hP9r/AL6o8hP9r/vqj2QHuX9vaV/z+x/99Uf29pX/AD+R/wDfVeG+Qn+1/wB9UvkJ/tf99UezA9w/t7Sv+fyP/vqj+39K/wCfyP8AOvD/ACE/2v8AvqjyE/2v++qPZAe4/wDCQaV/z+R/nSf8JBpX/P5H+deIeSno3/fVHkp6N/31R7ID2/8A4SDSv+fyP86P+Eh0n/n8j/OvEPKT/a/76o8pf9r/AL6o9kB3vxF1GzvrO0W1mWRlkyQD0FcRTFQL0z+Jp4rWMbIBaKXtSVYC0UUUCCijNFABRRRQAUUUUAFTr/q/wqCph/q/woGU+5oo7migQUUUUwClpKWgAooooGFFFFABRRRQIKKKKBi0UUUxBRSUUDKVFFFcogooooEFFFFABRRRQAUUUUDCiiigAooooAKKKKAFU7XVvQ1v3Y32Of8AZBrnzyK3rNxcWAB9MUgMrtRSsCrFT2NApgFFFJQIWiikoAKKKKACiiigYtFJS0AFFFFABRRRQAUUUUAFFFFABSUtFADaWlpKACloooAKMUUUAIRSYp1FMBmKUClopAFFFFACYpMU7FFMBmKXFOooAbilFLS0AFFFFABRRRQAUUtJQAUhFLRQA3FGKdRQAlLRRQIDRRRQMQ0lOoxQISgUuKKBhS0UUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigQUUUUDCiiigAooopgFTJwlRDk1JKdkRpAVG5YmkoopgFFFFABRRRTAKKKKACiiikAUUUUAFaujNmORfQ1lVc0qTy7vaejUMCS9XbdN71BV7U0wVkH41RpIAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUCCkpaSmAYpKdSUhhikIpaKAG4oxTqKLANpRS4ooAKKWigApaKKAFpKKKAEpKdSUANxSgUuKKACjFFLQA3FLiiloAQCjFLRQAmKTFOooAbijFOxSUwG4oxTqKAG4o206jFFgGgUuKXFLQA3FLiilpgIKWkpaACijNFAC0UlFAhaKKKACikooAKWkooGFFFFAC0lFLQAlFFFABRRS0AFLSUUAFLSUUATJwlU2OWJq3IdkRqpQgCiiimAUUUUAFFFFABRRRQAUUUUAFFFFAE1k228Q+pxWlqi5RW96yEbZIr/AN05rcnUTWe4c/LmpYGXSUClpgJS0lFAxaKSigQUUUtABRRRQAU3FLRQA0ik206igBmBS4p1GKLDG4oxT6KLCGYoxTqXFFhjMUYp+KMUAMxRg0/FGKdgGYoxT8UYpWAZijFPxRigBoFKKWimAYpTSUUCENJTqTFAxuKMU7FGKLCG4oxT8UmKLDGhaUCnUUAIBS0UUAJRilooAbilxS0YoATFJinYoosA3FLilxRTAKWiigAxSEUtFADcUhFPooAZilxTqKLAMxS4p1JigBMUYpaKLAJijFLiiiwCYoxTsUmKLANxRT6MUwG4pRRRQAUtFFAhKXtRRQAUUlLQMKKKKYgooopDCpm4iP0qJRlhS3DYTHrQBWHSloopiCiiigYUtFFAgooooAKKKKACiiigAooooAWiiimAUUUUDKNFFFcogooooAKKKKBBRRRQMKKKKACiiigAooooAKKKKAFrQ0ibbI0Td+lZ1ORijq69VNAGnqMOyXzAPlbg1UrVjdLy1+owfY1mSxNDIUb8KQDaKSlpgFFJRQAUUUUAFLSUUALRSUtABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKBRRQAUlLRQAUUUUAFJS0UAFFFFABRRRQAUtJRQAtFFFABRRSUAFFFFABRRRQAUUUUAFLSUUwFooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKUDJxTAdEMnNR3LZbaKmYiOOqZJJyepoASloooAKKKKYBRRRQAUUUUgCiiigAooooAKcrFHVh2NNooA6Btt1aZH8QzWSQVJB6ip9KudjGFzw3T2qa/tyD5yD/AHhSAo0UUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooopgFFFFIAooooAKKKKACiiigApKWkpgFLSUUALRSUtABS0lFIBaKSloAKSlpKAClopKAFopKWgAooopgFFFFABRRRQAUUlFAC0lAooAWikFLTAKKKKADFFFFABRRRQAUUUUALmikooAWikooAWikooAKKKWgBKKKKACiiigAooooAKWkooAWiiigAp8Yy2aaAScCpGYRpQBFcPkhR2qGgkk5PeimAUUUUAFFFFABRRRQAUUUUAFFFFABSUtFACda2NMmEkBjbqv8AKsipbaYwThx070mBYuYzFOR2PIqKtS4jW6hDJ16issggkHgjrQMKWkpaBBSUUtAwooooEFFFJQAtJRRQMKKKKYBRRRQAUUUUAFFFFABRQKKACiiigQUtJRQAUUUtABRRSUALRSUUAFLSUUDFopKM0xBS0UUDCkpaKAEooooAKKKKAFopKKBC0UUUDCiikoAWikozQAtFFFABRRRQAUUUUAFFBopgFFFFABRRRQAUUUUCCg0UUAFFJS0AFFFFABRRSUALRSUUALRRRTAKKKcg3H2oGOjXAzUEzbn9hU0z7FwOpqrSAWiiimAUUUtABRRRQIKKSloAKKKKACiiigAooooAKWkpaYwooooAo0UUVyiCiiigAooooAKKKKACiiigApaSloASilooASiiigApRSUtAE9nctbS56ofvCteWOO7hDKc+h9KwasWl29s3qh6igB8sbRPtcY96bWqrwXkXY/zFVJ7J0yY/mX9aAKlFBBB+YEfWjrQAUUUUAAooooEFLRRQMKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAtFJS96ACkoooAKKKKACiiigAooooAKKKKYC0lLRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUlLRQAlFLRQAUUY9KcqE9aYCAEnipQBGuTSFljHNVpJDIfagAkkMje3amUUUAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigAyQQQcEdK2rG7W4i2PjeBgg96xaVWZGDIcEUmBo3VqYyXjGU7j0qrV+0v0mGyXCv+hp1xZK/zRfKfTtSAzqWnPFJGcMp+opmRTAWiiigAoopaAEooooAKKKKACiiigAopKKBBRRRQMKWkopgLRRRQAUUUUAFJS0UAFFFFABRRRQAUlLRQAUUUUAFFFFIBaKKKYBRRRQAUUUUAFFJS0AFFJRQAUUUUAFFFFABRRS0AJS0UUxhRRRQIKKKKACiiigAooooASloooAKM0UUAFFFFABRRRQAUtJRQAUUtJQAUUtJQAUtJS/SgYUoBJ4pQhPWnkrGOaAAAIpJqtI+9vbtRJIXPt6UygQtFFFMAooooAKKKKACiiigAooooAKSlooAKKKKACiiigC7p935TeVIfkPQ+lXbq1Ew3x/e/nWL1q7Z3xixHKcp2PpSaAjIKnDAgjsaK1JIorpNwIz/AHhVGW1li7bh7UAQ0lB46jFFAwoFFFABRS0lMQUUUUDCkpaKACkpcUUAFJS0UAFFFFABRRRQAUUUtAhKM0UUDCiiimAUUlLSAKKKKYBRRRQAUUUUAFFFFABRS0UAJS0lFAC0lFFABRRRQAUtFFACUUUUxC0lFLQAlFLRQMKO9FFABRSUtABRmikoAWkpaSgAooooAKKKKAFopKKBC0UlLQAUlFFABS0lLQAUlLRQAlFLjPSnCP1oGIqlvpT2ZY1zSM6oPeq7MXOTQAMxZsmkoopgFL0pKWgAooooAKKKDQAUUUUCCiiigAooooAKBRRQAtFJS0xhRRRQBRooorlEFFFFABRRRQAUUUUAFFLSUALRRRQAUnelooAKSlooASloooAKKKKAFR3jbdGxU+1aEGqEcTr+IrOpKAN0S2twOqn60hsoH6HH0rDH5VIJpV+7IwoA1/7Oj7OaT+zU/vtWX9quP+erUv2u4/56mgDT/s1P77Uf2an99qzPtdx/z1NH2u4/56mgDU/s1P77Un9nJ/fasz7Xcf8APQ0v2u4/56GgDS/s5P7zUv8AZyf3mrL+13H/AD0NL9ruP+ehoA0v7OT+81H9nJ/easz7Xcf89DS/a7j/AJ6mgDS/s5P7zUf2cn95qzftdx/z1NJ9ruP+ehoA0/7OT+81H9np/easz7Xcf89DR9quP+ehoA0/7PT+81H9np/eas37Xcf89DSfarj/AJ6GgDT/ALPT+81H2BP7xrM+1XH/AD0NH2q4/wCehosBpf2en95qPsCf3mrN+1XH/PQ0farj/noaANL7An940n2BP7xrO+1T/wDPQ0fap/8AnoaYGj9gT+81H2BP7xrO+1T/APPQ0fap/wDnoaANH7An940n2BP7xrP+0z/89DR9qn/56GgDQ+wJ/eNH2FP7xrP+1T/89DR9qn/56GgDQ+wp/eNH2FP7xrP+1T/89DR9pn/56GgDQ+wr/eNH2Ff7xrP+1T/3zR9qn/56GgC/9iX+8aX7En941n/aZv75o+0zf3zQBf8AsS/3jR9jX+8aofaZv75o+0zf3zRYC/8AY1/vGj7Gv941n/aZv75o+0zf3zRYC/8AZF/vGj7Iv941R+0Tf3zR9om/vmgC99kX+8aPsi/3jVH7RL/fNH2ib++aAL32Vf7xpPsq/wB41S+0S/3zR9ol/vmgC79lX+8aPsq+pql583980efN/fNMC79lX1NH2VfU1S+0S/3zR9ol/vmgC59mX1NH2ZfU1TE8v980faJf7xoAufZh6mj7MPU1T8+X+8aPPl/vGgC39nHqaPs49TVTz5f7xo8+X+9QBb8gepo8gepqp58v96jzpP7xoAt+QPU0eQPU1U86T+9R50n96gC35A9TSeSPU1V86T+8aTzpP7xoAt+SPU0eSPU1U82T+9S+dJ/eoAteSPU0eSPU1V86T+9SedJ/eoAteUPU0eSPU1V86T+9R50n96gC15Q9TR5Q9area/8Aeo81/WgCz5a0bFFVfMf+8aQux7miwFoui9x+FRPcdkH41B9aWmAEknLHJooooABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiikAVat76WEAN86+9VaKANqK+t5hhjj2NSG3tpem38KwsA0oZl+6xH0pWA2f7PiPRjR/Zyf32rKFxOOkrUv2u4/56mgDU/s1P77Uf2an99qzPtdx/z1NH2u4/56miwGn/AGan99qP7NT++1Zn2u4/56Gj7Xcf89TQBp/2an99qT+zU/vtWZ9ruP8AnqaPtdx/z1NMDT/s5P77Un9nJ/fas37Xcf8APQ0fa7j/AJ6GgDS/s5P7zUf2cn95qzftVx/z1NH2q4/56GiwGl/Zyf3jR/Zyf3mrN+1T/wDPU0farj/noaANL+z0/vGj+z0/vGs37VP/AM9DR9qn/wCehosBo/2en940f2en941nfap/+eho+1T/APPQ0AaP9np/fNH9np/eNZ32qf8A56Gj7Vcf89DQBo/2en940fYE/vGs77VP/wA9DR9qn/56GgDR+wJ/eNH2BP7xrO+1T/8APQ0fap/+ehoA0PsKf3jR9hT+8az/ALTP/wA9DR9pn/56GgDQ+wp/eNH2FP7xrP8AtM//AD0NH2mf/noaAL/2Ff7xo+xJ/eNUPtM3980faZ/75oAv/Yl/vGj7Ev8AeNZ/2mf++aX7TN/fNMC/9jX+8aPsa/3jVD7RN/fNH2ib++aAL/2Nf7xo+xr/AHjVD7RN/fNH2ibP3zQBe+xr/eNH2Nf7xqj9om/vmj7RL/fNAF77Iv8AeNJ9kX+8apfaJf75o+0S/wB80WAu/ZF/vGj7Iv8AeNUvtEv980nny/3zQBe+yL6mj7Kvqao/aJv75o+0S/3zQBe+yr/eNJ9lX1NU/Pl/vmjz5f75oAufZV9TR9mHqap+fL/fNHny/wB80AXPsw9TR9mHqap+fL/eNHny/wB40AXPsy+po+zj1NU/Pl/vGjz5f75oAt/Zl9TR9nHqap+dL/eNL50v940wLf2cepo+zj1NVPOk/vGk86X+8aALn2ceppPs4/vGqnnSf3jR50n940AW/s49TR9nHqaq+dJ/eNHnSf3jQBa8geppPIHqaq+dJ/eNHnSf3jQMteQPU0eQPU1V82T+8aXzZP7xoAs+SPU0eSPU1W82T+9SebJ/eNAFryR6mjyR6mq3myf3qPNk/vUCLPlD1NHlD1NVvNk/vUebJ60AWfKHqaPKFVvNk/vUnmOf4qALWxR6UFkTuKqFmPVjSUATPcZ4QVESSck5pKKYXCloooAKKKKACiikoAWikpaACkpaKACiiigBKKWigAooooAKKKKACiiigCSGeWE5jY49Kvw6mh4lG0+1ZdLSsBuA20wyNh+tNNlC3IOPpWKODxxTxNMvSVqLAa39nx/3mpP7OT++1Zn2q4/56tS/a7j/AJ6GiwGn/Zyf32o/s5P77Vmfa7j/AJ6Gj7Xcf89DRYZpf2cn99qP7OT++1Zv2q4/56Gj7Xcf89DRYDS/s5P77Uf2cn99qzftVx/z0NH2q4/56GnYDS/s9P75o/s9P75rN+1XH/PQ0farj/noaLAaX9np/fNJ9gT+8azvtVx/z0NH2qf/AJ6GiwGj/Z6f3jR/Z6f3jWd9qn/56Gj7Vcf89DSsI0fsCf3mo+wJ/eNZ32qf/noaPtM//PQ07DNH7An940fYV/vGs77TP/z0NH2mf/noaLCND7Cv940fYE/vGs/7TP8A89DR9pn/AOehoA0PsK/3jR9hT+8az/tM/wDz0NH2mf8A56GgZofYU/vGj7En941n/aZ/+eho+0z/APPQ0WAv/Yk/vGj7En941Q+0Tf3zR9om/vmgC/8AYl/vGk+xp/eNUPtE/wDfNL9om/vmgC99jT+8aPsa/wB41R+0Tf3zR9om/vmmBe+xr/eNH2NP7xqj9om/vmj7RN/fNAF77Iv940n2Rf7xql9ol/vmj7RL/fNIC79kX+8aPsi/3jVL7RL/AHzR583980AXfsi/3jR9lX+8apefL/fNHny/3zTsBd+yr/eNJ9lX1NU/Pl/vmjz5f75osBc+yr/eNH2ZfU1T8+X++aTz5f7xoEXfsy+po+zL6mqXny/3jS+fL/eNAFv7OvqaPsy+pqn58v8AeNHny/3zQMufZ19TR9nX1NU/Ol/vGjzpP7xpgXPs6+ppPs6+pqp50n940edJ/eNIC39nX1NHkL6mqnnSf3qPOk/vUAW/IX1NHkD1NVPOk/vGjzpP7xpgWvIHqaPIHqaq+bJ/eo82T+9QBa8gepo8kepqt5sn96jzZP71AFnyR6mjyR6mq3myf3qTzZP71AFryR6mjyR6mqvmyf3qPNk9aALPlD1NHkj1NVvNk/vUvmv60AWPKHqaPKHrVbzX/vUeY/8AeoAseUPWjyx6mq/mP60eY/rQBY8setG1R6VX3v8A3jTSSepoAsmRFqJ5mPA4FR0UwDvzRRRQAUUUtACUUtFABRRRQAUUUUAFFFFMQUUUUgCiiigAoopaACiiigAopKKYylRRRXKIKKKKACiiigAooooAKKKKAFopKWgA70UlLQAUUGigAooooAKKKKACkpaKADFJS0UwEoFLQKACiiigQUUUUDCiiigAooooAKKKKACiiloAKKSigAopaKAEooooAKKKWgBKKKKYBRRRSAKKKKACilooASlpKWmAlLRRQAUUUlAC0lFLQAUUtFABSUtFABRRRQAUUUUAFJS0UwEpaKKAAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABS0UUAFFFFABRRRQAUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigAooooAKWkopAFFFLQAUUCimAUUUUgCiiigAooopgFFFFIAooopgFFFLQAlLRRQAUUlLQAUUUUAJRRS0AFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKWkoAWikpaACikooAWikooAWiiigAooooAKKKKBhRRRmmAUUUtAgooopDCkpaKAEpaKSmAtFJSigQUUUUAFFFFABRRRQAUtJRQAtFFFABRRSUALRRRQAlFLRQAUUUUAFFFFABRRRQAUUUUAFFFFABSUtFACUtFFABRRRTAKKKKACiiigYUUtJQAtJS0lAC0lFFAC0UlFAC0UUUAFFFFABRRRQIKKKKBhRRRQAUUUUAFFFFMAooozQAUUUUALRRRQAUUUlAC0UUlAhaKKKACiiigAooooAKKKKACkoooGLRRRQAUUUUAFFLRQAlFLRQAUUUUAFFFFAgooooAKKKKBhRRRQAUUUUAFFFFMAooooAKKKKACiiigApaSloAKKKKACiiigAooooEFFFFABRRRQAUUUUDClpKWgAooopgFFFFAFGiiiuUQUUUUAFFFFABRRRQAUUUUALSUUUAFLSUUALRRRQAUUUUAFFFFABRRRQAUUUlMBaKKSgBaO9FFABS0lLSASilooASiiimAUUUUAFFFFABRRRQAUUUUAFLSUUAFFFFAC0lLSUAFFLSUAFLRRQAlLSUtACUtJS0wEopaSgAoopaACiiigBaKKKACiiigAooooAKKKKACkpaKADFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAoxRRQAUUUtABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKACiiigApaKKQBRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAopaKACiiigAooooAKKKKACkpaKQCUUtFMAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUtFABRRRQAUUUUAJRS0UAJRS0UAFFFFABRRRQAUUtFACGiiigAooooGFAopaBBRSUtAwooooEFFFFMYlLRRQIKKKKACiiigYUUUUCCloooAKKKKAEpaKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFAwooooAKKKKACilooASilNFAAKKKKBBRRRQAUUUUAFFFFAwooooAKKKKYBRRRQIKKKKBhS0lLQAUUUUAFFFFABRRRQIKKKKACiiigAooooAKKKSgYUUtFABRRRQAUUUtABRRRQIKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFMAooooAKKKKACiiigQUUUUDCloooAKKKKACiiigAooooEFFFFABRRRQAUUUUDClpKWmAUUUUAFFJRQBSooorlEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALRRRQAUUUUAFFFFABRRRQAUUGimAUUUUAFLSUtIBKKKWgAopKWgApKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFLRQAUUUUAJS0UUAFFFFABRRRQAUlLSUwFpKWkoAKWiigAooooAWkopRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKWgUUAFFFFABRRRQAUUUUAFFFFABRRRQIKKKKYwooooAKKKKACiiigAooooAKKKKAClpKKQC0UUUAFFFFABRRRQAUUUUAFFFFMAooopAFFFFMBaKSloAKKKKACiiigAooooAKSlooGFFFFAgooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRS0lAC0UUUAFFFFABRRxRQAUlLRigQlLRRQMKKKKAFooozQAUUUUAFJS0UAJRRRTAKWkpaQBRRRQAUUcUUAFFFFMAooooAKKKKACiiigYUUUtAhKWiigAooooASloooAKKKKACiikoAWiikoAWiiigAooooAKKKKACiikoAKWiigAooopgFFFFABRRRQMKKKWkISloopgJS0lLQMKKSloASloooEFFFFABRRRQAUUUUAFFFFAwooooAKKKKYBRRRQIKKKWgYUUUUAFFFFABRRRQIKKKKACiiigAooooAKSlooAKKKKACiiigYUUUtACUtAooEFFFFABRRRQMKKKKBBRRRQAUUUUAFFFFAwooooAKKKKYBRRRQIKKKKACiigUALSUtFAwooooAKKKKACiiigAooooEFFFFABRRRQAUUUUwClpKKAFooooAOaKKKBlGiilrlEJRRRQAUUUUAFFFFAC0lFFABRRRQAUUUUAFLSUtABRRRQAUUUUAFFFFABRRRTAKKKKBBS0UUgEpaKKBhSUtJQAUUUUwCiiigAooooAKKKKACiiigAooooAKWiigAooooAKKKKBBRRRQMSilpKAClpKUUwEpaKKACiiikAUtJRTAWiiigAooooAKKKKACiiigAooooAKKKKAFpKWkpgFFFFABRRRQAUUUUAFFFFABRRRQAUUtJQAtFFFABRRRQIKKKKACiiigAooooGFFFFABRRRTAKKKKACiiigAooooAKKKKQBRRRQAUtFFAgooooAKKKKACiiigYUUUUAFFFFABRRRQIKKKBTGFLRRSAKKKKACiiimAUUUUgCiiimAUUUUAFFFFABRRRQAUUUUAFJS0lAC0UUUwCgUUUAFLSUtABRSUtABSUtFABRRRQIKKKKBhRRRQAUUUUAFFFFABRRR2oAWikpaACiiigAooooAKKKKAAdaKBRQAUUUUAFFFFABRRSUwFooooAKKKKACiiigYUUCigQtFJS0AFFFFABRRRQAUlLSUAFLSUUCCiiloGFFFFABRRRQAUUUUAFFFFABRRRQMKKKKYgo7UUtACUUtFIYUUUUCCg0UUxhSUUtIBKWgUUwCiigUCCiiigAooooAKKKKACiiigYUUUUCCiiimAUUUUgCiiimMKWiikAUUUUwCiiigQUUUUAFFFFABRRSUALRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQMKKKKYhaSiigYUUUUCCiiigBaKKKACiiigYUUUUAFFFFAgooopgFFFFIAooopgFFFFABRRS0AFFFFABRQRRQBRpaSiuUAooooAKKKKACiiigAooooAKKKKACiiigApaSloAKKKKACiikoAWiiigApKKKYC0CiikAUUtFABRRRQAlFLSUAFFFFMAooooAKKKWgBKKWkoAKKWigBKKKWgAooooASloooEFFFFACUtFJQMWkoooAWikooAWiiimAUUUUgCjtRRQAtFFFMAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUtACUUUZoAKKKWgAooooAKKKKBBRRRQMKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRSAKKKO1AC0UUUAFFFFABRRRTASloopAFFFFABRRRQAUUUUAFLSUUwFooooAKKKKQBRRRQAUUUlMBaKKKACiiigAooooAKKSloAKKSimAtFFFIAooopgFFFLQAUUUUAFFJS0AFJS0UAFFHFJQIKKKWgYlLRRQAUUUUAFFFFABS0lFABS0lLQAUlFLQAUUUUAFJRRTAWiiikAUUUUAFFFFMBKWiigApKWigAooooAKKKKAClpKKYBS0lLSAO1FFFABRRSUAFFLRQAlFFFMAooopAFFFFAC0UlFAC0lLSUALRRRQAUUUUwCiiigAoopaBhSUtFIAooopiEopaKAAUUUUDCiiigAooooAKKKKBBRRRQAUUUUAFFFFABRRRQAUUUUxhRRRSEFFFLQAUUUUxhRRRQAUUUUCCiikoAWikooAWiiigYUUUUCCiiloASiiloAKKSloAKKKKACiiigAooooAKKKKACiiigAooooGFFFFAgooooAKKKKYBRRRQAUUUUAFLRRQMKKKKACiiigAoooNABRRRQIKO9FFAwooooEFFFFMAooooGFLSUtAgooooAKKKKBlKkoorlEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUtABRRRQAUlFFMAzS0lFIAoopCwHemAtLTdy+v6Uu4e/5UgFopNw9/wAqNy+p/KgBaWm7h6/pRuH+RQA6kpNw/wAijcPX9KAFopNw9f0o3D/IpgLRSbh/kUu4f5FAhaSk3D/Io3D/ACKAHUUm4ev6Um4ev6UALRSbh6/pRuHr+lAxaKTcP8il3D1/SgAoo3D/ACKNw/yKAFpKTcP8ijcP8igBaWm5Hv8AlRuH+RQA6kpMj/IoyP8AIoAWikyP8ilyP8igBaSjI9/ypMj/ACKAFpaTI/yKMj/IpgLmikyP8ijI/wAikAtFJkf5FG4f5FADqKTcP8ijI/yKYC0UmR/kUZFAC0UmRRke9AC0UmRRkUALRSZFGRQAtFJkUuRQAUUZFJkUALRSZFGaAFooyKM0AFFGRRkUAFFGaM0wClpMijNABS0maM0ALmikzRmgBaKTIozQAtFJkUZHrQAtFFFABRRRTAKKKKQBRRRTAKKKKACiiigAooopAFFFFABRRRQAtJRRTAKWkooAKWiikAUUooOB1oASgUm5fWl3D1/SgBaSjI9f0oyPX9KACik3D/Io3D/IoBC0Um4f5FG4f5FADqKbuHr+lLuHr+lAC0Um5fX9KNw9f0oAWjNJuHr+lG5f8igBe9J3o3L6/pRuHr+lMBaKTI9f0o3D/IoAWim7h6/pS7h6/pQAtJSZHr+lLuH+RQAtFJuHr+lG4ev6UALRSZH+RRuH+RRcBaKTcP8AIo3D/IoAWik3D/IoyP8AIoAXtRRuHr+lGR/kUAFLSbh/kUmR/kUwHUlG4f5FJuH+RQAtLSbh/kUZFABRSZHr+lGR/kUALRSZH+RS5FAC0UmRRkf5FAC0Um4f5FG4f5FAC0UmRRkUALRSZFLkUAFLSZFGaACijNGRQAUUZFGaACgUZoyKAFopMijNABRRn60ZFMAoozRQMKKM0UCClpKXNABRRmigAoooyKACijI9aKACiiigAoooFAC0UUUAJRS0lABRRRTAKKKKAFopKWkAlFLRQAUUUUAFFFFABRRSUALRSUtMApaSjcvrSGFLTdy+v6Ubh/kUCHUUm4f5FJuX1/SncB1FJuH+RRuHr+lAwoo3D1/Sk3D1/SgBaWm7h/kUu4f5FAC0Um4ev6Ubh6/pQAtFJuH+RRuHr+lAC0Um4ev6Ubh6/pRcBaKTcP8AIo3D1/SgQtFJuH+RRuH+RQAtFJuH+RS7h6/pQMKKNw/yKNw9f0oEFFGR6/pRuX1/SgBaKTcP8ijI/wAigYtFJuH+RRkf5FAC0Um4f5FG4f5FAC0UmRRuH+RQIKWk3CjcKLgFLSZHrRkUwFopMj1oyPWkAtFJkUZH+RTAWikyKXIoAKKMijIoAKKMijIoAWikyKMigBaKTIoyKAFopMijIoGLRSZFGRQAtFJmjNAC0UmaM0ALRSZozQAtFJkUuaYgopM0uRSAKKM0ZFAwooyKMj1oELSUtJTGLRSUtABRRRQAUUUUAFFFFABRRSUALRRRTEFFFFABRRRQAUtJS0AFFFFAAaKKKBlGiiiuUQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUtABRRRQAlFFSQQS3Moit0Luew7UARkgcmr1lpN7ekGOMoh/jPSuh0vw7DaL596RJIOfYVPea1Bb/u7ZdxH93pSuBUtfCcKgG6lLn/ZNX00rSLcchcj+9WHcaleXBJaTaPReKrFnb70jH6mgDqfL0df+eH5UY0j/AKYflXK4FGBSA6rbpH/TD8qNuj/9MPyrlcUYFAHVY0f/AKYflSbdI/6Y/lXLYowKYHU7dI/6YflRt0j/AKY/lXLYowKQHU40j/ph+VGNI/6Y/lXLYFJimB1WNI/6Y/lSY0j/AKY/lXLYoxQB1ONI/wCmP5UY0j/pj+VctijFAHU7dI/6Y/lRt0j/AKY/lXLYowKAOo26R/0w/KjGk/8ATH8q5fFGBQB1GNJ/6Y/lSY0n/pj+VcxgUmKAOpxpP/TH8qQjSf8Apj+VcvgUYFAHUY0n/pj+VJjSf+mP5VzGBRgUAdPjSf8Apj+VGNK/6Y/lXMYFJgUAdPjSf+mP5UY0kf8APH8q5jFGKAOnxpP/AEx/KkxpX/TH8q5nAowKAOmxpX/TH8qT/iV/9MfyrmsCjAoA6X/iV/8ATH8qT/iV/wDTH8q5vAowKAOk/wCJX/0x/Kk/4ln/AEx/KubxS0wOj/4ln/TH8qMaZ/0x/KucwKMUgOixpn/TL8qP+Jb/ANMvyrnMUtMDof8AiW/9Mvyo/wCJb/0y/KudxRigDof+Jb/0y/Kj/iW/9MvyrnsCjAoA6DOm/wDTL8qP+Jd/0y/KufwKMCgDf/4l3/TL8qTOnf8ATL8qwcUYFAG8Tp//AEy/Kkzp/wD0y/KsHFFAG9mw/wCmX5UmdP8A+mX5Vg4FGKAN4mw/6ZflSZsP+mX5Vh4FJTA3c2H/AEy/KjNh/wBMvyrCxRigDdzYf9MvypM2P/TL8qw8UYoA3M2P/TP8qT/Qf+mf5ViUUAbebH/pn+VJmx/6Z/lWJRigDbzY/wDTP8qAbH/pn+VYnFFAG1my/wCmf5UZsv8Apn+VYtJQBt7rL/pn+VJusv8Apn+VY2BRgUAbH+hH/nn+VIbazl6BfwrIxSgkfdYigC/Lo8TAmJip96z57GeDkruX1FTxXU8XRs/Wr0OoRyfLKMH36UXYGDxRW3d6fFON8OFf9DWNJG8UhSQEGquA2iiigAooopgFFFFABRRRQAUUUUAFFFLSAKSlpKYBRRRSAKKSgmgBafDHJPJshjZ29BV/SdFm1AiST93B6n+L6V1CJY6TCAoVcev3jSuBh2Xhm4mw1xII1/u9614vDumwAGXcT6k1TuddmkJW3Xavq3Ws+S4mkOXlf86V2B0YtNHjGP3f4ineXo4/54/lXLZz1JNFIDqduj/9Mfyo26R/0x/KuWxSYoA6jbpH/TH8qNukf9Mfyrl8UYoA6jbpH/TH8qTbpP8A0x/KuXxRimB1G3SP+mP5UbdJ/wCmP5Vy2KXFIDqNuk+kP5UY0j/pj+VcvijApgdRt0j/AKY/lSbdJ/6Y/lXL0YoA6jbpP/TH8qNuk/8ATH8q5fAoxQB1G3Sf+mP5Um3Sf+mP5VzGBS4oA6bbpP8A0x/Kk26T/wBMfyrmcCjAoA6bbpX/AEx/KjbpX/TH8q5nAoxQB0u3Sv8Apj+VG3Sv+mP5VzWKMUWGdLjSv+mP5UbdK/6Y/lXNYFGBTsI6TGlf9Mfyoxpf/TH8q5vAowKLAdJjS/8Apj+VGNL/AOmP5VzdJiiwHSY0v/pj+VGNL/6Y/lXOUYFAHRY0z/pj+VGNM/6Y/lXOYoxQB0eNM/6Y/lSY0z/pl+Vc7QAM0AdFt0z/AKZflRjTP+mX5Vz2KTFAHRY03/pl+VJjTf8Apl+Vc9ijigZ0ONN/6ZflSY03/pl+Vc/ikxQB0ONN/wCmX5Un/Et/6ZflXP4oxTA6D/iXf9MvypCNO/6ZflXP0tIRv/8AEu/6ZflSf8S7/pl+VYGKXFMZvf8AEv8A+mX5Uf8AEv8A+mX5Vg0UCN7/AIl//TL8qQ/YP+mX5VhUUAbv+gf9MvypP9A/6ZflWHSUAbn+gf8ATL8qP9A/6ZflWHRQBt/6D/0z/Kj/AEH/AKZ/lWJRigZtZsf+mf5Uf6D/ANM/yrExRgUAbX+hf9M/yo/0L/pn+VYuKMU7AbX+hf8ATP8AKkzZf9M/yrGxRgUCNnNl/wBM/wAqP9C/6Z/lWPxSUAbP+hf9M/yozZf9M/yrGoxRYDZzZf8ATP8AKkzZ/wDTP8qx6KANjNn/ANM/ypCbI/8APP8AKsiigDVa2s5OgX8Kry6TEeYmKn3qjkjoxFSx3U8XRsj3oAhuLKeA5K7l9RVbrW3DqCSfLKMH36U26sI5xviwr+3Q07jMeilkR4nKSAgikpiDvS0lLQAUlLRTGJRRRQIKWkpaBhRRRQAUUUlAC0UUlIQtFJRQAtFFWLGynv5vLgXju3YUMCt3AHJPQVp2Wh3t1hmHkoe7d66Cx0m002PzJNrP3dulRXWthMpbLn3PSp5hjbfwzaRjNwzOfUHira6fpMPBCf8AAqwpr25mOXkI9lqElj1dj9TS1A6by9JH/PH8qNmk/wDTH8q5j86MUAdPt0n/AKY/lSbNJ/6Y/lXM4oxQB02zSf8Apj+VBXSfSH8q5nFGBQB023Sf+mP5UbdJ/wCmP5VzOBRiiwHTbdJ/6Y/lRt0n/pj+VczijFFgOm26T/0x/Kk26V/0x/KuawKMCmB0u3Sv+mP5Um3Sv+mP5VzeBRgUAdJt0r/pj+VG3Sv+mP5VzWBRgUgOl26V/wBMfyo26V/0x/KuaxRimB0u3Sv+mP5Um3S/+mP5VzfFFFgOk26X/wBMfyo26X/0x/KubxRgUWA6Xbpf/TH8qTbpf/TH8q5vFGKLAdHt0v8A6Y/lRt0v/pj+Vc3ijFFgOkxpn/TL8qTGmf8ATL8q5zFGKLAdGRpnpD+VGNM/6Y/lXOYoosB0WNM/6ZflRjTP+mX5VzuKMUWA6HGm/wDTL8qMab/0y/KuexRinYDoD/Zv/TL8qTGm/wDTL8q5/AoxRYDoP+Jd6RflRjTvSL8q5/FFFgN/Gnf9MvyoI0//AKZflWBRRYDe/wCJf/0y/Kj/AIl/pF+VYNFFgN7/AIl//TL8qT/iX/8ATL8qwqMUWA3f9A/6ZflR/wAS/wD6ZflWFikosBu/6B6RflR/oHpF+VYVFFgNz/QP+mX5Un+gekX5ViYoosBtn7D/ANM/ypP9B/6Z/lWLijFOwG1/oP8A0z/Kj/Qf+mf5ViUYosBtZsv+mf5UE2X/AEz/ACrGo4osBs5sv+mf5UmbL/pn+VY1FFgNnNl/0z/KkzZ/9M/yrHoosBsZs/8Apn+VGbP/AKZ/lWPijAosBr7rP0j/ACozZ/8ATP8AKsjFFFgNfNn/ANM/yozZ/wDTP8qyKTFFgNfNn/0z/KlzZ/8ATP8AKsfFFOwGvizP/PP8qRrS0k6AfhWTSgsOjMKQFybSUPMLbT71QntZoPvqSP7wqzHdzR98j3q7FexTDbIMH3p3YGGMUVq3emq+ZLfhvT1rLZSrFWBDDsaaYCUUUUwCiiigQUUUlAxaSlpKACloooAKKKKYgooooAWiiigAooooAKKKKYFGiiiuQAooooAKKKKACiiigAooooAKKKKACiiigAooooAKWkoz6ck9qAJbW3lu7hYIAS7H8q7ixsrXRbPcxG/Hzuev4VW8P6cmnWXnz4EzjLE9hWbqd+17OQpIiU8D1pAO1HU5bxiqEpEOgHU1n4A6UtJQAUUUUAFFFFABRS0lAC0lFFABRRRQAUUUUAFFFFACUUGigAzRRRQAUUUUAFFFJTAKKMUUAFFFFABRRRQAUUUUAGaKKKAEopaSgAooooAKKKKACiiigAooooAKSiigAooooAKKKKACkpaSgAooooAKKSlpgJRS0UAJRRRQAUc0UUAFFFFABSUUUwCiiigAooopAFFFFMApaSigAoNGaKAJ7a7eBgCSydxV+4givoMjGexrIqxaXBt5BnlD1FAFGSNopDG4ww/Wm1uajarcwebH99RkH1FYf1600AUUUUwCiiikAUUUUwCiiigAopaKACiiigBKKWigBpNbOg6Ob1xcXIIgU8D+9VTSdPbUbwRjiNeXNdVqN3Hp9qsMAAbGFHpUsBuo6jHZIIYAN44AHRa5+WR5nLysWY0jEsxZjlj1NJUgFFFFMBaKSloAKKKKACiiigAooooAKSlpKACig0UAFFFFABRSUtABRSUtMApKKWgBKWiigAoopKACiiigAooooAKKKKACiiigApKWkoGKKKKKBBRRSUxi0lFFABRSUUALmkoooEFFFFABRRRQMKKKKACikooELRSUUAFFFFMAooooAKSlxRQACkpaSgAoooNABRRRTAKKKKBhRRSUCFopKM0hhSUtJTATFT2108BwfmQ1DRQBqTwRXsORjP8AC1YskbRSGNxhhV21nMEmDyh6irl/bLcweYn31GQR3o2AxKWjHY9RRTEFFFFMYUUUUAFFFFIAooooAKSlpKAFpKKWgQlFFOijeaZIYxlnOBQBY06xl1C5EUfCj7zegrrx9l0mzCoAAPTqxqO2gh0nT8cAgZY9yfSsK5uXupjI54/hHoKh6jJby9lu3Jc4Tsoqt+lGaBQAlFLiigBKWiigAooooAKKKSmAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAlAoopgLRSUUAHNFFFACUtJRQAtFJRQAtFFJQAUUUUAFFFFABRRSUAFFFFMQUUUUAFFFFAwooooAKKKSgBaSiimAUUUUAFFFFABRRRQAUUUUAFFJRQAUUUUAFFFFABmiiigAooooAKKKKAEpaKSgAooooAt2t20RCSHKHv6VYvLRLqPzI8CQDII71mVbsbkxsI3PyHp7UAZpBVirDDDqKStXVbUMvnxjkfeA71k00wFopKWmAUUUUAFFFFABRRRQAUUUUwCiigUCFooooAKKKKACiiimMo0UUVyCCiiigAooooAKKKKACiiigAooooAKKKKACgUGigArV8O2P2zUQ7jMcPJ96yjwM12fhu2FrpAlcYZ8k/SkwG6/ebEFtGcE/e+lYQ4qW6lM91JITnnAqKgAooooAWkoooAKWkzRQAUZoooAKKKKACiiigAoopKACiiigAooooAKDRSUALSUdKKYBRRRQAUUUUgCiiigA6UUUUwCiijNABRR1pKACiiigAooooAKKKKACiikoAWkoooAKKKKACgUlLQAUUlFABmgmiimAlFFFABRRRQAUUUUAFFFFABRRSUAFFFFMAooooAKKDRSAKKKKYBmiiigAopKKACloooAKXqKSjNAGnps+VMLHkdPpWfqdv5FyWH3X5ohkMc6uPXmtDU4vOs946jmkBh0UgORmlqgCiiimAUUUUAFFFFABS0lLSAKKKKYBQeeB1PAoq9o9t9r1SKMj5RyaTA6nR7VNN0ne4w7Dc5rBupmubh5m7ngelbfiCfZAkCnG+sA+1QAyilNJTAKKKKACiiigAoopaACikooAKKKKACiiigAooooAKKKSgBaTNFLQAlLSUtMBKWiigAopKKACiiigAooooAKKKM0AFFJRQMWkpaKAEpaKM8UCCkoopjCikpaACikooAKKKKBBRRRQAUUUUAFFFJQAUUUUAFFFFAwooopgFFFFAgooooAKKKSgBaSiigAooopgFFFJQAtJRRSAWkoopjEpaSloAKSiloEFFFFAwrR02bIMLHkdKzqfC5imVx2NDELqdv5NxuUfI/SqdbmoxiayLDkqMisMdKEAUUcUVQwopKWgAoopKQC0UUUCCikpaACiiigBK6HwvYg77yQc/dXNc+FLsqDqxwK7dVWx0hQOCE/WpkwMrWbozT+Sp+ROv1rOpSdxLHqxzRSGFKKSimIWiiikMKKKSmAtJRRQAUtJS0AFJRRQAUUUUAFFFFABRRRQIKKKKACkpaSmMKKKKACiiigAoopKAFpKKKACiiigAzRRRQAUUUlAC0UlFMAooooAKKKKBBRRRQAUUUUDCkpaSgAooopgFFFFIAooopgHFFFFABRmkzS0AJRRS0AJRRRQAUUUUAFFFJQAtFJRTAWiiigBKWikoAWkoooAKKKKACg9KKSgDWsZRPAUbkjg1kXcP2e4ZP4f4as2Mnl3IB6NU+sxZjWUdV4pdQMmiigVQgooooAKKKKACiiimMKKKKACiiloEFFFFIAooopgFFFFAFGloorlAKSlpKACiiigBe1JRRQAUUUUAFFFFABS0lLQAUlLRQABd7qn94iu8vMW2i7V4/djFcRZjdfwD1au117jTEHsBSYHMDp9eaKOwpaAEooooAKKKKACiiigAooooAKKKKACikooAKKKKACiiigAoopKYC0lHWigAooopAHaiijNABRmjrQKYBRRQTQIKKSigYUUUUAFFFFABRRRQAUUlLQAGkoooAKKSigBaKSigBaSlpKAFpKKSgBaKSimAUtJR3oAKKKKADvRRRQAUUUUAFJRS0AJS0UUAFJRRQAUUUUwCiiigAooooAKKKKACiijNABRSUUAFFFFAAelbMJ82wIP93FY1a+mHNoR70gMEjazL6HFJT5xi5lH+1TKoAooooAKKKKYBRRRQAUUUUgClopKYC10Xg+INPNMR9w4Fc7XW+EFxZTv6mpYFbWpfM1JlzwnSs81Y1A51Cc+9V6kBKKKKYCUtFFABRSUtACUUUUAFFFFABRRRQAUUUUAFFFFMApKWigBKKWikAlLSUUwCiiigAooooAKKKKACikooAKKKWgYlFFLQIKKM0lMYUUGigApKWkoAKKKKBBRSUtABRRRQAUUUUAFFFJQAUUUZoAKKKKBhRRRTAKKKSgBaKSigQtFJRQAUUUUAFFFFMAoopM0gCiiiiwwooopgFFFFAgoopM0DFpKKKAClpKKBC0HpSUtAGvanzbIA9xWCw2uw9Ca29MObbH1rGuRi5ce9CGMopKWqEFFFFIYUUUUxBRRRSAKKSloAKKKKALelR+bqkC44DZrpvEEmy2WMd2rA8PDOrr7CtnxIf3yL+NS9wMfpSUp60lAwooooAM0UUUAFFFFABRRRQAUZoooAKKKKBBRRRQMKKKKBBRRSUwCiiigAooooGFFFFACUtJRQAUUUUALSUUUAFFGaKADikoopgFFFFAgooooGFFFFABRRRQAUUUlABS0UlMAooopAFFFFABRRRTAKKSigAoopaAEpaKSgApaSigAooooAKKKSgQUUUUxhRRRQAtJRRQAUUUUAFFGaTNAC0lFFABS0lFACqcOp9DWvdr51gfcZrHNbUXzWHP9w0mBzo6UtJ3I96WqEFFFFABRRRQMKKKKYBRRRQIKWiikAUUUUAFFFFABRRRTGUqKKK5RBRRRQAlFLSUAFFFFABRRRQAUUUUALRRRQAlLRSUAT2HOo2/+9XaeIf8AkGx1xdj/AMhG3/3q7PxD/wAg6OkwOZ7CijsKKACiiigAoopKAFopKKAFopKKACiiigAooooAKKKSmAtJRRQAUUUUgCiiigAooooAKKKM0wCijNJQIKKKKBhRRRQAUUUUAFFFFABRSZooAM0UUlAC0UUUAJRRS0AJRQaKACiiigBKKWkpgFFFFABRRRQAUUUlAC0UlFABS0UUAFFJS0AFFFJQAUUUUAFFFFMAoopKAClpBS5oAKKQ0UAFFFFABRRRQAUUUUAFa2lf8ep+tZNa2lf8ep+tIDFuf+PqX/eqOpLn/j6l/wB6o8VQBRRRQAUUUUwCilopAJS0UUAFFFFAB2Ndf4R/5Bstch2Ndd4R/wCQdNSYGVef8f031qGpbz/j+m+tRVIBRRRTAKKSigAooooAKKKKACiiigAozRSUAGaUUlLTAKKKKACiiigBKKKKACiiigAooooAKKSigBaSiigAooooAWiikoAWkoopgFFFJ3oAKKKKBhRRRQIKMUUUAJS4oooAKKKSgBaKSigBaSiigAooooAKKKKYBQaSigAooooAKKKKBhmkpaKYgoopKQC0lFFABRRRTGFFFFABRRSUALRmkooAO9FFFABRRRQAUUUUAFFFFAGppZ/cGsi7/wCPuStfSv8AUGsi7/4+3pLcRHRRRVAFFFFABRRRTAMUUUUgCiiigAooooA0/Df/ACFx/u1reJP+PpPpWV4c/wCQuP8AdrV8Sf8AH2n0qXuBk0maO9FAwooooEFFFFAwooopCCiiimAdqKKKACiiimMKKKKQgopKKYBRRRQAUUUUAFFFBoGJRRRQAUUUUAFFFBoAKKSimAUUUUCCiiigYUYopaACjFFJQAUUUUAFJS0UAFJRmimAUUUUAFFFFABRRRQAlLSUtACUtJRQAUUUUAFFFFABRSUUCCiiimAUUUUAFGaKSgYvWiiigQUUUUDCiikoAKKKKACiiigAooooEB6VtW//AB4f8ANYp6VtW/8Ax4f8ANJjOeP3j9aKD1P1oqkIKKKKACiiigAooopjCiiloEFFFFIAooooAKKKKACiiimMpUUUVyiCiiigBKWg0lAC0lFFABRRRQAUtFJQAtFFFAgpKWigZNY/8hG3/wB6uz8Q/wDINjrjLH/kI2/+9XZ+Iv8AkGx0mBzPYUUDoKKACiiigAopOlFABRRRQAUUUUAFFFJQAtJRRTAKKKKQC0lFFABRRRQAUUUUwCikJooAM0UUUAFFFFABRRRQAUUlFAC0lGaSgBc0UlGaAFooo/EfnQAlFBKjvTS6/wC1+VADqKb5i+j/APfNHmL/ALX5UAOzRTQynv8AnS59x+dAC0UmaWgApKKKACiiimAUUUlAC0lFFABRRRQAUtJS0AFFFJQAUUUUAFFFFABRRRQAUUUlMBaKSigApaSigQUUZooGFFFFABRRRQAUUmaKAFrW0r/j1P1rJrW0n/j2P1oYGLc/8fUv+9UdSXP/AB9S/wC9UdMAooopgFFFLSAKKKKYBRRRSASlpKKAF7Guu8I/8g6WuR7Guv8ACP8AyDZqTAyb3/j+m+tQVPe/8f031qCpAWikopgFFFFABRRRQAUUlLQAUUUlAC0lFFABRRRQAUUUUwCiiigAooooAKKKKACkoooGFFFFAgoopaACkoopgFFFFABSUtGKBiUlLmkzQIWkzSUZA7j86AFpabvQd/yo8xP9r8qAHUU3zE9G/wC+aPMT/a/KkA40lN3A9xRn3H50wHUUmaM0ALSUUtABRRiigApKWkpgFFFFABRRRQAUUUUAFJRS0AFJRmigYUUUlAhaKSimMWiikNAC0maKKAA0UUUCCiiigYUUUUAFFFFAgoopKAFooooGaml/6g1kXf8Ax9yVraX/AKg1k3X/AB9yUluIj7UUUlUAtFFFABRRRTAKKKKQBRRRQAUUUUAafhz/AJC4+la3iT/j7T6Vk+HP+QuPpWt4k/4+0+lS9wMiig0UDCiiigQUUUUDCiiigQUUUUAFFFHFABRSZooAKKKKYBRRRSAKWkooAKKKKYwopKKAFpKKWgBKKKKACkoooAKKKXimAlFFFAgooooGLSUGkoAWiiigAopKKACiiigAooopgFFFJQAtJRRQAUtJS0AFFJRQAuKSlpKACikzRQIKKKKYBRRRQMKKKKBBSUUUDCiiloAKKSigBaSiigAooooAKKO9FABRRRQIWkozSUDFNbVv/wAeH/ADWIelbVv/AMeP/ATSYHPnqfrRQfvH60VSEFFFFABRRRTGFFFFAC0lFLSEFFFFABRRRQAUUUUwCiiigClRRRXKAUUUUAFJS0UAHakpaKACkpaKAEopaKACiiigQUnelooGTWP/ACEbf/ersvEP/INjrjbH/kI2/wDvV2fiH/kGx0mBzPYUUg6CigBaSig0AFFFFABRRRQAUfSiigBKWkopgFFFFIAooooAM0UUUAFFJRTAKKKKACiiikAUUUUwCiikoAWkoooAKKKX3oASkqa3t57uYQ2kLSyHoAOPzrrNK+H13cASanN5C94hzmk2kBxZYZwAWP8AsjNadhoGsajg21k2w/xNxXqmmeGdJ0wA21qu8fxNzmtZUCjCgKPYYqXIDzaz+Hd7KA13fCL1UDNbVr8PNLTH2l2mP1xXY0cVN2MwIvBmgxY22Wfq1Wk8OaRGPls0rWzRRdgZ39h6Z/z5x/lTH8PaS4+azT8q1KKVwOfm8HaFN9+y/I4rMuvh3pMnNszQn65rs6Kd2B5nefDm8iBa0vhJ/skYrCvfDWtWALTWRKD+JTmvacUhUEYIBHvT5mB4EeG2srKw6hhijFe1aloGmamMXdqrH1XiuN1b4fSxBpNLn3jtCf8AGqUxHD0lWL2zurCYw3sDROO2OPzqvVAFJRRTAKKKKACiiigAopKWgAooooAKKKKACiikoAKWkopgL2opKKACiiigAooooEFFFFAwooooAKKKKACkpaKACtbSf+PZvrWTWtpX/Hsf96kBjXP/AB9y/wC9UVS3X/H1L/vVFVAFFAooAWiigUAFFFFABRRRQAUUUUAHY11/hH/kHTVyHY113hH/AJB01JgZV7/x/TfWoKmvf+P6b61DSAKKKKACkopaACikpaACikooAKKKKACiiimAUUUUAFFFFABRRRQAUUlFAC0UlFABRRS5oASg0UlAC0UlFMBaKSikAtFGaOAOTTAKMepFXtN0nUNVfbZW7Mvdm4xXZaX8P4E2vqcxmPUoOMVLkkB5+gaRtsSO7H0Wtaz8K63egMtp5cZ/iJr1Wy0qxsY9lrbIg9xk1cAx7fSocxnndr8OZmAN3qGPVQK17bwBo8YHnq0p+uK63FFTzMDDi8I6HF9yzH4mrKaBpSDizT8q080ZpXYGf/Yemf8APpH+VRSeHdJkHzWaVq5ozRdgc9N4L0GUHNoQfUNWVdfDuwfJtZzD9ea7akxT5mB5he+ANTgybW4FwOw6VgXulalYMReWbpjuBmvb8U1o0cYdFYf7QzTU2B4MpB9R9eKfXrep+EtJ1AMzW4jkPR14xXG6t4H1CyBksm+1RjnHQirU0wOVop0iPFIY5UZHHUMMUlWISiikpjCiikoELSUUUAFFFFAC0nFBpKBhS0UUwCiiigAopKKBBR2opO1AxaSlooAKKKKACiikoAWikooAKKKKACiiigQtFFFAGnpf+oNZN1/x9yVraX/qDWVdf8fclJbgRUUlFUAtFFFABRRRTAKKKKQBRRRQAUUUUwNPw7/yFx9K1vEn/H2n0rJ8O/8AIXH0rW8Sf8fafSoe4GT3pKKKBhmiiigQUUUUAFFFFABRRRQAUlFFAC0lFFAC0lFFABRRSU7DFpKWkoAKKKKACijNFABR1opKACiiigAooooAKWkopgLSUUUgCiikpgLSUUUwCiiigAooooAKKKKACikooAKKKKACilooAKKDgDJrpNA8I3eqoLi5Y29ufu5HL0nJIDm/xH50GvRpfAGntFiKVkfH3q43XdCu9DnCT/vIXPySjvUxqJgZVFFJVgFFFFMQUUUUDENLSUUAFFFFABRRSUALSZoooAWkoooAKUUUUAFFFFAgooooAKKSigYtJRS0AJW1b/8AHj/wE1inpW1b/wDHj/wE0mBz5+8frRQfvH60VQBSUtJQAtFJS0AFFFLTEFFFFIAooooAKKKKACiiimAUUUUxlKiiiuQQUUUUAFFFFABRRRQAUUUUAFJS0UCCiiigYUlLSUwJ7L/kIW/+9XZeIf8AkGxVxtl/yELf/ersvEP/ACDYqlgcz2FFHYUUAFBooNAB2pM0UUAFFGKKYC0lFLSASig0UAGaM0UUAJS0UmaYC0lFFABRRRQAUUlLQAlLSUUALRSUUAFFFFABRRirml6XeavdC3sYix/ic9F/Gi4FVQWYIil2PAVRk11+g+Bbm82z6qfJi6iIdWFdV4c8J2eioJHAnuiOZGHSui6Vm5DKWnaXZabCIrOBY1Htk/nV2imSSxxIWkdUUdSxxUgPppOOvFcrq/jrTbLMdp/pUw6qOBXGan4x1m/LLHL9mib+ACmotgep3Wo2dom65uI0H+9WFd+OtCtshbkysOwFeUylpnLzSyOx9WNNCqOij8qrlEehT/Em3UnyLFpPfNVJPiVdE/u9Nx+NcTj6UU+VAdl/wse//wCfAfnU0fxJuR/rNNz/AMCrh6UU+VAehw/Ei3JHn2LR++a1rTx1oVxhWuTGx7EV5P8AgKCFPVR+VLlC57pa6hZ3abre4jcf71Ws14FGzROGikkRvZjW7p3jDWrAqGn+0RD+AipcRnsFFcfpPj/TbsiO9U2sp6A8iurhninjEkMiup6FTmlYCG+sLS/gaG7gWRG65HP51wOveAZYN9xozb16+Qe3416PRQnYDwF0eKVopkMci8FWGKSvY/EPhiw1yImRBFcAfJKo6V5brWiX2h3PlXkZMZ+5KOjVopXEZ9GaSiqAKWkpaAEpaKKACiik+lAC0UmaKYBRRRQAtJRRQAUUUUAFFFFABRRRQAUUUUAFFFFABSUtJTAKKKKAAdK19J/49j9aya1tK/49j/vUmBjXX/H1L/vVFUt1/wAfcv8AvVFTAKKKWgAooooAKKKKACg0UUAFFFJQAvY11/hH/kGy1yHY11/hD/kGzUmBkXv/AB/TfWoamvf+P6b61BSAWikooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFJQAUUUUDCiig0CFpKWkpgFFFFIApKWkpgFFBpM0AFFNJwMk8VveHvC97rbCRgYLQHlyOW9qTdgMm0tri+uBb2cJllPYdPzrvNB8CQxBZ9XbzZOvldlrp9J0ey0m3ENnCq+rHqTWgKylO4xkMEUEYjhjVFAwABipKKr3d7bWcRluZljUep5qALFISAMkgD3ri9T8f2sTFNOhNwf7x4xXJ6h4l1fUMrJdFIz/AAKMVSi2B6neavp9kM3N1Gv0OaxLrx3okORHMZW9AK8vI3NudnY+7ZpMD+6Pyq1TA72b4jRA/udPL++aqP8AEa5P3NOx+NcaaSq5EB2A+Il73sB+dTR/EeUf6zTP/Hq4mijkQHolv8RLBiBcW7RfrWxaeL9DuyBHdgMexGK8jwO4H5UbVPUY+nFL2aA92huIZlDRSo4Po1S14Xa3NxZuHtLiSNh6tmuk03x1qVqQt4gul9emKhwYHqFFYGkeLdL1PaizCKY9Ubj9a3gwYZBBHqKi1gM3VtB0/Vo9t3ApbHDDgivPde8HX2lhprXNzbjknuor1WmnBGOPoapSaA8GyDnrn0PWkzXpniXwZb6juutPxBddSB0f2rze7triyuWt7yIxSqcEHoa2jK4EdFJmjNUIWiikNAxaKKKYgxRRR1oGFFFJQIWkoooGFFJRQAtFJS0AJRRRQAtFFJQAtJRRQAUUUUCCiiigAooooAWikpaANPS/9Qayrr/j7krW0r/UGsm7/wCPuSkgIqKKKoYUUUUAFFFJTELRRSUgFooooAKKKKANPw7/AMhdfpWt4k/4+0+lZPh3/kLr9K1vEn/H2n0qXuMye9JQetFABRRRQIKKSigBaKSigAooooAKKKKACiiimMKKSigAooooAKKKSgBc0maKKACiiigAooooAKWkooAWkoopgFFFFIBKWkopgFFFFABRRRTAKKKKACiiigApKKKQBRRRTAKKTNJmgB2aUkAc0zPIABLHgAdTXc+E/B7Epf6uuO8cJ7e9TKSQEfhDwo1yyahqceIhzHEe/vXoaqFUKoAA4AFCgKAFAAHAApSa5ZSbYBWT4nsUv9DuYnAOFyD6VrVV1P8A5Btx/uGhbgeIKcg57Ej8qWmr/F/vmnV2LYApKWkpgLSUUUALSUUUAFJS0lABRRRQAUUUUDClpKXNAgoopKBC0lFFAwooooAM0UUUCFo60lFAwNbVv/x4/wDATWKa2bf/AI8R/umkwMA/eP1ooP3j9aKoAoxRRTAKKKKQgooooAWiiigAooooAKKKKACiiimAUUUUwKVFFFcgBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFACUUtFMCay/5CFv8A71dl4h/5BsVcbZf8hC3/AN6ux8Qn/iWxVLA5rtRR2FJQAtJRRTAKKWigBKKKKQBRmiigAoopKAFopKKYBRRRQAUUUlAC0lFFABRRRQAUUUlAC0UlLQAUv1pM1u+F/Ds2vXW58pZRn53/AL3sKG7AM8OeHLrXrjjMVop+eQ9/YV6tpemWmlWi29nEEUdT3NT2lrBZ2yW9tGI40GABU1ZN3GFITgZJx9ahu7qCzt2nuZBHGoySTXmniXxnc6kzW2mkw2vQyd3oSuB1PiDxpY6WzQW+Lm5H8APA/GvPNV1zUdXcm8uGEf8ADGpxiswDHqT3J5NLVqNhCjA6CkozSVQAaSlooASilooAKKKM0ALSdaSigBc0lFFACEA9RV3TNY1HSZQ9jcsq90Y5zVOkxQB6h4e8c2eolbe/H2a5PAz0auuUhgCCCD3HSvASobr+Y6103hvxheaO6wXjG4s+nPVKhxGetYqvf2NtqFq9tdxCSNxyDSaff2uo2q3NpKJI2HbqKs5qAPIPFPhW40KQzwBprFjwQOU9q53jqORXvs0Uc8TRSoHjYYZSM5rynxh4WfRZTeWal7Fzyo/5ZmtIyEczQKQe3IpasAooooAKKKKACiiimAUtJRQAUUUZoAKQ0UUAFLSUtABRRRQAUUUUAFBoNJQAUUUUwCiiigArX0n/AI9j/vVkVr6T/wAezf71JgY11/x9S/71RVJdf8fcv+9UdMApaSloAKKKKYBRRRQAUUlFABRRRSAXsa67wh/yDpq5E9K67wj/AMg6akwMm9/4/wCf61BU97/x/wA31qCkAUUUUAFFFFABRRRQAUUUUwCiiigAoopKAFopKKACiiigYUUUUCCiiimAUUmaWgAooooAKSjNFACYpCdo5/TvTiQBk/l613Pg3wlvKanqie8UR/maluwFbwn4Oa72X+rKVi6xwnv716LHGkUapGoVFGABxinbcDjiisW7jDNRyzJDG0krhEUZJY4rP1vXLLRbYy3Ug3n7kY6sa8w1zxDf63KfOcxW+fliU04xuB1Wu+PI4y1vpCea/Qyn+E1w97fXeoTGW+naVz74FVx6AYpK1UUgH59hRTe1LmqELSGiigYUlLSUwDFFFFAgoopaQwopM0ZpgBwfUH1HFbmj+LNT0kqhc3FuP+WbdvxrCopNJgev6H4m0/WUxDKI5gPmjY4xW1Xgylo5BJG5R1OQynGK7fw144ZGS01k5B4Wf+lZShYD0LFZOv6BZ65amO4ULKB8koHKmtSKRJY1kjYMjDIYHNPqL2A8R1fSrvRrw214px/BJ2YVSr2zWNJtdYsmtrtAQfuv3U+teR61o9zol8ba5BKHmOTswraMrgZ9LRS1oAlFFGaBBSUUUDCiiigQUlLSUDCilpKAFpKWigBKKKKACiiigAooooAKKKKBBRRRQAUUUUAFLSCloA09K/1BrKu/+PyStXSv9Qayrv8A4+5KS3AipKWiqGFFFFMAooooEFJS0UgEooooAWiiigDT8O/8hdfpWt4k/wCPpPpWT4d/5C6/StbxJ/x9p9Kl7gZBooNFABSUUUAFFFFABRRRQAUUUZoAKKQ0UxhmiiigAoopKAFpM0UUAFFFFABRRRQAUUUtACUUUUAFFFFABSUtFACUUUdKACiiimAUUUUwCiiigBKKKKQBRRRTAMUtJRSAKSnYpDjOBkn0HJpgIafa21xfXK29nEZZWOOOg/Gug0Pwffaptluc21seQe7fhXomk6LZaRAIrOEKf4m7mspVEtgMPwx4Pg0zbdX2JrvqM9ErqqXpVe8u4LK3ee6kEcaDJJrC7YEzOqKWZgqgZJPQVwOveL2u9Vg0/S3Ig34kl/ve1ZXibxZPrDNbWZMNmDgkdZKwbABdTtAOm8VooaXYHuUX+qT/AHRUGp/8g64/3D/Kpoj+6T/dFQ6n/wAg64/3D/Ks1uB4gvRv99v50tIv8X++386WuxbAJRRRTAKKSloASloooATtRS0lABRRRQAUUUUAAooooAKKKKBBRRRQMKKKKACiiigAooooADWzb/8AHiP901jHpWzb/wDHiP8AdNJgYB+8frRQfvH60VSEFFFFABRRRQAtJRRQAtFFFABRRRQAUUUUwCiiigAooopjKVFFFcggooooAKKKKACiiigAooooAKKKKACiiigBKWiigCay/wCQhb/71dj4i/5BsVcdZf8AIQt/96ux8Q/8g2KkwOZ7ClpB0FLQAlLSUUAFBozRQAUUUUwCiikoAKKKKACiiigApKKKACiiigAooooAKKSloASiijNABSUGnwQy3VxHbW6lpZThQKAL+gaPNrmpLbRZESnMr+gr2SwsoNPs47W2QJHGMACs7w1okOh6YkCAGVhmR+7GtgGs27jHVT1PUrXS7N7q7kCRr+ZNO1G/t9Os5Lq6cJGg/OvIPEGuXGvXpmmJW3Q4ii7CklcB/iHxBda9clpCY7VT8kQNY9Gc0VolYQlFFFMAooooAKKKKACiiigBKKKKACiiigAooooAWkopaACiikpgaGi61eaHdie0cmMn95ETwwr1vQ9atNbsVuLVxnHzp3U14metXdI1S50a/W7tGI/56J2YVMo3A9yqO4t47mB4ZkDxuMMDVPRNXttZsEurZhyPmXuprSrMZ4z4q8PSaDqB2AmzlOY2/u+1YmK9z1jTINW06S0uEDK4+U+h7GvF9S0+fS9QlsrkHfGflb+8PWri7iKmKKWirEJiilpKYwooooAKSiigAooopAFFFFMAooooAWiiigAooooAKSiimAUUUUAFGKKKACtbSv8Aj3P1rJrW0r/j3b60mBjXP/H1L/vVHUlz/wAfUv8AvVHTAKWkpaYBRRRQAUUUUAFJS0UgEpaSlpgB6Guu8I/8g6auR7V13hH/AJB81SwMm+/4/pvrUFT33/H9N9agpAFFBooAKKKKYBRRRQAUUUlABRRRSAKKKKACjtRmimAZoopKYBS0hopAFFFFABRR0pM0wFopKWgAo6Ak9BR1ra8L6I+t6koYEWsRzI3qfSk3YDW8FeGTeyrqV+n7hDmND/EfWvSAMDAGBTYYo4IliiUKiDCqOwp9YN3GFc/4o8SwaJb7VxJduPkQdven+KPEEWiWRIIa5cYjT+teUXNxNd3L3N05eaQ5JPaqjG4Be3dzf3TXN5IZJW9ei+1QUppK2SEFGaKKACiiigAozRmigAooooGFFFFABRRRQAUlFFMAooooEFIQCMEcGloxQB0PhbxTPosy290xlsWOOeqe9eq2s8N1Ak8Dh43GVYV4Vj1ro/CXiSTRblbe4YtZSHnP8B9aynDqhnrFZuuaPbazYNbXCjPVH7qfWr8MqTRLLEwZHGQR3p9ZbAeHajp8+mX0lndLh0+6f7w9aq9K9X8Y+H11jTzLCMXUI3IR/F7V5QwZWZXXa6nDD0NdEJXQCUlLSVYBRRSUAFFFFABRS0ZoAKSiigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUAFLSUtAGppf8AqDWRd/8AH3JWtpf/AB7msm6/4+5KSAjopKWrGFFFFAhKWiigAooopAFFFFABRRRQBp+Hf+Quv0rV8Sf8fSfSsrw7/wAhdfpWr4k/4+0+lS9wMmkpTSUAFFFFABRRRQAUUZpM0AGaKKKYxKWjNFABSUUUAFFFFABRRRQAUUUUAFFFFABRRSUALRRRQAUUUUwCkoooADRRSUALRSUZoAWikoyKACijIozQAUUtGKAEpM0EqP4gfYGpba0u7xwlrbSOx9VwKLpARUoI6csfReTXUab4D1G5Ia/lFsnXC812GkeFNL0za6wiWYf8tGqHUSA4HSPDGqaqVZYvJgPV24P5V3mi+EdO0vbIyie4H/LRq6AAAYAAHtS1jKbYCAYHFLUF3eW9nC0tzKsaqMnJ5rhtd8dNJug0dSB0889R+FSotgdNr3iKx0aEmZw8x+5Gpzk+9eX61rV9rc++7crED8kQPAqpK7zStNM5klY5LE1Ga6IU7ANqay/5CVr/AL4qLFTWX/IStf8AfFW9gPb4v9VH/uCodS/5Btx/uH+VTxf6qP8A3RVfUv8AkG3H+4f5VyLcDxFOjf77fzpaah+9/vt/OnV2LYAooopgJS0UlAC0UUUAFFJRQAUUUUAFFFFABRRRQIKKKKBhRRRQAUUUUAFFFFABRRRQAHpWxb/8eP8AwE1jmtm3/wCPH/gJpMDAPU/Wig/eP1oqkIKKKKACiiigAooooAKKKKAFooooAKBRRTGFLSUUCCiiimMpUUUVyCCiiigAooooAKKKKACiiigAooooASlpKKYC0UUUgJrL/kIW/wDvV2PiH/kHRVx1l/yELf8A3q7DxD/yDoqTA5rsKKM8CigAooopgFFFJQAuaSiigAooooAKSiigAooooAKKKKACiiigBKKWkoAKKKKACkpaSgBCQASa9D+HugeVD/a92n7yT/VA/wANcf4c0ltZ1mK2wfJU5lPpXtEUaxRLGgwqDaAPaokwHnimSSJFG0jsFVRkk0+uB+ImvmNBpFo/zv8A64j+GpWoznPF3iGTXL8xxMRZQnCD+/WEDTQMDFOFaJCFopKWmAUUUUAFFFJQAtFJRQAUUUUAFFFFABRRRQAUUUUAFFFJmgBaKTNFMBaBRRQBreG9bl0LUVmQk27nEqdj717JaXMN5bJcW7h45BlTXg3sehrtPh9rxtbr+ybp/wB1JzESfu+1RJAel1ynjvQP7U0/7XbqPtVuMg+q966ukYBlIYZB6is07DPAAcjOMHuPSiui8b6P/ZOtNLGuLe5+Yein0rnTWydxCUUZpKYBRRRQAUUUUAFFFFMAooooAKKKKAFpKKKAFpKWkoAKKKKYBRRRQAUUtFACVraV/wAe7fWsqtXSv+Pc/WkwMa5/4+pf96o6kuf+PqX/AHqjpgFLSUUwFoopKAFooooAKKKKQBRRRTAOxrrvCP8AyD5q5HtXXeEf+QfNUsDJvf8Aj+m+tQVPe/8AH9N9agpAFFFFMAooooAKKSigAooopAFFFFMAzRRRQMKSlpKBBRRRTAKKKKACkoooAKKKKACjNJR0yT0FAEtvDJdXEdtCCZJW2jHb3r2TQdKi0jTI7WNRvxmRv7xrjvh1o255NWuF5HyxA9x616EKxm7jFqhrOqQaTp8l1OR8o+Vf7x9KuyOsaM7nCqMk+gryXxbrbazqjKh/0WA7VHZj60oq7AzNS1CfVL57y6JLMflX+6PSqtBNJW6VhC9qSiigAopQCSAqszH+FRk1JJbXMa7pLWZV9dhpXGQmkp3bjkUUwEooooAKKKKBBRRSUwCilpKBhRiloFACYop6KzuEjRnY/wAKjJp81tcQDM1tKg9SvFK4iKil69KDTASl4IwRkGkooA7PwN4jNrOul30mYXP7pz2PpXpAORXgnPG04Ycg+lep+CNeGq6cLedv9JgGCPVfWsZxtqM6ivNPiBoYs7pdTtkxFKdsijsfWvS6qanZRajYTWkw+WVcZ9KiLswPDTiipr60ksL6azmBDRsdue69qr5rpQC0UUUwClpKKACiiigAooooAKKKKACiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBqaX/AKg1k3X/AB9yVraX/qDWTdf8fclCAjopKWqAKKKKACiiikAUUUUAFFFFABRRRQBp+Hf+Quv0rU8Sf8fafSsvw7/yFx9K1PEn/H2n0qXuBk0UHrRQMKKKKBBRRSUDCiikpgLSUUUAFFFFAgooooGFFFFABRRRQIKKSloGFFFJQAtJRRTAKWk705EeWQRxI0jnoqjNIBKQlfUH2zXWaT4GvboCTUZPs6HkKvJNdXZeENGtVGbZZXH8TVDqJAeVKkj8JBK30WpPsl2Rxazf98GvaYbO2gXbDAiD2FS7R6L+VR7UDxH7Jd/8+k3/AHwaPsl3/wA+k3/fBr27aPQflRtHov5Ue1YHh5tLv/n0m/74NJ9kvP8An0m/74Ne47F/ur+VGxf7q/lR7VgeHfZLz/n0m/74NKLK9PS0m/74Ne37B/dX8qNo9F/Kj2rA8UXTNRf7tnJ+K1ag8Oa3P/q7MfjxXsOPYflS/lR7VgeXW/gfWpcecI4h7NWta/DteGur9z/sgV3gpal1GBz9l4O0a1AJthK4/iatuG3igQJDGqKOwFS1HLPDCCZZY0A/vNipu2A/FJXP6j4y0ex3L5xkkHQKMiuV1Lx5f3AKWUAgXtJnmqUGwPQ7q9trOMyXEyIoHOTzXH6v4/hTdFpUXnNjHmNxiuFvLq6vZDJeXDzN7nFQ/StI0u4Fu/1C81KXzb64aU9hnAFVs0gorVKwBRRRTAMVLZD/AImVr/vioqnsf+Qna/74pS2A9uiH7pP90VX1Mf8AEtuP9w/yqxF/qk/3RUGp/wDIOuP9w/yrkW4Hhqfxf75/nT6Yn8X++1OrsWwC0UlLTAKKKSgBaSiikAUUUGgAoopKYC0UUUAFFFFABRRRQIKKKKBhRRRQAUUUUCCiiigYdq2Lf/jx/wCAmsc1sW//AB4/8BNJgYJ+8frRQfvH60VQgooooAKKKKACiiigAooooAKWkpaACiiigYUUUUwEOaKWimBSooorkEFFFFABRRRQAUUUUAFFFFABRSUUAFLSUtMAooopATWX/IQt/wDersPEX/IOirj7L/j/ALf/AHq6/wAQ/wDIOipAc2OlFA6Cg0wCkoooAKKKKACig0lAC0UlFABRQaKACiiigAoopKBC0lFFAwooooAKKKKACkPAzS1f0SwOp61bWePlZsufTFAHoXw/0j7BpH2qVcT3PLZ9O1dZ2psUYiiSNeAihR+FONZNjM/W9Sj0nS57yU8IMAeprxO4nlvLmW6uGLSytkk+nauw+JOqme9j0yJvli5lHrXF4q4oQUtFFUAUtJRQAtJRRQAUtJRQAUUUUAFFFFABRRRQAUUUUAFJS0maADNJSVreHNCuNfvxDHlLdD+9k/wouBQtbO7vCRaW7y47gcUXFrc2j7bqB4j6kYFe36bp1rplottZxKiKOeOtRa5pltqumy29ygPykqccg1HOB4jRQ6GKSSJuqMRSVoA6lVmRlkjOHjO5SPUU2jNAHs3hXWF1nR4py2ZkG2UehraryTwFq39m659nkbFvc9f97tXrWaxkrDMHxlpK6toM0eP3kQ8xT9O1eNgsV+YYYcEelfQLKGBDcg9RXjHi7TTpfiKdAMR3BMiewqoMTMbNHakpa0AKKKM0wCiiigAooooAKKKKACiiigAooooABRRRTAKKKKACiiloASlpKWgArV0r/j3P1rJrW0r/AI9z9aTAxrn/AI+pf96o6luf+PqX/eqKmAtFJS0wEpaSigAopaSkAtFFFMAooooACOK63wl/yD5q5Lsa6zwl/wAeE1SwMq9/4/pvrUNTXv8Ax/TfWoaQBRRSUAFFFFMQUUUUAGaKKKBhRRSZoAWkoooAKKKKYBRRSUgCiiimAUlFGKACiiigAqS1tnvLyG1jBLSMAfpUVdh8N9N+0ahNqMi/LD8i5pSdkB6Dp1mlhYQ2kf3YlAqzRTJ5VgheVzhY1LGucZyfj/WzY2AsLd8T3A5x/d715quFAA6Crutag2q6xcXbtlNxEXstUq3irIQtFJRVALRnHJpKQjIwe9AHonw90aD7D/aVxGHndsLkdBXaSwxTRmOVFZGGCCOtch8PdXhn006fIwW4iOdp7iuzFYS3Geb+L/CRs92oaWhMPWSIfw+4rjeCMjpXvLKGUqwBB4INeb+MvCjWTvqOmoTA3MkQ/h9xVRn0YHG0UmQRkHijNagLSUUUwCiiikIWikooGLVixsrjULtLW0QvI35AU2ztbi/u0tbRC8rnt/CPWvWfDPh2DQ7MAAPcuMySH1qZSsIZ4b8M2uiwBiokumHzSEdPati6tILuBoLiNXjYYIIqemSyJDG0kjBUUZYntWN7jPGdf08aXrU9qn+rzmP6VnVp+JNQTU9dnuIuYlO1D6isuuiOwBRRSE1QC1d0bU5NI1WG8jJCAgSj1WqGTSHB4boaT1QHvFrcJdW0c8RBSRQwqQ81xHw21Uz2UmmzNmSE7hn+7Xbg5rmkrMDz74laTjytVhX5vuy47CuD4IBHQ17jq9iuo6ZcWjD/AFq4HtXiEkTQTywOMNE5X8BWtN3QCUlFFaiFooooGFFFFAgooooAKKKKACiiigAooooAKKKKACiiigAooooGFFFFABRRRQBqaX/qDWTdf8fUla2l/wCoNZN1/wAfclCERUtFFMAooopgFFFFIAooooAKKKKACiiigDT8Pf8AIXH0rU8Sf8fafSsvw9/yFx9K1PEn/H2n0qXuBknrRR3ooAKSiigAoopKYBRRRQMKKKKAClpKKACiiigAooooAKSiigBaTNFFABRRSUwClpKs6Zp8+qX6WdsDuY/M390etJuwE2kaVdazeC3tFO0ffk7KK9P0Pw7ZaNCPKQPPj5pWHJqxo2lW+kWKW1uoGB8zd2NaGa55TuAUZrN1fWrHSIPMvJQD2UfeP4VwmqeOdQumK2CC3j7OOppKLYHpbyxp9+RF/wB5sVH9sth1uYf++xXjFxqF9dsTc3ckhPviq/Pd3/76NX7Jge3fbbX/AJ+Yf++xR9ttf+fmH/vsV4jj/af/AL6NGP8Aab/vo0/ZAe3fbbX/AJ+Yf++xS/bbX/n5h/77FeIf8Cf/AL6NH/An/wC+jR7ID243tp/z8w/99ipUdZFDIwZT0IOQa8LcfIfmbp/eNeveEhjw1ZDJPyd6iUOUDXrKu/EWlWcrRT3QDpwwHatXHIrxfxPFH/wk+oHbyZPWiEbgegT+O9Dhztmkc+y1l3PxGTkWdn5noWOK4IBR0FOrVUkB0N5421q5J8oi2H+zzWLc3t5eNuvLp5SffFQYoq1FIBQAvT9eaDRR1qgDFJS0ZoAKM0maKAFopKUUwCp7H/kJ2v8AvioM1NZH/iZWv++KmWwHt8X+qT/dFQ6iM6fcAf3D/Kpov9Un+6KccEYIyDXGB4GMqXDqyfO33hjvTuvI5Fez3ug6XeqRcWcbE98VyWrfD4hWl0qclv8Ank3ArojVQHDYoqa6tbiynMF3E0Ug9RwfpUNap3AWikopgFFFFABRRRQIKSlooAKKKKACiiigAooooAKKKKBhRRRQAUUUUCCiiigYGti3/wCPH/gJrHNbFv8A8eP/AAE0mBgn7x+tFB+8frRVCCiiigAooooAKKKKBhS0UUAFFFFAgooooAKKKKYwooooEUqKKK5QCiiigAooooAKKKQ0ALSUUUwCiiigBaKKKQBRRRQBNZf8f9v/AL1df4h/5B0VchZ/8f8Ab/71dd4h/wCQdFSYHN9hRR2FFMAoopKACiiigBaSg0UAFFFFABRSUUAFFFFABRRRQAUUUUAJS0UlAC0UUUALXc/DOw3S3OoOuQcBD6VwjHapP4V7D4MsvsPhy3jIwW+Y/jUyegI3qgvrlLSzluJDhUUnNT1yXxEv/s2hfZlOHuTgVmhnmV5cve3s91KcvI559qhpSOAPakrYQUtJRQIKKKKBhRRRQAUUUUAFFFFABRRRQAUUUUAFFFFACGkNOqaysrjULxLS0QvLIcf7v1oAfpGl3Os6glnaKeT879lFey6NpVto+npaWqgBR8zd2NQeG9At9C08QxANMwzLJ3Y1r4xWbdwDNZHiTWYNF0ySeVgZGUqidzmrGsapbaRYSXd24CqPlHdj7V45rOrXWtX7Xd0SBn93H2UUJXGUWZpHeR/vOxY0UUVqIKKKKADe0bJKpw0TBx+Fe2+HdRGp6LbXWcuyfP7GvEvb1r0D4X3+YrqwduQ25B7VEloB6B14rhfifp/madDqCDLxNtP0ru6zfENot7ol3CwyfLJH1qIgeH57+tFIqlB5bfeQ4NLW4BS0UUAFFFFABRRRQAtJRmigAooooAKKKKACiiimAUUUUAFFFFAC0UlLQAla2l/8e5+tZNaul/8AHufrSYGPc/8AH1L/AL1R1Jc/8fUv+9UdMApaSigApaSloASlpKUUAFFFHSgAooopgHY11nhLiwmrk+xrrPCX/HhNUsDKvP8Aj+m+tQ1Lef8AH9N9ahpAFFFFMAooooAKSiigBaSiigAooooAKKKDQAUUlFABRRRTAKSlooAKKKKACm0tGKAGP9w46ngfWvX/AAbYix8O24K4eRdz/WvKtPtjd6na26jkyBj9K9wjjWONY1GFUAAVnNgOrlvH+pmw0IxRtiWY7ce1dT7V5V8Qb77Xr6QKfkt12ke9RFXYzmVG1Qo7UtAorcQUUtFAwooooAltria0uUuLZyk0ZyCK9W8LeJIdbtQjkJdoMOh7+4ryWpbW5ns7lLm1kKTIcgjvUSjcD3WmuqupVgCpGCD3rB8L+JYNctQrEJdoMPGe/uK3s5rHYDzPxl4UawkbUdNQtbMcyRj+H3FceCCMg8V726K6FHUMrDBB715j4x8KNp0rahpyFrVjl4x/B9K1hIDk80tIMEZHIpRWggopaKBjc1Ja2897dJa2iF5XOMDt70ttbT3lyltaoXmc4AHb616v4W8NQaHahmAe7cZdz29hUylYQ7wv4cg0O05Ae6cZkkPb2rfpOBTJJFjRndgqKMkntWF7jFllSGNpJGCooySe1eYeLvFUmqytZ2LlLRThmH8dJ4w8VPqczWNi5W0Q4dh/Ga5UcDA4ArWEOrAXgDA4FFGaK1EBpKKSmMDRRRQI0vDd+2m6/bTA4SRtsn0r2lGV1DqchhkV4E+dh28MORXs3hW/GoaBbTZyVUIfwrGohmwa8m8faeLLxEZ0XEU6jH1r1oVxvxJsfP0aK6Uc27bifapg7MDzPFLS9QD6jNFdACUUUUCCiiigAooooAKKKKACiiigAooooAKKKKACig0UAFFFFABRRRQAUtJS0Aael/6g1k3X/H09aumf6g1lXP8Ax9PSQEdFFFWAUUUUgCiiigAooooAKKKKACiiigDT8Pf8hZfpWp4j/wCPpPpWX4e/5Cw+laniP/j6T6VL3Aye9JQetFABRRSUAGaKKKYwooooEFFFJQMWikoFAC0lLSUALSUUUAFFFFABRRRQAlFLSGmAhJ/hGWPQetepeDNEXStNWWRQbmcbmbuB6Vw/hHTf7S1+MEZjt/3h969aAA4HAHQVjUl0Ad0rA8VeIotEtdqYe6k4RPT3rU1K+j0+wmu5iNsS5x6143qF/Nqd/Le3DFmc/LnsvaohG7AZd3VxfXLXF5K0krHqT09qjpKWuhKwC0UUtMApKKKAFopKSgBW+6a9f8J/8i3Zf7lePt9w16/4S/5Fqy/3Kyq7AbNeNeKB/wAVPf8A+/XsteN+J/8AkZr/AP36mluBk0tFLiugBM0UYooAWjpSUUALmkopaACkoooAWikooAKlsf8AkJ2v++KiqWx/5Clr/vilLYD3CL/VJ/uinZA5JwKbF/qk/wB0VX1TI0y4wcHYcGuMC2rBhlSCPUUteM6T4m1fTXLJctPGHOY36Yr0vw74ks9cg/dtsuFHzxn+lU4NAT65olprNqYrlAHA+SQDlTXkmr6Zc6PfNaXQPB+R+zCvbic1h+KtEj1rS3TAE8Y3Rv3HtVQnZgeQ0UMjxu0ci7ZEOGHpRXSAtFFFMQUUUUAFFFFAwooooEFFFFAwooooEFFFFABRRRQAUUUUAFFFFMYHpWvB/wAeX/ATWQela9v/AMeX/ATUsDCP3j9aKD94/WiqEFFFFABRRRQAUUUUDCloooAKKKKBBRRRQAUUUUxhRRiimIpUUUVyAFFFFABRRRQAUlLRQAlFFFMAoopaACiiikAUUUUAS2f/AB/wf71dd4hP/EvirkrP/j/g/wB6ut8Q/wDIPipMDnOwpKPSimAUGiigAooooAKKKSgBaSiigAooooAKKKSgBaSiigAooooAKKKKAFzRmm5ooAmt4jcXUMAHMjj+de6WqCO1hQDAVAP0rxvwtCLjxTYIegbJr2ofyrOQIK8v+I955+tQ2ueLcZx9a9PJwK8W8T3P2vxPezZyOFH4UR3BmXSGlpK0ASilooASijFFABRRRQAlLRRQAUUUlAC0lGaTNAC0UlLTAWiinIjySLHEheRzhVHU0AOt4Jrq4S3tkLzSHCqK9b8J+GodDtA7gPdyDLue3tVbwb4WTR7cXV0A97IMk/3B7V1VZSlcYVXvryCwtHublwkaDJJqWWVIYmklcKiDLMTwK8l8YeJH1y7MFuStjEcD/bPvSSuBQ8Sa7Pr+oGVyVtoziKPt9aycU44orVKwhtFLijimAlGaKKAENbvgi8+xeKYWJwki7D9TWFU1pKYdQtJQcFZlpPYD3z2prKHUq3Rhg0yCQTQpKOjqDUnesRnhmuW/2XX76HGB5px9Ko10nxAtvs/inIHEqbq5ut1sIKWkpaYBRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigBa1NM/wCPdvrWVWppn+oP1pAZFz/x8yf71R1Jcf8AHzJ/vVHTAKWkpaACkpaKACiiigAooooASloooAD3rq/Cf/HhNXKdjXV+E/8AjwmpMDKvP+P2b61DU15/x+zfWoKQC0UUUwCkoooAKKKKACiikoAWiikoAKKKM0AFFJRQAtJRS0wEooooAKWikoAWiigdaAOi8B23n+KUkIysUZ/OvV68/wDhlbgte3JHIbaK9BrCe4yKZxHE8h/gUtXh+qTm71e8uCc+ZJkV7Lr04t9Fu5On7sgflXiKZMYJ6nNVTAUUuaKK1AKWkooAWiikoAWkozSUATWt1PZXSXVo5SZDkEd69Y8L+I4NctByEukGJI/6ivIams7u4sbtLq0cpKhzx3qJRuI91zmkeNJEZHUMrDBB71i+GfENvrloCCEuUGJI/wCordrHYZ5j4v8ACTac73+nKWtmOZIx/B9K5IEEZHQ17zIiyIyOoZWGCD3ry3xh4Xk0mZr6yQvZucsoHKVrCXRgcxmpbS2nvbpLa1QvK5wAO31plrDNe3EdvaL5kshwAOcfWvWPCvhqHQ7UM4D3bjLue3sKqUrCF8L+GoNDtgzAPduPnkPb2FdBRSMwVSzEADkk9qweoxsjKqlmIVVGST2rzLxl4ra/kfTtOcrbqcSSD+KpvGnitrt303TZCIVOJZR39hXFgADAGBWsIdwEGAMAcUtGKK2EKKKSigYtJRmigAooooAB1r0H4YXebS4sSf8AVnf+defV1Hw8uPI8RumeJkC1E1oB6sOlZniO2+1aBeQ4zuTitOmSoJInQ9GUiudAeCjgFf7p2/lRU15H5N9cxntK1Q11LYBKKKKYgooooAKKKKACiiigAooooAKKKKACiiigYUUUUCCiiigAooooAKWkpaANLTP9Qayrn/j6krV03/Umsq5/4+pKSAjoooqgCiijimAUUUUgCiiigAooooAKKKKANLw9/wAhZfpWr4k/4+0+lZXh7/kLr9K1fEn/AB9p9Kl7gZBpKD1ooAKKKKYwooooAKSlpKBBmiiigYUUcUUAFFJRQAtFFJQAtFFFABRSCloAKO+aKRziNj7UAehfDazCadLeMPndyo+ldnjisfwlbi38P24A++N1bVc0nqBwHxJ1AhYNOjbBzuf3FcLjHFbvjG4+1eJpnzwi7RWJit4KyAbS0UVYBS0lFAC0maKKACiikzQAN9017B4S/wCRasv9yvHm+4a9h8Jf8i1Zf7lZVdgNmvG/E/8AyM9//v17JXjXif8A5Ge//wB+ppbgZlFGaK6ACigUUAFFFJQAtFJRmgAozSUtABRSUUwCprH/AJClp/vioamsP+Qpaf74qZbAe4xf6pP90VX1T/kGXH+4f5VYi/1Sf7oqDUx/xLbj/cP8q5FuB4Wn8X++f51PaXU9jdx3dq5SWM547ioVH3v99v507FdVroD2jQtUj1jS4ryPALD5l9DWlivPfhpeFJ7qzY8Mdyj0r0OuaSswPKvHmmix14ToMJdDcfY1zXSvRfiXbg6VFdY5jbH51523Wuim7oBKKKK0EFFJRQMWigUUAFFFFABRRRQAUUUUCCiiigAooooAKKKKBhRRRTEBrXg/48f+A1kHpWtb/wDHl/wGpYzDP3j9aKD95vrRVCCiiimAUUUUAFFFFAxaKKKQBRRRQIKKKKYwFLSUtACUUUUwKVFFFcggooooAKKKKACkpaSgAoopRQAYooooAKKKKACiiigCaz/4/wCD/errPEH/ACDoq5Kz/wCP+D/errfEH/IPipMDnOwoo7CimAUUlLQAUlFFABRRRQAUUlLQAUlFLQAlFLSUAFFFFABRRSUAFFFFACUZopDQB0ngCPzPE6t/zzGa9d9a8q+Gwz4guD6KK9VrKQxkzbYZG9FJrweeQyXlxIf4pG/nXuWoHbp85H9w/wAq8HU/NIf+mjfzqoCH0UlLVgFFFFABRjPagnAzXVaF4Gu9UtkubuXyIXGUA+9ik3YDlSDRXWa14EvdPhaexk+0RqMsp+9+FcnzkggqRwVPUUJ3AKSlpKYBSGig0wEooooAKKKQkAZPSgAzjGASTwAOpr03wN4W+xRrqWoIDdOMop/gHr9azfAnhUyMuranHgdYYyP1r0UVnJjH0jMqKWYgAckntRmvO/HPisyO2labJwOJpFP6CpSuBS8aeKW1KZtOsHItUOJHH8Z9K5LoMDpSDAGBRWqVhBSZoJopgFFFFACUUtFADaQnbhv7pzTsU1x+7f6UAe5eH5PN0KykPO6IVoN3rH8Ind4YsPaOtgjisXuM8y+KMeNWtJcf8s8VxZ613fxUX97aN+FcKeprSOwhKWkoqwFooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAVqaZ/qG+tZdammf8AHu31pAZFx/x8yf71R1Jcf8fMv+9UdMApaSigBaSlopgFFFFIAooooAKKKKYB2NdX4T/48Jq5Tsa6rwn/AMeE1SwMq8/4/ZvrUNTXn/H7N9ahpAFFFFMAopKKAFopKKACiig0AFFJRTAKKKKACijNFABRRQKACijNFABRRRQAUvY/SkoP3G+lAHpfw3j26NM/99812Fcr8PRjw8p9TXVVzy3Gc746lMXhmYg9WAryUDCgV6h8SGx4aI9ZRXmBrSnsADpSUtJWgBRRSUALRSUtACUtJS0CCiiigCxZXlxYXaXVo5SVD2/ir1nw14ht9ctAwIW4UYeOvH6sWV5cafdpdWjlJUP4MPeplG4z3OmTRJNG0cihkYYIPesjw34gt9csw6kJOoxJGeorarDYDG0fw3p2j3E09pH88pzkj7v0rYAxS0hIUEk4A5JoAGYKpZiABySe1ec+MvFrXLvpumSYiHEso7+wp3jLxabgvpulyYj6SzL39hXEYAXArWEOrATAAwOlFFFbAFIaWkoAKSiigApaSloAKKKKAFFa3haXyvE9if7z4NZIq9oh269YH/ppUy2Ee3+tFIOVBpa5hniviOPyfEN0nvmsytzxou3xXcgf3RWFXTHYBaSlpKoQUUUUAFFFFAwooooEFFFFAwooooAKKKKACiiigQUtJS0AJRRRQAUUUUAaWmf6k1l3P/H1JWrpn+oNZVz/AMfUlCAjooopgFFFFABRRRQAUUUUAFFJS0AFFFGaANLw/wD8hZfpWp4j/wCPtPpWV4f/AOQsv0rU8R/8fS/Spe4zJNFHekpgLRmkooAKKKKBBRRRQAUUUlAwooooAKKKKACiiimAUUUUALRSUZoAWkf/AFTUUj8xt9KTA9p0LA0Szx/zyFX+9ZHhaYTaBakHO1Ata9cr3A8X1wk63dk9fMNUDWz4ttza+JJ4yMbhvrGrpjsAUlFFUAUUlFABRRRQAUUUUwEb7hr2Hwl/yLVl/uV4833TXsPhL/kWrL/crGrsBs1414n/AORmv/8Afr2WvGvE/wDyM1//AL9TS3Ay6KKK6AFpKKKYBmjNJRSAKKKKACiiigAooooASp7D/kKWn++Kgqaw/wCQpaf74pS2A9yi/wBUn+6Kg1P/AJBtx/uH+VTxf6pP90VBqf8AyDrj/cP8q5FuB4co+9/vt/OlpF6N/vt/Ol712LYDpfh+SPEqgdChzXqorzf4bWjSahcXZHEXyg/WvSK5anxAcr8Rsf8ACLyZ/wCei15ea9F+JlwF0iK2zzI2fyrzo1tS2AKKSitQFpKKUUAFFFFABRRQaAEpaSloAKKKKBBRRRQAUUUUAFFFFMAooooAQ1rW/wDx5f8AAayTWtB/x5f8BpMZidz9aKP4j9aKYgooopgFFFLQAlLSUtAxKWiikAUUUUCCiiimAUtJRQAUUUUxlKiiiuQQUUUUAFFFFACUUUUwClpKWkAtJS0lABRRRQAUUUlMCaz/AOP+D/errPEH/IPirlLL/j/t/wDerq/EP/IPiqWBznYUUdqKYBRRRQAUUlFAC0lFFABS0lFABRRRQAUUUlAC0lFFABRRRQAUUUUAJSGlpDQB1fw2bHiGceqivVe5ryT4fPs8T7T/ABjFet9z9azluBX1EZ0+4H+wf5V4Mv3pP+ujfzr324Xfbyr6of5V4Ky7J51PUSt/OnEAoooqwClpKWgCW1CtfWyv90yrn8693iVViVUACgDGK8C9MHBByD6GvTfBXiqO+hTTr5wl1GMIxP3x/jUSQHZ4rhfGnhETq+paWgWdeZIx0f3+td1SGpTsM8AJ5IIKkHBB6ikzXoHjfwiZN+qaWn7wcyxAfe9688Bznggjgg9RWqdxD6KSlFMAxSUtFADTwCScAV13gnwq2pzrqOoIVtEOY0I++ff2qt4P8MSa3dC5uVK2URzz/GfSvWYoo4IkiiUKiDCgdqzlIBVVVUKoAUDAA7Clzig1yvjPxOmjWxtbVg17KMDH8HvUpXGU/HPiv7FG2madIDcuMSOD9we3vXmo4zySSck+poZ2kkaSVi0jnLMe5orVKwhc0UlLTAWikooAWikooAWiiigAof8A1bfSlFIwyuPXigD2jwiu3wxYj/pnWwRxWb4cTy/D9inpEK0u1YsZ5z8VD89otcIeprtvik2dRtI/9jNcSetaQ2EJRRS1YBRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAooooAKKKKACiiigArU03/j3P1rLrU0z/j3P1pMDIuP+PmT/eqOpLj/AI+ZfrUdMBaKKKACikpaYBRRRQAlLRRQAUUUUAHY11XhT/jwmrlT0NdV4U/48JqlgZN7/wAfs31qKpbz/j9m+tQ0IAooooAKKKKACiikoAWkpaSmAUUUUAGaKKKACkpaSgBaQUUUALRSUUALRSUUABpD9xvpRQeQfpQB6n8PDnw6vsa6quO+G0m7RJV/uPiuxrnluM5L4kDPhvPpKK8xPWvVfiBHv8MSY7ODXlPUZrWnsAUUUVYAaSijtQAUUUUAGaWkooELRSUooAWnRRyTTLDAhklc4VQOaI45JpkhgQySucKoFeo+EfC0ejwi5ugJL1xkn+59KmUrAO8H+GV0WFri4O68mHz+i+wrpqKQkCsHqMWmuodGVhlWGCPaq1pqdlezSxWtwkjxHDhT0q3QB5V4u8LyaRK13ZqXsnOSo6oa5jtkHIr3ieGOeJopUDo4wynvXlfi3wxJo07XNqpeyc54/grWE+jA5qij36ikNagLSUdqKYBRRRQAUUUUAApaKKBC1d0UZ16wH/TSqNafhpPM8S6evpJUy2Ge0r90fSloormA8e8and4ruf8AcFYVa3imTzPEt034Vk10x2AKKKKoAooooAKKKKACiiigAooooAKKKKACiiigANFFLQISlpKWgBKKKKACloooA0dN/wBSay7n/j6etTTv9Say7n/j6ehAR0UUUwCiiigAopKKAFoopKAFpKWkoAWiiimBpeH/APkLL9K0/Ef/AB9J9Ky9A/5Cy/StPxF/x9J9Kh7gZR60lKaKYxKKKKACiiigQmaKWkoGFFGKKACiiigAooooAKKKDTAOKM0lFABS0lFAC0o549abS5oA9I+HN4JtGe3Y/PHIePauv7V5N4K1Madr6o5xHcjZ7A16uG9Olc01Zgef/EmwKS2+ooM7v3be1cQTzXtOtacmq6ZNaOBl1wpPY14zc20tldSWlwpWSI4we49a0pvSwDM0Z4puaWtQFooopgFFFFABRRRigBG+6fpXsHhH/kWrL/crx9h8p+leweEf+Rasv9ysauwG1XjXif8A5Ga//wB+vZa8b8T/APIzX/8Av1NLcDKoooroAKKKSgBaSiigAooooAKKKKYBRRRQAlTWH/IVtP8AfFQ45qfT/wDkK2n+/Uy2A9xi/wBUn+6Kg1P/AJB1x/uH+VTxf6tP90VBqf8AyDrj/cP8q5FuB4enRv8Afb+dPSOSaVIYVLyyHCqKtabo+pajIY7W1kGXPzuPl616P4Y8J2+jgXE+Jrsjlj0X6V0SmkgL3hjSBo+kR27AGZhmQ+prYNFZPiLV4tH0qW4dhvI2ovcmufVsDgvH+oC811LdDlbUbW+tcxTpJJJpnnmbMkhyxptdcVZAFFFFUAlLRRQAUUUUAFFFFACUtFFAgooooAKKKKACiiigAooopjCiiigBD0rWg/48v+A1knpWrB/x5/8AAaTAxT94/Wij+I/WiqEFFFFAwooooAWkoooELSUUUhhRRRQAtFFFMAooooAKKKKYFKiiiuQQUUUUAFFFIaACiiimAUtJS0gFpKKKACiiigAooopgTWX/AB/2/wDvV1fiH/kHxVyll/x/2/8AvV1fiH/jwiqWBznaikopgLSUUUAFFFFABRRRQAUUUlAC5pKKKAFNJRRQAUUUUAFFFFABzSGlooASiloxQBseD5fJ8WWRPAdsGvZscmvCtOm+y6paXGf9W9e6Qtvhjf8AvKDWcgQ5hlSPUV4drdubXX72AjGGz+de5V5N8QLU2/iVp8YW4UYojuM5iilNJWggoopKAClV2R1kjYpIpyrDqKbRQB6b4N8XrqKLYai4S7UYVyeH/wDr12dfPwLK6vGxV1OVYdQa9M8F+L11BF0/UnCXaDCOTw4/xqJRGdoRXn/jXwfu36ppSYccyxAdfevQKOvapTsB8/jnPGCDgg9RTq77xr4PJL6ppKfMOZYgOvuK4BTuzwQRwQeoNaJ3EOra8L+HptfvhwVs4zmR/X2FVtA0W413UFtoARCOZZOwHp9a9k0zT7fTLKO1tUCxoPTqfU0pSAls7WGztkt7dAkaDAAqU0tZPiHXLfQ7BriYgyEYjTuxrMZV8V+I4dBsSQQ11IMRR+/qa8fuLia7uZLm5cvNIcsT29qm1K/udUvnvLxy0jngH+EelVq1irCEpRRRVAFLSUtABRRRQAUUUUALS0lFACinwqXubeMfxygUytTw3am98Q2sIGdrB/yoewHstlF5FlDF/cQCp6KDWAzyj4jzCXxJEgPEcWK5Q1q+J7n7V4kvXzkJIVFZZraOwhtLRRVAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACtPTP9QfrWZWppn+oP1pMDIuP+PmT/eqOpLj/j5k/wB6mUwEpaKKACiiigAoopKAFooopgFFFFAB2NdV4V/48Jq5XtXU+Ff+PGapYGVef8fs31qGpbz/AI/ZvrUVABRSUtABSUUUAFFFFMAzRRRQAUUUZoASlzSZooAKKSigBaKSigBaKSigBaSiigYtA60lKOtAjv8A4YTDyb2AnneCK72vLPh3ceT4ha3J4lQtXqYrCe4zJ8UW/wBp0C6TGcKW/KvF4z+6X8a94u4/OtJo+u9Cv6V4XcxG3vLi3PBicirpgNzRTaWtACiiigBKKWigBKBRSZoEOp8ccksqQwIXlc4VQOTTEDySLFEheVzhVHU16j4O8KppMIvLwB72QZ5/gHt71MpWAf4Q8Kx6RELq6Ae9cZJ7J9K6mijpWLdxjScV5/4y8YHL6ZpMnPSWZT09hTvGvi3BfS9Lk+bpLKp6ewrz/GP6+9XCPVgWtM1G50m9W8s3IcHLjPEnrmvX/D+vW2uWKzwMBIB+8jPVTXi1W9K1K50i+W7s2IYH507OKqUbge6A1HcQRXMLwzIHjcYKms3QNctdcsVuLdsOB88Z6qa1c1lawHkvi3wzLodwbi3UvZSHg/3K5zP4iveLq3iuoHguIw8bjDKRXkvivwzNoVyZoQ0ljIcq39z2Nawn0YGDmlpo/SlrUBaKBS0AJS0UUAFFFFAgrofAluZ/EytjiEbq5/Ga7j4Y2uZ7q8I4K7Qaib0GeiU1jhST0Ap1UNauPsmj3U+cbENc4HjOpS+fqt1L6yEVWo3Fmd/77lvzorqWwBRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAFpKKKBC0lFLQAUUUUDNDTv8AVGsy5/4+XrT07/Umsy4/4+XoQiOiiimMKKSloEFFJS0AFFFFABRRRQAUUUUwNHQP+Qqv0rT8Rf8AH0n0rM0H/kKr9K0/EP8Ax8p9Kh7gZXekJo70UxhSUUUALSUUUAFFFFABRRRQAUUUUwCikooAKWkooAKKKKACiikoAWkopKADLAhkOHU5U+hr1nwjrSaxpKEkefCNjr34715NV7RdWm0XUVu4SSh4lQdxUTjdAe09RXM+L/DC6zD9otsJeRjg/wB72Nbmm6hb6lZpdWrhkcc4PQ+lWq51dMDwmaKW3naC5jMcqnBVh1+lJXsGueHbDWo8XCbJe0qj5q4HVPBmq6eS0IFxD228tW8aie4HPCinyW9xAcTWssZH94UzPsa0ugFpKMj3oyvc0XAKKNy+tG5fWgAb7p+leweEv+Rasv8Acrx5mXaee1ew+ETnwzZH/YrKrsBs1414oP8AxU9//v17J3rxjxTIg8UX4LgHfU0twM2jNNBDdDml59K6AFzRSUUALSUtIaYC5opKWgAooooAKKBS0AFT6eP+Jtaf79Q1Pp//ACFbT/fqZbAe3Rf6tP8AdFOYBgQwyD2psX+rT/dFPPvXGAyOJIl2xoqL6AU7IFVJ9SsYATNdwpj1auZ1fx7Y2oaOwUzzdj/DTUWwOl1PUrbTLRri7kCqBwM8t9K8k1/W59cvjPLlYV4iT0FQanql7q85mvpS392MH5VqnXRCnbVgLRSUtagFFJRQAtFFFABRRRQAlLRRQAUUUUCCiiigAoopKAFooopjCiiigAooooAQ1qwf8ef/AAGso9K1IP8Ajz/4DSYGMfvH60UH7x+tFNCCiiimMKKKKBBRRRQMKKKKQBRRRTAKWk70tABRRRQIKKKKBlKiiiuUQUUUGgApKKKYBRRRQAtFFFIAopaSgAooopgFFFFAE1l/x/2/+9XVeIf+PCKuVsv+P+3/AN6up8Qf8eEVSwOd7UUdhRTAKKKKACig0UAFJmlpKACiiigAooooAKKKKACiiigAooooAKKKKAClpKKAB87OOoINe2eG70X+hWs4OflC/lXilejfDK/8zT5tPY/NAdw/GplsCO5rhvibZGSwt71R/qG+Y/Wu5rO1+wGpaNc2h/jXI/CoW4zxE88+tJQQUZkYYZGK4+lJWogopDRQAUlLRQAlALK6ujFXU5Vh1BpaKAPTPBfi5dQRdP1Jwl2gwjno4/xrtBXgClldXRijqcqw6g16b4M8XLqCLYai4S7QYVyeHH+NRKIHZ4z1rhvE/gT7deC60lkheRsSqeBj1Fd1RUXGZmg6NbaLp6W1uoz1du7GtOiq97eQWNrJc3LhI0GSTQBFq2p22k2L3d04CKOB3Y+grxvWtXudbv2u7kkLn93H2Uf41Y8S6/Pr9+ZGJW1jOIo+31rHPrWkY2EBpMUtJVgJRS0lAAKWkooAWikpaACiiigApc0lFADh1Fdp8MrEzajPfkcRAxg1xDHCnHU8D617D4H03+z/AA9CWGJJxvf61M3oCOiqpqk/2bTLmYnGyMkflVuuU+Il/wDY/D5jRv3kzhce1ZrcZ5U8pnlknbrKxY0lGNoCjtRW4hKKKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFFAC1pab/qD9aza0tN/49z9aQGTcf8fMn+9TKfcf8fMn+9TKYCUtFFMBKWkpaAEpaKKACkpaKACiiigAPQ11Phb/AI8Za5btXUeFv+PGWpYGVef8fs31qKpbz/j9m+tRUAFFFJTAKKKKACiiikAUUlFABSUUUwCiiimAUUUUAFFGKKQBRRRQAUUUUALRSUZoAv6JeGx120uQcfMEP417aGBAI5BFeAsxADDqh3D8K9o8M339oaDaT7ssUw31rKaGavevH/GtkbLxLKcYW5+cV7B2riPiVppm0+PUI1zJCdpx6UoPUDzmlpMg8joaWthBS0lLQMKKKKBCGlRHkkWKJC8rnCqOpp0ccksqRQoZJXOFUdTXp/g/wnHpUa3l6BJeuM89E+nvUylYYzwd4STSoxe3yh71xkA9E+nvXXUU01i3cBc1wnjTxf5e/TNLkzIeJZVP3fYUvjTxb9nV9M0uQGZhiWVT90e3vXnYzkkkkk5JPc1cIdQHevOSeSfU0lApa2AT60nenYpMUCLek6nc6PfLd2jEEH507MK9f0LWrbWrFbm2YbsfOh6qa8Uq9o+rXOi3y3Vqxxn94nZhUSjcZ7fUV1bQ3du9vcRh4nGCpFVNF1e11mxW5tW7fOndT71pCsXoB5D4p8MTaFcGWEGSyc/K39z2NYOPyr3e7tYbu2eC4QPG4wQRXk3ijw3NoVyZIwZLJz8rf3fY1rCfRgYNFLSVqAZopKKAFpM0UlAhWbapb0r1rwJYmy8ORblw0rb/AMDXlmnWrX2p21moz5rgN7CvcbWIW9tFAvSNQv5VjUfQZNXK/EK8Ft4deHOGuDsFdVXmPxKvvtGqw2KnIgAc496iKuwOOAwij0FLQetFdIBRRRQAUUUUAFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigApaKKYwooopAaGn/6o1mXH/Hy9aWn/AOqNZtx/x8vQgI6KSlpiEpaKSgBaKKKACiiigAooooAKKKWmBoaD/wAhVfpWl4h/4+V+lZug/wDIVX6VpeIf+PlKh7gZXeijvRTAKSiigYUUUUAFFFFABRmkopgLRSUUALSUlLQAlLRSUALRSUUAFFFFABRRRQAlJS0UAaWg65daDdeZAS1ux/eRHp9RXquj6zZ6xbLNaSAkj5oyfmWvGKltLm4sZxNZytE45+U/e+tRKFwPchzTgPSuD0fx+hCw6tEVfoHTp+NdlYahaahF5lpOko77T0rBxaAkmtLefmaCOQ/7QqD+x9NP/LjB/wB81eoqbgUf7H03/nxg/wC+aQ6Nph62EH/fNX6T8Kd2BQ/sXS/+fCD/AL5o/sTS/wDnwg/75q/Rmi7Aof2Jpf8Az4Qf981chhjgiWKFAiL0UdBTiaM0agKayLzw3pV7M809qhkc5Zscmtfml/ChOwHJ3HgLR5s7TLH/ALtZN18OSuTZXZPp5hr0LikIFNTaA8jvfB2t2eTsSZf9jk1jT2t1bEi5tZYsd2HFe6DI6VDcWtvcjFxAko/2hVqq+oHhasG+6QfpS16lqfgjSb4lo1a3c9PL4FclqfgbVbIF7VlniHQD71aKomBzVFOnjltn2XULwMOzikx+RrRO4BS0lLTAO9FFFIBan0//AJCtr/vioKm0/wD5Clr/AL4pS2A9vi/1af7oqLUCV0+4KnBCHB/CpYv9Un+6Kg1PjTbj/cP8q5FuB4cS828zSvL87fePvQAFGFGB7U1Ojf77fzp1daSsAUUUtUAUUUUAFJS0lAC0UUUCCiiigYUUUUAFFFFABRRQKACiikoAWiiimAUUUUAFFFFACGtSD/jz/wCA1lnpWpB/x6f8BpMRjH7x+tFB+8frRTAKKKKYBRRRSAKKKKBhRRRQAUUUUwClpKWgAooooEFFFFMZSooorkEFFFJQAUUUUwCiiloAKKKKQBRRRQAUUUUwCiiigCay/wCP+D/erqPEH/HhFXL2f/H/AAf71dRr/wDx4xVLA57tRRRmmAUZpKKAFooooAKKKSgAooozQAUUUlAC0UUUCCikooGLRSUUAFFFFAC0UmaKAFzW14R1L+zfEUDs22KY7ZDWLSHOODgg5B+lJq4H0ApDAEdDyKDyKw/CGrDVtChlJHmoNrjuMVuVkM8f8c6YdN8QPIi4gueU+veudzXr3jbRv7W0R/KXNxD80Z/nXkAJPUYIOD9RWkXoIdRSUtUAUUUUAFFFLQAClVmV1dGKupyrDqDSUUwPT/Bni1dRjWw1Bwl2gwrE8SD/ABrss18/qzK6ujFHU5Vh1FdZp3xA1G0gWG6iSYKMB+5+tZuPYD1GWVIomkkcKijLMegFeTeMfEr63dm2tmK2URx/vn/CoNd8W6lrUfkMRBb91Tgt9awRgAAdBTjEAoopKsBaTNFFABSUtJQAtFJS0AJS0UUAFFFFABSUUhO0E/l70AanhvTG1fXre2A+RD5jHtxXtqKqKEQAKBgCuQ+Heimx0o3s64mufmGeqiuwrKTuAteT/EHUxfa8LeNsx2w2sP8Aar0bX9RTS9InunbBCkJ/vdq8ReV55XuJTmSVtzU4IAzmkoorUAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUtABWlpv+oP1rNrR03/UH60mBl3H/HxJ/vVHUlx/x8yf71R0wCiiimAUUlLSAKKKKYCUUtJQAtFFFAAeldP4X/48Zq5jtXT+F/8AjylpMDLu/wDj9l+tQ1Ld/wDH5L9aipAFFFFABRSUtABSUUUwCiikoAKKKKACiiigBaSiigAooooAKKKKACjNJRQAGiikoAOvXvXefDPUcfaNNkbkndGPauDq9ol+2l6zb3anA3BG+hqZK6A9vHIqvf2iXtlNbOARIhXntUsMiTRLLGco4yCKkxWIHg95ZvYX09nIMGFyoz3FRCu++I2ik7NWt0yV+WRQOvvXBcdR0raLugEpaKKoAoJwCTRmmt8yketAHo/w80OKKz/tS4QNcSHCZ/hHqK7iuR+H+qw3ejLa7gJ4Tgp3xXW5PpWEtximuF8Z+LhbB9N0uQNOwxJIDwv096d4y8XrbK2naXIGnYYkkU8IPb3rzo5JJYlmJySe5qoR7gM5JJYksTkk9zRTqSthBS0UUAFFFFACUoFFLQBe0XVrrRL0XNqx25/eR9mFev6Lq9rrFitzasOR8yd1PvXiVaGi6vdaJei5tWJX/lpH2Yf41E43Ge21BeWkN7avb3CB45Bggj9aq6NrNprNmtxauCSPmTPKn3q9JIsSGSQhUUZLE8AVgB4nrVgdK1aezBJRG+QnuKoGtXxNfJqOv3E8JzGrbVPqKy66Y7AJRS0VQCUgpcVJb20l5cxWkIJkmbaMdqTdgOv+G2l+fdzapKv7tfkjz616QKo6Np0el6ZDaRgDao3Y7t3q/XPJ3YEdxOlvbyTyHCRrkmvENSvGv9Tubtzks5APtXonxD1b7FpAs42/eXJ2nHUCvMAMAL6VpTXUB1FFFagFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigAooooGFFFFMQUtJRQMWiiikBf0//AFJrNuP+Ph60tP8A9Uazbj/j4ehCI6KKKYBRRRQAUUUUAFFFFABRRRQAtFFFMDQ0L/kKj6Vo+If+PlKzdC/5Cg+laXiD/j5T6VD3AyyeaSg9aKYwooooEFFJS0DCkoo7UAFFFFMAooooAKKSigAooooAKKKM0AFFJS0AFFFFMAooooAMUUdqKAF7YPSun+Ht4LTXJLbolwMAe9cuKs6fcmy1O2uwceU9RJXQHuFc/wCNUuD4emmtJpIpYuQUPJrcgkEsMcg6OoNMvIFubSWFxlXUiuZbgeMrq2qbR/xMbjp/epf7Y1Uf8xGf/vqqk8TW9zNA4wyOePamZrqSTQFz+2NV/wCgjcf99V2Pw71e4uZLq0vLh5WBzGXOTXA1r+E7z7D4ktpGOEYFW/GpnFWA9iri/iMLyC1t7y1uZYUj+VghxnNdqOeayvE9iNQ0K4gxkgbx+FYJ6geUrq+qp93UZ/xarcHinW4D8tzv/wB81jL935uoJFBrp5UwOrg+IGrR4E8ELAdx1rWtPiLZvhbu2lU+qjivPc0ZJ71LpoD2Cy8T6ReAbLuOMn+FzzWtHIkqhomDr6g14PsXOQoB9RV2z1XUbJg1teS4H8JbiodLsB7bxSgelec6b8QbmIqmpwB06Zj612Gl+JNL1JB5NwqOf4HODWbi0Bbv9KsdQQrd20chI+8RyK4vWPh+ybpdIm9ykn9K9BzRQpNAeGXdpc2MhjvYHhYd2HBqHtXt99p1pqERju4EkBGMkciuE13wJLbhp9IYunUxtyfwraNW+4HF0UsivFI0UqGOReqMORTc1re4BmprA/8AE1tP98VBU2n/APIVtP8AfFKWwHuMZ/dp/uiq+qH/AIltx/uH+VTx/wCqT/dFQan/AMg24/3D/KuRbgeHJ0b/AH2/nTu9MTo3++386fXYtgClpKWmAUUUUCCiiigYUUUUAFFFFAgooooGFFJS0AFFFFABRRRQAUUUUwCiiigQUUUUAIelacP/AB6f8BrMPStOD/j0/wCA0mBjn7x+tFB+8frRTAKKKKACiiigAooooAKKKKBhRRRTAKWkpaACiiigQc0UUUxlKiiiuQQUlLSUAFLSUtMAoopaQCUUUtABSUUUAFFFFMAooooAls/+P+D/AHq6jX/+PGKuYs/+P6D/AHq6fX/+PGKpYHPUtFFMBKWkooAWikooEFFFFABRRRQAUUUhoGLSUUUAFFFFABQaKSgBc0lFFMApc0lFIBc0maSigDpfAetf2ZrX2aVsQXfBJ6Ka9cByOOlfPhB6qSGByCK9d8Ea+usaUI5W/wBKtxiQe3aokgOmIGCD34NeSeOtCOlaobuFP9FuTngcIa9bqlq+mwatp8tncrlHHB9D2qYuwHhdKKtarplxpGoyWV0p3Kfkbsw9qqitQDFLSUUALSUUtMBKKKKACjOKKSgApaSloAKKKKACjNFFABRRRQAUUUUAFFFFABRRRQIStrwloj65rCKyn7NAd0jdsjtWXZ2c9/eR2dqheWQ447D1r2bw7osOh6YlrEAXIzI/941MnYZpxoqKEQBVUYAHal9qXpWP4m1iLRtIluHP7wjbGvfJ71nuM4j4j6wLu+TS4HzFFzLj+9XG0skkk80k8x3Sync596StkrIQUtJS0wCiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUtJRQAtFFJQIWkpaSgYtaOnf6g/Ws2tLTv8AUn60gMqf/j4k/wB6mU+f/j4k+tMpgFFJRTAKWkopALRRRTAKKKKACiiikAHoa6bwv/x4y1zJ6V03hj/jxloYGXd/8fkv1qKpbv8A4/JfrUNJALSUUUwCiikoAKWkooEFFFFAwooooAKSlpKACloooAKKSigAooooAKKKKBBRRRQMSgruBHrRSgUAem/D3Wftmm/YJm/fW3Cg9WX1rsa8O0nUZdK1OK9hJG04ceq969psLyG/s4rq3bMci5HtWM1ZjHXdvHd20kEwykilTXi+t6XJo+qS2cgITJMRPda9urnfGOgLrWnExAC5hG5D3PtSjKzA8jopzq6O0cqlZEOHU9jTM1uIXNJRSUwJLe4ntZhNazNFIO6nGa0p/E2tXEHkvdFV6bkPNZVJSsgFHUkkknkk9TS0lAoGLRRRTAKKKKBBRRS0AFFJRQAtJRRQBNZXt1YTebZTvE3cA8H61bvde1a/j8u4umCdwh61m0tKyABgcClpKWqAKBRSgUDDgAk9BXffDzQWRW1e7jw78RKw6D1rm/C2hSa5qSggi0hOZG7N7V6/DEkEKRRjCIMKPasakugDsUyV0ijaSRtqKMsT2qQ1xPxC137Laf2ZbN+/mH7zHZazSuwOJ8Saq2sa3Nc5/dofLQfTvWXQFwAB2pcc10pWQC5ooopgFFFFAgooozQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKWgBKWkpaBhRRRSAvWH+pNZtx/x8PWlYf6o1m3H/Hw9CEMooopgFFFFABRRRQAUUUUAFGOaKKYhaKKKBl/Qv8AkKD6Vo+IP+PlKz9C/wCQoPpWh4g/4+UqHuBlmig9aKYBRRRQMSiiigAo70UUAFFJRQAZooopgFFFHegAopKKACiiimACloopAFFJRTAKWkooAWikzRQAuaRhuUj8aKM0gPXPBl/9v8OwSMcsuVP4Vuk1558NL3bNd2LNgAAoK9BrlkrMDybxzZ/ZPEskgGI5wNv1rAr0L4mWJl0+3vVH/HufmP1rz0dAfUZrem7oBcUqsY3SRTgo4bP0pKXqCPUYq2B7bpN2t9plvcr0kQVakUPGyHowIrkfh1fGfR5LUnm2bArrya5GrMDxTXLU2OuXltjCq/y1Qrr/AIkWQh1S3u0HEinefeuPrpg7oApaSl7VYBRS0UAHNKvytuQlG/vL1pKKLAb+keLtU0whHf7RB33nLfhXe6L4q03VlCrJ5Mx/5ZyHBNeSA4NAOH3qSrjow6is5U09gPeaK8s0LxreaaVhv83Ft69XFeiaZqtnqluJrOZXGOVB5X61hKLQFHxB4ZsdaiJZBFcDlZF4JPvXl+saReaNcmG9T5M/LKB8rV7XuzVTUNPttStWt7uMOjDuOV+lVCbiB4jU1h/yFLT/AHxWv4j8L3OhymSMGayJ4YdU+tZFkf8AiZWhB43jmtr3QHuEf+rT/dFQap/yDbj/AHD/ACqxEP3Sf7oqDVP+Qbcf7h/lXMtwPDE6N/vt/On0iDhv98/zp1di2AKSlopiCiiigAooooGFFFFAgooooGFFFFABRRRQAUUUUCCiiigAooopgFFFFABRRRQAh6Vpw/8AHp/wGs01pQ/8en/AaTAxz94/Wig/eP1opgFFFFABRRRQAUUUUAFFFFAwooopiClpKWgYUUUUCCiiimMpUUUVyCCkpaSmAUtJS0gCigUUAFFFFABRRRQAlFLRTAKKSloAls/+P6D/AHq6fXv+PGOuYs/+P6D/AHq6bXf+PKKpYGD2oo7CkpgFFFFABRRRQAUUlFAhaKSigAooooGFFFFABSUUUAFFFFABRRRQAUUUUAFJS0lABV3RdVm0XVYr2EnaDiRf7wqlSUAe9WF5Df2cd1buGjkGQR61ZxXkvgjxIdHuxZXbk2Ux+Un+A16yjK6hlIKkZBHesmrDMTxT4dg16xKEBblBmOSvIbu1nsrp7W7jKTIcEHvXvdc/4p8M2+u224AR3SD5JB3+tNOwHjtFWL6zubC7a1vIykqnuPvfSq9aCCiiimAUUmaKACiiigAooooAKWkpaAEpaKKACiiigAooooAMUd6KXNABinRxyTSpDAheVzhVFEMcs8yQW6GSZzhVFepeD/CUekxi7vAJLxxnnolJysBL4O8MJolp584DXsoyzf3fYV0+KWkPArK4yOaRIYnlkbaiAlifSvG/FmuNrurFkP8AosPyxj+971vePvE/2l20nT5MRqf3zqep9K4cAAYAwKuERBSikpa0AKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAtFJRQAtJRRQAtJRRQAVpad/qD9aza0dP/wBQfrSYGXcf8fEn1plPuP8Aj4k+tR0wFpKKKYBS0UUgCiiimAUUUUgFopKKACul8M/8eUtc0e9dL4Z4spaTAzLv/j8l+tQ1Lef8fkv1qKhAFFFFMAoozSUAFFFFABRRSUALSUUUAFFFFABRRRTAKKKKQBRRRQAUUUmaAFpKKKAFopKWmIX69K63wL4h/s66/s67f/R5j8jH+FvSuRpDzjnBHQ+hqZK6Ge/A5FBrivA3icX0K6bfOBcxDCMf4xXaA1g1YZw/jnwsbpW1TTkxOo/eIB94ev1rznOc5BBBwQeoNe/EcVwPjLwcZC+paSmJOskQ/i+lXCQHn9LTeckEFWU4ZT1BpRWohaKKWmMSloooAKKKKBBS0lLQAUUlFAC9aSiigAooooASlopKACiijmmMXNXdJ0y51i/WztVPJ/eP2UUzSdMu9YvFtbFCST88mOFFeu+H9DtdDsVgt1BkPMkndj/hWc5WAsaNpVvpFglrbqAFHzH+8fWr9AqK5uIrW3eedwkaDLE1gBT1zVYdH02S7mYZAwi/3j6V4ze3c1/eSXlwxMkhzz2HpWn4n12TXdQLAkWsZxGvr71jE5reEbagJRRRWggooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUtFAwooooAKKKKAL1j/qjWbP/wAfD1pWP+qNZs//AB8PSQhlFFFMAooooAKKKKACiiigQUUd6WmMKKSigDQ0P/kJj6Vo6/8A8fK1naH/AMhMfStDXv8Aj4Woe4GX3ozRRTGFFFFABRRRQAUlFFABRRRmmAUUlFABRRRQAUUUUAFFFFABRRRTAKKKSgApaSigBaKKKAEooooA1PDF59h8R2kpOEY4evZVIIBHQjNeCklcOp+ZWBFe2aJdrfaTb3C90A/KsKq1uA3X7Eahotzan+Ncj8K8WQEAqwwVYr+Ve9MAQQe/FeN+JrMaf4guoAMITuX8aVJ62Ay6TOKM0ldAHS+AL77L4g+zs2EuFJP1r1SvCrSdrW9t7hDgq4/KvcLWZbi3jmQ5V1BFc1RWdwOd8f2P2vw87quZImBH0rysHcoIr3W+gFxZTQsMh0I/SvEJ7drW6mtmGDExBq6T6AR0tFFbAFFFFMBaSiigApKKKAEqxYX13ptwJ7GUxuDnbn5W+tQUlJq4Hqfhrxba6uiwXJEN2OCD0f6V04PY14MMhg6MUdejDqK7zwn4xLFLDV3AbpHN2/GuedO2wHdTQxzRNFKivGwwVI4Neea74Pl0/Uob3TFMltvBaPuvNejqQQCDkHkGlIB4IrNNoBkP+qT/AHRVfVP+Qbcf7hq3VTVP+QZcf7hoW4Hh69G/3z/OlpF6N/vml712LYAooopiCiiigYUUUUAFFFFAg70Ud6KBhRRRQAUUUUCCiiigAooopgFFFFABRRRQAUUUUgA9K0of+PT/AIDWaelaMP8Ax6f8BoAyD94/Wig/eP1opgFFFFAwooooEFFFFABRRRTGFFFFABS0lLQIKKKKACiiimMpUUUVyCEopaKYBRRRSEFFFFAwopaSgANFFFABRRRTAKKKKAJbT/j+g/3q6bXf+PKOuZtP+P6D/erpdd/48o6lgYPaijtRTAKKKKACikooAKKKKACiiigAozRRQAlFFFABRRRQAUUUUAFFFFABRRRQAUUUlAC0lFFAAQCMHpXd+B/FnkFdK1ST5OkMrdvY1wtB5/ofShq4H0CCCOKWvNvB/jM2xTTtXfMfSKY9vY16OjrIgdGDKwyCO9ZNWGZPiHw9Z67amOdQsoHySDqDXk2taNe6JcmG+QlM/LMB8rV7jVa/sba/t2gu4lkjYY5HT6U1KwHgxpK7DxD4CurItcaQTNB18nqwrjnDRymKVGjkHVGHStE7iClpKKYC0UUtACUtFFACUtJRQAtFFFABRRS0AFGKMUo5cIoLOeigdaAEqxp2n3eq3QtrCJpHJwzY4X610eg+Br3Uts2o5trf/nmfvNXpGl6TZ6VbLBZQhABy2OT9TUOQWMrwv4UtdChEjgS3jD55D2+ldHSUE4HXpWYxa4bxv4tW1R9M02QNcsMSODwg9Ki8YeNVhD6fpDhpjw8o6LXneSWLMxZ2OWY9TVxiIMYySSSeST1NFFFagFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKAClpKKACiiigAooooAK0dP8A9SfrWdWjp/8AqT9aQGXP/wAfEn1qOpJ/+PiT60ymAUlLRQAUUlLQAUUUUwCiiikAUtJRQAdjXS+Gv+POSua7V0nhv/jyloYGZd/8fkv1qKpbz/j8l+tQ0gCiiimAUUUmaAFpKKSgBc0ZpKKAFooooAKKKSgBaKKSgBaKSjNAgooooAKKKKBhRRRTAKKKKAAmkpaSgB0ckkMyTQuUlQ5VhXq3hHxPHrVsIJ2CXsYwyn+P3FeT1JbzTW1wlxbSGOZDlWB/nUSjcD3oGjGa5fwp4rg1iJbe6IivVGCD/H7iupFYvQZyHinwZBqe67sAIbwDkD7r/wD1681urWezuGt7uJopl7Edfeveqy9b0Gx1q3Md3EN3aReGH41UZ2A8Vorf1zwnqWjsWVDc2w5DoPuj3rBBDZ2nOOtbJpgFJS0lMQUUUUAFLSUUAFFFFAC4ooooAKKKKACkNO5PSnW8M11MIbSFppT/AAqKNhkJ4BLHArb8PeGb3XZQQphtB96Rh94e1dJ4e8A/Ml1rTbiOVhXt9a7yKGOCIRwoqIvRVGBWcp9gKWkaRZ6PaC3s4wo/ibu1aHSiop5ooImmmkCRqMsx7VkA6SVIo2kkcIijLMegFeWeMfFDaxcGzs2K2UZ5I/5aH/Cl8XeLJNWdrKwYx2Sn5mHWT/61csAAMAYFawh1YDs9hwBRSUVqAtFFFAgooooAKWkooAKKWkoAKKKKACiiigAooooAKKKKACiiimAUUUUALRRRQMSloooAKKSloAu2X+qNZ0//AB8PWjZf6o1nT/8AHw9JCGUUUUwCiiigAooopgFFFFABRRRQAUUUtAF7RP8AkJj6Voa9/wAfC1n6J/yEh9Kv67/x8LUvcZm0Ud6SgBaKSloAKKSigAooooAKSiimAUUUUAFFFFMAooopAFFJRTAWkoNFABRRRQAUUUUAFLSUUALRRRQAoHUe1ej/AA4vfN0l7InLW5z+decCuk8BXv2TxCIScLcjH5VnUV0B6rXnfxLswk1peoOuQ5r0Sud8b2X23w5cBVzImCtYRdmB5NRTVOVB9OKdXWAjDKn6V6z4GvRd+HYVJy8I2tXlArsvhtemK/uLFjxL84/Csqq0A9IryfxzY/Y/ETSAYW5+evWK4j4mWoawhvB1jO386ypu0gPPKKU0ldQBSUUUwCiiigA5ooooAKKKXpQAUuARz/8AqpKWgDs/CHixrZk0/VJMxHiKU/w/WvRVYMoZSCCMgjvXg5wRg9K6vw34yl0yMWmoK00A+4w6r9awnT6oD06sjxRdrZaDdTOcfLge9UpPG2jJB5gl3HH3B1rhfE3iSbXpVRVMVoh+VD1P1rOMG2BhLwv1JNLSUV1ALRRRTAKKKKACiiigAooooAPpRRRQAUUUUCCiiigYUUUUCCiiimAUUUUAFFFFABRRRQMDWjD/AMev/AazjWhD/wAev/AaQGSfvH60UH7x+tFMQUUUUDCiiigAo5oooAKKKKYgooooAKWkpaACiiigAooopjKVFFFcghKWkpaYBRRRSELSUUUDFpKKKACiiigAooopgFFFFAEtp/x/Qf71dLrn/HlHXNWn/H7B/vV0uuf8eUdSwMGik7CimAtJRRQAUUUUAFFFJQAtFJRQAUUUUAFFFFABRRRQAUUUUAFFFGaACikooAKKKKYBRRRQAtFJRQAvBGCMium8MeMLnRnFvdlp7Mn6stcxSUmrge9WGoWuo2y3FnMssZ7g9KsZrwrStXvtHuBNYTFR3jP3T+Fel+H/ABtp+qBYbo/ZbnoQ/Rj7Vm42GdVWRrHhvTNXjK3Vuqt/fQYNawIIBBBB6EU6lewHler/AA/1Cz3SadILiLtGB81crc21zZuUvLd4GHZhXv2PSoLmxtbpStzbxyZ7soJqlIDwVfm6HNLivVb/AMA6NdFniV4nPoeKwbr4cXiZNpfRkdlIqlJCOHoropvBOuxE7YfN/wB2qj+GddTg6ZKad0Bj0VpHw/rYP/ILmpyeGtefppkoouBl0V0EHgrXpiN0Plf71atr8N758G6vowO4A5pcyA4o8ck4qS3ilunCWsLzP6KK9PsPh9o1thphJK49W4rpbTT7K0ULb20SY7hRmk5hY8y0nwJqt9h7wi0i7qw5Nd5ovhbS9HX9zAJJO7yc/lW5RUOTYwxRRmub13xjpukqURxcXA48tD0PvSSA3ru6gs4GnuZVjiUcsxrzLxT42m1DfZaUTFbdGk7t9Kw9a1zUNbmL3cpEY+7GvA/EVmYAGAMCtIx7iGgYz3J6k9TS0tFaAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFBooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAKK0NP/wBSfrWcK0LD/Un60gMyf/j4k+tR1JP/AMfEn1qOmAUtJS0AJS0UUAFFFFABRRRTAWkoooAOxrpPDf8Ax5yVzfauj8Of8eclJgZt3/x+S/WoalvP+PyX61FQAUlFFABRRRQAUUUUAFFJmjNAC0lFFABS0lFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAKMUUUAORmjkWSN2SRTlWU4Ir0Dwx44V9lnrBCuOFm7H6159RweDyKmUUwPfEdZEV0YMrDII706vHdB8U6horhd5nts8xscn8K9K0TxHp+sxL9nlCy94mPzCsXFoZrsqupVgCp6gjg1zOteCtM1MtLEn2efsU4X8q6eiknYDyLVPB2s6cSUj+1xjvGOlc9JmJ9kytG46qRXv2Kz77R9PvkK3NrGc9woBq1UYHiHXpRXpV98O9Mmy1o8kTn1ORWJc/DvUo8m3vI3HpirU0I5CjNbk3g/XoulsZPcVWbw5rqHnTJTVcyAzM0ZrRHh7XCf+QZLUyeFtff/AJh8i/WjmQGTmlHPSujt/AutzEb2WH/eFbFp8OGyDfXocdwgxS50BwZIX7xxVqy06/v2AsrOSUeoFep6f4N0WxYMkBkYf3zkVuxQQwrthiSMf7IxUOp2GedaV8PbmfbJqs4WM9Y1+9Xc6Xo1hpUIjs4FXH8RGW/OtDFFQ5NgFIabLLHDGZJXCIoyWJ6VxniDx5Bb7rfSQJpenm9VFJJsDpdX1ey0i2M15KF4+VM8t9K8r8ReJrzXZSuWhtAfljHBP1rNvbu5v52nvZWlkJzgngfSoK2jC24DcdgMCloorQAooooAUUUUUCClpKWgBKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFMQUUUUDFooooASloooAKKKKBF2y/wBUazp/+Ph60LL/AFVZ8/8Ar3pIBlFFFMAooooAKKKKACiiimAUUUUALRSUtAF3Rf8AkJD6Voa7/wAfC1Q0X/kJD6Ve1z/j4Woe4zOopKKYBS0lFABmkpaSmAtFJRQAtJRRQAUUUUAFFFJQIKKKKBhRRS0AJRS0lABRS0lMAooooAKKKKAFopKWgQVLa3LWl5BcqfmjcVDSHkYpNaDPdbWZbi2ilU5DIDmlmjEsLxsMh1Irlfh/q63mlfYpG/f2/XPcV1orlaswPD9WszpWq3FlPldrEqcdc1WDx/3v0r2+90qxvyDdW0bsP4iOapnwvo5/5dFrRVbIDx3en979K6n4e2stxr/2uMHyYUKs2OM12/8Awiujf8+orTtLO3sYfKtYljT0UdaUql1YCxXI/EiVV8O7CfmaQYFdYTXl3xA1Vb7VUs4WzHbjD+5qYK7A5c0UlFdQC0UUUwCiiigAooooAKKKKAFopKWgAzRnikooAML/AHV/KjNFFIAooopgFLSd6KAFooooAKKKKACikpaAEpaSloAKKKKBBRRRQAUUUUAFFFFMAooooAKKKKACjpRRQMD0rQi/49f+A1nnpWhD/wAe3/AaTAyT94/Wig/eP1opgFFFFABRRRQAUUUUIAooopgFFFFAC0UUUAFFFFAgooopjKVFFFcogooopAFFFFABRRRQAUUUUwCiiigAooooAKKKKAJbT/j9g/3q6PXP+PKOuctP+P2H/ero9b/48o6kDCoo7UUwCiiigAoopDQAtJRRQAUUUUAJS0UUAFFFFABRRRQAUUlFABRRRQAUUUUAFFFFMAooooAKKKKACiiigBMUhGTnkHsw6inUlAG5ovizVtHwiy+fbjqj8tXf6N430rUgEmf7LMf4JO9eSgUYBOcc+vepcbhc+gUdZEDIwZT0INOFeHadrmqaY4azu2wO0hyK6nT/AIjzR4TUbMyHu0fFQ4sZ6RRXNWXjfRLoAPciBj/C9bVvqdjcjMF1G+fQ0rAW6KQMD0IP40Z+lIBaTn1paKAEoxQTgckfnVa41Gzthme5jQD1NFgLOKTOK5298a6Hag7btZmH8K1zl/8AEksCunWTKezvyKaiwPRC4VSzEADqSa57V/Gmk6apVZRcyj/lnGea8z1LxDq+qE/a7oqvpGcVlgDOcZPqetWoCOj1rxjqmrbo0f7Pbnoq8N+dc/33ElmP8R60maKtJIBc0lFFMAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUtFACUUUtACCr9j/qT9ao1esf9UfrQBmT/wDHw/1plST/APHw/wBajoAWikpaACiiimAUGiigAooooAKKKKADtXR+HP8AjzkrnO1dF4d/485KTAzbv/j8l+tQ1Nd/8fcv1qGhAFFJRQAtFJRQAtIaKKACiijpQAUUUUAFFFFABRRRQAUUUZoAKKKSgBaKKKYBRRRQAUUUUAFFFFABRRRQAUUUUALnFOjd4nEkLtG4/iQ4plLRYDrdG8eXtkFi1GP7TEOAV6iu50vxJpeqKv2e6QSH/lmx5FeMZoHynKFkPqhwazdNDPfs8Uma8a07xRrOm4ENyJIx/DJya6ew+I0ZAS/snDd3XpUODA7+jHpWDZ+LtEuwMXqIx/hateC8tZ1zDOjj2NTYCfB9aMH1NGfp+dLSAbz60YPqadSH8KADn1oxSFgBkkD8ap3Oq2FqD9ou40x6miwF2lrlr7xzo1rnyZftDDslc5f/ABEvJsrp1t5A9ZBmqUGwPSJpooIzJNIqIOpJrldZ8eadZZjsgbuXplOgNedX2qahqLl726dieynAqoMD7oAq1T7gaer69qesSZu5ysfZIzjj3rMGAPlAAoorVJIAooopgFJRRQAUtJSigAooooAKKKKACiiigQUUUUDCiiigAooooAKKKKACiiigAoFFFMQtJS0UDCiiigQUUUUAFFFFAFyz/wBUaz5/9e1X7P8A1RqhP/r3pIBlFFFMBKWiigAooopgFFFFABRSUtAC0UlFAF7Rv+QkPpV/XP8AXrVDRv8AkIj6Vf1z/XrUvcDNzRQaKBhRSUtACUtJRTAKKKKACiiigApKWigApKWkoAKKKKACiiimAUUmaKAFopKBQAtFHaigAopaSgAooooAKSlooAsaffXGmXqXdo22RDyOzD3r1DQfFun6rEqySCC46GNz1NeUUDg5BKn+8OtRKCkB7yDkZBGPrS14lBq+p2y7YbyTH+02an/4SPWu15WXsmB7LTZHVELOwVR1JPSvGz4i1o/8vlVrnU9Quxi4u5CPRWxR7Jgdx4o8ZwwQvZ6S4lncYaUdEFedncWLOxZmOWJ6mlAA6D6n1oraMFEBKWiirAKKKWgBKKKKACiiigAooooAKKKKACiiigAooooAKKKWgBKKKKACiiloASilooASiiigBaKSloAKKKKBBRRRQAUUUUwCiiigAooooAKKKKBhRRRQAGr8X/Ht/wABqhV6L/j3/CkwMs/eNFHc/WimAUUUUAFLSUtACUUUUIAooopgFFFFAhaKKKBhRRRQAUUUUwKVFFFcogooopAFFFAoAKKKKACiiimAUUUUAFFFFABRRRQBJa/8fkP+9XSa2f8AQo65y1/4/Yf96ui1v/jzjqWBh9qKO1FMAopKKACkpaKAEpaKKACiiigBKWikoAWikooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRk+tFFACYU/wAK5+lKrOpyk0qf7rYpKKQFqLUdQh/1V7MMermraeItbT7t6fxrLoosBsjxRr3/AD+Cmt4k1tx817+VZNFFkBcl1PUZv9bey/g2KqO0j8vPM/8AvNmkzRTsA3ao/hX64pfpS0lABRRS0wEpaKKACiiigAooooAKKKKACiiigAooooAKKKKAClpKKACiiigAooopgFFFFABRRS0AJRRS0AFFFFACUUtFACUooooAKvWP+qP1qjV2y/1R+tIDOn/17/WmU+f/AF7/AFplMAoFFFABRRRTAKKKKACiiigAooooAO1dD4d/485K57tXQ+Hv+POSkwM27/4+5frUNS3f/H3L9aioAKM0UUAFFFFAwopKKBBRRRQAtFFFABRRRQAUUlFAC0lFFMAopaKACiiigAooooAKKKKACiiigAoopaAEooooAKKKKBhRRRQIKPrzRRQABV/ugfQVIkkifcuJl+j0yiiwFyPU9Ri/1d7L+LVZTxHrSfdvPzrKopcqA2D4o13H/H4tRP4h1mQfNen8KzKKOVATy39/N/rL2b8HNV2LMfnlkf8A3mzRRjinZDECr2VR+FL+NL1pKACjNFFMApaSigQtJS0lAC0UlFAC0UUUDCiiigQUUUUDCigUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUtFMQUUUUDCiiigQUUlFABS0UlAy5Z/wCrNUJv9e1X7P8A1RqhN/r2oER0tFFABRRRQAUUUGgAopaKYCUUUtACUUUUDL2j/wDIRH0q/rn+vWqGj/8AIRH0q9rn+uWpe4Gd3pKXvSUAGaKKSgBaKSimAUtJS0AJRS0lABRS0lABRRRQAUlLRTAKSlpKQBRS0UAFFFAoAKKKKACiiimAUUCigAopaKAEpaKSmAUUtFACYopaKAEooooAKBS0UAJS0UUAFJS0UCEopaKBhSUtJQAUUUtACUUUtABSUtFACUUUtACUtJS0AFJS0UAJS0UUAFFJS0AFFJRQAtFFJQAtFFFAgooopjCiiigQUUUUDCiiigAoopO9ABV+L/j2/CqBq/F/x7fhSYGX3NFB6mimAUUUUALSUUtACUtJRTAKKKKBBRRRQMWikpaACiiigAooopgUqKKK5BBRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAktf+P2H/AHq6LW/+POOuetf+P2H/AHq6HW/+POOpYGHRR2pKYBS0lLQAUUUUAFFJRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAooooAKKKKACiiigAooooABRRRQAUUUUAFFFFAAKWkpaACiiigAooooAKKKKAEpaKKYBRRSUALRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFLRQAmKWiigAooopgFFFFABRRRQAUUUUAFFFFABV2y/wBUfrVKrtn/AKo/WkBnTf69/rUdPm/17/Wm1QBRRRSAKKKKYBRRRQAUUUUAFFFFAB2roPD/APx5yVz9dD4f/wCPOSkwMy7/AOPuT61FUt3/AMfcv1qKhAJRS0lAC5pKKKACiiigApaSigBaKM0lAC0lFFABRRRQAZooopgFLRRQAUUUdqACiiigAooooAKKKKAFopKKACiiigYUCiloEFFJS0AJS0lLQAlLSUtABRRRQAUUUUwEpaSloGFFFFIQUUUUDE+lFLRTASloooEFFFFAwo7UUCgApaSigAooooEFFFFAwooooAKKKWgApKWigBKKKKACiiloAKKKKYgoopKAFooooAKSlpKACiiloGW7T/V1nzf69qv2h/d1Rm/1zUkIjpaKKYBRRRQAUUUUwFooooAKKKKBid6WkooAvaP/AMhEfSr2uf65KoaP/wAhEfSr+t/69al7iM00UGigYlFLSUwCiijFABRRRQAUUUUAFJS0UAFFFFABRRRQAUUlLQAUUUUAFFFFABRRRTAKKKKACilooASlpKWmISilooGJRS0UCCiiigYlLRRQAUUUUCCikooAWiiigAooooAKKKKBiUUUUAFFFLQAlFFLQAlFFLQAlLRRQAlLSUtABRRRQAUUUUAFFFJQAtFFFABRRRQAUUUUwCiiigQUUUUDCiiigAooooAQ1ei/49/wqlV2P/j3/ChgZh6mig9TRQAUUUUAFFApaYCUUUUAFFFFABRS0lAC0UUUAFFFFAgooopjKVFFFcggooooAKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAS2v/H7D/vV0Ot/8ecdc9a/8fkP+9XQa1/x5x1IGHS0nalpgJRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigBaKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUCgBaKSigANLSUtABSUtJQAUUUUAFAoooAWiiigAooooAKKKKYBRRRQAUUUUAFFFFABRRRQAVctP8AVH61Tq5af6s/WkBnTf65/rTKfN/rn+tMqgFopKWgAooooAKKKKACiiigAooooAO1dB4f/wCPR65+ug0D/jzkpMDNu/8Aj8l+tQ1Ndf8AH3L9ahoAKKKKACiiigAooooAWkoooAWkoooAKKKKACiiigAoFFFAC0UUUwCiijFABRRRSAKKKKAClpKKACiig0wCiiigYUtJS0CEooooAKWiigBKWiigAoopKBhRS4opiEpaKKACiiikAUUUUAFJS0UwEopaKACikpaAE6UtFFAwopaSgQUUtFAxKKWk70AFFGKWgAooooEFFFFACUUtJQAtFFFABRRRQAnejvS0UAFFFFMAoopKBhS0UUAWrX/V1Qm/1zVftf8AV1Qm/wBc1IQ2iijvTAKKKKAA0UUUALRRRTAKKKSgBaKKSgZe0j/kID6Vd1r/AFy1R0j/AJCA+lXta/1y1L3EZx60UUUDEopaSgAoozSUwFoopKYAKWiikAUUUUAFFFFABRRRQAUUUUAFFFJQAtJRRTAKUUlLQAUCiloASloooAKKKKYBRRRQAlLRRQIKKKKACiiigAooooAKKKKACiiigAooooASjtRRQMKKWigBKKKKACilpKAAUUvakoAWiiigAooooASloooAKKKSgBaKKKACiiigBKWkpaACiiimAUUUUCCiiigAooooAKKKKACrsf8AqPwqlVyP/UfhQwM09TRQepooGFFFFAC0lFFMAooooAKKWigAooooAKKKKACiiigAopaKYFGiiiuQQUUUUAFFFFMAooooAKKKKACiiigAooooAKKKKAJbX/j8h/3q6DWv+POOuftf+PyH/erf1r/jzjqWBi0lHYUUwCiiigAooooEFFFFABRRRQAUUUUAFFFFAwooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFJRQAtFFFMAooooAKKAKMUAFFGKKACiiigQUtJRQAUUUUDCiiigAooooAKKKWgAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFW7T/AFZ+tVKt2v8Aqz9aQFCb/XP9aZTpv9c/1plMBaSlopgFFFJQAtFFFAB2ooopABooooAK39A/49JKwDW/oH/Hm9DAzbr/AI+5frUVTXX/AB9y/WoaACiiigAooooAKKKKACiiigAooooAKKKKACiiigApaSigBaKKKACiiigQUUUUAFFFFMYUvakFLSASiijvTAKKWigBKKKKAFopM0ZoAWkozRmgAoo7UUAFLSUCgBaKKKACilooASloxRQAlFLSUwCiikoAKKKKBhRRRQAtFFFABRiiloAKKKKBBRRRigAooooAKKKKACiiigAooooAKKKKACiiigAoopKYxaSlpKAFpKWigAooooAs23+rqjN/rmq/bf6uqE3+uakhDaKKWmAUUUYpgJS0UUAFFFFABQaKKACikpaALmkH/iYD6Vd1r/XLVLSf+QgPpV3Wv9ctSwM+ijvSGgYUUUUwCkoooAKKWigAooooAKKKKACiiigAooooAKKKSgAopaSmAUUUUAFLRRQAUtJS0AFFFFABRRRQAUUlLTEFFJRQAtJRS0AFFFFABRRRQAUUUUAFFFFABRSUUDClpKWgBKWiigAooooASiiigBaKSigBaSlooASlpKWgAooooAKKKKACiiigBKWiigAooopgFFFFAgooooAKKKKACiiigAooooAKtx/6j8KqVbj/ANR+FJgZx6mig/eNFMYUUUUwCiiigApaSigAxRRRQAtFFFABRRRQAUUUUAFFFFMClRRRXIIKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFAEtr/AMfkP+9W/rX/AB6R1gWv/H5D/vVvaz/x6R0mBi9qKO1FABRRRQAUUUUAFFFFABRRRQAUUUUAFFJu5xsc/Rc0oDn7sMx/4AaADNJmpks7yXiO0lP1U1ch8O65cY8rTzz6nFFwM6iujt/AuvSkebEsQ/3q1rX4azNg3WolfVQKXMgOFLAdx+dIDnpXq9j4B0W2wZ4zOw7k1Q8YeDIpbT7XpEXlyxD5ox/EKXMB5xmgU0kgkOCrKcMD2NTR21zIu6O2kZfUKaq4DKKmFlff8+kv/fJpwsL4/wDLpL/3yaLgV6Ks/YL7/n0k/wC+TUckE8LBZoWRm6AjrRcCOkqz9ivD0tJf++TR9gvf+fSX/vk0XArUVZ+wX3/PpL/3yaPsF9/z6Sf98mi4FakpzKVJVhgjqKSmAlGaKsWFhealcCCwhMsh6+gpAV8jucUoIPQ13+kfDpcLLq1xuPeEdvxq34o8HWKaG8mlweVNAN2c9QKXMB5tRSK25QcY9qWqAKMUoyWVR1Y4FbieENfkRXSzBVhkHd2pXsBhUVvf8Ib4i/58R/31Sf8ACG+Iv+fEf99UcyAwaK3T4O8Q/wDPiP8Avqk/4Q7xF/z5D/vqjmQGHRW5/wAIb4ix/wAeI/76o/4Q3xF/z4j/AL6o5kBh0Vt/8Id4i/58h/31S/8ACHeIf+fEf99U+ZAYlFbg8G+IT/y5D/vqsq8tJ7C7e1u02TJ95fShMCGiipba2nvJxDax+ZKei0xEVJWn/wAI/rX/AD4Gj/hH9a/58DSuhmZRWn/wj+tf8+Bo/wCEf1v/AKB5ougMykrTPh/W/wDoHmk/4R/W/wDoHmi6AzqStE+H9a/58DVS8sruwdUvYfKZuQPWi4EWaKltLO7vnZLKAysvJHpVsaDrJ/5cGouBn0taI0DWv+fA0v8Awj+tf9A80XQGbRWl/wAI/rX/AEDzR/wj+tf9A80XQGbRmtE6BrX/AD4Gk/sHWf8Anwai6Az80VNdWd1YyCO8h8pz0HrUFNALRRRQAUUUUwCiiigAooooAKKKKACrVr/q/wAaq1atfuH60AUJf9c/1ptOm/1z/Wm0AFJRS0wCkopaACiijFABRRRSAKKWkoAK39B/49HrA7Vv6D/x6PSYGbdH/S5PrUVS3X/H1J9aipoAoopKAFoFFFABRRRQAUUUUAFFFFABRRRQAUUHHcgfjSZH95fzoAWikyP7y/nS5H95fzouAtFJkf3l/OjI/vL+dAC0UZH95fzo4/vL+dABRRx/eX86Tj+8v50ALRRx/eX86OP7y/nQAUUmR/eX86XI/vL+dABS0ZH95fzo4/vL+dMAoo+X+8v50v0OaAEpKWkJH91z9FoAKSnBZGPyQzH/AIAamjsL+X/V2ch+qmldAV6K04vDmuzH93p/5mr8HgfXpf8AWQrF/wACpcyA52kyB/EPzruLX4bzvg3WoFfVQK3LLwDo1vhp4zOw7k4pOaA8sDA9KUV6F4v8HRNa/bNIi2SRD54x/GK87HfPBBwQexpxlcB4oqRLW5dQy28hU9CFNO+yXf8Az6y/98mndAQ0tTfY7v8A59Zf++TS/Yrz/n1l/wC+TRdAQUVIYJhIIzC4kPRcc1J9gvv+fST/AL5NF0BXpKnNleDraS/98mmG1ux/y6y/98mndARUVIbe6AybaUAdflNRZp3AWikooAWijmkzjqQPqaBi0tM3qO+fpzUsEFxcsFt7aVyfVSKVxDRR9SBW/Y+DNcu8F4BAh/iJrptO+HdrEytqFybj1XGKlzSGedjB6HNLiuw8Y+El02MX2lxn7OOJIxzt9644kFcg01K4Bkf3gPxoyv8AeH516D4Z8KaRqWhQXVzAWlfOWzWmfAehn/liR+NT7RAeV5H94fnRXqZ8B6GoJMJOAe9eZahGlvqVxBGMIjYUU4zTEQ0lNLAdfyHWtEaHqx07+0PsjfZx37/lVXQFClpo56U4UxhRRRQIKKKKACiiigYlLRRTEJS0UUDCiiigQUUUUDLNv/q6pTf65qu2/wDq6oy/65qSENopaKYCUtFFMAooooAKKKSgAopaDQAUlFFAy7pH/IQH0q7rP+uWqWkf8hAD2q5rP+uWpe4GeTSUtJTAKKKKAClpKKAFooo60AFFFFACUtFFABRRRQAUUUUAJSmkpRQAlFFFMAooooAWlpKKAFooooAKK1/D+gT6886wy+UIhndjr7VLfeE9asgWNsJIx/EDzU86vYDDNJSyK8TlZY3Qj1XFMyOxBp3AdRWpovh691uF5bRgqp1zWgfAmtDoV/OlzpAc1S10n/CCa16r+dKPAmtf3l/Ol7RAc2KMV0w8B6x3dfzrF1TTLjSbz7Ndff7H1pqaewFPFGKWkqgCiijNABRSUtMAoooyByegoAaWA65oBHqPzrvfAmgW13pElxqEAk8xiFB9K0b3wDpU3Nrut29uaxdVXsB5kKWuuvfh/fw5NpcicdgeK5m9s7jT7pra7QJMvJANXGaewFek3Y/hf/vmkkPyfiK9j0vSrCXS7Z2tUJMYycUpz5QPHdw/uv8A980m4f3X/wC+a9s/sXTf+fSP8qin0bTVgkItE4Q9vao9qB4zmikfiaYDoJGA/OgVsmAtFFFMAoopaACiiigAooooAKKKKACiikoAKWkpaACiiimIKKKKACiiigAooooAKKKKACiiigAq1H/qfwqrVpP9T+FDGZ/c0UHqaKACiiimAUUUUAFFFFABRRS0AJS0UUAFFFAoAKWikoAWikooApUUUVyiCig0UwCiiigAooooAKKKKACiiigAooooAKSlooAltf8Aj8h/3q3dZ/49I6wrX/j8h/3q3tZ/49Y6QGL2ooooAKKKKACiiigAooooAKKKKACgjII6ZoooA9I+HrWOoaY0UttGZ4G25I+8PWu1W1tlHywRD/gAryHwVq8Wka6HuX228i7SfevVodVsp2VYZw+7pjms5bjLYiQdEQf8Bp2AOgA+lZmra/pujso1CcxbhkcZzVbRPFWna5fSW1gzMY13EkYqQNvHuaKd1FU9VumsdMubtU3mFCwX1oAtUh+oH1rzO08eazrF2llp9gqSyHG4HO33rZ1rQPEstqJbPWWecD5oiMA/jTsBnfEDwxE0T6rYGNWXmdA3UetdB4EdZ/Cts5VDyRyteTahLqqSvb6jJMJAcMGzg16t8OgR4Sg/32pvYR0wRT/An/fNL5a/3F/Kl/A0uakYmxf7q/lXAePAB4m0YBV+/wBhXoNef+Ov+Rr0Uf7VNAd4iLsX5E6D+GneWv8AcX8qE+6v0FOpAN8tf7q/lTJUXyn+RfuntUtMl/1T/wC6aAPCdT/5C13/AL9Vat6p/wAhi8/36qfU4rZCENdJ4C1YaZroglIEFxxk/wB6ubOPUfnTS5jZZVYbom3jnuKGroD6DpHVXQo4yrDBHtWb4c1Aanodrc7gzlBvwehrTPTrishnifijTX0nxBPARiOUl4/pWVmu8+KP2OSK3mjnja7Vtu1Wz8tcFkHuPzrSL0ESRnE0P++K9303/kHWx/6Zj+VeCKf30XIPzCvedMP/ABLLX/rkP5VMwLlJTScAn0FcfY+PLe710aWLUqxlMe7NQM7KikrH8Sa/F4fs47iaMuHbbigDZoxWN4Z8QReILNriGIxhTjBraoATHvRXI6346tNI1aSwlgLOmOfrXVwSiaCOUDAdQw/GgB1eN+OD/wAVdd/hXsprxrxz/wAjdefhVw3EzCBrpfh8f+KsiH/TNq5muj+Hx/4q6H/carlsB7Bgeg/KjA9B+VApaxGJgeg/KjA9BWNrPijStFuUgv5mSRxlQFzxWf8A8LB8Pf8APy3/AHzTsB1OB6D8qMD0H5Vzdp440K7uo7eG4YySHCjb3rpaQDSBg8D8q8x+KPGrWmP+edennpXmXxSH/E1tP+udVHcBvwt51S6/3K9OCjA4H5V5n8LR/wATS7/3K9OHSiW4CYHoPyowPQflRXL3XjvQrW6kt5p3EkbFWG3vUoDqMD0H5UYHoPyrk/8AhYXh/wD5+H/75q3pnjHR9UvVtLSZmlboCuKdgOhwPQflRgeg/KgUUgPLficf+J1bj2rkK674nf8AIct/oK5LvW0dhBRRRVgFFFFABRRRQAUUUUAFFFFABVq2/wBX+NVas233D9aQFGX/AFz/AFptOl/1r/Wm0wCiiimAUUCikAUUUUAFFFFABRRRQAdq3tB/49HrCNbuhf8AHo9DAzbr/j7k+tR1Ldf8fUn1qGhALSUUUwFooopAFFFFABRRRQAUUUUAFFFJ3oA7DwV4b0/XNOee9QmQNjOa6T/hX2if3D+dU/hd/wAgeX/fNdwKxk3cZyX/AAr7RP7jfnR/wr3Rf7rfnXXUVPMwOR/4V9ov91vzo/4V9ov91vzrrqKOZgcj/wAK+0X+4350v/CvtF/uN+ddbRRzMDkv+Ff6L/cb86P+Ff6J/wA82/Outoo5mByX/Cv9E/55t+dH/Cv9E/55t+ddbRRzMDkv+Ff6J/zzb86X/hANE/55n866yijmYHKf8IBon/PJvzo/4QHRP+eTfnXV0UczA5T/AIQDRM/6s/nXAeJLGDTNbezthiIDIFe014943P8AxVUn+7VwbuBhkbhtzjNeleBGsdT0kebax/aITtOR1HrXmma6LwRrMGj6rKbxykEqYz6GrnsI9XS1t0HyQRj/AICKfsUdET8qo6frVjqLlLSRnI7lSBUWq+ItM0l9l/OY2xx8vWsBmoB6AD6UY9z+dc/oHiyz17Up7WzQ7Yl3bz3rosUANoqlrd4+naRcXcab2iXIX1rg9K8Z63rOpRWVtbLGZD8zZ+6KaVwPSPw/OvOPHfhlI2bVNPMYXcDNHu6e9bevaZ4p8ktpmqeYccxlcV5pqB1SOdk1FpxJuG7Odpqogey+HVVtDtTtT7n92tLy0/uL+VZvhsFdBtAQc7K1c/Wpe4DfLT+4v5Unlr/cX8qfSUgOB1NgnxLs/lX7hGMcV3mxf7q/lXnuusV+JNjjuteh9+hqmAmxf7i/lSeWn/PNP++adS1IFLU0UaZdEIn+qb+H2rwv+J/9417xqf8AyDLr/rk38jXhGPnf/eNaUwJIIJrmXyrdPMk/u1sWvhHXrkArZhVPctis7TL2TTtSgvIzjYw3+617fZ3CXVpFcR/dlUMMVU5NAeb2vw71CXH2m7EPqBzWxafDmwjIN1cNP+ldvRWfOwMG08JaHaEGKyG4dyc1sRW8MKhYoo1A9FqakpXYCYp1NpGdVGWZV+pxSAJY0mjaORQyMMEHvXkXi/QJNDvS8YJs5iSjf3T6V6jd6tYWaFri6iUD0YGuU8Q+MfD13YS2bZuQ6kY29DVRugNjwGd3hW1Pua6KuX+HjZ8KQY6B2x9K6ek9wGyfcb/dNeIXttc3niK6gtIjLKz8ACvb3+63+6a8Y/tW70TxTdXVoQRv/eIR94VUAO08MeB4LEJd6rie56hO0ddnsXbt2rtxjGOKz9C1i11qwS6tWHI+Ze6mtOpbYHn3jHwft36jpMeCOZYh3964PqPT1B7V74QCCCMivOPHPhj7M7app8f7sn97GOx9a0hPowOKopoORkUtbALRRRQAUUUlAC0UUUAFFFFMAopKWgAooooAsW/3Kpy/65quW/3Kpy/61qQhlFLSVQBS0lLQAUd6SloASloFFABRRRQAUUlFAy7pP/IQH0q5rX+uWqWk/wDH+PpV3Wv9ctT1EZ560lKaKYxKKKWgQlLSd6WgYUUUUAFJS0UAFFFFABSUtJQAUUUUAFFFFMAooooAKKKKAClopKAFprNtXP4UGtfwppbavrsaFcwwHdJ7ipk7IR6L4K0v+zdBiDriWX52PseldBikRQqhVGFUYH0p1crd2MpX2lWN+my6tkcfTFebeN9G0zRXhFj8skxOY85xXqjsEQsxwFGSa8W8SakdW1u4uR/qx8iD0xVwvcDsPhhzp9z9a7nHufzrh/hgMafc8dxXc5qZbgGPc/nS4pM0tSAYrhPiXYM1vBqKL/qflb8a7uqWr2Kajps1rIMh1OPr2pxdmB4lmilnhktLmW1mGJIWINMzXYncBaKKKAClpKUUAFSW1s97dw2cYJaZgv0qPjqeldv8OtGLyPq1wnH3IwR+tTOVkB3Om2i2NhBbIABGgB+tWqKK5AIbq4jtbaS4lOEjXcTXiuqXr6hqU905J3MQv0r1vxFpU2sacbSK6NuGPzkDqPSuS/4Vu4AA1M4HH3a0pyS3A4OU/u/xH869w0U50e1/65iuIf4bSsuBqZ6jtXe2Fv8AZLKG3LbjGoXPrTqSTAsVDcf8e8v+4f5VNUcq743TONykVkB4O/8Ar5/+ujfzoFdy3w2YyOw1RvmYt931pP8AhWz/APQUP/fNdKqIDh6X8K7j/hW0n/QUP/fNKPhu/fU2/wC+aftYgcPiiu6/4Vuf+gm3/fNYviXwx/YFvHN9rM+84IIximqiegHP0UGirASlopKAFpKWkoAWiikoAWiiimAUUUUCCiiigAooooAKKKKACiiigAooooAKtR/6n8Kq1ZT/AFP4UDKB6mig9TRQAUUUUwCiiigQUtJS0DCiiigBKWikoAWigUUALRQKKBBRQcUUDKVJRRXMIKKKKACiiigAooooAKKKKACiiigAooooAKKKKAJLX/j8h+tb2sf8eiVg2v8Ax+Q/71b2sf8AHpHUgYwooopgFFFFABRRRQAUUUUAFJS0lABSGlpp45oAkt2jW8tzOu6LzBvX2r2/SbDToLWOSxgCI4DLzmvF10XVbiEPDaMyuPlOOteueDzeL4fghv4vLlhGwD1FRIB3i3RodZ0aVHTMsal4z3z6Vw/wtDR+ILuKRcSJFhhjvmvT55PKhaTbu2jOPWuG8H2N6njLUL6a08i3nUlanoM78dKy/Ev/ACL1/wD9cTWoOlZniSOWXw9fRwrukaIhR6mkBxXwm04eRcaky/M/yAmvR8cVgeCNNk0zw3BBMmyQ/Mw9M1uzMyRMyLuYDIX1psDhfiXf2CWa2LQJJeSdCOqD1ribDxBrWm2i2llemOFei471e1zRfEVzfXGo31pwWOCDnC9qzrPRtVvZBHbWbkn++MVatYRYk8X+Iwmf7RYfhXq3hm6mvNBt57h98rD5m9a8l8ReHb7QoIHvMfvjgYOcGvVfB/8AyLVr9KUrdBm3mvM/ifLJFrGnyxNtkQ5VvSvTMV5j8UUkfVrBIkZ3boqjNStwMP8A4S3xH/0Ejx/s1oeHvE+u3fiG0t7m/Z4n+8uOtGm+AdWvI/Nu3FspGVAOSazvD9tLZ+Nra2nG2SMkEVegj2wU2X/Vv/umnDpTZf8AVP8A7prMZ4Vqn/IZvf8Afq94V0mHWtaFncsVj2FuPaqGqf8AIZvf9+t34c/8jWP+uTVq9hHUf8K60v8A56vQfhzpR6yPXZ0tZ3YzJ0DQ4NDtHt7d2aNm3YPart7btc2zwrKY9wwWFWaKQHES/DjTpZDI9zKzt1JOaYfhtpva5cV3NIelPmYHjnjDw/b+H7q1W2lL+YwzmvWNM/5Bdp/1yH8q4D4qf8flj/vCu/0z/kF2n/XIfypsCy/KNj0NeS6JoWqx+NlupLQrCLgsW9s164KXA9qm9gE7muT+IemXup6VbxWEIldXJIPpiusp1O4HI/DvSr3StIeK/hEUhYkD2rrqKKQHlnjHwxrOo+J5rq0tFkhO3DZ9K9MsUaOygRxhljAI98VPRTuAV4345H/FW3f4V7JXjnjj/kbLv6CnDcDniK6H4ff8jfD/ALhrnyK6DwB/yN8H+4a0lsI9hFFFLWIzkvFXgxfEV/DdG8MBjXbjGc1ij4WoP+Ymf++a9FLKOrAfU0b0/vr+dO7A4TTvhwtlqEF1/aJfymDbdvWu+pgZT0ZT+NOFK9wFPSvMvij/AMhS0/6516YeleZfFH/kK2n/AFzpx3AX4Xf8hO7/ANyvTR0rzL4Xf8hO6/3K9NHSiW4CHrXAaj8N473UZrv+0SnmsW27eld/TSy5+8v50kB53/wq6P8A6Cbf981paD4DTR9UjvhfGUp/DtxXY71/vr+dKCpPDKfoadwHClooNIDyv4m/8hy3+grku9db8Tv+Q3b/AErke9bR2EL2oooqwCikpaACiiigAooooAKKKKACrNt9w/Wq1WLf7n40AUpP9a/1ptOl/wBa31plABS0UUwCiiigAooopAFFFLTAKKKKAA9DW5oX/Ho9YfatzQ/+PR6TAzrr/j7l+tRVLdf8fcn1qKhAFLSUtMAooFFIAooopgFFFFABS0lLQAlJ3p1GKQHonwwZU0eXc6r854JxXb+dF/z1j/76FeCpLPEMQzNGPQGlN3e/8/cn5msnC7A9486L/nrH/wB9Cl86L/nrH/30K8E+13v/AD9yfmaPtd7/AM/cn5mlyDPe/Oi/56x/99Cjzo/+esf/AH0K8E+13v8Az9yfmacLy9/5+5P++jRyAe8+dF/z1j/76FKsiOcK6k+xzXgpvbwDJu5PzNek/D3SbqCybUb6R2e45RGP3RScbAdnRS0VICZphljU4aRAfQtVbVr+LTNOmu5jgIpx7ntXiV3ql/f3kt3LcurStnaD92mo3A928+L/AJ6x/wDfQo86L/nrH/30K8E+03n/AD9Sf99Gl+13n/P1J/30arkA9686L/nrH/30KPOi/wCesf8A30K8F+13n/P3J/30aPtd7/z9yfmaOQD3nzos/wCtj/76FeP+NiD4pkIYEbeoOaxhd3v/AD9yfmajZndt0jF29TVRjYBc1NZyRJqFs1wu+EON6+oquTgZNXV0XVpo1eKzZlbBU461bA9s0+K2S0iNrEscbKCAB2rO8V6JFrWkSxGMGdAWib0NS+GZbmXRYBeQeTLGoQr9O9akr+XGz4ztGcVh1A8v+F0bReINQikGHRNrcdxXqVcF4Nsb6Lxdqd3c2vkxTD5ffmu9obuBj+LP+Ravv+udcn8LdN/d3GpOufM+VD6V13imGa48PXkNum+V0wq+tR+EbCTTfD1tbSpskHLD3ovoBtVwfxH1eCO2XTI4kkmkOXbHMfpXcTuY4HdRuKg4Hqa8f1fSPEN7fXOo3VmBvJyc9AOlOO4EMPivXraBIYb3EaDCjb0ofxp4jAH+n4yQPu1U07RdU1V9tlasVzgs4wKj13R7vRLmO3vB8zkFSOhq9APaNDuJbnR7aadt0jrlj61oelZXhvnw/Z/7lamMYrJgeU/EGaW38V289u+yVV+VvSss+KvER5OpN/3zWn8RIZZvFNvHBG8jsvAUZplj4B1a6gaW6YW/y7lUHJb2rVWsBb8Ea/rF94mjtr29MsJjJK4716jXj3gOGS38ceRMu2SNCrCvYKiW4FbU/wDkG3X/AFyb+Rrwn+N/9417tqX/ACDbr/rk38jXhJ/1j/7xqqYC4DAqeh4r074d6qLnRmtppAJIGwoJ/hrzDvUkcs8JJglaMng471co3A91e7tkGXuIlHu4rNu/E+jWgJmvV467ea8ZfzJOZZZG/wCBGmGJAhwG/E5qPZge7aZqVtqloLq0YtETgEjrVyuZ+H//ACLEX++a6as2BFOSkEjL1CkivEdS1zWL26nW4v3KLIVCjjAr265/49pf90/yrwOb/j8uf+uzVpBXAjK7jl3dj7saMDBwB09KfikI4P0rQR6p8NX3eFo1/uua64VxHwvkzoskf91q7cVhLcYjD5W+hrw7XV2+Ib1fRq9yNeJeJ12eKb4e4qoAJoOs3GhagtzAS0LH99H2I9a9l06/t9Sso7q1cPHIMj2rwofSt7wl4gk0G+EcpLWMxww/umqnHqgPYaiufKMDicqIypDbumKpajrdhp2nfbridfJIyuDktXlviHxTf67I0YZoLPPEY/i96iMWwKOsQ2kGr3Edg++23fL7VTFIF2jA6UtdC0AWiiimAUUUUAFFFFMApKWigBKWiigAooooAnt/uVUm/wBc1W7f7lU5f9a1CENpaSimAUUUUDFoooxQAUUUUCCiiigBKKKKALmk/wDH+PpV3Wf9ctUtJ/4/x9Ku6x/rlqXuBn0lL3pKYwpaSigQUtJS0DCiiigApKDS0AFJRRQAtJRRQAtJRRQAUUUUwCiiigApaSlFABSGlo9ycCgBh3ZAQbnY4UeteteDNEGj6SnmL/pEw3Oe+D2rlvAXh43l0NVu0xBGf3QPc16ZiuepK+gAKWkqK4njtoHmmYLGgySe1ZAcz4/1n+z9J+yRNie5+UEHlRXlo4GO/f61f17VZNZ1ea7cnYDtRfQDvVCuiEbIDT0rX9Q0eJ4rJ8K3Jq9/wm2u/wDPYflXP0Yp8iYHQHxtr2DiYdPSu98G642taQHnI+0x8SivIjW/4H1b+y9dEUjbbe54b69qicFbQD16kNJnNHWsRnnfxF0Ro5F1e2T5ekwHc+tcQORkcg17rdW8d1bvBMoZHBBBrx3xDos2hak0DAm3c5iftit6cugjNpaQU6tgCiipbW2nvblLW1QvK5wAO3vSbsBa0TSpta1SO0iHyZzK3oteyWdrHZ2kdvCAEjUKKy/DGgxaHYCMYad+ZH9/StuuacuZgFFITUEl5bRPsluIkb0ZgDUAWKKq/wBoWf8Az9w/99ikOoWX/P3D/wB9inYC1RVX+0bEdbuH/vsVZVldQykFT0IpWAWiig4AyeAKADFGKr/b7P8A5+of++xR9vs/+fqH/vsUAWaKrfbrT/n6h/77FL9utP8An6h/77FFgLHauH+Jv/IPg+tdh9ttf+fmH/vsVxXxKmilsIBFKjnPRWzVwXvAcFRSUtdYBRSUUAFLSUUALRSUtABRRRTEFFFFABRRRQAUUUUAFFFFABRRRQMKKSlpgFWU/wBT+FVqsp/qvwpMCh3NFHc0UwCiiigAoopRQAUUUUCCiiigYUUUUAFFFFAC0UUUCCiiigCjRRSVzALRRRQAUUUUAFFFFABRRRQAUUUUAFFFJQAtFFFAElr/AMfcP+9W7rH/AB6pWFa/8fcP1rd1j/j0SpAx+1FHaimAUUUUAFFFFABRRRQAUUUUAJTJPuH6VJimOMoQO9AHtvhY7vDdgSB/qgOla9cRoPjHQ7LRbW1nuHWWJNrDb3rQPjvw73u3H/AayaA6jFA/Cs7R9Zs9YhaaxZ2jU43MuM1o5pDFooooAKKjnk8qF5NpbaM4Fcs/xC0SGVorgzxSKcFWTFAHTXswt7SWZgCEGcGltJUnt0ljAwwzwK5S58eeHrm0li8+TLKQBsq94G1AajoAkB6SMMe3aiwDfHumjUPDU+1cyw/Mn1qx4PGPDVoD1AwfrW3JGssbI4yrDBFQWFnHY2wgi+6CT+dAFisZ/sj+IgJkUzqP3RPatgiuI8RXf2Lx7pDE4R8hqYHbYwfxrgvEWmi28faZqEY4uMhz9K7339eao6jpyX0sDtwYjkGi4F8Hp9KSX/VP/umgcAfSkk/1b/7ppAeE6p/yGr3/AH6s+H9YbQtU+3LAJjsK7Scdaqaof+J1ej/bqtmtegjuj8T5wf8AkFL/AN9Uf8LRl/6BQ/76qb4d6Vp+o6VcPeWqSusmAW9K63/hGdE/6B0X5VDsM4v/AIWlL/0Ch/31S/8AC0pP+gUP++q7T/hGdE/6B0X5Uf8ACM6J/wBA6L8qV0BxR+KUn/QLH/fVJ/wtJz/zCx/31Xbf8Ixoh/5h0X5Vna94c0eLRLyWKwjWRIiVYdjRoB554n8Tf8JFJbs1qIPJbJ5zmuqtfiPpcFrDC0bZjQKeK85iGY1z1rr/AAV4Wsdf06W5u2IdX2gAdqtpWA3f+Fm6T/zzf8qP+Fm6V/zzb8qnHw40fuSfwpf+Fc6MPX8qnQBlr8R9LuLmOBY23SNtHFdojbkVh3Ga5O28A6TbXUVxHndEcjiusUbVAHQDFJgOrnPEHi+x0G7W3u1bcwyMCuirB17wrY67cpPdscoOABSAx/8AhZmj/wB1/wAqP+Fl6N6P+VSf8K50jszflR/wrnSf77flT0AZ/wALK0Y9n/KuG8Q6lFq2tTX0AIjkAABrsdT8BaZaadNOkjbo1yOK87TkH6kVcUhCkVv+AR/xV8H+4awDXQeAv+Rug/3DVS2A9goNLRWIzzP4kXWqQazarYPOsZjOfLUkZrkPt/iH/ntd/wDfBr3lkVuWRT9Rmm+VH/zzT/vkU0wPFtEv9e/tq0WSa6MbSAMChxiva+5poiQHIjQf8BFOobAUnivMfiicata/9c69NPSvMfin/wAhW1/65047gL8LTnVbof7FenA8V5d8K/8AkLXX/XOvUF6US3AXuK8X8QXuurr92kEt0Iw527UOMV7TTTGhOSin8KSYHhP2/wAQf89rv/vg1v8Agi71iTxLEl1JcGI9Q6kCvVvKj/55p/3zSiNAchFB9hTuA6iiipA8r+J3/Ibt/pXJd6634nf8hu2+lclW0dhBRRRVgFFFFABRRRQAUUUUAFFFFABVi3+5VerFv9ykBSl/1rfWm06X/WN9abTAKKKKYBRRRQAUUUUALRRRQAUUUUAHatzQ/wDj0esPtW3of/Ho9JgULr/j7k+tRVLdf8fcn1qKgAoopKYC0UlLSAKWkpaAEopaKACiiigAooooASkxS0tADcUmKfQRQAzFJ0p+M1oaDolxruoLbQAiFT+9k7AUm7AX/Bnh19b1ATzqRZQHJJ/jNevoqogVQAoGABVXTrCDTbKO0tUCRxjHHc1brFu4xaKKKkDzf4m6q0ksWkxN+7PzSfWuFArpfiDhfExJ7jp61Lofge91Swa7uJDbZGYV7t9a1i0kBy2OKMVc1PTL3SbgwahCY2HRh90/jVOrWoCYpQKUUYpgJijFLilxQBDN/qmr3Lw/zoNjkD/Ur2rxCRdyEDrXp2leN9DtdLtreaaRZIowrDb3rOaA7QUVy48feHu9zIPqlbGkavaaxbmeyLtEDjcy4zWVgNCiiigAooqC8uPsts8xUsEGSB1oAnqOZUeJlkA2nrmuaXx7oDMUknlikBwUdMGquueM9JfR5/sd0zTfwjFOwHWwwxQxCOFFRB0AFcj8SdJ+26It3GCZrVht+lb3hu+/tHQra6zkuOa0J4UuIWikUFWGCKNgM7wz/wAi9Z/7lauelQWVqlnapbx/cTpU/cUmBiW8kH/CSzRPGpnIyjEcgVtdDXF3919l+JVkrHCSxMD9a7XFMDhf7P8AsXxRimUYW6iZz9a7yqc+nxTahDet/rIlKirlDYFXU/8AkG3X/XJv5GvCP43/AN417vqf/INuv+uTfyNeEH77/wC8aumBNbxGe6ht1ODM4QH0rsP+FcXZ/wCX41yWmH/ib2P/AF2Fe7J90fSnOTQHm/8Awre6/wCf80h+G10QR9vPNel0VHOwMnw1pLaNpKWbyeYVYndWtRRUgMlTfEyf3gRXmkvw4vmuJXW84dy3516dSYpp2A8yHw3vMf8AH6aQ/De+xxeV6dikxT52BzXg3w9PoFvNFPL5m/pXTUmKWpeoCE1wWu+BLnU9auL6O72LL/D6V32KTFNOwHmg+G973vad/wAK4vCMG+r0oClp8zA83f4e6hLFHFNqbvHEMIpPSmf8K3ux0via9LpKOdgeL+INCm0C5ignl8zzV3A1lV2fxQ41Sx/65muLBreDugHUUlLVAFFFFMAooooEFFFFAwopKWgAopKWgCeD7lVJf9a1WoD8lVJf9a1CAbRRRTAKBRRQAtFJS0AFFFFAgpDS0UAJRRRQBc0r/j/H0q7rP+uWqWlf8f4+lXdZ/wBctS9wM/vSUtFMBKKKWgAooooGFFFJQAtJRRQAUUUUAFFFFABRRRTAKKKKACiiigApaSjNAC1r+GtCl16/C4K2kZzI/r7Uzw9oVzr12I4gUtlP7yT29q9c0zT7bTLNLW0QJGo7dz61jUnbRATWtvFa26QQqFjQYAFS0UVgAhrz34ja/gDR7V/mbmYj0rq/EutRaJpb3DEGVhiJfU141LLJcTvcTsWkkO4k/wAq0hG7uBGowAOwp4pOgrpPCvhQ69azXM8rQRdImA6mtm0gOdorotT8F6vYbnhUXEI/izz+Vc9IrxOVmieNh/fGKakmAhpjZ4ZThlO4fUU7IPQg0Gm9QPW/ButrrGjp5jf6TCNsoroRXifh/V5ND1aO6TJiY7ZV9fevZ7S4iu7ZLi3YPHIMqRXNJWYE2Kz9a0i21mwe1uVBz91u6mtGiouB4lrGk3Wi3htrtTt/5ZydmFUq9t1XS7TVbNre8jDqeh7qfWvLvEHhi90SQuA09pn5ZAMkfWuiFS+jAxK6jwVrel6RIyX0WyaQ8T4zx6Vy2cinw21xd5FtC0mOuBmrkk0B7lb3ENzEJIJFkUjIKnNS143pLeJNHmEljBLs/iiIJDV6XoesSajp7TXVs9vKn30x/KuZxsBPrmqw6Pp0l5MR8o+Rf7x9K8Tv7ufU72S8uZH3yHIAbG0eldL4sl1rXdQOLKRbSI4jTB5PrWH/AGPq3/Pi/wCVXBIDO2H+/J/30aXZ/tv/AN9GtA6Rqneyk/75pP7J1P8A58ZP++avQDOkT5PvP1H8Rr3XQuNFtOc/uxXjMmk6oUx9ik6j+GvadGRo9JtkcYYRjI9KidgLwqG9/wCPKf8A3D/KpxUN2pa0mVRklCB+VZAeCFcyS/M/+sb+I+tG3/af/vo1oHRtUWSXNk/MjHp70n9k6n/z4yf9810q1gKAX/bf/vo0oH+3J/30avf2Tqf/AD5Sf980f2Vqf/PlJ/3zTugKeD/z0f8A76NGD3Zj9Tmrv9lan/z5Sf8AfNJ/Zepf8+Un/fNCaAqg0tTS2N9Chea1dEHVitQZz0q07gLS0lFMBaKSigBaKKKYBRRRQIKKKKACiiigYUUUUAFFFFABSUtFABRSUooAKsJ/qvwqvVhP9V+FDAo9zRR3NFMAooooAKWkpaACiiigAooooAKKKKACiiigQUtJS0DCiiigCjRQaK5hBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQBJa/wDH5D9a3dX/AOPVKwrX/j8h+tbur/8AHqlJgY/aiiigAooooAKKKKACiiigAoopaACkpadHFLPII7eJpJG6KozQBCwAGTiuk8KeDZ9blW5vFaGyU9xgv7fSt/wv4DwUvNbAJ6rAOn1NegxxrGipGoVVGAB2qHICG0tILK2S3tYljiQYCgVOKDxXPW3iOO+8VPpFph44Y90kg9fSoGdFSZpKzvEE81tol3cWzbZYk3KfegDSzXP+JfCtjr0JZlEV0B8syjn6VF4Q8UQeILFQxCXkYxIh7+4rpAaewHg2r6Re6LeG2v4tv92QD5WH1rtfhVdYa8sifuAMBXa61Y2F/p0kepIhhUElm/h968j0PVovD/iGW4t2aW1DFc9yKrdAe2UteWap8R7+aFksbVYQSMSZ5r0LQbmS70e3nmbdI65Y+tS1YDRry/4nuYdb06ZTgo4r1CvL/iqu7UbMUID0eymWeyhlU5DIP5VPXlWmeP59N0mGyWzWaWIY3E1Z0Hxjquq+KLWC4Iht3zmMdDT5QPTKa/8Aq3/3TT8UyT/Vv/umpA8G1b/kOXw/26q1a1j/AJDt9/v1UrVbCPTPhYwGkXmWH+tFd0HX+8v514Rpi64Y3OkPKsefm2DvV7Z4x/563P5VLQz2rcP7w/OjcPUV4sF8Y/8APW5/KnBfGX/Pa4/KlYD2fcPUVna/htCvgCP9S1eV7fGX/Pa4/KorhfFggc3E05hx84I4xRYDFhHyCvS/hYQNFnyQP3p4rgNN0u+1QsNPiEm3rmta28P+LLVSLQGFT1CnrVPYD2Hcv94fnS7l/vD868j/ALK8bD/ls/507+y/G/8Az3f86iwHrWQe4ory7TdO8ZJqlq1zM5hD/Pz2r1AZ2jPXFIBaTIHUgUVwXji38Rzaoh0d3WELztPegDvNw/vCl3D+8K8eFj44/wCe0v508WHjg/8ALeX86qwHpuvsDol1gj7hrw5D8p/3jXRSaX41ljMcksjI3VSetZt5oWqabame9txHEDyRVR0EUq6DwD/yN8H+4a52ui8A/wDI3wf7hqnsB7DRRRWIzO1DW9P02ZYry4EbsMgHvVT/AISzRP8An8WuG+KSBtetM/8APM96w9F8KX+uQPPYhPLjbadz45qraAeqf8Jbon/P4KP+Eu0T/n7Fee/8K41r0h/7+Uf8K41sdof++6LID0A+LtDx/wAfdcF8QtTs9Tv7eSyl8xVTBPpTP+Fc63/0y/77pR8OtcHTyf8AvumrIQ74eajaabqVxJey+WrJgGvQF8V6Lj/j8Feff8K61s9TD/33S/8ACuda/wCmP/fyh2Yz0H/hK9F/5/Fpf+Eq0b/n8WvPf+Fc6z/0x/7+Uf8ACudZ9If+/lKyA9B/4SrRf+fxansdf0y/uPItbkPJ6V5Hrnhi80GKOW/VNsrbV2tnmrvw9RV8VrgEcetHLpcD2OiiipA8r+J//Iat/pXI96674n/8hq2+lcj3raOwhaKSirAWiiigAooooAKKKKACiiigAqxB9yq9TwfcNICnJ/rG+tNp0n+sb60ymAtFFFMAooooAKWkpaACiiigAooooAT1rc0T/j0esM9DW5of/Ho9JgZ91/x9yfWo6kuv+PqT61FQAtFFJTAWiiikAClpKWgAooopgFJS0UAFFJRSGFLmm0tAhaKbWnoOiXmu3QitlKwg/vJT0A9qTdgGaPpN1rV6traKdufnk7KK9g0TR7XRrFba1UcD5m7saNF0e00ayW3tUA/vN3Y1o1jKVxhRRRUgFFFFAHOSeFoLvxG2q3x8wJ/qoz0+tdEAAMDge1MEqGQoHUuOq55p+aAKmpadaanatb3sKyIfXt715b4l8I3eis1xbbriyJ6gfMv4V67TXRXUqyhlYYIPenGVgPAQQRkf/qp1d/4p8Dhg97owCuOXg7H3FcCVZHMcilJFOGVhgitoyuISlooqgEpDgDJwKfHHJNII4UZ3bgBRmu98M+BACl5rQDN1WDsPrUykkBheFvCM+tSrcXaGKxU55GC/tXq9rbQ2lukFvGEjQYAAqSNFjRURQqqMADtQTWLdxi5oJrnrPxEt94ouNKtsNFAgZpB6+ldBSAKKyPFV1PZ+H7q5tW2zRDKmqvhLxNb+ILFeQl3GAJIz/MU7AHiHwjpuuIztGILo9JkHNeV61oV5odyYb2M7Twko5B+te6iqOtw2E2mTDVERrYLlt3amnYDlfhffebpc1iTzbnj8a7mvF/D2twaFr0tzFuNkxIx39q2NV+It7dYh02AQqzgebnnFDTA9RoqrprO+nwPI+9mQEn1q1UgeX+PZjbeMdOnU4KsM/SvTLeVZ4ElQ5V1BFeW/EwZ1+3x12HFWdM8ff2fosNp9m864iXbycA1fLdAem0V5/wCFvFeo6v4nWC6xHC8ZIjB4zXf1LVgKuqH/AIll1/1yb+RrwbPzv/vGveNU/wCQZd/9cm/lXgw++/8AvGrpgW9NP/E2sSSAPOHJ7V7mlxCEX9/F0/vivA8dwcHsaf591/z9SfnVSjcD3r7TB/z3i/77FH2qDvcRf99ivBfPuv8An6k/M0ySW48tibiT/vo1PswPoFXV13IwYeoOaWub8BM7eGIS7Fjnqa6Q1nYBM0wzRdPNj/76pLjP2eXHB2n+VeFzXd8bu4/02UYlYdapRuB7sJo/+esf/fVHmx/89U/76rwkXmoDpfS0v23UP+f6Wq9mB7sroxwrqT6A0/FeN+FL+9XxRYiW7d42J3KT1r2OoasAtNZ1X7zKv1NOrgvic1zFDZz207xbSQ23vQtQO582P/nqn/fVHnR/89U/76rwc3+ohf8Aj/l6Z61t6XoXiPVtMa+tL18fwKxxuFU4WA9c86L/AJ6x/wDfVBnix/ro/wDvqvDLqTWbKYxXk9xE47kHb+dRC8vWGftshHrmmoAdf8T2V9SsSjq37s/dOa40UO8spBmlaQjoT2oFaxVkA6lpKWqAKKKKACiiimIKKKSgBaSlpKBhS0lLQBND9yqkn+tarcP3aqS/61qEA00UUUxBS0lLQMKKSloEFFFFABRRRQMSiiigRb0v/j/H0q7rH+uWqWl/8f4+lXdY/wBctSwKBpKWkpgFL2oFFABRRRQMKKMUlABS0lFABS0lFAC0UlFABRRRTAKKKKACiiigBDWl4f0qLWNSW3nuRBGOSScbvas0005DBlYqynKsD0qXsI900+xt9OtEtrSMJGo7d/erdeeeEfGvMen6w/zHiOY9/Y16ArBgCCCD0I6GuZprcY+obu5htLZ7i4cJHGMsTT2cKpZmAA5JPQV5X428TnVrk2Fk5FnEcOw/jPpSSuBmeJNbk13U3nJIt0OIk7fWswUxRgcdBTwQASeldKVkBLZ2cuoX0NlbjLyt+Qr2zS7GLTtPhtIVAWNf171yHw70MxQtqtynzy8Rg/w+9d1isJyuwExVG/0mw1BcXlqkv4Vfqrf3kNhZy3U7BY4xnPvUq4zzPxlommaM6fZJiJZekPYVzNWNU1CXVtTmvZicucKvoKr10xvbUQhAIOehrrvAniP+zrgaZev/AKNIf3bH+A+lcnikYZHoRyD6GiUboD3wEEZByKWuD8D+KxKE0vUpMSqMRSMfvCu8rmasAVm67qNppmmS3F7taPGAh/iPpV25nitoHmmcJGgyxPpXjninXpNf1EspItIjiNf73vTjG7AzbicXN1JOkQiWRsiMdFrtvhaTi6HGN3pXBjrXd/C3/l6/3q2n8IHoQ6dB+VLj6flQOlL+Fc4CY9h+VGPYflS0tADcew/KjHsPyp1FADcew/KlApaKACiiigBMfT8qMfT8qKKADH0/KjH0/KjNGfagAx9PyoI9h+VGfakP0oAwvGX/ACLVz0/KvH4/9WteweM/+Raua8fi/wBWtdFECQUtIKK2AWiiigAooopgFFFFAgooooAKKKKBhRSUtABRRRQAUUUUAJS0UUAFWE/1X4VXqwn+r/ChgUT1NFB6mimAUUtFABRRRQIKKKKBhRRRQAUUUUAFFFFAhaKSloAKKKKYFE0UUVygFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAElr/wAfkP1rc1f/AI9UrDtf+PuH61uav/x6pSYGR2oo7UUAFFFFABRRRQAUUUUAFLSUtABVrS9SuNIv0vbU/Mn3hj7w9Kq0UAe5aJqsGsadHeQEfMPmX+6fStAnAryb4e6rNZ659gyTbzjcR6NXpuqWcl9ZPbxXUlsW4Lp1xWTVhnHeOfGC2yPpmlvuuGGJJB0UVh/C4Y8R3DM2SYslj3NbL/DC2Ylv7RmJY5JPUmnW/wAODayGS11eeFzxuUdqelhHd71x94fnWZ4kw3h2/AYf6k1z58Eah/0MV3TJPAt/JG0b+IbpkYYZT3pKwHmVhdXNhcx3dnIY5ozwR/F7V7B4W8VW2u2eHZYbuMfvEY4H1rnf+FXxjj+0ZKcvwz8tt0eqTIxGCRVOzAzvHnis6jM2l6c5FtGcSuP4z6Vxy4AAHQV0nibwa3h+yW7W5aYO2CDWHpen3Wr3qWlihZ2+83ZRTVkBJpmmXOs3qWVmmWYjc3ZRXuOl2YsNOhtQ27y1AJ96oeGvD9roFiIYQGmYfvJD1Y1tVDdxhXmPxSH/ABMrL8a9OrzL4qHbfWB5Oc4A6miO4HEEgD8eB616H4D8JtE8esagCsmP3MZ7CofBXg4ysmqatH8vWKE/zNejAADAqpSAWmyfcb/dNOpsn3G+hqAPBdY/5D1//v1UPer2sr/xPr//AH6pkda1Wwj0v4W/8gW6/wCuvpXcfl+VecfD3WdP03S7iK8mEbtJkD1FdZ/wlei/8/YqHuM3Pyo/KsT/AISvRf8An7Wl/wCEq0X/AJ/F/OpswNus3xDxoN90/wBS1Vv+Eq0X/n8WqWteJdIm0e7iS6DM8RCgdzTsBzvwnPNyO2P616UOleP+BtfsdAaY35cbxgbRmuvHxH0Pv53/AHzTaYHY0VhaT4pstXINnHMyE43leM1u1IBRWZrmtQaJai5uo5GiJwSi5x9awh8R9CIyDN/3zRYDrjRXIH4i6F/02/75rW0LxJZa80gsVkKx9WYYFOwGzS/lSVi6v4kttJuRBNbXMjHnMaZFIDaP4VyfxIP/ABTEn+8Kjb4i6IGIYTAjqCvSsPxd4u0zWtFazs/M81iD8wwBVJAcSOg+grofAB/4q+D/AHDXO9h9K6HwB/yOEH+4at7CPZKKQUtZDPLPif8A8h+z/wCuZrX+FZzpt6P+mtZHxQ/5D1n/ANczS+BPEWm6JZ3MWoSMjySbl2jPFX0A9SoxXO2vjTQ7u5jghuHMkhwoK966KoATbRilrE1PxVpGl3n2W7mZZcZwFzQBtYpcVjaV4n0rV7s21lKzyhdxBXHFbNABiimu4RGZuijJrm28daAsjxm4fcjbT8vegDF+K/FlYn/prXOfD8/8VYv0rQ8f69putWNqthKzyRyZYEY4rN+H5/4q1fpWi2A9lpKTuaWswPLPih/yGbb6VyHeuv8Aif8A8hq2+lcgOtbR2ELS0lLVgFFFFABRRRQAUUUUAFFFFIAqeD7lQVND9ygCpJ/rG+tMp8n+sb602mAUUUUAFFFFMBaKKSgBaKSigBaKKKAENbmif8er1hmtzRP+PVqTAoXX/H1J9aiqS6/4+pPrUVAC0UUUAFFFHSgApaSloAKKKKACiiimAUUUHA5JA9zSAQ00nHXqeg7mr2naZfarOIbGAsT/ABkfL+deieHvA1np+2fUMXNx1w3RD7VMpJAcn4b8G3msMtxehrezznphm/CvUtPsbbTrVbaziWONR0HerAXAAAAA6ClxWLlcYtFJVPVNTtdKs3uryQJGo6d2+lIB2p6hb6ZZSXV1IEjRSeT1+lVPDWrPrWlLfMmwOTtHtXk/iXxFc+ILlnkJS1TPlxA8fWvR/h5/yKdv+NNqyA6ekPWlpD1pAeW+NNSvNL8YxXFjKyMF5XPDD0rrvDHiu016IRsRDdr9+MnqfauI+Iv/ACMyf7prmIpJYJknt5DHKhyritOW6A+gQaWuM8H+Mo9TC2WpMIrxRhWJ4k/+vXZis2rALXLeKvCMGrxtc2oEV4oyCP4/Y11NFCdgPBbmCa0uXtrqMxzIcFT3+lR16/4o8M22u2xYAR3aD93IO/sa8kvrS5068e0vYzHKhx7N9K2jK4hbK+n029ivbY/PEckf3h6V7Roeqwaxpsd5AQdw+cejdxXhxauq+HGpy2mtPYgkwSrkL6H1pTVxnq5PpXDeOvFwsYm0zTnDXbjDuvSMf411+oW0t1ZyQQXDQOwwJF6iuHk+GEbuztqczuxyzHqTUKwGV8LQf7fvCzFmKZJPUmvVa4Sz+HclhKZbPWJ4ZCMFlHWrh8Jar/0Ml1+VDswNHxudvhS9zx8teO2N1c6fdRXllIY5kwcj+KvSrrwRf3cJhudfuZYz1Ujg1Q/4VguMf2jJVKwHUeFvEtvr2n+bkR3EY/exnt71w3jrxM2q3jadZuRZxH5yP+WhrWtvh1PaOz2mrTRM4wxXuKwPE3hA+HrRblblpg5+YHsaStcDnP6Vo6Fo11ruoJb2ykIjBpJey4pNC0W7169W3tVIj/5aS9lFeyaLo9ro1itraIAAPmbuxpykBctYBb20cKnIRQM1LRRWYHlXxK/5GG3/AN01ygySFVSzMcBQOSa6v4lAnxHaqqlmZSFUdSa3PBfhAWwTUtUQNOwzHGeiD/GtVKyAf4H8KPYFdUvsi5dcIn90V2ppaKzbuBU1P/kGXX/XJv5GvBgP3kn+8a951L/kG3X/AFyb+VeDE/vZf941dMB9JT4YZrhtlvC8reijNadv4Z1y6x5VpjP9/itboDJpsmfLbjtXVwfD/W5cecYox7NWlB8M9y4udQkXPULUuSA6HwF/yK8P1rpaoaNpkWkaclnC5dU/iPU1fFYvcBk4zBIB/dP8q8ai8Na1e3lwYLTAMzYL8V7TRQnYDy22+HmqyY+1TRx/7rZqTWfA6aTos999sklkiGdhHFenVk+J4vO8PXsfqlPnYHj+mTeTqlpP02sP1r3OI7o0PqoNeAAlVBHVXH6GvdtIuBdaXbzKc7kA/SqmBcrkviTb+Z4ZklHWJhXW1meIrBtS0W4tEALSDjPrULcDx7RdMm1nU4bOEHBwXbsB3r2yytIrKzitoFCxxjAArE8HeHF0HT/3oBu5eZG9K6OnKVwKd/p9pqEXlXtukyehFcD4j8CQ2kMl7p9zsjXkxucKPpXpJrzL4ia59ruRpFs58qM/viD1NON7gcYpyKcKQUo610IQtLSUtAwooooEFFFFMYlLRSUCFpKKKACloooGTQ/cqpL/AK1qtQ/dqrL/AK1qEIbRRS0AJS0lLTASilooAKKKKACiikoAWiikoAt6X/x/D6Vd1f8A1y1S0v8A4/h9Kuav/rlpPcZRooooEFFFFAwooooAKKKKAEpaKSgBaKKKAEopaSgAooopgFFLRQAlLSUtACGmmn0hoAiZQRg9K7Hwb4vkspE07U3LwNxHKeq+1ciRTdtRKNwO18beLftAbS9KkOwj97KO49K4dRgACnhQBgUYpKNgFFa3hzSH1rVo7cA+ShzM3oKykR5JFjiUtI5wqjvXsHhLQ00XSkRgDcSDdI3f6UpysgNqCJIIUijGEQAAVJRRXOAhrzH4g699suv7KtX/AHMR/fEHhq6vxn4gXRtNKRMDdzDbGvpXkg3MS7nLscsT61rTjfUAApwoOAMk4FW9P02+1OYRWNuzk/xEfL+dbXSArY49KSu8034eApv1O5YOf4E5ArmPEOjS6JqJt3y0LcxOe496Smm7AZBByCpKspyrDsa9A8H+M/MCadq77ZBwkx/i+tcERTWAIwaJQTA67x74jN/OdMspP9HQ/vWU/eNccF7AcU7GOlKKIxsA3FaWh6/eaA0htIlkEnUN2qgRTSKbjcDp/wDhYesDpaR/nR/wsXWf+fOL865YitDQNGbXNTNkshi2ru3CocUgNofETWf+fOP86d/wsTWP+fOP86uj4aP/AM/70v8AwrR/+f8Aeo90Cl/wsTWP+fOP86P+Fiax/wA+Uf51d/4Vo/8Az/vR/wAK1f8A5/3o90Cn/wALE1j/AJ8o/wA6P+Fiav8A8+Uf51c/4Vq//P8AtR/wrWT/AJ/3o90Cn/wsTV/+fKP86Q/EXV/+fKP86un4ayf8/wC1NPw1l/5/2o90CkfiLrH/AD5R/nSf8LE1n/nzj/Orv/CtJf8An/ej/hWkn/P+9HugUv8AhYes/wDPnH+dH/Cw9Z/59I/zq7/wrST/AJ/3o/4VpJ/z/vR7oFH/AIWHrH/PpH+dL/wsPWP+fWP86tT/AA5eGCSX7c52KWx64rilGc+xIqlGLA6LUfGmpanYvZz2yIj9WBrAUYUD0oUU/FaxikACiloqgCiiigQUUUUwCiiigAooooAKKKKBhRRRQIKKKKBhRRRQISlpKWgAqwn+q/Cq9Tp/qvwoYyn/ABGijuaSgApaKKYgooopDCiiimAUUUUAFFFFABRS0UAFFFFABRRRQBRooormEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAEtp/x+RfWtvVv+PVKxLX/j8i+tberf8eyUmBkUUUUAFFFFABRRRQAUUUUAFFFFABRRSGgDZ8HnHiq2+le19zXifhA/8VXbfSvbO5rOW40FFZ+vXEtpol5cQttkjjLKfevOvBfirWdT16K2urlnjIyQakD1Wkpa5fx7qt3pOipcWT7JDJgn2oA6bFAFcJ8Odf1HWZ7kX8xkVFyue1d5QBzPjnSbrWNNgs7QDe0nJPYVd8OaBaaBYLBbqDKR+8kI5Y1B4v1aTRrO2vF5RZMSD1FbNrMl1axXEZBWVAwx70wJqWsTxPr9v4f0xriUgzMMRR92NO8K3k9/oMN1dHMkhJPt7UgNqse/0Sz1LV4Lu7w7W3KRnpmtauT1zVW0vxjpqM+2C5yH/pQB1oGBxS00MPz5/CuJ1nxY0vii10bTZMqG/fyD+QoA7imv9xvoaVeg+lI/3G+hoA8K1n/kP3/+/VM1c1n/AJD9/wD79Uz0P0rZbCNzQ/CFzr9q1zDIqhW28nFaP/Csb0/8vC/99V0Xwx/5Ac//AF0rtKzbdxnlP/CsL3/n4X/vqj/hWF5/z8L/AN9V6tSYpXA8q/4Vjff891/76qO4+HN9bQSTmZCI1Lfer1jGap6qP+JVd/8AXI00wPB84LD+6cVPDY31xH5tvaySJ/eVciqzfel/3q9Z+GzE+Fxz0kNW3YRzfgS/1bSL8WNzZz/Y5zx8n3G9a9TpozS5rNu4yC+tUvbKa2kGVlUqa8Rn0nUILueEWM5CSEKQnGK92zSYoTsB4OdN1LHFhPnp9yvWvBmknSdAhhkXErfM3HPNb2KWhu4CYrlPHWrX2n2SW+mWzyz3GQXVM7BXWUEUgPAWsNRBaSazuMnlmKVD2zXumuHbo90ePuHtXhSj5W/3zWsXcQV0PgD/AJG+3/3DXPV0PgH/AJHC3/3DRLYD2IdaWk9aKyGeWfFE/wDE+tP+uZrkBHKRn7LMffbXW/E//kYrHPpXoekWls+lWrG3jJMY7Vd7IDxzRIpRrtkfss4HmryV4Fe79zUK2tuhBWBAR3AqapbuAV4/8QYpT4qLLbSuNnVVyK9gqN7eGRtzxIx9SKE7AeWfDVJR4lkL20sY8nqy4r1eo0hijOUjVT6gVJQ2BDcjdbyADqprwS7tbhdQus2NwczNj5D619AEVGYYieY0P4ChOwHgHkTrk/Ypx77K3/h9/wAjYv0r1e+ghFlOfKT7h7V5T4E48Zyf9dD/ADqr3QHsfc0Unc0tQB5Z8T/+Q1b/AErkB1rr/if/AMhq2+lcgOtbR2EOopKWqAKKKKYBRRRQAUUlFAC0UlLQAVND92oamh+4aAKkn+sb602lk/1jfWkpgFFFFABRRRQAGiiikAUUUtABRRRQAlbmif8AHq9YZrc0T/j1ehgZ91/x9SfWoqluf+PmT61HQAUUUUCCiiigApaQUtAwooooATvS0UUwCrekvYpqcR1RS1qT8wFU6OvXpUsD3XTYLOCzjGnxxpAwyuzoat15l4D8SmznXS76TMEh/cuT0PpXpgORkGsGrMYtFIWABJIAHUntXHeKPHFvp+6003E910JH3V/GklcDY8Q+IrLQrUyXDhpT9yMdSfevI9b1q91278+8chAf3cQPC1Wu7i4vrlrm8maWZv4mPSosVqo2AYfut9DXr3w7P/FKW/415Gw+Vvoa9c+Hf/IqW/40TA6ik7ilo71kB5J8RP8AkZk/3a5jFdT8RB/xUqf7tcwBxW8dhDQCGVlYqynKsOoNekeDPGQuQmnas4WcDEcp/jHv7151ikI6EEgg5BHY+tEo3A+gQc0V574L8ZFnTS9Wf5+kUp/i9jXoOeKxasMDXP8AizRbDVtNdrx1hkjGUmPGPatPU9TtdLtGubyUIijOM8n6V5J4m8T3fiGcqCYrJT8sYP3vc04pgYrp5crx7gwQ4Df3vet3wN/yNkX0FYKqAMAYFdB4HGPFcX0FaPYD2PuaWisnxRdz2Ph28ubZissaZUjtWIGrSGvMfBHifVtT16O2u7lnjKglTXp5p7AJij8K474i6zfaPY2slhKY2kchiKh+HGt3+sJeG/mMvl4257UAdvxXP+LdEl160htI32Luy7egrf7Vz3i3V5NFFndA/u9+1x65pIDU0fSrTR7JLWzjCqo+Y92NaFQxSrLCkiEFXUMCPesjxR4ig8P2HmuQ07nEcfc0wN2iqWj3EtzpcE8/+skXJq7SAzJtFs7nV49SnQSTQjCA9BWnXNz6w1n4zg02RsRXMZYE9jXRMyqpZiAFGST2FADqK4u08UvqvjdNPs3/ANDiQ7iP4zXaUAVNT/5Bt1/1yb+VeDN/rZP94171qX/INuv+uTfyrwVh+9lz/eNaUwOl+H139l8T7CcLLHt/GvXh6V4PpE5tdbsZs4HmgH6V7qjq6h1PDDIpTWoD8CjijPtTXkVBl2CD1Y4qAHUtRxSxzJvikV16ZU5FSUAFFI7bUZvQZryvUfHmsy3E0UMUcKxuUBU88U0rgeqFgoyxAHqay9X1DTk0+dJ7uIblIwGGa8ludd1m6z5+pS7T/DWe48xt0hLt6k1apgIQP3g7b2I+les/D+7Fz4YgUnLxkg15P7V3HwvvALi9smb0KiqmtAPR6KKKxAKKM1na5rFtounvd3TYAHyr3Y0AZ/jDxAmi6YwQg3Uo2xr/AFryElnZpJGLSOcsx71oXNxqHiTVnmILzEFkj/ur7VQA5KkEMvBU9RW8FYQYoApaK0AKKKKACkoopjFopKWgQUUUUDEpaSloEJS0UUDJYvu1Vl/1jVai+7VWT/WNQgG0UUUxC0lLRQAUUUUDCiiigQUUUUAJRS0UDLWl/wDH8PpVzV/9ctU9L/4/h9Kuat/rVqXuBRooopgFFFFAgooooGFFFFABRRRQAUUUUAJRRRQAUUUUwCiiigApaBS0AJRiiloAbikxTqKAGYpCQASTxTsVteFNAk13UQXUizhOXbs3tUydkBvfD7w8Wb+2L1MD/lipHT3r0QVHDEkMSRRqFRBhQPSn1zSdwHVU1PUINNsZbu5cKkYz9ass4VSzEAAZJPavJPG3iM6xqH2W2Y/Y4DjI/jPeiMbsDI1fU59Y1KS9uCfmOEXso7VWSohUimuhKwG/4Q02w1TVvJ1CQgjlI+z16xaWlvZwiK1iWKMfwqK8NgnltriO5gYrJGcgj0r2Xw9q8es6VFdxkBiMOv8AdNZVE7galY/iXRYta0x4GGJVGY37g+lbFBrJaAeDzwy208lvcLtljOGFRGvQPiFoJkjGrWifOnEqgfe968+BBGRyK6oSugFxS4pKdVgFIRS0tADCK6b4d/8AIzv/ANcq5siuk+Hn/I0P/wBcqiewHq9FFFcoBRUU1xBBjzpo489N7AZqP+0LL/n7g/77FAFmiq32+z/5+4P++xR9vs/+fuD/AL7H+NAFmk4qv9vs/wDn7g/77FJ9vsv+fuD/AL7FAFjFGKr/ANoWX/P5B/32KT+0LL/n8g/77FPUC1ik20KwZQykEHoRS0rgVtQH/Evuf+uTfyrwlRy3+8f517vf82Fx/wBc2/lXhKj5n/3j/OtqQDwKdikFLXQAUUUUAFFFFMQUUUUAFFFFAwooooEFFFFABRRRQAUUUUAFJS0UAFFFFMAqdP8AV/hUFTr/AKr8KQyn3NFHc0UwCiiikIKKKKYwooooAKKKKACiiigBaKKKBBRRRQAUUUUxlGiiiuUQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAS2v/H3F9a2tW/49krFtf+PuL61tat/x7JSAyaKKKACiiigAooooAKKKKACiiigApKWmmgDX8If8jXbfSvbe5rxLwj/yNdt9K9sz8xrOW40JNFHPE0UyB0YYZT0NUrXRdMs5hNa2MMUgGNyjmrN1dRWlu887bY0GSazLHxRpOoXS29rcb5G6CpA2qr3lla30Qiu4EmQHO1hxViqGratZ6RbrPeuURm2gj1oAfY6XYaeWNlaxwFuuwdauVl6Pr+n6yXFjIXKdcitSgDjficM+HAD/AH6xPBfjCHTtBlttSYlrZd0fq3tW58TP+ReT/fryoKCq8dBVpXQFnxBq11rl695dsQM/u4+yivXfBX/Ir2leLyj5DXtPgr/kV7WiSsBu15j8VSy6hp7IcMrZBr06vMvip/x/2H40kAaj4/Mnh+K2sQy30i7JWP8AAPauZ8KqT4rsixJYkkk9zWbWt4SGfFtiPrVWsI9wHQfSkf7jfQ0Ch/uN9DWYzwrWf+Q/f/79Uz0NXNZ/5D9//v1TPQ1sthHp3wx/5Ac//XSu1rivhh/yBLj/AK612p6Vk9xmN4g8SWHh9Ea+34fptGazLHx9o1/fR2kAm8yQ4GV4rG+K3+rta5LwqP8AiqLTj+IU0tAPcap6t/yCrv8A65GrnrVPVf8AkF3f/XI1IHgbsqyShmA+Y9a7rwd4w0fRtE+yXkjiXeT8oyMVU8AaNY6xPerfwiTaxxntXZjwNoOP+PVatsCr/wALH8Pf89Jf++aP+Fj+Hv8AnpL/AN81a/4QXQf+fUUf8INoP/PsPyqdAIbb4g6Dc3UVvE8peVtq/LxmusUhlBHQjNc5B4M0S3uI5o7cB4zlTiujUYAA6ChgLWNrXibTNElSO/lZGfpgZrZrK1bw/p2sSI99CHZOhpAZH/Cw/Dv/AD8Sf980f8LD8Pf895P++alPgXQj/wAu9H/CC6F/z709AM3VvHehXWmTwQzSGR1woIrzSMqVIDAkknivXD4E0M/8sawPGfhnTdI0NrmzTbKCAKqLQjgyK3/AI/4rC3/3DWCOg+ldD4BH/FXwf7hqnsB6/R3paMVkM8o+J/8AyMVj9K9J0UEaPaZB/wBWK82+J4/4qC0PomRXPrr2tKiompzKi8KAelXa4HvFFea/Di81XUdVlmu72Wa3jUqVbpmvSM8VLQDqKjZwilm6KMmvF9W8Saw+s3vkajKkSykIoPAFCVwPa80ZrzD4e6tqd74jkhvL2WaMRZ2seM16bQ1YB2aTn0pr/cbHXBrxO+17W11K7jXU5gqykAZ6ChK4Hs98CbKcY/gNeS+BB/xWcn/XU/zrNOu62ylW1Ocg9RnrWl8PwT4sVj1PJPqaq1kB7F60UtFQB5Z8UP8AkM230rjx1rsPih/yGbb6Vx/eto7CHUlLRVgFFFFABRRRQISilpKBhRRQaAFqaH7pqCpofuGgCo/32+tJTpP9Y31ptMAooooAKKKWgBKKKKAFooopAFFFFAhDW3on/Hq9Ypra0T/j1ehjKFz/AMfUn1qOpLn/AI+ZPrUdABRRSUALRRRQAUtJRQAtFFFMBKWikoAKKWkpANOT0OCOQR2r0Xw544totFI1Z2E8AxxyWHavOyKQgZzjkVLjcDpNf8aahq+YbUm1tT3U8sPeubAA6d+vvRSihRsAYoxS0VQEbj5G+hr1v4dj/ilLf8a8mb7jfQ1618PP+RUt/wAaznsB09FLRWQzyf4iD/ipY/8AdrmBXUfET/kZI/8AdrlxW8NhARSGnU2qAawzjBwQcgjsfWu88PeO0t9GeLVNzXMAxER/y0+tcIaQjPWk43AuazrF7rt2bi9chM5SIHhapYpcUuKEhiVv+CP+Rri+grANb3gg/wDFVxfQUpbCPZKjnhiuIWhnjEkbjDKehp+eTUF5dw2VpJc3DbYoxlj6VgMr2mi6ZZz+fa2MUUnTco5rQxWJpninStUuhb2kxaQ9sVt0AVb7TrPUEVL23SZVOQGHSm2Ol2OnBhZWyQB/vbB1pNT1Wz0uJJL2URq5wCabpmsWOqh/sUwk2fex2oAu4riPiiu7RIh713NcT8Tv+QNF/vU1uBleGvGkOneGTBe7nuLYYjH9/wBK4vV9Ru9Xv/tl65Ls42p2Qe1MKjg45Apkg+5/viteUD3TQh/xJbX/AHBWhVDQxjRrX/cFX6xe4HlnxKkkh8S2M0TFXQbgR7VW1jxxfappq2UK+QSMTSDqfap/id/yHrT/AHDXI1qkmhHRfD7C+LYQOnlmvYq8c8Af8jfF/wBczXsQqZ7jK+o/8g+5/wCuTfyrwhx++l/3jXvGof8AHhcf9c2/lXhMv+vm/wB806YEZBBVh1U5FdpH8RZreyhggs1d40Ckt61x2KCK0cbgdBeeO9duc7FSAHuhrFu9X1e7VvtGpTMD/Dniq+KRh8hpcqA9Z+HZJ8LxliSfMPJrqa5X4df8itH/ANdDXVVi9wGyfcb6GvB7vjULsf8ATZv517y/3T9K8Hvv+Qnd/wDXZv51dMCGikLAMF5LHoo6mul0DwZf6uyzXYa2te+eHNaNpCMGwsrrU7pbaxiMjscFgPlX616r4V8LQaDCZGPmXcg+dz2+lamk6NZaRbCGyhVOPmfHLfWr+KylK4wopaKgCve3UVnaSXMxOyMZOK8X8Ra9N4h1EzuSttGSIkr2m7t0ubaSCQBlkUjBrwrU7B9L1W4sZAQY24PqDVwsBvfD0FvFsZHaJq7TxH4Ms9V3T2uLe665XgOfeuV+GMBk1qefHEYK5r1MUSbTA8R1TR9Q0mYx3sB4/jQfL+dUa9t10QnRrozoroIyQD64rw+NtyZ9Sa1hK4D6KKK0AKKKSgQtJS0UDCkpaSgApaKKACiiigRLF92qsn+sNWYvu1Wk/wBYaEA2iiimAtFFFABRRRQAUUUUDCiiigQUUUUDLWl/8fw+lXNW/wBatU9L/wCP4fSrmrf65aT3Ao0UUUAFFFFABRRRQAUUUUAFFFFABRRSUAFLSUtACUUUUwCiiigBaKKKAClpKKAFopK19F8N6jrUoEcZhgz80jjHHtSbsBV0jSrjWr5bW2B25/eSdlFew6TplvpVhHaWygKo5PqfWo9F0a00azEFqgz/ABP3Y1o1zSlcBKSnVWvoppbSSO3fZIwwG9KgDiviD4m+zxHSbGT99IP3rD+AV5wuFGM//Xr0j/hW1rJK01zqM8kjnLMe9WI/h1pKfeuJm+taxkkB5iGX+8KeHT++v516tH4D0VOqs31q3H4O0NP+XNG+tP2iA8f8xOzA10PgrXG0rVxE4c2twcPjse1ekJ4a0VOmnQfiKnj0bTYiDHYwqR6LUyncC8DkZHfmloAAGBRWYEc0ayxNG4yrgqw9q8c8UaM2iau8QGLaY7oj6CvZ6w/FeiprOkSRY/fINyMOvHaqhKzA8eFOFJseN2jlXbIhwyntTq60AUUUUAIa6P4e/wDI0N/1yrnDXR/D3/kaW/651E9gPWKQ0tIa5QPOviqTtssEj5+xrgvmz/rH/M13vxVztsuCfn7VweRn7jflW8LWAAW/vt+ZpQW/vv8AmaNw/ut+VLkf3W/KrsgEy399/wAzRhv77fmadkf3W/KnDH91vyosgIyG/vv+ZprK3y/O33x3NTcf3W/Kmtj5flb747UmkB7lpA/4lVt/uCrtUtI/5BVt/uCrtczAgvR/oU//AFzb+VeED77/AO+f517vef8AHnP/ANc2/lXhOPnf/fP862pAKOtOpop1dABRRRQIKKKKYBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFTr/AKv8KgqZf9X+FAyp3NFHc0UxBRRRSAKKKKYwooooAKKKKACiiloEFFJS0AFFFFABmiiigZRooormEFFBooAKKKKACiiigAooooAKKKKACiiigAooooAltf8Aj7i+tbWrf8eyVi2v/H3F9a2tW/490pAZPalpKKACiiigAo70UUAFFFFABRRRQAU006mmgDX8I/8AI1W1e2H7xrxPwj/yNVt9K9s/iNZy3GjN8QWMupaLcWcGPMkQhd3SuH8KeBdW0fW4ry5MJjQc7Wya9KoAFSAveuc8b6Dc6/pcVtaMgdJNx3nAxXSUcUAcb4E8L33h6W4a8aMiQYGw5rsqKKAON+Jhx4fT/fry1FcoNsbsMdQK9S+Jv/IAj/36wvhte2kjPpV5bxuw+aN2HJPpVp2QjimguHXC2szEnoFr2rwlby23h21imQo4GSD1FaSWVqhyltGCPRasVLdxjT1rzL4qAnUNPAGSc4Ar0415Z8ULnGt2QjPzwkMfahAch9nu+hsrj/vmt/wRpt5L4ptpzayxxRA7mdcCvT9Glt9S0u3uxCh3oB09K0kRUGEUKPQCm5CAUOfkb/dNLUdywjt5XJwFQn9Km4zwvWGzr18f9uqueDTruX7RqFzOOjucfnUWeDWq2EeofC7/AJAlz/11rt64b4WH/iTXX/XWu5zWb3Gc/wCKPC8PiNYhNO8Xl9NvesnTPh7badqcV6l5KxjOdp7121FFwCqmq/8AILuv+uR/lVvNU9WP/Equ/wDrkf5UgOC+FikXV+ccZP8AOvSO3SvCtM1m/wBHlm+wMBvY5zWh/wAJr4gH/LUfnVOLEeyc+lHPpXjR8a+IP+ew/Omnxn4gP/LYfnRysZ7Nz6UoryPRfFuuXOuWVvPNmOSTDDPUV65SasAuaM0lcB4/8QanpOoQR2Em1W60kB6D+FH4V4t/wmviD/nsPzpR408Qf89h+dVysD2iuS+JA/4piQ44DCuHHjTX/wDnsPzqtqPiPVNUtTbXrgxHqM01F3Ayx0H0FdF4B/5G6H/cNc77V0PgH/kboP8AcNXLYR7DRRRWIzyn4oD/AIn1r/uGuOCu7LHGCZHO1QO5rs/iapk8Q2aLjcyYGfWtrwb4MSxKajqW2S4YZjQcqo9a0TsgNfwXon9i6IiOMTzfPL7GuhxxQOvNBPYVm2Bj+Kb9NN0C6mY4LIUX614ehZhvf7z8mu3+JetC8vE0q3fMcXzSY7N6VxeM1pFaCOr+Gf8AyNMv/XGvWa8n+Ggx4ol/6416zipluMRvun6GvG7Xw3f67fX8+nvFhbhgyuea9kf7rfQ15V4N1MWHja9gc4jupSgHvmhARj4f696wfnW/4P8ABt/pWrG+1CSPgfKqGu8HpS0nJgFFFBpAeV/E4g63bj0Arkcc10Hj27W78VSLG25IlAyPWsCto7CCiiirAKSiigBaKSigAooooAKKKO9AC1LD92oqlh+5QBVk/wBY31ptOf77fWm0AFFFFMAooooAKWiigAooopAFFFFAARW1ov8Ax6PWJW1ov/Hq9DAo3P8Ax8yfWo6kuP8Aj5k+tR0AFFFFABRRRQAUopKKYC0UUUAFFFFABRRRSATFGKWigBMUYpaKACikozQAjfcb6GvWvh5/yKlt+NeSN9xvoa9b+Hn/ACKlv+NZz2BHT0UUVkM8o+In/IyR/wC6a5bNdP8AEU/8VJH/ALprlxW8NhDqSiirAMUmKWigAooooAa1bvgf/ka4voKwjW74H/5GuP6VEtgPY+5rM8R2M2paDd2Vvt82VMLu6Vp9zRWAzzrwZ4K1TRdaW7vPJ8sDHytk16PSCloYHL+OfD91r9hBDZ+Xujck7zioPAnhm88Prc/bDHmXGNhzXX0UXAK4n4nHGjw/71dtXEfE/H9jwg92px3A84CSFQywyMpH3lHFM8m4kdFS1mY7xwFrtvhvdwSNNpl1Gj9DFuHX1r0KOytYm3R28an1Aq3LoBFpMbRaXbI4wwQZBq6aSis2B5d8TI5Jdfs0hRpHKH5VGTXKfY73/nwuf++K6zxlqhtfHFnPEebchH+hr0mAw3EKTIqlXUEHFac1kB5p8O9Jvf7f+3TW8kMMaFfnGCTXqYpqjAwBgU+obuBW1DjT7n/rm38q8JkP7+U/7Zr3HWZVg0i7kYgARN/KvDFbcN/945q6YD60dO0LUdVt2nsFR0U4IPWs8etdH4E1M6frwhdsQ3A2gf7VaSdkBTPhTXQf+PYflSHwnrzAhbZcn1Fez0VlzsDF8KaVLpGhxWs5BkzubHY1tUUhNQA2Q4jY+gNeJ2Wmz63r11bW0kaMZm5c4zz2r17W7tbLR7q4c4EaE14bHNKlwbmBykvmGRWHUVpAD1zQfBmm6UBJKv2mf+9IM7T7V0wArmPBviePXLMRTEJewjDr/eHqK6cGod+oC0UyWWOGMySuqIOrMcAVwHifx6BvstF+Z+jzHoPpQk2B3A1GzN79jWdDcdSgPIq1Xg1hqNzYaxFqnmNJOrfOzH7wPWvb9NvoNRsYru3YNHIMg+9OUbAWSK5fxd4Rh16MTQt5V5GPlbs31rqaCM1KdgOW8EeGpvD9pP8Aa2Vp5mydvQV0/SlxikNO9wOd8d3n2TwxOwOC7BfzryJPlUL6V3XxRvwxttORs7hvYDtXCA81tTWgElLSClrUAooooAKKSloASloooAKKKSgBaKKKAJIvu1Wk/wBYasxfdqtJ/rDQgG0UUCmIWiiigAooooAKSiloGFFFFABRRRQBa0z/AI/R9Kuat/rVqnpv/H6PpVzVf9ctJgUaKKKACiiigAooooAKKKKACiiigAopM0tABSUUUwClpKKAFpKKBQAtFFFABRRRQA1huUr616t4D1T+0dCRGwJIDsI9q8pPWum+H2ofYvEDW7n5LhQqj3rKoroD1eiiiucAoxRketFACYoxS/hR+FACUtRtLGv3nVfqaha+tE+9dQj6sKLAWqKz31nTI/v6hbj/AIGKqTeKtGh63kbf7pp2YG1RXKzePdDiJy8jf7oqz4f8Wadr1zLBah0kToH43CizA6KikFLSA8z+IOi/Y70apAv7qY/vcf3q5GvbdXsItS02a1mAKspwT2PrXis8DWtzLbOwcxNt3Doa6KUr6AMooFGa2Aaa6P4fH/iqW/651zjGuh+Hv/I0t/1zrOewHrRpKO9LXKBXurG1vMfardJdvTcOlV/7F0v/AJ8IP++a0KM0XAz/AOxdL/6B8H5Uf2Jpf/PhB+VaFHFFwKH9i6X/AM+EH/fNL/Yumf8APjB/3zV+ii7Aof2Npv8Az4w/980f2Npn/PjD/wB81foouAiIqIFQAKOgFLRRQBDef8ec/wD1zb+VeEH77/75/nXu95/x5z/9c2/lXhH8cn++f51tS3AUdaWkFOroAKKKKYgooooGFFFFABRRRQIKKKKBhRRRQIKSlooAKKKKACiiigAooooGFTL/AKv8KhqVf9X+FAFXuaKO5opgFFFFABRRRQAUUUUCCiiigBaKSloAKKKKBhRRRQIKKWimMoUUUVyiCiiigAooooAKKKKACiiigAooooAKKKKACiiigCW1/wCPuL61tar/AMeyVhwHFzGfetzU+bVTSYGTRRRQAUUUUAFFFFABRRRQAUUUUAFIRS0UAa3hAZ8VW2PSvbMcmvBLG8m0+9S7tseanTPSt8+Pdd67YqmSbA9O1rURpOmy3jRlxGMkCuKHxQgIBFm/PtXO3/jHWNQtHtZ/LEbjBxWAMgAA8Cko9wPQv+Fn2/8Az5yflSj4n2vezk/KvPNx9aTcfWnyoD2Lwt4tg8RyzRwwPGYhk7q6T8K+f7a9vLNy1ncyQM3BKHrVg65rP/QTuP8AvqpcRnonxPP/ABIIv+uleb6JenT9ZsrsHAjYbqZc6jqF5GI7u8lmQchXPAqsVyMVSWgj6FhcSwpIOjqG/On147B4512CCOGMRFYwFBPXFJL431+Vcb0T3U1PKxnquq6raaVatcXkqoFGQueW+leKazqMms6rcX8ox5nCr6DtUV5eXd/J5l9cyTsOgY8CoapRsI9R+Gd95+gG1Jy1u2D+NdnXhmj63f6I8rWBX9794N0rSbx1r7DrGPoaTi7jPXydoJPA9TXEeOfFdva2L6fYSiS6l4JU5CjvXD3niPXL1Sst9JGp6hD1rJ2/MWJJY9WPehRENQbVwPrTj0NKBQRVgemfCwH+xbo4/wCWtdxz6V4RYaxqWmxNFY3DRoxyQDVv/hKte/5/X/OocWM9F8XeLf8AhG5oIzbmTzVzmsAfE9e9k1cXqGp32pur30xkZBhcnpVXLZ601ER6CPidH3s2pl18R4bm0mg+yODIhUHHrXBAn1pwJ9afKgEOSzH1OaSn0lVYBuKMU6ilYC7oAx4j08/9NK90GMCvBbK5NlfwXaruaFtwB712P/Cyrn/nzSpkmM9JOK8u+KIzqltVj/hZd1/z5pXOeI9ek8QXEU0kQjMfYUlF3EY2OaWnY5oxWlgEpwopRQAV0PgH/kboP9w1z1XtG1SXRtSS+hRXdQRtbpSewHueR60ZHrXmP/CydR/584fzo/4WTqH/AD5Q/maz5WMr/E4k6/bYOCEPI7VF4X8a3OkulrqTNPaE4Dn7yf8A1qydf1mbXr5Lq4jWNkGAF6VmEdqtR0Ee+Wl3Be2yXFrIskTjgg1g+MfEsWh2DJEwa8lGEUdV96840DxHf6B5i2p8yKQYEbHhT6is26uZ726e5u5Gklc9SensKSjqBEzSSyvLMxaSQ7nY9zSgUAUtWB1Xw3KJ4llZ3CjyepNep/aYP+e0f514IkkkbbopGjb1XrTjd3v/AD+zfnUONwPeGuICD+/j6eteE3cxtdenuoz80NyWBH1phu73/n9m/OoHG8Hcck9T601GwHvem3K3enW84bO+ME/WrYIrx/TPHOqaZYR2cNvC6x9GbrVl/iNrbDC21uKnlYz1j8K5vxT4ptNGs3WKRZbtwQiKc4PvXnF54u128yGuTCD/AM8zWOzNJIZJXaSQ9WY8mmoADvJLK8spy8jFifrTqSlrQQUUUUwEooooAKKKKACiiigAooooAWpYvumoqli+7QBVf77fWm0rH52+tJQAUUUUwClpKWkAUUUUwCiiikAUUUUwCtnRv+PVqxa2dGP+juKTApXP/HzJ9ajqW6GLqT61FQAUUUUAFFFFABRRS0AFFFFABSUtJQAtFJS0CCiiimMKKKKQBSEUtBoAY/3G+hr1r4d/8ipb/jXkxGQR61s6X4q1XSbJLO0WMxJ03daiSuB7RmkryP8A4T3Xv7kNKPHuvZ+5DUcjGO+Iv/IyR/7tcwKuarqd1rF4Lq9CiQDA21UArWKshBRS0VQCUUtFACUUtJQA01u+Bv8Aka4/pWGRVrStRn0nUFvbZFeRezdKmSA91zyapaxqSaVpc966llhXJA715z/wsXWc/wDHtBVXU/G2qanp8tlPBCscowxXrWXIxm8vxOtWAP2STn2p3/CzLb/n0k/KvO1BVQAeAMU7J9avkQj0P/hZlr/z6SflW94Y8UQ+IvP8mF08nGdw6149k+tT21/e2Zb7HdSQbvvbD1pOAz3r8K4b4o5/sq34/iNcL/b2tf8AQTn/ADqC71HUL5FS9u5JlXoGPSkoMCbw/fHTdes7rPyg7T+Ne4qdyhh3Ga+fvQ91IIroF8a68kaorR4UYBzTlG4HsXPpVDV9WtNJtHnupVXaMqmeWNeUyeMNflGDOE91NZN1dXV9J5l7cyTt23npSUGAaneSanqF1fS8NMcgelet+Brz7Z4atgTl4htavIK1dJ8Q6lo0Lw2JUo5yd3aqlHQD2yms4RSzkKo6k9BXkTeN9fYY3Rj3BrOvNe1i9UrPfShD1VTwaj2bA6nx74mjuY/7J0+QMM5kkU8fSuGxjgdKQADp3606toxsgF6UqTNbzR3CfehYMKSmnoQabQHuekXi32l29yGB8xAT9au5HrXj+i+Mb/RbH7JDDHKgOQX6irp+I+rdrSCsHBgepZ9KjmljhjLzOsaDqzHivKZ/iBrkqkLFAme4rEvtX1TUci7vZCh6xg8U1BgdD468ULqjjTdPY/Z4z+8f+/7VyG3AAHSlVQBgDAp2K0UbCH2d3cafeR3lo5WaM5/3vrXpDfELTotKjnZHa6YYMYHQ15pikwAc459aHFMZq634l1PXXPnyGGA9IkPBHvWSqhRhRgU4CnCmlYBuK6Lwh4mfQbnyLjL2Mp5H/PM+1YBpCO1DVwPebW6gu4FmtpFkjboVOamzXhumaxqOktmxuGCdfKJ+Wulg+JF+oAubOIn1WsXBgem5HrVLVdRt9LsnurpwqqMgE/ePpXAz/Ei8ZcW1nHu/2q5bVNWv9Xl8y/mZh1EYPyimoMCLVdQl1bVJr6XI8xvkU/wiqw60oFKBW6VgFFLRiimAtJRS0AJS0lLQAUUUUCCiiigYUUUUgJIulVpP9Yasx/dqq/3zTQCUUUUwFpKWigBKWiigBKKKWgBM0tFFABRRRQBa03/j9H0q5qv+tWqWnHF6tXtVHzKaT3AoUUUUAFFFFABRRRQAUUUUAFFFJQAtJRRTAKKKKACiiigAoFFLQACiiigAooooAQ0+0na1v7a6TrC+6mUlJq4HqEnj/R4IkLmV3KgkKM81nT/Eyzzi3tJT/vCvPtoHQUvPrWXs0B2cvxJvD/qLKP8A4FVOX4g65JwkECfSuX5oxT5EBty+Mdfl/wCWwT/dNU5df1yX72pTD6GqOKMU+RASPf6lIfn1GdvxqBnnf79xI31NPxRinygQmJT97J+poEMf9wVNilxT5QIdijotW9KvpdL1SC9hJBRgrf7veocUFQQR6jFDiB7vZ3Md3aRXERyki5Wqmsa5YaPAZLyYAjpGD8x/CvNLDxhf6doy6dbIpKjCu38IrCuZ57uczXczzSH+Jj0rFU3cDode8aX+qlorQm2tj0K/eYe9c5k9zk9ye9J1pRW0YpALRRSVYCGui+H3/I0t/wBc6501f0LVn0TUjexxiQlduDUTV0B7bQTXnI+JE/8Az5rR/wALIn/581rn5JAdR4k8UWvh4RfaY3fzDgbKw/8AhZWm/wDPvP8AlXK+JvET+IRCHhEflHOR3rEwc1pGnpqB6L/wsrTv+fef8qUfErTf+fef8q855peafs0B6OPiVpn/AD7z/lTh8SdL/wCfef8AKvN+aOaPZID0j/hZGl/88J/yoPxK0oY/cT8kDpXmxJpDk49jmj2SA96tbhLm2jnThXGRmpdwryq28f6pa20cCW0JWMYBPepP+Fjat/z6wVl7Ngel3hH2Ofn/AJZt/KvCM/O/++f511cnxD1SSJ42tYcOpU/jXKD7xb1JNa04tbgOFLSClrYBaKKKYBRRRQAUUUUAFFFFAgooooAKKKKACkpaKACiiigAoopKAFooopjCpV/1f4VDUy/6v8KQFXuaKSlpgFFFFABRRRQAUUUUCCilooAKKKKACiiigAooooAKKKKYFGiiiuUAooooAKKKKACiiigAooooAKKKKACiiigAooooAAdrK3oa6G6Hm2GR/dFc6elb2myCex2nqODSYGWOlFLKhjlZT2NNoAWkxS0UwCkpaMUAJS0UYoAKKKWkAlFFFMANJilopANopaKAEopaKYDcUYp2KMUgG4pcUuKKYCUuKWigBMUYpaWgBuKKdSUANpcUtFACUYpcUUANxSYp9JigBuKXFLS0AJilFFFAC0d6SloAKKKKAEoyaKMUAGaKKKYBilooFABRRRQAUlLRQA2jmlxRigBKKdRQAmKMUtLQA3FGKdSUAJikxT8UUAMxRinUUAMxS4pcUUAJilFGKXpQAClpKKYBRRRQAUtJRQAUUUtACUUtJQAUUUtABUo+WMn2qIDJFOnbbHj1oArE8k+tFFFABRRS0wCiiigAooooAKKKKACiiigArS0V+ZE9azasafL5N2pPQ8GkBZ1Bdt0T/eqrWnqceUWQfw1m0IBKKKKACg0UUAFLSUtABRRRTASilzSUAFKKSigApaKKACiiigAooooAbSEU+kxQA3FGKdijFIBBSilxRTAKKKKACiiigBKWiigBMUmKdRigBmKWlxRSATFLS0UwExSYp1FADcUmKfRigBmKMU/FGKAGbaUCnYooGJikxT6SmIbilxS4paAG4paWigYUhFLRQIZikxT8UYpAMxS4p2KAKYxMUtFLQAmKMUtFAgo7UUUDCkNLRQA00c0uKMUAN57mlxS4oxQAYopaBTAKWikoAKKKKAFoo70UAFFFFABRRSUALRRQBk4pASLwlVCckmrMx2x49arU0AUCigUwCiiloAKKKKAEpaSloAKKKKACiiigCW2bZdRt71r6ku+AMOxzWHnHI6it5CLixGOflx+NSwMmilIIJU9RxSUwCiiigApKWigAooooASig0UwCiiigAooooAKKKKAClFJS0AFFFFABRRRQAUhFOpKAExSYp1FACYoxS0UAJijFLRQAUmKWigAooooAKTFLRQA3FGKdRQAgFFLRTAKKKKAExSYp1FIBvNHNLilosA3HrS4FLS0AJikp1FMBKSlooATFIRTqKAGYpMU/FGKLANApQKXFKBQAlLS0UwCiiigAooooEFFFFAwooooEFFFFABRRRQAUUlLQAUUUUAFJRS0DEpaSlpgFSOdsRpiDLUXDcBaQEFFLRQAUUUUwCiiigQUUUUAFLSUtAxKWkpaACiiigApaSimIWikooAo0UUVygFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABVzSp/JuNjHCvxVOjpyOooA2NShz++UfWs+tWxuFu7fa/3wMMKoXVubeTH8B6GkBDS0lLTAKKKKAClpKWkAlLRSUALSUUUAFFFFMApKWigBKKWigAooooAKKKKACiigUAFLRRQAlFFFABRRRQAUUtJQAUUUUAFFFFABRRRQAtFFFACUtFFABRRRTAKKKKACiiigAooooAKKKKACiiigAooooAWkopaACkoopgFLSUUALRSUUgCiiigAooopgFFLRQAlFFFABRS0UAJS0UUAFFFFACUtJTkG4+1AD417moJ23SewqaZ9i4HU1V+tABRS0lMApaKKACiiigAooooAKKKKACiiigApOmCOo5paSgDetpFu7PnrjBrLkjMcjI3UUWFybefk/I3WtO9txMgkj+8B+dIDKoo9qKACiiigApaSloAKKKKAEpaKKACiiigAooopgFFFFABRRRQAUUUUAFFFFAwooooEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFLSUtABRRRQAUYoooGFJS0gpgFLQaSgBaKSloEFFFFABRS0lABRRRQAUUUUAFFFFAwooooAKKKKACiiigAooooAKKKKACiiimAUUUUAL2oopKAFooooAKKKKQBRRSUALT4x3pqjcaWV9i4HU0wIZm3PjsKZRRTAMUUtFAgooooGJS0UUAFFFFABRRRQAUUUUALWjpE+C0DH3FZtOR2Rw69VOaANC/g8uTeB8rdfrVWtaN0vbX68H2NZksbQyFG/A1IDKKKKYBRRRQAUUUcUAJRS0lMAooooAKKKKACiiigBaKBRQAUUUUCCiiigYtJS0UAJRS0UAAooooASloooAKSlooASilooASilooASilopgFFFFABRRRQISilooATFLRRQAUUUUDCiiigBKKKKACiiigAopaKAExRS0UwEpaKKACiiigAooooAKKKKACiiigAooooAKKKKAEpaKKACkpaSmIKWkpaBiUtFORc8mkA5BtXJqu7bnJqWZ8fIPxqCgBaKSl7cUwHpDLIjPHGzKvUjtTK63RrXy9KKdPPHNcvdReRdSxY+42KlSuwIqKKKoQUUUtACUtJRQMWkpaSmAUtFFIQUUUUDCilopgUKKKK5RBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABSUtFABRRRQA+GV4JRIh5H61uRSxX0Hv3HcVgU+KV4XDxnB/nQBcuLV4Gz1Tsah+laVtfRXK7ZMK3cHvST2AJ3QnHsaQGdRT3ikjPzqR70zIoAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFLSUAFFFFABRRRQAUUUtACUUUUAFFFFABRRS0AFFFFABRRRTAKKKKACkpaKACiiigAooooAKKKKACiiigBaKKSgAopaKACkopaAEooopgLSUUUAFFFFABRRRQAUUUYoAKKKKACilooAKKKKACigAnoKesfc0ANVd30p7sIl9+wpHlVBheTVZmLHLUADMWbJopKWmAUUUUAFFFFABRRRQAUUUUAFFFFACUUtFABRRRQAlaWm3uzEEx4/hNZ1JQBtXlnvPmRYDenrWcQQcEYI7VPZaiYwI58lezelXpYYbpNwIPuKkDJoqeWzlj5A3D2qA5B+YYpgLRSUtABRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooGFFFFAgoopKAFooooAKKKKACiiigAooooAKKWkoAKKO1FAwpaSloEFFFFABRRRTGJS0UUAFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigYUUUUAFFFFABRRRQAUUUUAFFFFABRRRTAKKKKQC0UUUwCiiikAUUlKAT0oAKUKW+lOCdzSPKqcLyaAFZhGvvVZiWbJNDEsck0gqkAUUUtAhKKKWgApKKKBhRRRQIWiiigYUUUUAFFFFAhe1JRRQBNa3LWsu4cqeorYdIryAEH6H0rBqa2uZLZsryvdaTQyWaJ4W2uOOxqOtWKaG6jxwc/wnqKrzWBGTEePQ0gKdFKyOhwykU3imAtJS0UAJRS0UwEopaKAEopaSgAooooAXFFFFABRRRQAUtJRQAUUUUAFLRSUAFLRRQAlFLSUwClpKKACiiloASlopKBC0UUUDCikpaACikpaBBRRRQAUUUUAFFFJQMWiiigBKWiigAooopgFGKKKACiiigBaSiigAooooAKKKKACiiigAooooAKKKKYBRRRSEHekpaKBiUtFJQAUtKFJp4QDk0ANVM8miSQIMDrSSTADCVB15PWgA6nmiiigAq1p1s11epEvrk/SqwBJAUZJ6Cut0Sw+x2od1/eycnPUe1KTsgNFVCKFT7qjArnvEloVlW5RfkPDfWujqK5gW5t2hfow49qyTswOFpKnu7Z7S4aFwRjofUVBW4gooooAKKKKACiiimAtFFFIAooooAWikzRTGUaKKK5RC0lFFABRRRQAd6KKKACiiigAooooAKKKKACiiigAooooAKKKKADvnofWrlvqMsI2t86/rVOloA247+3lHzYB/2qeY7WXoV/CsHigFh91iPpSsBuGwgPTNJ/Z8Pv8AnWOJpR0lb86Xz5v+erfnRYDX/s+H3/Ol/s+H3/Osbz5v+erfnR583/PVvzoA2f7Ph96P7Ph9/wA6xvPm/wCerfnR583/AD1b86LAbP8AZ8PvR/Z8PqfzrG8+b/nq350efN/z1b86LAbP9nw+9H9nw+prH8+b/nq350efN/z0b86dgNj+z4fej+z4fU/nWP58v/PRvzo86X/no350AbH9nw+9J/Z8PqayPOl/56N+dHnS/wDPRvzosBr/ANnw+9H9nxep/Osjzpf+ejfnR50v/PRvzosBr/2fD6mk/s+H1NZPnS/89G/Ojzpf+ejfnQBrfYIfU0fYIfU1k+dL/wA9G/Ojzpf+ejfnRYDW+wxeppPsMXqayvOl/wCejfnR50v/AD0b86LAav2GL1P50fYYvU1lebL/AM9G/OjzZf8Ano350AapsYvU0fYYvU1lebL/AM9G/OjzZf8Ano350WA1TZReppPsUXqayvNk/wCejfnS+bJ/z0b86LAan2KL1NH2OP3rL82X/no350edL/z0b86dgNT7HF70n2OP3rM82X++350ebJ/z0b86LAaf2SP3o+yR+9Znmyf32/OjzZP77fnRYDS+yx+po+yx+prN82T++aPNk/vn86LAaX2WP3pPssfvWd5kn99vzo8yT++aLAaP2WP3o+zR+9Z3mSf3zR5kn980WA0fs0fvSfZkrP8AMk/vmjzH/vmiwGh9mj9TR9mj96z/ADH/AL5o8x/75osBofZ4/ej7Onqaz/Mf++aN7/3zRYC/9nT3o+zp71Q8x/75o3yf3zQBf+zp70eQlUN7/wB80b3/AL5oAv8AkJ60nkJ71R3v/eNG9/7xoAveQvvR5K+9Ud7/AN80b3/vGiwF3yVo8lapb3/vGje/940wLnkr70eSvvVPe/8AeNG9/wC8aALnkrR5K1S3v/eNLvf+8aALnlL70nlLVTe/940b2/vGgC35S0eWue9VN7f3jRvb+8aALflL70eUtVN7/wB40b2/vGgC15Yo8sVV3t/eNG9v7xoAtbFpcIvXFVNzeppDk9zQBaMyL05qF5mbgcCo8UUAFFLRQAlLRRTAKKKKACiiigApKWigAooooAKKKKACiiigAooooAKKSigA+tSQzSQnMbH6dqjpaANSHVFPEyke4qyJLWYdUrCoHHTilYDdNnbt0/Sk/s+H3rEEkg6SN+dO8+b/AJ6t+dFgNn+z4fej+z4fU1j+fN/z1b86PPm/56N+dFgNf+z4fU/nR/Z8PqayPPl/56N+dJ503/PRvzosBsf2fD6n86T+z4fU1kefN/z0b86Xzpf+ejfnRYDW+wRepo+wRep/Osnzpf8Ano350edL/wA9G/OiwGt9gh9T+dH2CL1NZPnS/wDPRvzo86X/AJ6t+dFgNb7BD6n86PsEPqfzrI86X/no350edL/z0b86LAa/2CH1P50fYIvU/nWR50v/AD0b86POl/56N+dFgNf7DF6mk+wxe9ZPnS/89G/Ojzpf+ejfnTsBrfYYvU0fYYvesnzZf+ejfnSiaX/no350WA1fsMXvR9ii96yvOl/56N+dJ5sv/PRvzosBrfYovej7FF71k+bL/wA9G/OjzZf+ejfnRYDV+xRe9H2OL3rK82X/AJ6N+dHmy/8APRvzoA1fsUXqaPsUXvWV5sv/AD0b86PNk/56N+dAGp9ji96PscXqazPNl/56N+dJ5sn99vzpgan2OP3o+xx+9Znmyf32/OjzZP8Ano350gNP7JH70n2SP3rN82T++aPNk/vt+dFgNL7JH70fZI/es3zZP77fnR5sn99vzpgaX2WP3pPssfvWd5kn98/nR5kn99vzoA0fsiepo+yx+9Z3mSf3z+dHmSf3z+dAGj9lj96PssfqazvMk/vn86TzJP75oA0fsye9H2ZPes/zJP75/OjzJP75osBofZk96Psye9Z3mSf3zR5kn980WA0Ps6e9H2dPes/zJP75pfMf++aLAXvs6epo8hPeqPmP/eNJ5j/3zRYC/wCQnvR5C+9UPMf++aPMf++aLAX/ACE96TyF96peY/8AfNJ5j/3jTsBe8lfejyF96peY/wDeNHmP/eNKwF3yVpPJWqe9/wC8aN7/AN40AXfJWk8lap73/vGje/8AeNOwFzyVo8lap73/ALxo3v8A3jSAueSvvSeUtVN7/wB40bn/ALxoAt+UtHlLVTe/djRuf+8aALflCjyxVTc3940bm/vGnYC35YpPLFVdzf3jRub1NFgLXlijYKq7m/vGjc3qaLAWvkHUimmZF6c/Sqx+tHSiwx7ys3sKZRRTEFFFBoGFFFFAgpaKKACiiigYUUlLQIKKKKBh3ooooAKKKKBBRRRQAUUUUDFBKnKkg+1XINSkTCyjcPUdapUUAbaXltKPmwP96l8i1k5BH4Vh4FKGYdGYUrAbX2GA9CaP7Ph9TWP5kn/PRvzpfOl/56t+dFgNf+z4fU0n9nxeprJ86b/no350edN/z0b86LAa/wDZ8Pqfzo/s+H1P51kedN/z0b86POl/56N+dFgNf+z4fU0n2CH1P51k+dL/AM9G/Ojzpf8Ano350WEa/wBgh9TSfYIfU/nWT50v/PRvzo86X/no350WA1vsEXqaPsMXqfzrJ86X/no350ebL/z0b86dhmr9hi9TR9hi9TWV5sv/AD0b86PNl/56N+dFgNX7DF6mj7DF6msrzZf+ejfnSebL/wA9G/OiwGt9ii9TSfYovU1l+bJ/z0b86PNk/wCejfnRYDU+xReppPsUXvWZ5sv/AD0b86TzZP8Ano350WA1fsUXqaPsUXqay/NkP/LRvzo82T/no350WA0/sUXqaPscXqazPNk/56N+dHmyf89G/OiwGn9ji9TR9jj9TWZ5sn99vzo8yT++350WEaf2OPvmj7HH71meZL/z0b86PMk/56N+dFgNL7JH70v2OP3rM8yT/no350ebJ/fb86LAaX2SP3o+yJ71m+ZJ/fb86PMk/wCejfnRYDS+yR+9H2SP1NZvmSf32/OjzJP75/OnYDR+yx+9H2WP3rO8yT++350eZJ/fNIDR+yx+9J9mj96z/Mk/vmjzJP75p2A0Psyepo+zJ71n+ZJ/fNHmSf3zQMv/AGaP3o+zR+9UPMf++aN7/wB80AX/ALOnvR9nT3qh5j/3zS+Y/wDeNAF37OnqaPs6e9Ud7/3zS+Y/940WAu/Z09TR5CetUt7/AN40b2/vGiwFzyE96PIWqW9/7xpd7/3jTAueQvvSeQtVN7/3jSb3/vGgRc8laPJWqe9/7xo3v/eNFhlzyVo8lap73/vGje/940AW/KWjylqpvb+8aNzf3jQBb8paPKWqm5v7xo3N/eNFgLXlrR5a1V3N/eNG5v7xoAteWtHlrVXc3940bm/vGmBa8taPLWqu5v7xo3N/eNICz5Yo8sVW3N/eNG5v7xoAs7Foyg7iq2Se5opgTtMAPlFQs7N1NJSUAFFFLQAlLznGMk9BUttazXT7YELe/pXSaZosdriWfDy/oKltICvoekmNhdXK8/wqe1b9FFZN3AKKKKQFHVNOS/gx0kX7rVyM8EtvKY5kKsP1rvKqX1hBfR7ZR8w6MOtVGVgOKoq9f6Xc2bnK74+u4dBVHrWqdxBRRS0wEpaSimAtFJS0gCiiigAzRRRTGUqSiiuUQUUUUAFFFFAB3ooooAKKKKACiiigAooooAKKKKACiiigAooooAKBRQKAFoopKBC0lFFAwopaKACkopaAEpaSloAKKKKYC0UUUCCiiigAooooASloooAKKKKBhRRRQAUUUUAFFFFABRRRTAKKKKACiiigAooooAOtLSCloASilpKACloooASiiigApaKKACijFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJRTAWkpaKACkoooAWiiikAUUUUCCiiimMWiiikAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigBKWiigAooooAKKKKACiiigAooFFABRRRQIKKKKBhRRRQAUUUUwCiiigAooooAKKWk70ALRSUtACUtJS0AJRRRQAtJRRQAUUUtABRRRQAUUUd6AFooooAKKKKACiiigAooooEFJRRTGFFFLQAlFFFAC0UlFABRRS0AJS0UUwCiiigAooooAKKKKQBRRRQAUUUUALSUUtABRSUtABRRRTAKKKKACiiigAooooGFFFFAgoNFFAwpaSloASloooEFFFFABRRRQAUUUUDCiiigAooooEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAC0UUUwCiiigAooooAKKKKBhRRRQIKKSloAKKKKACiiimAUtFFABRRRQAUUUUAFFFFABRRRSAKKKKACiiigAooooGFFFFABS0UUwEopaKBBSUUtACUtFFABSUtFABRRSUALRRSUwFopKKQC0UlLQAUUUUwCiiigAooooGFFFFAgooooAKWkpaBhRRSUAT2Vq95dLbxsFZucmuhtvD1vGQ1wxdh6HiuYVmRtyMVI7itG11q8t8DcGXvnrUyT6AdZFFHEAIkVPoKfWZZa1bXOFkPlP0+bvWlWTTW4C0UUUgCiiigAooooAQgMMMAR6HpWddaLZ3BLbSjf7PSrF5f29mpMrgt/dHWsO68QzyEi2UIh/vdaqKfQCtqmlPpyrI0isjnAA6is+nzTyztmWRm9ieKjrZX6iClpKKYC0UUUgCiiimAtFJmigCjRRRXKAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUCiigAoooFABRRRQAUCigUALRRRQIKKKKBhSUtFACUUUUAFFFLQAUtIKKYhaKKKACiiigAooooAKKKKBhRRRQAUUUUCClpKKBhRRRQAUUUUwCiiigAooooAKKKKADtS0UUAFFFFABRRRQAUUUUAFFFFABRRRQIKKO1FAwooooAKKKKACiiigAooooAKKKKACiiigAFLSUtMAooxRSAMUUUUxhRRRQIKKMUtACUUtFABRRRQAUUUUAFFFFABRRRQAlLRRQAUUUUAFFFFIAooopgFFFFABRRRQAUUUUAFFFFABRRRQACiiigAooooAKKKKACiiigAooooAKKKKYBRS0UAJRS0UAFFFJQIWkpaSgYUUUUAFFFFABSikpaACiiigAoFFLQAUUUUAFJS0UAFBoooAKKKKACiikpgFFLRQAlLRRQAlFFLQAlLSUtABSUUtABRRRQAUd6BRQAUUUUAFFLRQISloooGFFFFABRRRTAKKKKACiiigAooooAKKKKACiiigBaSiigYtFFFABRRRQIKKKKACiiigAopaKBiUUd6WgBKKKKBBRRRQAUUUUAFFFFABRRRQAUUUUAFFFLQAUUlLTAKKKKACikpaACiikoAWiikoGLSUUtABRRRQIKKKKACloooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKBhRRRQAUtJS0wCiiigAooooASloooEFFFJQAtJRRQAtFJRQAtFJS0AFFFFMBKKWigAooooAKKKKACiiigAooooAKKKKBhS0UUAFFFFABRRRQAfzrV0zWZbZhHOS8PcnqKyqKTVwO9hljnjEkTBlPcU+uP0nUXsZwrEmFjgj0rrkdZEDoQVYZBrKUbAOoooqQDp1rD1XWxGTBaEF/4n7Cm67qm0G1tm+b+Nh29q52rjHqwHO7SOWkYsT60lJRWogpaSimAUUtJQAtFFFIYUUUUCCiiimMo0UUVyiCiiigAooooAKKKKACkpaKACiiigAooooAKKKKACiiigAooooAKUUlLQAZoopKACloooAKSlooAKKKKACiiigBaKKKYgopaSgAooopAFFFFABRRRQMKKKKYBRRRQAUUUUAFFFFABRRRTAKKKKACiiigApaSigBaKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKYBRRRQAUUUUgCiiigApaSloAKKKKYBRRRQMKKKWgTCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKQBRRRTASloooAKKKKACiiigAooooAKKKBQAUUUUAFFFFABRRRQAUUUUAFFFLTASlpKKAFopKWgAooxRQAUlFFABRRRQAUUUtACUvaiigAFFFFABRRRQAUtJ3paACiiigAooooAKKKKYBSUtJQAUUUUALRRSUAFLSUUALRRRQAUUUGgAooopgFFFFIAopaSgApaKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUUAFFFLQMSl70lLQIKKKKBhSUtJQIWikpaACiiigYUUUUAFLSUtACUUUtACUUUtACUUUUAFFFFABRS0UAJRRRQAUUtFACUtJS0xBSUtJQAtFJS0AFFFFAwooooABRRRQAUUUUCClpKWgAooooAKKKKACiiigAooooAKKKKACiiigAooozQAUUUUDCiiloAKKKKYBRRRSAKKKKYBRRRQIKKKKAEpaSloASlopKAFopKWgYUUUUAFFFFMAoopaAEopaSgQUUUUDCiiigQUUUtABRRRQAUUUUAFFFFABRRRQAV0Hh2+OTaSt7qa5+nwyGGZJQcbTk0pK6Gd7VLVb1bKzZ+rtwo9PerNvKLiCOYcBxmuV126NxflQflj+XFZRV2BnMzMxZzlmOSaSiitgCilpKAClpKKBBRRS0AFFFFAwooooEFFFFMCjRRRXKAUUUUAFFFFABRRRQAUUd6KACiiigAooooAKKWigBKKKKACiiigAoopaACiiigAooooAKKKKACiiigAoFFFAC0UUUCFpKKKACiiigAooooAKKKKYwooooAKKO9FABRRRQAUUUUwCiiigAooooAKKKKAClpKWgAooooAKKKKACiiigAooooAKKKKACiiigAooopgFFFFABRRRQAUUUUgCiiigAooooAKKWkoAKWkpaYBRRRQAUUUUDClpKWgTCiiigAooooAKKKKACkpaKACiiigAooooAKKKKACiiikAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiloASiiigAoopaAEooooAKWkopgFFLSUAFLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUtACUtFFABRRRQAUUUUAFFJRTGFFFFAhaSlooAKSiigAoopaACiiimAUUUUDCgUUUgCloooEFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACikpaACiiigAooooAKKKKQBS0UUwEpaSloGFFFFAhKKKKACloooGFFFFABRS0UAJS0lFAC0lLRQAUlLRQAUUUUAFFFFABSUtFABRRRQAlLRRQAlFLRTEFFFFABRRRQAUUUUDCiiigAooooEFFFFACiiiigAooooAKKKKACiiigAooooGFFFFAgooooAKKSigYtFFFABS0UUAFFFFMAooooAKKKKBBRRRQMSilpKBBS0lFAxaKSimAtFFFIAooooAKKKKYC0UlLQAUUUUhBSUtFABSUtJTGLRRRQAUUUUCCiiigAooooGFFFFAgo60UUAb+lanDDpksc0m2RRhB61hO5kdnPVjk03juKKSVgCiiimAUUUUxhRS0UAJS0lLQIKKSlpDCiiigQUUUUxlGiiiuUQUUUUAFFFFABRRRQAUUUUAFFFFABRRS0AFJRRQAtJRRQAUUUUAFLSUtABRRRQAUUUUAFFFFABRRRQAUtAooEFFFLQAlLSUUAFFFFABRRRQAUUUUxhRRRQAUUUUAFFFFAC0UUUwEopaKAEopaKAEooooAKWkpaACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKWkooAKKKKAFpKKKAClpKKAFpKKWgAooooAKKKKYBRRRQMKXpSCg0CFopKWkAUUUUwCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAAUUUUAFFFFABRRRQAUUUUAFFFFMAooooAWkoooAWiiigAooooAKKKKACiiigAooooAKKKKACiiigBaSlpO9AC0UCigAooooAKKKKACiiigYlFLSUxC0UlFABS0UlAC0UUUAFFJS0wCiiigAooooAWiiikAUUUUAFFFFABRRRQAUUUUwCiiigAooooAKKKKACiiigAoopaAEopaSgAooooAWiiigYUUUUAFFJS0CCgUUUDCilpKAFooooAKKSloASloooAKKKKACiiigAooooAKKKKACiikoAWiiimAUYoooEFFFFABRRRQMKKKKBBRRRQMKKKWgQlLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAwoFLRQAUUUUAFFFFMAooooAKKKKACiiigApKWkpgLSUUUgCilopgFFFFIQUUUUDFopKWmAUUUUgCiiigQUUUUAFJS0lAC0UUUxhRRRQIKKKBQMKKKKACiiigQUUUUgCiiigAooopgFFLRQAlFFLTASilopDCiiigQUUUUAFFFFMCjRRRXKAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRS0AFFFFABRRRQAUUUUAFFFFABRRRQAtFFFAgooooAKKWkoAKKKKYBRRRQAd6KKKBhRRRQAUUUUAFLSUtABRRRTAKKKKACiiigAooooAO1FFAoAKKKKACiiigAooooAKKKKACiiigApaSigAooooAKKWkoAKWiigApKWigBKWiigAooooASlopKAFooopgFFFFAwooooAWiiigQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRS0UAJRS0lAC0lLRigBKKWimAlFLRQAlFLRQAlLRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUALSUtFABRRRQAUUUUAFFFFABRRRQAUlLSUwCloooASloooAKKKKACiiigAooopgFKKSlpAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUwCiiigAopaSgYUUUUCCloopDEopaKYhKWiigYUUUUwCg0UnegApaSigQtFFLQMKKKKQBRRRQAUUlLQAUUUUAFFFFABRRRQAUUlLQAUlLRQIKKKKACiiimAUUUUAFFFFABRRRQMKKWjFAhKKWigAooooGFFFFABRRRQIKKKKACiiigAooooAKKKKACiiigAopaKAEooooGLRRRQAUUUUCCiiimAUUUUAFFFFAwooooAKKKKACjFFFABRRRQIKKKKACiiloGFFFFAgooooAKKKKBhRRRQAUUUUxBRRRQMKKKKBBRSUtAwooooEFFFFABRRRQAUUUUDCloooASilpKBBS0lLQAUUUUAFFFFAwpaKKBCUUtFMChRRRXKAUUUUAFLSUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUtFACUtAooEFFFFAwooooAKKKKACiiigAoopaACiiigQUUUUAFFFFABRRRQAtJRS0AJRS0UxhSUUUAFFFFAC0UUUAFFFFMAooooAKKKKACiiigAooooAKKKWgBKKWigBKKKWgBKKWkoAKKKWgBKKWigBKWiigApKWigAooooAKKKKAEpaKKACiiigAoopKAFooopgFFFFAwpRSUtABRRRQIKKKKACiiigAoopaAEooo70DCiiigQUUUUAFFFFABRRRQAUUUUAFFFFAB2oxRRQAUUUUAFFLSUAFAopaACiiigAooooAKKKKACiiigAooooAKKKKACiiimAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUtJS0AFFFFABRRRQAUUUUAFFFFABRRRTAKKKKACiiimAUUUUAFFFFIAooopgFFFFABS0CikAUUUUAFFFFABRRRQMKKKKBBRRS0DEoopaYBRRRQISilooASloooASilooASloooAKKKKBhRRRTAKKDRQIKKKKAFooopDCiiigAooooAKKKKACiiigAooooEFFFFABRRRQMKKKSgBaKKKYBRRRQIKKWkoAKKWigAooooAKKKKACiiigAooooAKKKKBgKKKKACiiigQUUUUALSUtJQAtFJS0AFFFFABSUtFACUUtFAwooooAKKKKACiiimAUUUUCCilooAQ0UUUDCiiigQUUUUDFooooEJS0UUAFFFFMYUUUUAFFFFAgooooAKKKKBhRRRQIKKKKBhRRRQAUUUtACUUtFAhKKWigYlLRRQAUUUUCCijFFAwooooEFFFFABRRRQAUUUUwFooopAAopaKYFAUYoorlAMUYoooAMUYoooAMUYoooAMUYoooAMUYoooAMUYoooAMUYoooAMcUYoooAMUuKKKBBijFFFAwxRiiigAC+9G2iimAYoxRRSAMUAUUUALijFFFAgxRiiigYYoxRRQIMUYoooGBFGKKKADFAGaKKYBijFFFABijFFFAC4oxRRQAmKMUUUwDGaMUUUALijFFFACYpcUUUAGKMUUUAGKMUUUAGKMUUUAGKMUUUAGKMUUUAGKMUUUAGKXFFFAxMUYoooELijFFFAARRiiigBMUuKKKYxMUYoopCDHFGKKKAF20mKKKBi4pMUUUALijFFFABijFFFABijFFFMQYooooAMUuKKKADFGKKKADFGKKKAExS496KKAExS4oooAMUY5oooAMUmKKKBi4oxRRQAYoxRRQIMUbfeiigYYoxRRQIMUYoooAMUuMCiigAxRiiigAxQRiiigAxSYoopjFxRiiigQYoxRRQAYoxRRQAYpMUUUDFxRiiigQYoxRRQAYoxRRQAY4oxRRQMMUYoooEGKMUUUAG2giiigYu2jbRRQIMUYoooAMUYoooAMUYoooAMUYoopgGKMUUUDDFGKKKYBjNGKKKADFGKKKADFGKKKBBijFFFACgUYoopAG2jFFFMBMUuOKKKQwxRiiigAxRiiigQYoxRRTAMUYoooGGKMUUUCFxRiiigBMUYoooAXFGKKKBhijFFFAgxRiiimMMUUUUAGOKMUUUgDFGKKKYgxRiiigYuKMUUUAGOKTFFFAC4oxRRQAYoxRRQAYoxRRQAYoxRRQAYoxRRQAYoxRRQIMUYoooGGKMUUUAGKMc0UUCFxSYoooGLikxRRQAuKTFFFABijFFFABilxRRQIMUYoooAMUYoooGAFGKKKBBijFFFABijHNFFMYEUYoooAMUYoooEGKMUUUDFxRiiigAxRiiigAxQBmiigAxRiiigAxRiiigAxRiiimIMUYoopAGKMUUUwAijFFFABijFFFIAxRiiimMKXFFFABikxRRQAYpcUUUAJijFFFABijFFFABilxRRQAYoxRRQAYoxRRQAYoxRRTAMUYoopAGKXFFFACYoxRRQAYpcUUUCDFGKKKBhijFFFABijFFFAgxRiiigAxRiiimAuKMUUUhj1XI60UUUwP/2Q=="

def _upload_rich_menu_image(api_client, rich_menu_id):
    """Upload the rich menu image."""
    try:
        img_data = base64.b64decode(_RICH_MENU_IMG_B64)
        blob_api = MessagingApiBlob(api_client)
        blob_api.set_rich_menu_image(rich_menu_id, body=img_data, _headers={'Content-Type': 'image/jpeg'})
    except Exception as e:
        logger.warning("Rich menu image upload failed: %s", e)


def delete_rich_menu():
    """Delete ALL rich menus."""
    deleted = 0
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            try:
                api.cancel_default_rich_menu()
            except Exception:
                pass
            try:
                existing = api.get_rich_menu_list()
                # Try multiple attribute names
                menus = getattr(existing, 'richmenus', None) or getattr(existing, 'rich_menus', None) or []
                if not menus and hasattr(existing, '__iter__'):
                    menus = list(existing)
                logger.info("Found %d rich menus to delete (type: %s)", len(menus), type(menus).__name__)
                for rm in menus:
                    rid = getattr(rm, 'rich_menu_id', None) or getattr(rm, 'richMenuId', None)
                    if rid:
                        try:
                            api.delete_rich_menu(rid)
                            deleted += 1
                            logger.info("Deleted rich menu: %s", rid)
                        except Exception as e:
                            logger.warning("Failed to delete rich menu %s: %s", rid, e)
            except Exception as e:
                logger.warning("Failed to list rich menus: %s", e)
    except Exception as e:
        logger.warning("Delete rich menu error: %s", e)
    logger.info("Deleted %d rich menus", deleted)
    return deleted


def get_sender_object():
    """Build Sender object for customized bot display name/icon."""
    if not MessageSender or not sender_name:
        return None
    try:
        kwargs = {"name": sender_name}
        if sender_icon and sender_icon.startswith("http"):
            kwargs["icon_url"] = sender_icon
        return MessageSender(**kwargs)
    except Exception:
        return None


def get_insight_followers():
    """Get follower demographics from Insight API."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            # Get follower demographics
            demo = api.get_follower_demographics()
            return {
                "ages": getattr(demo, 'ages', None),
                "genders": getattr(demo, 'genders', None),
                "areas": getattr(demo, 'areas', None),
                "available": getattr(demo, 'available', False),
            }
    except Exception:
        return None


def get_message_delivery_stats(date_str=None):
    """Get message delivery stats for a specific date (reply/push/multicast/broadcast)."""
    try:
        if not date_str:
            date_str = time.strftime("%Y%m%d", time.gmtime(time.time() - 86400))
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            result = {"date": date_str}
            try:
                s = api.get_number_of_sent_reply_messages(var_date=date_str)
                result["reply"] = getattr(s, 'success', 0)
            except Exception:
                result["reply"] = None
            try:
                s = api.get_number_of_sent_push_messages(var_date=date_str)
                result["push"] = getattr(s, 'success', 0)
            except Exception:
                result["push"] = None
            try:
                s = api.get_number_of_sent_multicast_messages(var_date=date_str)
                result["multicast"] = getattr(s, 'success', 0)
            except Exception:
                result["multicast"] = None
            try:
                s = api.get_number_of_sent_broadcast_messages(var_date=date_str)
                result["broadcast"] = getattr(s, 'success', 0)
            except Exception:
                result["broadcast"] = None
            return result
    except Exception:
        return None


def get_message_interaction_stats(request_id):
    """Get user interaction statistics (opens, clicks) for a sent message."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_message_event(request_id=request_id)
            return {
                "overview": getattr(resp, 'overview', None),
                "messages": getattr(resp, 'messages', None),
                "clicks": getattr(resp, 'clicks', None),
            }
    except Exception as e:
        logger.debug("get_message_interaction_stats failed: %s", e)
        return None


def get_statistics_per_unit(date_str=None, num_days=7):
    """Get daily follower statistics (new followers, blocks, unblocks) for recent days.
    Uses LINE Insight API: /v2/bot/insight/followers?date=YYYYMMDD"""
    results = []
    try:
        for i in range(num_days):
            d = time.strftime("%Y%m%d", time.gmtime(time.time() - 86400 * (i + 1)))
            try:
                with ApiClient(configuration) as api_client:
                    api = MessagingApi(api_client)
                    resp = api.get_number_of_followers(var_date=d)
                    results.append({
                        "date": d,
                        "followers": getattr(resp, 'followers', None),
                        "targeted_reaches": getattr(resp, 'targeted_reaches', None),
                        "blocks": getattr(resp, 'blocks', None),
                    })
            except Exception:
                results.append({"date": d, "followers": None, "targeted_reaches": None, "blocks": None})
    except Exception as e:
        logger.debug("get_statistics_per_unit failed: %s", e)
    return results


def upload_rich_menu_image_custom(rich_menu_id, image_bytes, content_type="image/png"):
    """Upload a custom image to a rich menu (from admin panel upload)."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            blob_api.set_rich_menu_image(rich_menu_id, body=image_bytes, _headers={'Content-Type': content_type})
            return True
    except Exception as e:
        logger.warning("upload_rich_menu_image_custom failed: %s", e)
        return False


def send_imagemap_message(to, base_url, alt_text, width, height, actions):
    """Send an Imagemap message to a user or group.
    actions: list of {"type": "message"|"uri", "text"|"uri": str, "x": int, "y": int, "w": int, "h": int}
    """
    if not ImagemapMessage:
        return False
    try:
        imap_actions = []
        for a in actions:
            area = ImagemapArea(x=a["x"], y=a["y"], width=a["w"], height=a["h"])
            if a.get("type") == "uri":
                imap_actions.append(URIImagemapAction(link_uri=a["uri"], area=area))
            else:
                imap_actions.append(MessageImagemapAction(text=a.get("text", ""), area=area))
        msg = ImagemapMessage(
            base_url=base_url,
            alt_text=alt_text or "圖片選單",
            base_size=ImagemapBaseSize(width=width, height=height),
            actions=imap_actions,
        )
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.push_message(PushMessageRequest(to=to, messages=[msg]))
        return True
    except Exception as e:
        logger.warning("send_imagemap_message failed: %s", e)
        return False


def get_all_follower_ids():
    """Get all follower user IDs (paginated)."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            follower_ids = []
            token = None
            for _ in range(50):  # safety limit
                try:
                    if token:
                        resp = api.get_followers(start=token)
                    else:
                        resp = api.get_followers()
                except AttributeError:
                    if token:
                        resp = api.get_follower_ids(start=token)
                    else:
                        resp = api.get_follower_ids()
                ids = getattr(resp, 'user_ids', None) or getattr(resp, 'follower_ids', None) or []
                follower_ids.extend(ids)
                token = getattr(resp, 'next', None) or getattr(resp, 'next_token', None)
                if not token:
                    break
            return follower_ids
    except Exception as e:
        logger.warning("get_all_follower_ids failed: %s", e)
        return []


# ---- Room (multi-person chat) support ----
def get_room_member_count(room_id):
    """Get number of users in a multi-person chat."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            return api.get_room_members_count(room_id)
    except Exception:
        return None


def fetch_all_room_members(room_id):
    """Fetch all member IDs in a multi-person chat."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            member_ids = []
            token = None
            while True:
                try:
                    if token:
                        resp = api.get_room_member_ids(room_id, start=token)
                    else:
                        resp = api.get_room_member_ids(room_id)
                except AttributeError:
                    break
                ids = getattr(resp, 'member_user_ids', None) or getattr(resp, 'member_ids', None) or []
                member_ids.extend(ids)
                token = getattr(resp, 'next', None) or getattr(resp, 'next_token', None)
                if not token:
                    break
            return member_ids
    except Exception as e:
        logger.warning("fetch_all_room_members failed: %s", e)
        return []


def get_room_member_profile(room_id, user_id):
    """Get profile of a member in a multi-person chat."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            profile = api.get_room_member_profile(room_id, user_id)
            return {
                "display_name": getattr(profile, 'display_name', ''),
                "user_id": getattr(profile, 'user_id', ''),
                "picture_url": getattr(profile, 'picture_url', ''),
            }
    except Exception:
        return None


# ---- Rich Menu enhanced ----
def get_rich_menu_by_id(rich_menu_id):
    """Get a single rich menu by ID."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            rm = api.get_rich_menu(rich_menu_id)
            return {
                "id": getattr(rm, 'rich_menu_id', ''),
                "name": getattr(rm, 'name', ''),
                "size": {"width": getattr(getattr(rm, 'size', None), 'width', 0), "height": getattr(getattr(rm, 'size', None), 'height', 0)} if getattr(rm, 'size', None) else None,
                "chat_bar_text": getattr(rm, 'chat_bar_text', ''),
                "selected": getattr(rm, 'selected', False),
                "areas_count": len(getattr(rm, 'areas', []) or []),
            }
    except Exception as e:
        logger.warning("get_rich_menu_by_id failed: %s", e)
        return None


def get_default_rich_menu_id():
    """Get the ID of the current default rich menu."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_default_rich_menu_id()
            return getattr(resp, 'rich_menu_id', None)
    except Exception:
        return None


def get_user_rich_menu_id(user_id):
    """Get the rich menu ID linked to a specific user."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_rich_menu_id_of_user(user_id)
            return getattr(resp, 'rich_menu_id', None)
    except Exception:
        return None


def validate_rich_menu_obj(rich_menu_dict):
    """Validate a rich menu object before creating it."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.validate_rich_menu_object(rich_menu_dict)
            return {"valid": True}
    except Exception as e:
        return {"valid": False, "error": str(e)}


def download_rich_menu_image(rich_menu_id):
    """Download the image of a rich menu."""
    try:
        with ApiClient(configuration) as api_client:
            blob_api = MessagingApiBlob(api_client)
            content = blob_api.get_rich_menu_image(rich_menu_id)
            return content
    except Exception as e:
        logger.warning("download_rich_menu_image failed: %s", e)
        return None


# ---- Rich Menu Alias enhanced ----
def get_rich_menu_alias(alias_id):
    """Get rich menu alias info by ID."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_rich_menu_alias(alias_id)
            return {
                "alias_id": getattr(resp, 'rich_menu_alias_id', ''),
                "rich_menu_id": getattr(resp, 'rich_menu_id', ''),
            }
    except Exception:
        return None


def list_rich_menu_aliases():
    """Get list of all rich menu aliases."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            resp = api.get_rich_menu_alias_list()
            aliases_raw = getattr(resp, 'aliases', []) or []
            return [{"alias_id": getattr(a, 'rich_menu_alias_id', ''), "rich_menu_id": getattr(a, 'rich_menu_id', '')} for a in aliases_raw]
    except Exception:
        return []


def update_rich_menu_alias(alias_id, new_rich_menu_id):
    """Update an existing rich menu alias to point to a different menu."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            from linebot.v3.messaging import UpdateRichMenuAliasRequest
            api.update_rich_menu_alias(alias_id, UpdateRichMenuAliasRequest(rich_menu_id=new_rich_menu_id))
            return True
    except Exception as e:
        logger.warning("update_rich_menu_alias failed: %s", e)
        return False


def list_rich_menus():
    """List all existing rich menus."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            result = api.get_rich_menu_list()
            raw = getattr(result, 'richmenus', None) or getattr(result, 'rich_menus', None) or []
            menus = []
            for rm in raw:
                menus.append({
                    "id": getattr(rm, 'rich_menu_id', '') or getattr(rm, 'richMenuId', ''),
                    "name": getattr(rm, 'name', ''),
                    "selected": getattr(rm, 'selected', False),
                    "chat_bar_text": getattr(rm, 'chat_bar_text', ''),
                })
            return menus
    except Exception as e:
        logger.warning("list_rich_menus failed: %s", e)
        return []


def link_rich_menu_to_user(user_id, rich_menu_id):
    """Link a specific rich menu to a user."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.link_rich_menu_id_to_user(user_id, rich_menu_id)
            return True
    except Exception:
        return False


def unlink_rich_menu_from_user(user_id):
    """Unlink rich menu from a user."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.unlink_rich_menu_id_from_user(user_id)
            return True
    except Exception:
        return False


def multicast_message(user_ids, text):
    """Send a message to multiple users at once."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.multicast(MulticastRequest(
                to=user_ids,
                messages=[TextMessage(text=text)]
            ))
            return True
    except Exception as e:
        logger.warning("Multicast failed: %s", e)
        return False


def build_confirm_template(text, yes_label, yes_data, no_label, no_data):
    """Build a ConfirmTemplate message for yes/no interactions."""
    if not TemplateMessage:
        return None
    try:
        return TemplateMessage(
            alt_text=text,
            template=ConfirmTemplate(
                text=text,
                actions=[
                    PostbackAction(label=yes_label, data=yes_data),
                    PostbackAction(label=no_label, data=no_data),
                ]
            )
        )
    except Exception:
        return None


def build_carousel(columns):
    """Build a CarouselTemplate message.
    columns: list of {"title": str, "text": str, "actions": [{"label": str, "text": str}]}
    """
    if not TemplateMessage or not CarouselTemplate:
        return None
    try:
        cols = []
        for c in columns[:10]:  # max 10 columns
            actions = [MessageAction(label=a["label"], text=a["text"]) for a in c.get("actions", [])]
            cols.append(CarouselColumn(
                title=c.get("title", "")[:40],
                text=c.get("text", "")[:60],
                actions=actions[:3]  # max 3 actions
            ))
        return TemplateMessage(
            alt_text=columns[0].get("title", "選單"),
            template=CarouselTemplate(columns=cols)
        )
    except Exception:
        return None


def broadcast_message(text):
    """Broadcast a message to all bot followers."""
    if not BroadcastRequest:
        return False
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.broadcast(BroadcastRequest(
                messages=[TextMessage(text=text)]
            ))
            return True
    except Exception as e:
        logger.warning("Broadcast failed: %s", e)
        return False


def manage_rich_menu_alias(alias_id, rich_menu_id, action="create"):
    """Create or delete a Rich Menu alias."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            if action == "create" and CreateRichMenuAliasRequest:
                api.create_rich_menu_alias(CreateRichMenuAliasRequest(
                    rich_menu_alias_id=alias_id,
                    rich_menu_id=rich_menu_id
                ))
                return True
            elif action == "delete":
                api.delete_rich_menu_alias(alias_id)
                return True
    except Exception as e:
        logger.warning("Rich menu alias %s failed: %s", action, e)
    return False


def batch_link_rich_menu(user_ids, rich_menu_id):
    """Link a rich menu to multiple users at once."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.link_rich_menu_id_to_users(
                rich_menu_id=rich_menu_id,
                user_ids=user_ids
            )
            return True
    except Exception as e:
        logger.warning("Batch rich menu link failed: %s", e)
        return False


def batch_unlink_rich_menu(user_ids):
    """Unlink rich menu from multiple users at once."""
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.unlink_rich_menu_id_from_users(user_ids=user_ids)
            return True
    except Exception as e:
        logger.warning("Batch rich menu unlink failed: %s", e)
        return False


def add_line_emoji_to_text(text, emoji_product_id="5ac1bfd5040ab15980c9b435", emoji_ids=None):
    """Build a TextMessage with LINE emojis embedded.
    This is a helper - LINE emojis use $placeholder in text + emojis array."""
    # LINE emoji format: text has $ placeholders, emojis array maps each
    # For simplicity, return a TextMessage with emoji support
    try:
        if not emoji_ids:
            return TextMessage(text=text)
        emojis = []
        for i, eid in enumerate(emoji_ids):
            emojis.append({
                "index": text.index("$", sum(1 for c in text[:text.index("$")] if True)),
                "productId": emoji_product_id,
                "emojiId": eid
            })
        return TextMessage(text=text, emojis=emojis)
    except Exception:
        return TextMessage(text=text)


def build_translation_flex(original, translated, src_flag, tgt_flag, sender_name_display=None, quoted_text=None):
    """Build a Flex Message for translation with original + translated text."""
    try:
        body_contents = []
        # Quoted message context (if replying to another message)
        if quoted_text:
            qt = quoted_text[:50] + "..." if len(quoted_text) > 50 else quoted_text
            body_contents.append({
                "type": "text", "text": "↩ " + qt,
                "size": "xxs", "color": "#6a6a7a", "wrap": True, "margin": "none",
                "style": "italic"
            })
            body_contents.append({"type": "separator", "margin": "sm"})
        # Sender name
        if sender_name_display:
            body_contents.append({
                "type": "text", "text": sender_name_display,
                "size": "xs", "color": "#8a8a9a", "margin": "none"
            })
        # Original text
        body_contents.append({
            "type": "text", "text": src_flag + " " + original,
            "size": "sm", "color": "#b0b0b0", "wrap": True, "margin": "sm"
        })
        # Separator
        body_contents.append({"type": "separator", "margin": "md"})
        # Translated text
        body_contents.append({
            "type": "text", "text": tgt_flag + " " + translated,
            "size": "md", "color": "#ffffff", "wrap": True, "margin": "md", "weight": "bold"
        })

        flex_obj = {
            "type": "bubble",
            "size": "kilo",
            "body": {
                "type": "box", "layout": "vertical",
                "contents": body_contents,
                "backgroundColor": "#1a1a2e",
                "paddingAll": "16px",
                "cornerRadius": "12px"
            }
        }
        return FlexMessage(
            alt_text=tgt_flag + " " + translated,
            contents=FlexContainer.from_dict(flex_obj)
        )
    except Exception as e:
        logger.warning("Flex message build failed, falling back to text: %s", e)
        return None


def build_quick_reply(group_id=None):
    """Build Quick Reply buttons for translation messages, respecting per-group command toggles."""
    try:
        # Core buttons always shown
        items = [
            QuickReplyItem(action=MessageAction(label="📖 說明/Info", text="/help")),
        ]
        # Command-linked buttons: only show if that command is enabled for the group
        if is_cmd_enabled(group_id, 'qry'):
            items.append(QuickReplyItem(action=MessageAction(label="🔍 儲區/Gudang", text="/qry ")))
        items.append(QuickReplyItem(action=MessageAction(label="❌ 不翻我/Skip", text="/skip")))
        items.append(QuickReplyItem(action=MessageAction(label="✅ 翻譯我/Unskip", text="/unskip")))
        if is_cmd_enabled(group_id, 'pw1'):
            items.append(QuickReplyItem(action=MessageAction(label="🔑 班長密碼/PW1", text="/pw1")))
        if is_cmd_enabled(group_id, 'pw2'):
            items.append(QuickReplyItem(action=MessageAction(label="🏭 儲運密碼/PW2", text="/pw2")))
        if is_cmd_enabled(group_id, 'pkg'):
            items.append(QuickReplyItem(action=MessageAction(label="📦 包裝碼/Kemas", text="/pkg ")))
        if is_cmd_enabled(group_id, 'scrap'):
            items.append(QuickReplyItem(action=MessageAction(label="🎨 廢料色/Warna", text="/scrap")))
        # Camera quick reply button (opens camera directly)
        if MsgCameraAction and get_group_feature(group_id, 'camera_qr'):
            try:
                items.append(QuickReplyItem(action=MsgCameraAction(label="📷 拍照/Foto")))
            except Exception:
                pass
        # Clipboard quick reply button (copy useful text)
        if MsgClipboardAction and get_group_feature(group_id, 'clipboard_qr'):
            try:
                items.append(QuickReplyItem(action=MsgClipboardAction(
                    label="📋 複製儲區指令",
                    clipboard_text="/qry "
                )))
            except Exception:
                pass
        # Camera Roll quick reply button (opens photo album)
        if MsgCameraRollAction and get_group_feature(group_id, 'camera_roll_qr'):
            try:
                items.append(QuickReplyItem(action=MsgCameraRollAction(label="🖼️ 相簿/Album")))
            except Exception:
                pass
        # Location quick reply button (share location)
        if MsgLocationAction and get_group_feature(group_id, 'location_qr'):
            try:
                items.append(QuickReplyItem(action=MsgLocationAction(label="📍 位置/Lokasi")))
            except Exception:
                pass
        # URI-based buttons (open external links)
        try:
            items.append(QuickReplyItem(action=MsgURIAction(
                label="💡 提案/Saran",
                uri="https://app-walsin-crm-improvement.azurewebsites.net/improvePropose/personalList"
            )))
            items.append(QuickReplyItem(action=MsgURIAction(
                label="📅 差勤/Absen",
                uri="https://hrm.walsin.com/servlet/jform?file=hrm8w.pkg,hrm8aw.pkg,BPM_JS.pkg,hrm8w_walsin.pkg,hrm8w_walsinhrisp.pkg&locale=US&init_func=人事_WS"
            )))
        except Exception:
            pass
        return QuickReply(items=items)
    except Exception:
        return None


# ─── Admin Panel ────────────────────────────────────────

ADMIN_HTML = '''<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="UTF-8">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<meta name="viewport" content="width=device-width,initial-scale=1,user-scalable=no">
<title>翻譯Bot 管理後台</title>
<link rel="manifest" href="/manifest.json">
<meta name="theme-color" content="#7c6fef">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<meta name="apple-mobile-web-app-title" content="Bot管理">
<link rel="apple-touch-icon" href="/icon-192.png">
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#0a0a0a;color:#e0e0e0;min-height:100vh;padding-bottom:env(safe-area-inset-bottom)}
.header{background:linear-gradient(135deg,#5b6abf,#8b5fbf 50%,#b07cc3);padding:18px 16px;position:sticky;top:0;z-index:100;box-shadow:0 2px 12px rgba(91,106,191,.3)}
.header h1{font-size:18px;color:#fff;display:flex;align-items:center;gap:8px}
.header .platform{font-size:13px;font-weight:400;opacity:.8}
.login-wrap{display:flex;justify-content:center;align-items:center;min-height:80vh;padding:20px}
.login-box{background:#1a1a2e;border-radius:16px;padding:32px 24px;width:100%;max-width:360px;box-shadow:0 4px 24px rgba(0,0,0,.4);border:1px solid #2a2a3e}
.login-box h2{text-align:center;margin-bottom:24px;color:#7c6fef}
.input-field{width:100%;padding:14px 16px;border-radius:12px;border:1px solid #2a2a3e;background:#0d0d1a;color:#fff;font-size:16px;margin-bottom:16px;outline:none;transition:border .2s}
.input-field:focus{border-color:#7c6fef}
.btn{display:block;width:100%;padding:14px;border:none;border-radius:12px;font-size:16px;font-weight:600;cursor:pointer;transition:all .2s}
.btn-primary{background:#7c6fef;color:#fff}
.btn-primary:active{background:#6358d4;transform:scale(.98)}
.btn-red{background:rgba(240,71,71,.15);color:#f04747;border:1px solid rgba(240,71,71,.3)}
.btn-red:active{background:rgba(240,71,71,.25);transform:scale(.98)}
.btn-sm{padding:8px 14px;font-size:13px;width:auto;border-radius:8px;display:inline-block}
.btn-dark{background:#2a2a3e;color:#e0e0e0;border:1px solid #3a3a4e}
.tabs{display:flex;background:#0a0a0a;border-bottom:1px solid #2a2a3e;position:sticky;top:56px;z-index:99;overflow-x:auto;-webkit-overflow-scrolling:touch;scrollbar-width:none}
.tabs::-webkit-scrollbar{display:none}
.tab{flex:none;padding:12px 16px;text-align:center;font-size:13px;font-weight:400;color:#8a8a9a;cursor:pointer;border-bottom:2px solid transparent;transition:all .2s;white-space:nowrap}
.tab.active{color:#7c6fef;font-weight:700;border-bottom-color:#7c6fef}
.panel{display:none;padding:16px}
.panel.active{display:block}

/* Stats grid */
.stats-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}
.stat-card{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:12px;padding:16px 12px;display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:80px}
.stat-value{font-size:28px;font-weight:700;color:#e0e0e0;font-family:ui-monospace,SFMono-Regular,"SF Mono",monospace;letter-spacing:-.5px}
.stat-value.highlight{color:#7c6fef}
.stat-label{font-size:12px;color:#8a8a9a;margin-top:4px;display:flex;align-items:center;gap:4px}

/* Cards */
.card{background:#1a1a2e;border-radius:12px;padding:16px;margin-bottom:12px;border:1px solid #2a2a3e}
.card-title{font-size:15px;font-weight:600;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}
.card-sub{font-size:12px;color:#8a8a9a}
.badge{display:inline-block;padding:3px 10px;border-radius:6px;font-size:12px;font-weight:600}
.badge-on{background:rgba(67,181,129,.15);color:#43b581}
.badge-off{background:rgba(240,71,71,.15);color:#f04747}
.badge-yellow{background:rgba(250,166,26,.15);color:#faa61a}

/* Feature badges */
.feat-badges{display:flex;gap:8px;flex-wrap:wrap;margin:10px 0}
.feat-badge{display:inline-flex;align-items:center;gap:4px;padding:4px 10px;border-radius:6px;font-size:12px;font-weight:500}
.feat-badge.on{background:rgba(67,181,129,.12);color:#43b581;border:1px solid rgba(67,181,129,.25)}
.feat-badge.off{background:rgba(100,100,120,.12);color:#8a8a9a;border:1px solid rgba(100,100,120,.2)}

/* Select */
.sel{appearance:none;background:#2a2a3e;color:#e0e0e0;border:1px solid #3a3a4e;border-radius:8px;padding:6px 28px 6px 12px;font-size:13px;cursor:pointer;outline:none}
.sel-wrap{position:relative;display:inline-block}
.sel-wrap::after{content:"▼";position:absolute;right:8px;top:50%;transform:translateY(-50%);pointer-events:none;color:#8a8a9a;font-size:10px}

/* Channel dropdown (full width) */
.ch-select{width:100%;appearance:none;background:#1a1a2e;color:#e0e0e0;border:1px solid #3a3a4e;border-radius:12px;padding:14px 40px 14px 16px;font-size:15px;font-weight:500;cursor:pointer;outline:none}
.ch-select-wrap{position:relative}
.ch-select-wrap::after{content:"▼";position:absolute;right:16px;top:50%;transform:translateY(-50%);pointer-events:none;color:#8a8a9a;font-size:12px}

/* Toggle */
.toggle{position:relative;width:48px;height:26px;cursor:pointer;display:inline-block;flex-shrink:0}
.toggle input{display:none}
.toggle .slider{position:absolute;inset:0;background:#3a3a4a;border-radius:13px;transition:.2s}
.toggle .slider:before{content:"";position:absolute;height:20px;width:20px;left:3px;bottom:3px;background:#fff;border-radius:50%;transition:.2s;box-shadow:0 1px 3px rgba(0,0,0,.3)}
.toggle input:checked+.slider{background:#7c6fef}
.toggle input:checked+.slider:before{transform:translateX(22px)}

/* Whitelist / user items */
.wl-item{display:flex;align-items:center;justify-content:space-between;padding:14px 16px;border-bottom:1px solid #2a2a3e}
.wl-item:last-child{border-bottom:none}

/* User card */
.user-card{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:12px;padding:16px;margin-bottom:12px}
.user-name{font-size:16px;font-weight:700;margin-bottom:6px}
.user-id{font-size:12px;color:#8a8a9a}
.user-admin-row{display:flex;align-items:center;justify-content:flex-end;gap:8px;margin-top:10px}
.admin-label{font-size:13px;color:#faa61a}

.empty{text-align:center;color:#5a5a6a;padding:32px 16px;font-size:14px}
.toast{position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:rgba(30,30,50,.95);color:#fff;padding:10px 24px;border-radius:10px;font-size:14px;z-index:200;opacity:0;transition:all .25s;pointer-events:none;box-shadow:0 4px 20px rgba(0,0,0,.4);border:1px solid rgba(124,111,239,.3)}
.toast.show{opacity:1}

/* DM section in groups panel */
.dm-section{background:#1a1a2e;border:1px solid #2a2a3e;border-radius:12px;padding:16px;margin-bottom:12px}
.dm-toggle-row{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px}
</style>
</head>
<body>
<div id="app">

<!-- Login -->
<div id="loginPage">
<div class="header"><h1>🤖 翻譯Bot 管理後台 <span class="platform">LINE</span></h1></div>
<div class="login-wrap">
<div class="login-box">
<h2>🔒 管理員登入</h2>
<div style="font-size:11px;color:#666;margin-bottom:8px">v2.6-0412c</div>
<input class="input-field" id="pwInput" type="password" placeholder="輸入管理密碼" autocomplete="off" onkeydown="if(event.key==='Enter')document.getElementById('loginBtn').click()">
<div id="loginMsg" style="color:#f04747;font-size:12px;min-height:18px;margin-top:4px"></div>
<button class="btn btn-primary" id="loginBtn" type="button">登入</button>
</div>
</div>
</div>
<script>
document.getElementById('loginBtn').addEventListener('click',function(){
  var m=document.getElementById('loginMsg');
  var k=document.getElementById('pwInput').value.trim();
  if(!k){m.textContent='請輸入密碼';return}
  m.textContent='登入中...';
  m.style.color='#aaa';
  fetch(window.location.origin+'/api/admin/status',{headers:{'X-Admin-Key':k}})
  .then(function(r){if(!r.ok)throw new Error('HTTP '+r.status);return r.json()})
  .then(function(d){
    if(d&&d.ok){
      m.textContent='';
      document.getElementById('loginPage').style.display='none';
      document.getElementById('mainPage').style.display='block';
      window._ADMIN_KEY=k;
      try{localStorage.setItem('bot_admin_key',k)}catch(e){}
      if(typeof KEY!=='undefined') KEY=k;
      if(typeof loadAll==='function') loadAll();
    }else{
      m.style.color='#f04747';
      m.textContent='登入失敗: '+(d?JSON.stringify(d):'no response');
    }
  })
  .catch(function(e){
    m.style.color='#f04747';
    m.textContent='連線錯誤: '+e.message;
  });
});
document.getElementById('pwInput').addEventListener('keydown',function(e){
  if(e.key==='Enter') document.getElementById('loginBtn').click();
});
</script>

<!-- Main -->
<div id="mainPage" style="display:none">
<div class="header"><h1>🤖 翻譯Bot 管理後台 <span class="platform">LINE</span></h1></div>
<div class="tabs">
<div class="tab active" onclick="switchTab('overview')">總覽</div>
<div class="tab" onclick="switchTab('groups')">群組</div>
<div class="tab" onclick="switchTab('skip')">白名單</div>
<div class="tab" onclick="switchTab('users')">使用者</div>
<div class="tab" onclick="switchTab('names')">保護名單</div>
<div class="tab" onclick="switchTab('storage')">儲區</div>
<div class="tab" onclick="switchTab('packaging')">包裝碼</div>
<div class="tab" onclick="switchTab('passwords')">密碼</div>
<div class="tab" onclick="switchTab('scrap')">廢料色</div>
<div class="tab" onclick="switchTab('insight')">數據</div>
<div class="tab" onclick="switchTab('settings')">設定</div>
</div>

<!-- Overview Panel -->
<div class="panel active" id="panel-overview">
<div class="stats-grid" id="statsGrid">
<div class="stat-card"><div class="stat-value" id="st-uptime">0h 0m</div><div class="stat-label">⏱ 運行時間</div></div>
<div class="stat-card"><div class="stat-value" id="st-text">0</div><div class="stat-label">💬 文字翻譯</div></div>
<div class="stat-card"><div class="stat-value" id="st-image">0</div><div class="stat-label">🖼️ 圖片翻譯</div></div>
<div class="stat-card"><div class="stat-value" id="st-voice">0</div><div class="stat-label">🎤 語音翻譯</div></div>
<div class="stat-card"><div class="stat-value" id="st-wo">0</div><div class="stat-label">📋 工單偵測</div></div>
<div class="stat-card"><div class="stat-value" id="st-cmd">0</div><div class="stat-label">⌨️ 指令</div></div>
<div class="stat-card"><div class="stat-value highlight" id="st-cust">0</div><div class="stat-label">👥 客戶</div></div>
<div class="stat-card"><div class="stat-value highlight" id="st-groups">0</div><div class="stat-label">💬 群組</div></div>
<div class="stat-card"><div class="stat-value" id="st-dm-users">0</div><div class="stat-label">👤 DM使用者</div></div>
</div>
<!-- API Usage Card -->
<div class="card" style="margin:16px 16px 0" id="apiUsageCard">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🔑 OpenAI API 用量</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">
<div style="font-size:13px;color:#8a8a9a">Tokens（本次啟動）</div>
<div style="font-size:13px;text-align:right" id="st-tokens">0</div>
<div style="font-size:13px;color:#8a8a9a">預估花費</div>
<div style="font-size:13px;text-align:right" id="st-cost">$0.00</div>
</div>
<a href="https://platform.openai.com/settings/organization/billing/overview" target="_blank" style="display:block;margin-top:12px;padding:10px;text-align:center;background:#2a2a3e;border:1px solid #3a3a4e;border-radius:8px;color:#7c6fef;font-size:13px;font-weight:600;text-decoration:none">💳 查看 API 餘額</a>
</div>
</div>

<!-- Groups Panel -->
<div class="panel" id="panel-groups">
<!-- DM Section -->
<div class="dm-section">
<div class="dm-toggle-row">
<span style="font-weight:600;font-size:15px">📨 私訊 DM 翻譯</span>
<label class="toggle"><input type="checkbox" id="dmToggle" onchange="toggleDM()"><span class="slider"></span></label>
</div>
<div class="card-sub">總開關關閉時，只有白名單內的人可以私訊翻譯</div>
<div id="dmWlList" style="margin-top:10px"></div>
</div>
<div id="groupList"><div class="empty">載入中...</div></div>
</div>

<!-- Whitelist/Skip Panel -->
<div class="panel" id="panel-skip">
<div class="ch-select-wrap">
<select class="ch-select" id="skipGroupSelect" onchange="loadSkipList()">
<option value="">選擇群組...</option>
</select>
</div>
<div class="card-sub" style="padding:8px 4px;font-size:12px">開啟 = 不翻譯該成員訊息</div>
<div class="card" style="padding:0;overflow:hidden">
<div id="skipListContent"><div class="empty">請先選擇群組</div></div>
</div>
</div>

<!-- Users Panel -->
<div class="panel" id="panel-users">
<div class="ch-select-wrap">
<select class="ch-select" id="usersGroupSelect" onchange="loadUsers()">
<option value="">全部使用者</option>
</select>
</div>
<div id="usersList"><div class="empty">載入中...</div></div>
</div>

<!-- Protected Names Panel -->
<div class="panel" id="panel-names">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:4px">🛡️ 翻譯保護名單</div>
<div class="card-sub" style="margin-bottom:12px">名單內的名字翻譯時會保持原樣不翻（人名、公司名皆可）</div>
<div style="display:flex;gap:8px;margin-bottom:12px">
<input id="newNameInput" type="text" placeholder="輸入名字..." onkeydown="if(event.key==='Enter')addName()" style="flex:1;padding:10px 12px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:14px;outline:none">
<button class="btn btn-primary btn-sm" onclick="addName()">新增</button>
</div>
<div id="namesList"></div>
<div id="namesCount" style="font-size:12px;color:#8a8a9a;margin-top:8px"></div>
</div>
</div>

<!-- Storage Panel -->
<div class="panel" id="panel-storage">
<div class="card">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;font-weight:700;font-size:15px">📦 儲區資料更新</div>
<div class="card-sub" style="margin-bottom:14px">上傳 Excel 檔案自動更新儲區查詢資料</div>
<input type="file" id="storageFile" accept=".xlsx,.xls" style="display:none" onchange="previewStorage()">
<button class="btn btn-primary btn-sm" onclick="document.getElementById('storageFile').click()">選擇 Excel 檔案</button>
<div id="storageFileName" style="margin-top:8px;font-size:13px;color:#8a8a9a"></div>
</div>
<div id="storagePreview"></div>
<div id="storageActions" style="display:none;margin-top:12px">
<button class="btn btn-primary btn-sm" onclick="uploadStorage()">確認更新</button>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:6px">目前資料</div>
<div id="storageStats" style="font-size:14px;margin-bottom:14px">載入中...</div>
<button class="btn btn-dark btn-sm" onclick="downloadJson()">下載 JSON</button>
</div>
</div>

<!-- Packaging Panel -->
<div class="panel" id="panel-packaging">
<div class="card">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;font-weight:700;font-size:15px">📦 包裝碼資料更新</div>
<div class="card-sub" style="margin-bottom:14px">上傳 Excel 檔案（第一列標題列，含代碼/Code欄位）</div>
<input type="file" id="packagingFile" accept=".xlsx,.xls" style="display:none" onchange="previewPackaging()">
<button class="btn btn-primary btn-sm" onclick="document.getElementById('packagingFile').click()">選擇 Excel 檔案</button>
<div id="packagingFileName" style="margin-top:8px;font-size:13px;color:#8a8a9a"></div>
</div>
<div id="packagingPreview"></div>
<div id="packagingActions" style="display:none;margin-top:12px">
<button class="btn btn-primary btn-sm" onclick="uploadPackaging()">確認更新</button>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:6px">目前資料</div>
<div id="packagingStats" style="font-size:14px;margin-bottom:14px">載入中...</div>
<button class="btn btn-dark btn-sm" onclick="downloadPackagingJson()">下載 JSON</button>
</div>
</div>

<!-- Passwords Panel -->
<div class="panel" id="panel-passwords">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">🔑 密碼設定</div>
<div style="margin-bottom:16px">
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">班長工號密碼 (使用者傳 /pw1 時顯示)</div>
<textarea id="pw1Input" rows="3" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical"></textarea>
</div>
<div style="margin-bottom:16px">
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">儲運工號密碼 (使用者傳 /pw2 時顯示)</div>
<textarea id="pw2Input" rows="3" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical"></textarea>
</div>
<button class="btn btn-primary btn-sm" onclick="savePasswords()">儲存密碼</button>
<div id="pwSaveResult" style="margin-top:8px;font-size:13px"></div>
</div>
</div>

<!-- Scrap Panel -->
<div class="panel" id="panel-scrap">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">🎨 廢料鋼種顏色</div>
<div class="card-sub" style="margin-bottom:14px">使用者傳 /scrap 時顯示的內容</div>
<textarea id="scrapInput" rows="16" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical;font-family:monospace"></textarea>
<div style="margin-top:12px">
<button class="btn btn-primary btn-sm" onclick="saveScrap()">儲存</button>
<div id="scrapSaveResult" style="margin-top:8px;font-size:13px"></div>
</div>
</div>
</div>

<!-- Insight Panel -->
<div class="panel" id="panel-insight">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">📊 好友趨勢（近7日）</div>
<div id="insightTrendChart" style="min-height:160px;position:relative">
<canvas id="trendCanvas" width="600" height="180" style="width:100%;height:180px"></canvas>
</div>
<div id="insightTrendData" style="font-size:12px;color:#8a8a9a;margin-top:8px"></div>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">👥 好友人口統計</div>
<div id="insightDemoData" style="font-size:13px;color:#8a8a9a">載入中...</div>
</div>
<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">📈 昨日發送統計</div>
<div id="insightDelivery" style="font-size:13px;color:#8a8a9a">載入中...</div>
</div>
</div>

<!-- Settings Panel -->
<div class="panel" id="panel-settings">
<div class="card">
<div style="font-weight:700;font-size:15px;margin-bottom:12px">⚙️ 功能設定</div>
<div class="ch-select-wrap" style="margin-bottom:12px">
<select class="ch-select" id="settingsGroupSelect" style="font-size:13px;padding:10px" onchange="loadFeatureSettingsForGroup()">
<option value="">全域預設</option>
</select>
</div>
<div id="settingsCustomBadge" style="display:none;margin-bottom:10px"><span class="badge badge-on" style="font-size:11px">已自訂</span> <span style="font-size:12px;color:#8a8a9a;cursor:pointer;text-decoration:underline" onclick="resetGroupSettings()">重設為預設</span></div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">👋 歡迎訊息</span><br><span style="font-size:12px;color:#8a8a9a">新成員加入時自動發送</span></div>
<label class="toggle"><input type="checkbox" id="welcomeToggle" onchange="toggleFeatureSetting('welcome_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div style="padding:12px 0">
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">中文歡迎詞</div>
<textarea id="welcomeZh" rows="2" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical" onblur="saveWelcomeText()"></textarea>
<div style="font-size:13px;color:#8a8a9a;margin:8px 0 6px">印尼文歡迎詞</div>
<textarea id="welcomeId" rows="2" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical" onblur="saveWelcomeText()"></textarea>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🎨 Flex 翻譯卡片</span><br><span style="font-size:12px;color:#8a8a9a">關閉後用純文字顯示</span></div>
<label class="toggle"><input type="checkbox" id="flexToggle" onchange="toggleFeatureSetting('flex_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">⚡ Quick Reply 按鈕</span><br><span style="font-size:12px;color:#8a8a9a">翻譯後顯示快捷操作</span></div>
<label class="toggle"><input type="checkbox" id="qrToggle" onchange="toggleFeatureSetting('quick_reply_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🔇 靜音模式</span><br><span style="font-size:12px;color:#8a8a9a">翻譯訊息不震動手機</span></div>
<label class="toggle"><input type="checkbox" id="silentToggle" onchange="toggleFeatureSetting('silent_mode',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🎬 影片 OCR 翻譯</span><br><span style="font-size:12px;color:#8a8a9a">影片截圖自動 OCR 翻譯</span></div>
<label class="toggle"><input type="checkbox" id="videoToggle" onchange="toggleFeatureSetting('video_ocr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📍 位置訊息翻譯</span><br><span style="font-size:12px;color:#8a8a9a">翻譯地點名稱和地址</span></div>
<label class="toggle"><input type="checkbox" id="locationToggle" onchange="toggleFeatureSetting('location_translate_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">👁️ 標記已讀</span><br><span style="font-size:12px;color:#8a8a9a">處理訊息時顯示已讀標記</span></div>
<label class="toggle"><input type="checkbox" id="markReadToggle" onchange="toggleFeatureSetting('mark_read_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🔄 防重複發送</span><br><span style="font-size:12px;color:#8a8a9a">X-Line-Retry-Key 冪等性</span></div>
<label class="toggle"><input type="checkbox" id="retryKeyToggle" onchange="toggleFeatureSetting('retry_key_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📷 拍照快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入拍照按鈕</span></div>
<label class="toggle"><input type="checkbox" id="cameraQrToggle" onchange="toggleFeatureSetting('camera_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📋 複製快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入複製指令按鈕</span></div>
<label class="toggle"><input type="checkbox" id="clipboardQrToggle" onchange="toggleFeatureSetting('clipboard_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🖼️ 相簿快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入開啟相簿按鈕</span></div>
<label class="toggle"><input type="checkbox" id="cameraRollQrToggle" onchange="toggleFeatureSetting('camera_roll_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">📍 位置快捷鈕</span><br><span style="font-size:12px;color:#8a8a9a">Quick Reply 加入分享位置按鈕</span></div>
<label class="toggle"><input type="checkbox" id="locationQrToggle" onchange="toggleFeatureSetting('location_qr_enabled',this.checked)"><span class="slider"></span></label>
</div>

<div class="wl-item" style="border-color:#2a2a3e">
<div><span style="font-weight:600">🗣️ 翻譯口吻</span><br><span style="font-size:12px;color:#8a8a9a">控制翻譯的語氣風格</span></div>
<select id="toneSelect" style="padding:6px 10px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px" onchange="toggleFeatureSetting('translation_tone',this.value)">
<option value="casual">日常口語</option>
<option value="natural">母語自然風格</option>
<option value="formal">正式書面</option>
</select>
</div>
<div style="padding:4px 0 12px">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:6px">自訂語氣指令（填寫後覆蓋上方選項）</div>
<textarea id="toneCustom" rows="2" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical" placeholder="例如：用最口語的印尼文翻譯，像當地人聊天" onblur="toggleFeatureSetting('translation_tone_custom',this.value)"></textarea>
</div>

<div style="border-top:1px solid #2a2a3e;padding-top:12px;margin-top:4px">
<div style="font-weight:600;margin-bottom:8px">🤖 AI 模型自動切換</div>
<div class="card-sub" style="margin-bottom:8px">訊息超過指定字數自動升級為 GPT-4o（翻譯更流暢但較貴）。設為 0 表示全部用預設模型。</div>
<div style="display:flex;gap:8px;margin-bottom:8px;align-items:center">
<div style="font-size:13px;color:#8a8a9a;white-space:nowrap">字數門檻</div>
<input id="modelThreshold" type="number" min="0" value="0" style="width:80px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;text-align:center">
<div style="font-size:12px;color:#8a8a9a">字</div>
</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<div style="flex:1">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">預設模型（短訊息）</div>
<select id="modelDefault" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">
<option value="gpt-4o-mini">gpt-4o-mini</option>
<option value="gpt-4o">gpt-4o</option>
</select>
</div>
<div style="flex:1">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">升級模型（長訊息）</div>
<select id="modelUpgrade" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">
<option value="gpt-4o">gpt-4o</option>
<option value="gpt-4o-mini">gpt-4o-mini</option>
</select>
</div>
</div>
<button class="btn btn-primary btn-sm" onclick="saveModelSettings()">儲存模型設定</button>
<div id="modelSaveResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📊 LINE 配額 & 統計</div>
<div id="lineQuotaInfo" style="font-size:13px;color:#8a8a9a">載入中...</div>
<div id="lineInsight" style="font-size:13px;color:#8a8a9a;margin-top:8px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🔗 Webhook 狀態</div>
<div id="webhookInfo" style="font-size:13px;color:#8a8a9a">載入中...</div>
<button class="btn btn-dark btn-sm" style="margin-top:8px" onclick="testWebhook()">🧪 測試 Webhook</button>
<div id="webhookTestResult" style="font-size:12px;color:#8a8a9a;margin-top:6px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📈 送出統計（昨日）</div>
<div id="deliveryStats" style="font-size:13px;color:#8a8a9a">載入中...</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">👥 好友清單</div>
<button class="btn btn-dark btn-sm" onclick="loadFollowers()">載入好友列表</button>
<div id="followersList" style="font-size:13px;color:#8a8a9a;margin-top:8px;max-height:200px;overflow-y:auto"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🎨 Rich Menu 管理</div>
<div id="richMenuList" style="font-size:13px;color:#8a8a9a;margin-bottom:8px"></div>
<div id="richMenuDefault" style="font-size:13px;color:#8a8a9a;margin-bottom:8px"></div>
<div style="margin-bottom:8px">
<span style="font-size:13px;color:#8a8a9a">查詢用戶綁定選單：</span>
<div style="display:flex;gap:6px;margin-top:4px">
<input id="rmUserIdInput" type="text" placeholder="user ID" style="flex:1;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px">
<button class="btn btn-dark btn-sm" onclick="checkUserMenu()">查詢</button>
</div>
<div id="rmUserResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
</div>
<div style="font-size:13px;font-weight:600;margin-bottom:6px">Alias 列表</div>
<div id="rmAliasList" style="font-size:13px;color:#8a8a9a"></div>
<button class="btn btn-dark btn-sm" style="margin-top:6px" onclick="loadAliases()">重新載入 Alias</button>
<div style="margin-top:12px;border-top:1px solid #2a2a3e;padding-top:12px">
<div style="font-size:13px;font-weight:600;margin-bottom:6px">📤 上傳選單圖片</div>
<div class="card-sub" style="margin-bottom:8px">選擇 Rich Menu 後上傳圖片（建議 2500x1686 或 2500x843）</div>
<div style="display:flex;gap:6px;margin-bottom:6px;flex-wrap:wrap">
<select id="rmUploadSelect" style="flex:1;min-width:120px;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:12px"></select>
<label class="btn btn-dark btn-sm" style="cursor:pointer;display:inline-flex;align-items:center">
選擇圖片 <input type="file" id="rmImageFile" accept="image/png,image/jpeg" style="display:none" onchange="previewRmImage(this)">
</label>
</div>
<div id="rmImagePreview" style="display:none;margin-bottom:6px"><img id="rmImagePreviewImg" style="max-width:100%;max-height:120px;border-radius:6px;border:1px solid #3a3a4e"></div>
<button class="btn btn-primary btn-sm" onclick="uploadRmImage()" id="rmUploadBtn" style="display:none">⬆️ 上傳</button>
<div id="rmUploadResult" style="font-size:12px;color:#8a8a9a;margin-top:4px"></div>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🗺️ Imagemap 圖片選單</div>
<div class="card-sub" style="margin-bottom:10px">發送可點擊區域的圖片到群組（圖片需為 HTTPS URL，建議 1040px 寬）</div>
<div class="ch-select-wrap" style="margin-bottom:8px">
<select class="ch-select" id="imapGroupSelect" style="font-size:13px;padding:10px"></select>
</div>
<div style="margin-bottom:8px">
<div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">圖片 Base URL（不含 /1040 等尺寸後綴）</div>
<input id="imapBaseUrl" type="text" placeholder="https://example.com/images/menu" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<div style="flex:1"><div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">寬度</div><input id="imapW" type="number" value="1040" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px"></div>
<div style="flex:1"><div style="font-size:12px;color:#8a8a9a;margin-bottom:4px">高度</div><input id="imapH" type="number" value="1040" style="width:100%;padding:6px;border-radius:6px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px"></div>
</div>
<div style="font-size:13px;font-weight:600;margin-bottom:6px">點擊區域</div>
<div id="imapActions"></div>
<button class="btn btn-dark btn-sm" style="margin-bottom:8px" onclick="addImapAction()">＋ 新增區域</button>
<div style="margin-top:8px">
<button class="btn btn-primary btn-sm" onclick="sendImap()">📤 發送 Imagemap</button>
</div>
<div id="imapResult" style="font-size:12px;color:#8a8a9a;margin-top:6px"></div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">🤖 Bot 顯示設定</div>
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">名稱</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<input id="senderNameInput" type="text" placeholder="翻譯小助手" style="flex:1;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
<button class="btn btn-primary btn-sm" onclick="saveSenderSettings()">儲存</button>
</div>
<div style="font-size:13px;color:#8a8a9a;margin-bottom:6px">圖示 URL（選填）</div>
<input id="senderIconInput" type="text" placeholder="https://example.com/icon.png" style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px">
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📢 推送訊息</div>
<div class="ch-select-wrap" style="margin-bottom:8px">
<select class="ch-select" id="pushGroupSelect" style="font-size:13px;padding:10px"></select>
</div>
<textarea id="pushText" rows="3" placeholder="輸入要推送的訊息..." style="width:100%;padding:8px;border-radius:8px;border:1px solid #3a3a4e;background:#0d0d1a;color:#e0e0e0;font-size:13px;resize:vertical;margin-bottom:8px"></textarea>
<div style="display:flex;gap:8px">
<button class="btn btn-primary btn-sm" onclick="pushMessage()">📤 推送到群組</button>
<button class="btn btn-dark btn-sm" onclick="broadcastMessage()">📣 推送全體好友</button>
</div>
</div>

<div class="card" style="margin-top:12px">
<div style="font-weight:700;font-size:15px;margin-bottom:10px">📋 Rich Menu 選單</div>
<div style="display:flex;gap:8px;margin-bottom:8px">
<button class="btn btn-primary btn-sm" onclick="createRichMenu()">建立選單</button>
<button class="btn btn-red btn-sm" onclick="deleteRichMenu()">刪除選單</button>
</div>
<div id="richMenuListOld" style="font-size:12px;color:#8a8a9a;margin-top:6px">LINE 底部常駐按鈕</div>
</div>

</div>

</div><!-- mainPage -->
</div><!-- app -->

<div class="toast" id="toast"></div>

<script>
window.onerror=function(msg,url,line,col,err){
  document.body.innerHTML='<div style="color:red;font:16px monospace;padding:20px;white-space:pre-wrap">JS ERROR:\\n'+msg+'\\nLine: '+line+'\\nCol: '+col+'</div>';
  return false;
};
var KEY=window._ADMIN_KEY||'';
var API=window.location.origin+'/api/admin';

function toast(msg){var t=document.getElementById('toast');if(!t)return;t.textContent=msg;t.classList.add('show');setTimeout(function(){t.classList.remove('show')},2000)}

function api(path,method,body){
  method=method||'GET';
  var opts={method:method,headers:{'X-Admin-Key':KEY,'Content-Type':'application/json'}};
  if(body)opts.body=JSON.stringify(body);
  return fetch(API+path,opts).then(function(r){
    if(r.status===403){toast('密碼錯誤');return null}
    var ct=r.headers.get('content-type')||'';
    if(!r.ok||ct.indexOf('application/json')<0){toast('伺服器錯誤('+r.status+')');return null}
    return r.json();
  }).catch(function(e){toast('連線失敗: '+e.message);return null});
}

function doLogin(){
  KEY=document.getElementById('pwInput').value.trim();
  if(!KEY){toast('請輸入密碼');return}
  api('/status').then(function(d){
    if(!d)return;
    document.getElementById('loginPage').style.display='none';
    document.getElementById('mainPage').style.display='block';
    localStorage.setItem('bot_admin_key',KEY);
    if(typeof loadAll==='function') loadAll();
  });
}
</script>
<script>
var FEAT_KEYS=['translation_on','image_on','voice_on','work_order_on'];

var TAB_KEYS=['overview','groups','skip','users','names','storage','packaging','passwords','scrap','insight','settings'];
function switchTab(name){
  document.querySelectorAll('.tab').forEach(function(t,i){
    t.classList.toggle('active',TAB_KEYS[i]===name);
  });
  document.querySelectorAll('.panel').forEach(function(p){p.classList.remove('active')});
  document.getElementById('panel-'+name).classList.add('active');
  if(name==='overview') loadStats();
  if(name==='groups'){loadGroups();loadDM();}
  if(name==='skip') loadGroupSelect();
  if(name==='users'){loadUsersGroupSelect();loadUsers();}
  if(name==='names') loadNames();
  if(name==='storage') loadStorageStats();
  if(name==='packaging') loadPackagingStats();
  if(name==='passwords') loadPasswords();
  if(name==='scrap') loadScrap();
  if(name==='insight') loadInsightTab();
  if(name==='settings') loadFeatureSettings();
}

function loadAll(){loadStats();loadGroups();loadDM();loadGroupSelect();loadUsersGroupSelect();loadUsers();loadNames();loadStorageStats()}

async function loadStats(){
  var d=await api('/stats');
  if(!d)return;
  var h=Math.floor((d.uptime_seconds||0)/3600);
  var m=Math.floor(((d.uptime_seconds||0)%3600)/60);
  document.getElementById('st-uptime').textContent=h+'h '+m+'m';
  setStatVal('st-text',d.text_translations||0);
  setStatVal('st-image',d.image_translations||0);
  setStatVal('st-voice',d.voice_translations||0);
  setStatVal('st-wo',d.work_order_detections||0);
  setStatVal('st-cmd',d.commands||0);
  setStatVal('st-cust',d.customers||0);
  setStatVal('st-groups',d.groups||0);
  setStatVal('st-dm-users',d.dm_users||0);
  var tt=d.tokens_total||0;
  document.getElementById('st-tokens').textContent=tt.toLocaleString();
  document.getElementById('st-cost').textContent='$'+(d.estimated_cost_usd||0).toFixed(4);
}
function setStatVal(id,val){
  var el=document.getElementById(id);
  el.textContent=val;
  if(val>0)el.classList.add('highlight');
  else el.classList.remove('highlight');
}

var _groupList=[];
async function loadGroups(){
  var d=await api('/groups');
  if(!d)return;
  _groupList=d.groups||[];
  var el=document.getElementById('groupList');
  if(!_groupList.length){el.innerHTML='<div class="empty">尚無群組紀錄<br>Bot 收到群組訊息後會自動記錄</div>';return}
  var html='';
  for(var i=0;i<_groupList.length;i++){
    var g=_groupList[i];
    var skipCt=g.skip_count||0;
    var memberCt=g.member_count?g.member_count+'人':'--';
    html+='<div class="card">'+
      '<div class="card-title"><div><span style="font-weight:700;font-size:16px">#'+(g.name||'(未知群組)')+'</span><span style="font-size:12px;color:#8a8a9a;margin-left:8px">👥'+memberCt+'</span></div>'+
      '<span class="badge '+(g.translation_on?'badge-on':'badge-off')+'" style="cursor:pointer" onclick="toggleFeat('+i+',0)">'+(g.translation_on?'翻譯開':'翻譯關')+'</span></div>'+
      '<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin:8px 0">'+
      '<span class="card-sub">中文 ⇄ 🇮🇩 印尼文</span>'+
      '<span class="card-sub">｜跳過: '+skipCt+'人</span></div>'+
      '<div class="feat-badges">'+
      '<span class="feat-badge '+(g.image_on?'on':'off')+'" style="cursor:pointer" onclick="toggleFeat('+i+',1)">🖼️ '+(g.image_on?'圖片開':'圖片關')+'</span>'+
      '<span class="feat-badge '+(g.voice_on?'on':'off')+'" style="cursor:pointer" onclick="toggleFeat('+i+',2)">🎤 '+(g.voice_on?'語音開':'語音關')+'</span>'+
      '<span class="feat-badge '+(g.work_order_on?'on':'off')+'" style="cursor:pointer" onclick="toggleFeat('+i+',3)">📋 '+(g.work_order_on?'工單開':'工單關')+'</span></div>'+
      buildCmdBadges(g, i)+
      '<div style="display:flex;align-items:center;justify-content:space-between;margin:10px 0;padding:10px 12px;background:rgba(124,111,239,.08);border-radius:8px;border:1px solid rgba(124,111,239,.2)">'+
      '<div><span style="font-size:12px;color:#8a8a9a">累計花費</span><br><span style="font-size:18px;font-weight:700;color:#7c6fef">NT$'+(g.cost_twd||0).toFixed(1)+'</span></div>'+
      '<button class="btn btn-dark btn-sm" style="font-size:12px" onclick="resetCost('+i+')">歸零</button></div>'+
      '<button class="btn btn-red btn-sm" onclick="leaveGroup('+i+')">退出群組: '+(g.name||g.id.substring(0,12))+'</button></div>';
  }
  el.innerHTML=html;
}

function toggleFeat(idx,keyIdx){
  var g=_groupList[idx];if(!g)return;
  var key=FEAT_KEYS[keyIdx];
  var cur=g[key];
  var body={group_id:g.id};body[key]=!cur;
  api('/groups/settings','POST',body).then(function(d){if(d){toast('已更新');loadGroups()}});
}
var CMD_DEFS=[["pw1","🔑密碼1"],["pw2","🏭密碼2"],["pkg","📦包裝"],["scrap","🎨廢料"],["qry","🔍儲區"],["notice","📢公告"]];
function buildCmdBadges(g,idx){
  var ce=g.cmd_enabled||{};
  var h='<div class="feat-badges" style="margin-top:4px">';
  for(var c=0;c<CMD_DEFS.length;c++){
    var key=CMD_DEFS[c][0],label=CMD_DEFS[c][1];
    var on=ce[key]!==false;
    h+='<span class="feat-badge '+(on?'on':'off')+'" style="cursor:pointer" onclick="toggleCmd('+idx+',&apos;'+key+'&apos;,'+(!on)+')">'+label+(on?'開':'關')+'</span>';
  }
  return h+'</div>';
}
function toggleCmd(idx,key,val){
  var g=_groupList[idx];if(!g)return;
  api('/groups/settings','POST',{group_id:g.id,cmd_toggle:key,cmd_val:val}).then(function(d){if(d){toast('已更新');loadGroups()}});
}
function leaveGroup(idx){
  var g=_groupList[idx];if(!g)return;
  if(!confirm('確定退出「'+(g.name||g.id)+'」？'))return;
  api('/groups/leave','POST',{group_id:g.id}).then(function(d){if(d){toast(d.message||'已退出');loadGroups();loadGroupSelect()}});
}
function resetCost(idx){
  var g=_groupList[idx];if(!g)return;
  if(!confirm('確定歸零累計花費？'))return;
  api('/groups/reset-cost','POST',{group_id:g.id}).then(function(d){if(d){toast('已歸零');loadGroups()}});
}

var _dmUsers=[];
async function loadDM(){
  var d=await api('/dm');
  if(!d)return;
  document.getElementById('dmToggle').checked=d.master_enabled;
  _dmUsers=d.known_users||[];
  var el=document.getElementById('dmWlList');
  if(!_dmUsers.length){el.innerHTML='<div style="padding:8px 0;font-size:13px;color:#5a5a6a">尚無人私訊過 Bot</div>';return}
  var html='';
  for(var i=0;i<_dmUsers.length;i++){
    var u=_dmUsers[i];
    html+='<div class="wl-item" style="border-color:#2a2a3e"><span>'+u.name+'</span>'+
    '<label class="toggle"><input type="checkbox" '+(u.whitelisted?'checked':'')+
    ' onchange="toggleDmWl('+i+',this.checked)"><span class="slider"></span></label></div>';
  }
  el.innerHTML=html;
}
async function toggleDM(){
  var on=document.getElementById('dmToggle').checked;
  var d=await api('/dm','POST',{master_enabled:on});
  if(d) toast(on?'DM 已開啟':'DM 已關閉');
}
function toggleDmWl(idx,on){
  var u=_dmUsers[idx];if(!u)return;
  api('/dm/whitelist','POST',{user_id:u.user_id,action:on?'add':'remove'}).then(function(d){
    if(d) toast(on?'已加入白名單':'已移出白名單');
  });
}

async function loadGroupSelect(){
  var d=await api('/groups');
  if(!d)return;
  var sel=document.getElementById('skipGroupSelect');
  var cur=sel.value;
  sel.innerHTML='<option value="">選擇群組...</option>';
  var groups=d.groups||[];
  for(var i=0;i<groups.length;i++){
    var g=groups[i];
    var opt=document.createElement('option');
    opt.value=g.id;opt.textContent='#'+(g.name||g.id.substring(0,16));
    sel.appendChild(opt);
  }
  if(cur)sel.value=cur;
}

var _skipUsers=[];
async function loadSkipList(){
  var gid=document.getElementById('skipGroupSelect').value;
  var el=document.getElementById('skipListContent');
  if(!gid){el.innerHTML='<div class="empty">請先選擇群組</div>';return}
  var d=await api('/skip?group_id='+gid);
  if(!d)return;
  _skipUsers=d.users||[];
  if(!_skipUsers.length){
    el.innerHTML='<div class="empty">尚無成員紀錄<br>成員在群組發訊息後會自動出現</div>';
    return;
  }
  var html='';
  for(var i=0;i<_skipUsers.length;i++){
    var u=_skipUsers[i];
    html+='<div class="wl-item"><span style="font-size:15px">'+u.name+'</span>'+
    '<label class="toggle"><input type="checkbox" '+(u.skipped?'checked':'')+
    ' onchange="toggleSkip('+i+',this.checked)"><span class="slider"></span></label></div>';
  }
  el.innerHTML=html;
}
function toggleSkip(idx,on){
  var gid=document.getElementById('skipGroupSelect').value;
  var u=_skipUsers[idx];if(!u)return;
  api('/skip','POST',{group_id:gid,user_id:u.user_id,action:on?'add':'remove'}).then(function(d){
    if(d) toast(on?'已跳過翻譯':'已恢復翻譯');
  });
}

async function loadUsersGroupSelect(){
  var d=await api('/groups');
  if(!d)return;
  var sel=document.getElementById('usersGroupSelect');
  var cur=sel.value;
  sel.innerHTML='<option value="">全部使用者</option>';
  var groups=d.groups||[];
  for(var i=0;i<groups.length;i++){
    var g=groups[i];
    var opt=document.createElement('option');
    opt.value=g.id;opt.textContent='#'+(g.name||g.id.substring(0,16));
    sel.appendChild(opt);
  }
  if(cur)sel.value=cur;
}

var _allUsers=[];
async function loadUsers(){
  var gid=document.getElementById('usersGroupSelect').value;
  var path=gid?'/users?group_id='+gid:'/users';
  var d=await api(path);
  if(!d)return;
  _allUsers=d.users||[];
  var el=document.getElementById('usersList');
  if(!_allUsers.length){el.innerHTML='<div class="empty">尚無使用者紀錄<br>使用者互動後會自動出現</div>';return}
  var html='';
  for(var i=0;i<_allUsers.length;i++){
    var u=_allUsers[i];
    var langBadge=u.line_lang?'<span class="badge badge-on" style="font-size:11px">'+u.line_lang+'</span>':'';
    html+='<div class="user-card">'+
      '<div style="display:flex;justify-content:space-between;align-items:flex-start">'+
      '<div><div class="user-name">'+u.name+'</div><div class="user-id">ID: '+u.user_id+'</div></div>'+langBadge+'</div>'+
      '<div class="user-admin-row">'+
      '<span class="admin-label">🔑 管理員</span>'+
      '<label class="toggle"><input type="checkbox" '+(u.is_admin?'checked':'')+
      ' onchange="toggleAdmin('+i+',this.checked)"><span class="slider"></span></label>'+
      '</div></div>';
  }
  el.innerHTML=html;
}
function toggleAdmin(idx,on){
  var u=_allUsers[idx];if(!u)return;
  api('/users/admin','POST',{user_id:u.user_id,is_admin:on}).then(function(d){
    if(d) toast(on?'已設為管理員':'已取消管理員');
  });
}

var _protectedNames=[];
async function loadNames(){
  var d=await api('/names');
  if(!d)return;
  _protectedNames=d.names||[];
  var el=document.getElementById('namesList');
  document.getElementById('namesCount').textContent='共 '+_protectedNames.length+' 個保護名稱';
  if(!_protectedNames.length){el.innerHTML='<div style="padding:8px 0;font-size:13px;color:#5a5a6a">尚無保護名稱</div>';return}
  var html='<div style="display:flex;flex-wrap:wrap;gap:8px">';
  for(var i=0;i<_protectedNames.length;i++){
    html+='<span style="display:inline-flex;align-items:center;gap:6px;padding:6px 12px;background:#2a2a3e;border:1px solid #3a3a4e;border-radius:8px;font-size:13px">'+
    _protectedNames[i]+'<span style="cursor:pointer;color:#f04747;font-weight:700;font-size:15px" onclick="removeName('+i+')"> ×</span></span>';
  }
  html+='</div>';
  el.innerHTML=html;
}
async function addName(){
  var inp=document.getElementById('newNameInput');
  var name=inp.value.trim();
  if(!name){toast('請輸入名字');return}
  var d=await api('/names','POST',{action:'add',name:name});
  if(d){toast('已新增: '+name);inp.value='';loadNames()}
}
function removeName(idx){
  var name=_protectedNames[idx];
  if(!name)return;
  if(!confirm('確定移除「'+name+'」？'))return;
  api('/names','POST',{action:'remove',name:name}).then(function(d){if(d){toast('已移除: '+name);loadNames()}});
}

async function loadStorageStats(){
  var d=await api('/storage/stats');
  if(!d)return;
  document.getElementById('storageStats').innerHTML='客戶數: <strong style="color:#7c6fef">'+d.count+'</strong>';
}

var storageFileData=null;
function previewStorage(){
  var f=document.getElementById('storageFile').files[0];
  if(!f)return;
  document.getElementById('storageFileName').textContent='📄 '+f.name;
  storageFileData=f;
  document.getElementById('storageActions').style.display='block';
  document.getElementById('storagePreview').innerHTML='<div class="card"><div class="card-sub">點「確認更新」上傳並解析</div></div>';
}

async function uploadStorage(){
  if(!storageFileData){toast('請先選擇檔案');return}
  var fd=new FormData();
  fd.append('file',storageFileData);
  try{
    var r=await fetch(API+'/storage/upload',{method:'POST',headers:{'X-Admin-Key':KEY},body:fd});
    if(!r.ok){toast('上傳失敗('+r.status+')');return}
    var d=await r.json();
    if(d.error){toast(d.error);return}
    toast(d.message||'更新成功');
    var ghStatus=d.github?'✅ 已推送 GitHub，Render 將自動部署':'⚠️ GitHub 推送失敗，僅暫時生效';
    document.getElementById('storageActions').style.display='none';
    document.getElementById('storagePreview').innerHTML='<div class="card"><div style="color:#43b581;font-weight:600">✅ 已更新 '+d.count+' 筆客戶資料</div><div class="card-sub" style="margin-top:4px">'+ghStatus+'</div></div>';
    loadStorageStats();
  }catch(e){toast('上傳失敗: '+e)}
}

async function downloadJson(){
  try{
    var r=await fetch(API+'/storage/json',{headers:{'X-Admin-Key':KEY}});
    var blob=await r.blob();
    var url=URL.createObjectURL(blob);
    var a=document.createElement('a');
    a.href=url;a.download='storage_data.json';a.click();
    URL.revokeObjectURL(url);
    toast('JSON 已下載');
  }catch(e){toast('下載失敗')}
}

// ─── Packaging ───
async function loadPackagingStats(){
  var d=await api('/packaging/stats');
  if(d) document.getElementById('packagingStats').textContent='共 '+d.count+' 筆包裝碼';
}
function previewPackaging(){
  var f=document.getElementById('packagingFile').files[0];
  if(!f)return;
  document.getElementById('packagingFileName').textContent='📄 '+f.name;
  document.getElementById('packagingActions').style.display='block';
}
async function uploadPackaging(){
  var f=document.getElementById('packagingFile').files[0];
  if(!f){toast('請選擇檔案');return;}
  var fd=new FormData();fd.append('file',f);
  try{
    var r=await fetch(API+'/packaging/upload',{method:'POST',headers:{'X-Admin-Key':KEY},body:fd});
    if(!r.ok){toast('上傳失敗('+r.status+')');return}
    var d=await r.json();
    if(d.ok){toast(d.message);loadPackagingStats();document.getElementById('packagingActions').style.display='none';}
    else{toast(d.error||'上傳失敗');}
  }catch(e){toast('上傳失敗: '+e);}
}
async function downloadPackagingJson(){
  try{
    var r=await fetch(API+'/packaging/json',{headers:{'X-Admin-Key':KEY}});
    var blob=await r.blob();
    var url=URL.createObjectURL(blob);
    var a=document.createElement('a');
    a.href=url;a.download='packaging_data.json';a.click();
    URL.revokeObjectURL(url);
    toast('JSON 已下載');
  }catch(e){toast('下載失敗')}
}

// ─── Passwords ───
async function loadPasswords(){
  var d=await api('/passwords');
  if(!d)return;
  document.getElementById('pw1Input').value=d.pw1||'';
  document.getElementById('pw2Input').value=d.pw2||'';
}
async function savePasswords(){
  var pw1=document.getElementById('pw1Input').value;
  var pw2=document.getElementById('pw2Input').value;
  var d=await api('/passwords','POST',{pw1:pw1,pw2:pw2});
  document.getElementById('pwSaveResult').textContent=d&&d.ok?'✅ 已儲存':'❌ 儲存失敗';
  if(d&&d.ok)toast('密碼已更新');
}

// ─── Scrap ───
async function loadScrap(){
  var d=await api('/scrap');
  if(!d)return;
  document.getElementById('scrapInput').value=d.text||'';
}
async function saveScrap(){
  var text=document.getElementById('scrapInput').value;
  var d=await api('/scrap','POST',{text:text});
  document.getElementById('scrapSaveResult').textContent=d&&d.ok?'✅ 已儲存':'❌ 儲存失敗';
  if(d&&d.ok)toast('廢料色資訊已更新');
}

// ─── Insight Tab ───
async function loadInsightTab(){
  // Load trend
  var t=await api('/insight/trend?days=7');
  if(t&&t.trend&&t.trend.length){
    var trend=t.trend.reverse();
    var labels=[];var followers=[];var blocks=[];
    for(var i=0;i<trend.length;i++){
      var d=trend[i].date;
      labels.push(d.substring(4,6)+'/'+d.substring(6,8));
      followers.push(trend[i].followers||0);
      blocks.push(trend[i].blocks||0);
    }
    drawTrendChart(labels,followers,blocks);
    var html='<table style="width:100%;font-size:12px;border-collapse:collapse">';
    html+='<tr style="color:#8a8a9a"><td>日期</td><td style="text-align:right">好友數</td><td style="text-align:right">封鎖</td></tr>';
    for(var i=0;i<trend.length;i++){
      html+='<tr><td>'+labels[i]+'</td><td style="text-align:right">'+(trend[i].followers||'-')+'</td><td style="text-align:right;color:#f04747">'+(trend[i].blocks||'-')+'</td></tr>';
    }
    html+='</table>';
    document.getElementById('insightTrendData').innerHTML=html;
  }else{
    document.getElementById('insightTrendData').textContent='無趨勢資料（需至少20名好友且帳號開通超過7天）';
  }
  // Load demographics
  var ins=await api('/insight');
  if(ins){
    var dhtml='';
    if(ins.demographics&&ins.demographics.available){
      if(ins.demographics.genders){
        dhtml+='<div style="font-weight:600;margin-bottom:4px">性別</div>';
        var g=ins.demographics.genders;
        if(Array.isArray(g)){
          for(var i=0;i<g.length;i++){
            var pct=g[i].percentage?Math.round(g[i].percentage*100)+'%':'';
            dhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+(g[i].gender||'-')+' '+pct+'</span> ';
          }
        }else{dhtml+=JSON.stringify(g)}
        dhtml+='<br>';
      }
      if(ins.demographics.ages){
        dhtml+='<div style="font-weight:600;margin:8px 0 4px">年齡</div>';
        var a=ins.demographics.ages;
        if(Array.isArray(a)){
          for(var i=0;i<a.length;i++){
            var pct=a[i].percentage?Math.round(a[i].percentage*100)+'%':'';
            dhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+(a[i].age||'-')+' '+pct+'</span> ';
          }
        }else{dhtml+=JSON.stringify(a)}
        dhtml+='<br>';
      }
      if(ins.demographics.areas){
        dhtml+='<div style="font-weight:600;margin:8px 0 4px">地區</div>';
        var ar=ins.demographics.areas;
        if(Array.isArray(ar)){
          for(var i=0;i<Math.min(ar.length,10);i++){
            var pct=ar[i].percentage?Math.round(ar[i].percentage*100)+'%':'';
            dhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+(ar[i].area||'-')+' '+pct+'</span> ';
          }
        }else{dhtml+=JSON.stringify(ar)}
      }
    }else{dhtml='人口統計需至少20名好友'}
    document.getElementById('insightDemoData').innerHTML=dhtml||'無資料';
    // Delivery
    if(ins.delivery){
      var s=ins.delivery;
      var dv='日期: '+(s.date||'-');
      if(s.reply!==null)dv+='<br>Reply: '+s.reply;
      if(s.push!==null)dv+=' ｜ Push: '+s.push;
      if(s.multicast!==null)dv+=' ｜ Multicast: '+s.multicast;
      if(s.broadcast!==null)dv+='<br>Broadcast: '+s.broadcast;
      document.getElementById('insightDelivery').innerHTML=dv;
    }
  }
}
function drawTrendChart(labels,followers,blocks){
  var canvas=document.getElementById('trendCanvas');
  if(!canvas)return;
  var ctx=canvas.getContext('2d');
  var W=canvas.width;var H=canvas.height;
  ctx.clearRect(0,0,W,H);
  var pad={t:20,r:10,b:30,l:50};
  var cw=W-pad.l-pad.r;var ch=H-pad.t-pad.b;
  var maxF=Math.max.apply(null,followers)||1;
  // Grid
  ctx.strokeStyle='#2a2a3e';ctx.lineWidth=1;
  for(var i=0;i<4;i++){
    var y=pad.t+ch*(i/3);
    ctx.beginPath();ctx.moveTo(pad.l,y);ctx.lineTo(W-pad.r,y);ctx.stroke();
  }
  // Labels
  ctx.fillStyle='#8a8a9a';ctx.font='11px sans-serif';ctx.textAlign='center';
  for(var i=0;i<labels.length;i++){
    var x=pad.l+cw*(i/(labels.length-1||1));
    ctx.fillText(labels[i],x,H-8);
  }
  ctx.textAlign='right';
  for(var i=0;i<4;i++){
    var y=pad.t+ch*(i/3);
    var val=Math.round(maxF*(1-i/3));
    ctx.fillText(val,pad.l-6,y+4);
  }
  // Follower line
  ctx.strokeStyle='#7c6fef';ctx.lineWidth=2;ctx.beginPath();
  for(var i=0;i<followers.length;i++){
    var x=pad.l+cw*(i/(followers.length-1||1));
    var y=pad.t+ch*(1-followers[i]/maxF);
    if(i===0)ctx.moveTo(x,y);else ctx.lineTo(x,y);
  }
  ctx.stroke();
  // Dots
  ctx.fillStyle='#7c6fef';
  for(var i=0;i<followers.length;i++){
    var x=pad.l+cw*(i/(followers.length-1||1));
    var y=pad.t+ch*(1-followers[i]/maxF);
    ctx.beginPath();ctx.arc(x,y,3,0,Math.PI*2);ctx.fill();
  }
}

// ─── Rich Menu Image Upload ───
var _rmImageData=null;
function previewRmImage(input){
  if(!input.files||!input.files[0])return;
  var file=input.files[0];
  var reader=new FileReader();
  reader.onload=function(e){
    _rmImageData=e.target.result;
    document.getElementById('rmImagePreviewImg').src=_rmImageData;
    document.getElementById('rmImagePreview').style.display='block';
    document.getElementById('rmUploadBtn').style.display='inline-block';
  };
  reader.readAsDataURL(file);
}
async function uploadRmImage(){
  var sel=document.getElementById('rmUploadSelect');
  var rmId=sel.value;
  if(!rmId){toast('請選擇 Rich Menu');return}
  if(!_rmImageData){toast('請選擇圖片');return}
  document.getElementById('rmUploadResult').textContent='上傳中...';
  var d=await api('/richmenu/upload/'+rmId,'POST',{image:_rmImageData});
  if(d&&d.ok){
    document.getElementById('rmUploadResult').innerHTML='<span style="color:#43b581">✅ 上傳成功</span>';
    toast('圖片已上傳');
    _rmImageData=null;
    document.getElementById('rmImagePreview').style.display='none';
    document.getElementById('rmUploadBtn').style.display='none';
  }else{
    document.getElementById('rmUploadResult').innerHTML='<span style="color:#f04747">❌ 上傳失敗: '+(d&&d.error||'unknown')+'</span>';
  }
}

// ─── Imagemap Send ───
var _imapActionCount=0;
function addImapAction(){
  var div=document.getElementById('imapActions');
  var idx=_imapActionCount++;
  var html='<div id="imapAct'+idx+'" style="background:#0d0d1a;border:1px solid #3a3a4e;border-radius:8px;padding:8px;margin-bottom:6px;font-size:12px">';
  html+='<div style="display:flex;gap:6px;margin-bottom:4px;align-items:center">';
  html+='<select id="imapType'+idx+'" style="padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px"><option value="message">文字訊息</option><option value="uri">開啟網址</option></select>';
  html+='<input id="imapText'+idx+'" placeholder="訊息文字 或 URL" style="flex:1;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<span style="color:#f04747;cursor:pointer" onclick="document.getElementById(&apos;imapAct'+idx+'&apos;).remove()">✕</span>';
  html+='</div>';
  html+='<div style="display:flex;gap:4px">';
  html+='<input id="imapX'+idx+'" type="number" placeholder="X" value="0" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<input id="imapY'+idx+'" type="number" placeholder="Y" value="0" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<input id="imapAW'+idx+'" type="number" placeholder="W" value="1040" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='<input id="imapAH'+idx+'" type="number" placeholder="H" value="1040" style="width:60px;padding:4px;border-radius:4px;border:1px solid #3a3a4e;background:#1a1a2e;color:#e0e0e0;font-size:12px">';
  html+='</div></div>';
  div.insertAdjacentHTML('beforeend',html);
}
async function sendImap(){
  var sel=document.getElementById('imapGroupSelect');
  var to=sel.value;
  if(!to){toast('請選擇群組');return}
  var baseUrl=document.getElementById('imapBaseUrl').value.trim();
  if(!baseUrl){toast('請輸入圖片 Base URL');return}
  var w=parseInt(document.getElementById('imapW').value)||1040;
  var h=parseInt(document.getElementById('imapH').value)||1040;
  var actions=[];
  for(var i=0;i<_imapActionCount;i++){
    var el=document.getElementById('imapAct'+i);
    if(!el)continue;
    var type=document.getElementById('imapType'+i).value;
    var text=document.getElementById('imapText'+i).value.trim();
    var x=parseInt(document.getElementById('imapX'+i).value)||0;
    var y=parseInt(document.getElementById('imapY'+i).value)||0;
    var aw=parseInt(document.getElementById('imapAW'+i).value)||1040;
    var ah=parseInt(document.getElementById('imapAH'+i).value)||1040;
    if(!text)continue;
    var act={type:type,x:x,y:y,w:aw,h:ah};
    if(type==='uri')act.uri=text;else act.text=text;
    actions.push(act);
  }
  if(!actions.length){toast('請新增至少一個點擊區域');return}
  if(!confirm('確定發送 Imagemap 到此群組？'))return;
  var d=await api('/imagemap/send','POST',{to:to,base_url:baseUrl,width:w,height:h,actions:actions});
  if(d&&d.ok){
    document.getElementById('imapResult').innerHTML='<span style="color:#43b581">✅ 已發送</span>';
    toast('Imagemap 已發送');
  }else{
    document.getElementById('imapResult').innerHTML='<span style="color:#f04747">❌ 發送失敗: '+(d&&d.error||'unknown')+'</span>';
  }
}

// ─── Feature Settings ───
var _settingsGid='';
async function loadFeatureSettings(){
  await loadSettingsGroupSelects();
  _settingsGid='';
  await _loadFeatures('');
  // Load LINE quota
  var s=await api('/stats');
  if(s){
    var info='';
    if(s.line_quota){
      var q=s.line_quota;
      if(q.quota)info+='配額: '+q.quota.toLocaleString()+' 則/月';
      if(q.used!==null&&q.used!==undefined)info+=(info?' ｜ ':'')+'已用: '+q.used.toLocaleString()+' 則';
    }
    if(s.followers)info+=(info?'<br>':'')+'好友: '+s.followers;
    if(s.unfollowers)info+=' ｜ 封鎖: '+s.unfollowers;
    document.getElementById('lineQuotaInfo').innerHTML=info||'無法取得';
  }
  // Load Insight API data
  var ins=await api('/insight');
  if(ins){
    var ihtml='';
    if(ins.delivery&&ins.delivery.reply)ihtml+='昨日回覆: '+ins.delivery.reply+' 則';
    if(ins.demographics&&ins.demographics.available){
      if(ins.demographics.genders){ihtml+=(ihtml?'<br>':'')+'性別分布: '+JSON.stringify(ins.demographics.genders)}
    }
    document.getElementById('lineInsight').innerHTML=ihtml;
  }
  // Load Webhook info
  try{
    var wh=await api('/webhook');
    if(wh&&wh.webhook){
      var whInfo='URL: '+(wh.webhook.endpoint||'(未設定)');
      if(wh.webhook.active!==null)whInfo+=' ｜ '+(wh.webhook.active?'✅ 啟用':'❌ 停用');
      document.getElementById('webhookInfo').innerHTML=whInfo;
    }
  }catch(e){document.getElementById('webhookInfo').innerHTML='無法取得';}
  // Load Rich Menu list
  var rm=await api('/richmenu/list');
  if(rm&&rm.menus&&rm.menus.length){
    var rmhtml='目前選單: ';
    var rmSel=document.getElementById('rmUploadSelect');
    rmSel.innerHTML='<option value="">選擇 Rich Menu...</option>';
    for(var r=0;r<rm.menus.length;r++){
      rmhtml+='<span class="badge badge-on" style="font-size:11px;margin:2px">'+rm.menus[r].name+'</span> ';
      var opt=document.createElement('option');
      opt.value=rm.menus[r].id;opt.textContent=rm.menus[r].name||rm.menus[r].id.substring(0,16);
      rmSel.appendChild(opt);
    }
    document.getElementById('richMenuList').innerHTML=rmhtml;
  }else{
    document.getElementById('richMenuList').textContent='尚未建立 Rich Menu';
    document.getElementById('rmUploadSelect').innerHTML='<option value="">無選單</option>';
  }
  // Populate imagemap group select
  var imSel=document.getElementById('imapGroupSelect');
  imSel.innerHTML='<option value="">選擇群組...</option>';
  if(_groupList&&_groupList.length){
    for(var g=0;g<_groupList.length;g++){
      var opt=document.createElement('option');
      opt.value=_groupList[g].id;opt.textContent='#'+(_groupList[g].name||_groupList[g].id.substring(0,16));
      imSel.appendChild(opt);
    }
  }
  // Load delivery stats, rich menu default, aliases
  loadDeliveryStats();
  loadRichMenuDefault();
  loadAliases();
}
async function loadFeatureSettingsForGroup(){
  _settingsGid=document.getElementById('settingsGroupSelect').value;
  await _loadFeatures(_settingsGid);
}
async function _loadFeatures(gid){
  var path=gid?'/features?group_id='+encodeURIComponent(gid):'/features';
  var d=await api(path);
  if(!d)return;
  document.getElementById('welcomeToggle').checked=d.welcome_enabled;
  document.getElementById('welcomeZh').value=d.welcome_text_zh||'';
  document.getElementById('welcomeId').value=d.welcome_text_id||'';
  document.getElementById('flexToggle').checked=d.flex_enabled;
  document.getElementById('qrToggle').checked=d.quick_reply_enabled;
  document.getElementById('silentToggle').checked=d.silent_mode;
  document.getElementById('videoToggle').checked=d.video_ocr_enabled!==false;
  document.getElementById('locationToggle').checked=d.location_translate_enabled!==false;
  document.getElementById('markReadToggle').checked=d.mark_read_enabled!==false;
  document.getElementById('retryKeyToggle').checked=d.retry_key_enabled!==false;
  document.getElementById('cameraQrToggle').checked=d.camera_qr_enabled||false;
  document.getElementById('clipboardQrToggle').checked=d.clipboard_qr_enabled||false;
  document.getElementById('cameraRollQrToggle').checked=d.camera_roll_qr_enabled||false;
  document.getElementById('locationQrToggle').checked=d.location_qr_enabled||false;
  document.getElementById('toneSelect').value=d.translation_tone||'casual';
  document.getElementById('toneCustom').value=d.translation_tone_custom||'';
  // Model settings (global only, not per-group)
  if(!gid){
    document.getElementById('modelDefault').value=d.model_default||'gpt-4o-mini';
    document.getElementById('modelUpgrade').value=d.model_upgrade||'gpt-4o';
    document.getElementById('modelThreshold').value=d.model_threshold||0;
  }
  document.getElementById('senderNameInput').value=d.sender_name||'翻譯小助手';
  document.getElementById('senderIconInput').value=d.sender_icon||'';
  var cb=document.getElementById('settingsCustomBadge');
  if(gid&&d.is_customized)cb.style.display='block';
  else cb.style.display='none';
  if(d.bot_info&&d.bot_info.name){
    var qi=document.getElementById('lineQuotaInfo');
    if(qi&&!qi.innerHTML.includes('Bot:'))qi.innerHTML+='<br>Bot: '+d.bot_info.name;
  }
}
function toggleFeatureSetting(key,val){
  var body={};body[key]=val;
  if(_settingsGid)body.group_id=_settingsGid;
  api('/features','POST',body).then(function(d){if(d)toast(_settingsGid?'群組設定已更新':'全域設定已更新')});
}
function saveModelSettings(){
  var md=document.getElementById('modelDefault').value;
  var mu=document.getElementById('modelUpgrade').value;
  var mt=parseInt(document.getElementById('modelThreshold').value)||0;
  api('/features','POST',{model_default:md,model_upgrade:mu,model_threshold:mt}).then(function(d){
    if(d){
      toast('模型設定已儲存');
      var info=mt>0?'≥'+mt+'字用 '+mu+'，其餘用 '+md:'全部用 '+md;
      document.getElementById('modelSaveResult').innerHTML='<span style="color:#43b581">✅ '+info+'</span>';
    }
  });
}
function saveWelcomeText(){
  var zh=document.getElementById('welcomeZh').value;
  var id=document.getElementById('welcomeId').value;
  var body={welcome_text_zh:zh,welcome_text_id:id};
  if(_settingsGid)body.group_id=_settingsGid;
  api('/features','POST',body).then(function(d){if(d)toast('歡迎詞已儲存')});
}
async function resetGroupSettings(){
  if(!_settingsGid){toast('請先選擇群組');return}
  if(!confirm('確定重設此群組為全域預設？'))return;
  var body={group_id:_settingsGid,reset:true};
  var d=await api('/features/reset','POST',body);
  if(d&&d.ok){toast('已重設');loadFeatureSettingsForGroup()}
  else toast('重設失敗');
}
function saveSenderSettings(){
  var name=document.getElementById('senderNameInput').value.trim();
  var icon=document.getElementById('senderIconInput').value.trim();
  if(!name){toast('請輸入名稱');return}
  api('/features','POST',{sender_name:name,sender_icon:icon}).then(function(d){if(d)toast('已更新')});
}
async function broadcastMessage(){
  var text=document.getElementById('pushText').value.trim();
  if(!text){toast('請輸入訊息');return}
  if(!confirm('確定推送給所有好友？'))return;
  var d=await api('/broadcast','POST',{text:text});
  if(d&&d.ok){toast('已推送全體');document.getElementById('pushText').value=''}
  else toast('推送失敗');
}
async function testWebhook(){
  document.getElementById('webhookTestResult').innerHTML='測試中...';
  var d=await api('/webhook/test','POST',{});
  if(d){
    var r=d.success?'✅ 成功':'❌ 失敗';
    if(d.status_code)r+=' ('+d.status_code+')';
    if(d.reason)r+=' - '+d.reason;
    document.getElementById('webhookTestResult').innerHTML=r;
  }else{document.getElementById('webhookTestResult').innerHTML='測試失敗';}
}
async function loadFollowers(){
  document.getElementById('followersList').innerHTML='載入中...';
  var d=await api('/followers');
  if(d&&d.followers){
    var h='總計: '+d.count+' 人<br>';
    for(var i=0;i<d.followers.length&&i<100;i++){
      var f=d.followers[i];
      h+='<span style="font-size:11px">'+(f.name||f.user_id.substr(0,12)+'...')+'</span> ';
    }
    if(d.count>100)h+='<br>...僅顯示前100人';
    document.getElementById('followersList').innerHTML=h;
  }else{document.getElementById('followersList').innerHTML='無法載入';}
}
async function loadDeliveryStats(){
  var d=await api('/delivery');
  if(d&&d.delivery){
    var s=d.delivery;
    var h='日期: '+(s.date||'-');
    if(s.reply!==null)h+='<br>Reply: '+s.reply;
    if(s.push!==null)h+=' ｜ Push: '+s.push;
    if(s.multicast!==null)h+=' ｜ Multicast: '+s.multicast;
    if(s.broadcast!==null)h+='<br>Broadcast: '+s.broadcast;
    document.getElementById('deliveryStats').innerHTML=h;
  }else{document.getElementById('deliveryStats').innerHTML='無法取得';}
}
async function checkUserMenu(){
  var uid=document.getElementById('rmUserIdInput').value.trim();
  if(!uid){toast('請輸入 user ID');return}
  var d=await api('/richmenu/user/'+uid);
  if(d){
    document.getElementById('rmUserResult').innerHTML=d.rich_menu_id?('綁定: '+d.rich_menu_id):'無綁定（使用預設）';
  }
}
async function loadAliases(){
  var d=await api('/richmenu/alias/list');
  if(d&&d.aliases){
    if(!d.aliases.length){document.getElementById('rmAliasList').innerHTML='（無 alias）';return;}
    var h='';
    for(var i=0;i<d.aliases.length;i++){
      var a=d.aliases[i];
      h+='<div style="padding:4px 0;border-bottom:1px solid #2a2a3e"><b>'+a.alias_id+'</b> → '+a.rich_menu_id+'</div>';
    }
    document.getElementById('rmAliasList').innerHTML=h;
  }
}
async function loadRichMenuDefault(){
  var d=await api('/richmenu/default');
  if(d){
    document.getElementById('richMenuDefault').innerHTML='預設選單 ID: '+(d.default_rich_menu_id||'（未設定）');
  }
}
async function pushMessage(){
  var gid=document.getElementById('pushGroupSelect').value;
  var text=document.getElementById('pushText').value.trim();
  if(!gid){toast('請選擇群組');return}
  if(!text){toast('請輸入訊息');return}
  if(!confirm('確定推送到群組？'))return;
  var d=await api('/push','POST',{group_id:gid,text:text});
  if(d&&d.ok){toast('已推送');document.getElementById('pushText').value=''}
  else toast('推送失敗');
}
function createRichMenu(){
  toast('建立中...');
  api('/richmenu','POST',{action:'create'}).then(function(d){
    if(d&&d.ok){toast('Rich Menu 已建立');loadFeatureSettings();}
    else toast('建立失敗');
  });
}
function deleteRichMenu(){
  if(!confirm('確定刪除所有 Rich Menu？'))return;
  toast('刪除中...');
  api('/richmenu','POST',{action:'delete'}).then(function(d){
    if(d&&d.ok){toast('已刪除 '+(d.deleted||0)+' 個選單');loadFeatureSettings();}
    else toast('刪除失敗');
  });
}
async function loadSettingsGroupSelects(){
  var d=await api('/groups');
  if(!d)return;
  var groups=d.groups||[];
  var sels=['pushGroupSelect','settingsGroupSelect'];
  for(var s=0;s<sels.length;s++){
    var sel=document.getElementById(sels[s]);
    if(!sel)continue;
    var cur=sel.value;
    sel.innerHTML=sels[s]==='settingsGroupSelect'?'<option value="">全域預設</option>':'<option value="">選擇群組...</option>';
    for(var i=0;i<groups.length;i++){
      var g=groups[i];
      var opt=document.createElement('option');
      opt.value=g.id;opt.textContent='#'+(g.name||g.id.substring(0,16));
      sel.appendChild(opt);
    }
    if(cur)sel.value=cur;
  }
}

window.addEventListener('load',function(){
  var k=localStorage.getItem('bot_admin_key');
  if(k){document.getElementById('pwInput').value=k;doLogin()}
});
if('serviceWorker' in navigator){navigator.serviceWorker.register('/sw.js?v=54').catch(function(){})}
</script>
</body>
</html>'''


SW_JS = '''const CACHE='bot-admin-v54';
const URLS=['/admin'];
self.addEventListener('install',e=>{self.skipWaiting();e.waitUntil(caches.open(CACHE).then(c=>c.addAll(URLS)))});
self.addEventListener('activate',e=>{e.waitUntil(caches.keys().then(ks=>Promise.all(ks.filter(k=>k!==CACHE).map(k=>caches.delete(k)))).then(()=>self.clients.claim()))});
self.addEventListener('fetch',e=>{const u=e.request.url;if(u.includes('/api/')||u.includes('/health')){e.respondWith(fetch(e.request));return}e.respondWith(fetch(e.request).then(r=>{if(r.ok){const c=r.clone();caches.open(CACHE).then(ca=>ca.put(e.request,c))}return r}).catch(()=>caches.match(e.request)))});'''

MANIFEST_JSON = json.dumps({
    "name": "翻譯Bot 管理後台",
    "short_name": "Bot管理",
    "start_url": "/admin",
    "display": "standalone",
    "background_color": "#0a0a0a",
    "theme_color": "#7c6fef",
    "icons": [
        {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png"},
        {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png"}
    ]
}, ensure_ascii=False)


def commit_storage_to_github(json_data):
    """Auto-commit storage_data.json to GitHub repo."""
    return _commit_file_to_github("storage_data.json", json_data, "Update storage data via admin panel")


def commit_packaging_to_github(json_data):
    """Auto-commit packaging_data.json to GitHub repo."""
    return _commit_file_to_github("packaging_data.json", json_data, "Update packaging data via admin panel")


def _commit_file_to_github(filename, content_str, message="Auto-update", branch="main"):
    """Generic: commit a file to GitHub repo."""
    if not GITHUB_TOKEN:
        logger.warning("No GITHUB_TOKEN, skipping GitHub commit for %s", filename)
        return False
    try:
        api_url = "https://api.github.com/repos/" + GITHUB_REPO + "/contents/" + filename
        req = urllib.request.Request(api_url + "?ref=" + branch, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        sha = None
        try:
            with urllib.request.urlopen(req, timeout=8) as resp:
                existing = json.loads(resp.read().decode())
                sha = existing.get("sha")
        except urllib.error.HTTPError as e:
            if e.code != 404:
                raise
        content_b64 = base64.b64encode(content_str.encode("utf-8")).decode("utf-8")
        body = {"message": message, "content": content_b64, "branch": branch}
        if sha:
            body["sha"] = sha
        data = json.dumps(body).encode("utf-8")
        req = urllib.request.Request(api_url, data=data, method="PUT", headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json"
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            logger.info("Committed %s to GitHub (%s)", filename, branch)
            return True
    except Exception as e:
        logger.error("GitHub commit %s failed: %s", filename, e)
        return False


def _ensure_data_branch():
    """Create 'data' branch if it doesn't exist."""
    if not GITHUB_TOKEN:
        return
    try:
        # Check if branch exists
        url = "https://api.github.com/repos/" + GITHUB_REPO + "/branches/data"
        req = urllib.request.Request(url, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        urllib.request.urlopen(req, timeout=5)
        return  # branch exists
    except urllib.error.HTTPError as e:
        if e.code != 404:
            return
    except Exception:
        return
    try:
        # Get main branch SHA
        url = "https://api.github.com/repos/" + GITHUB_REPO + "/git/refs/heads/main"
        req = urllib.request.Request(url, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        with urllib.request.urlopen(req, timeout=5) as resp:
            main_data = json.loads(resp.read().decode())
            sha = main_data["object"]["sha"]
        # Create data branch
        url = "https://api.github.com/repos/" + GITHUB_REPO + "/git/refs"
        body = json.dumps({"ref": "refs/heads/data", "sha": sha}).encode("utf-8")
        req = urllib.request.Request(url, data=body, method="POST", headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json",
            "Content-Type": "application/json"
        })
        urllib.request.urlopen(req, timeout=5)
        logger.info("Created 'data' branch on GitHub")
    except Exception as e:
        logger.error("Failed to create data branch: %s", e)


def _load_file_from_github(filename, branch="main"):
    """Load a JSON file from GitHub repo. Returns parsed dict/list or None."""
    if not GITHUB_TOKEN:
        return None
    try:
        api_url = "https://api.github.com/repos/" + GITHUB_REPO + "/contents/" + filename + "?ref=" + branch
        req = urllib.request.Request(api_url, headers={
            "Authorization": "token " + GITHUB_TOKEN,
            "Accept": "application/vnd.github.v3+json"
        })
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read().decode())
            content = base64.b64decode(data["content"]).decode("utf-8")
            return json.loads(content)
    except Exception as e:
        logger.warning("Load %s from GitHub (%s) failed: %s", filename, branch, e)
        return None


_last_save_time = 0
_save_lock = _threading.Lock()
_save_scheduled = False

def save_settings():
    """Persist all bot settings to GitHub (background, throttled to max once per 30s).
    If throttled, schedules a delayed save so no changes are lost."""
    global _last_save_time, _save_scheduled
    now = time.time()
    with _save_lock:
        remaining = 30 - (now - _last_save_time)
        if remaining > 0:
            # Throttled — schedule a delayed save if not already scheduled
            if not _save_scheduled:
                _save_scheduled = True
                _threading.Timer(remaining + 1, _flush_pending_save).start()
            return
        _last_save_time = now
        _save_scheduled = False
    _threading.Thread(target=_do_save_impl, daemon=True).start()


def _flush_pending_save():
    """Called by timer after throttle window expires."""
    global _save_scheduled
    with _save_lock:
        _save_scheduled = False
    save_settings()


def _do_save_impl():
    global _last_save_time
    try:
        data = {
            "group_settings": group_settings,
            "group_target_lang": group_target_lang,
            "group_img_settings": group_img_settings,
            "group_audio_settings": group_audio_settings,
            "group_wo_settings": group_wo_settings,
            "group_cmd_enabled": group_cmd_enabled,
            "group_skip_users": {k: list(v) for k, v in group_skip_users.items()},
            "group_tracking": group_tracking,
            "group_user_names": group_user_names,
            "dm_master_enabled": dm_master_enabled,
            "dm_whitelist": list(dm_whitelist),
            "dm_known_users": dm_known_users,
            "dm_target_lang": dm_target_lang,
            "admin_users": admin_users,
            "bot_stats": bot_stats,
            "extra_customers": extra_names_by_group,
            "group_api_usage": group_api_usage,
            "user_languages": user_languages,
            "welcome_settings": welcome_settings,
            "group_welcome_settings": group_welcome_settings,
            "flex_enabled": flex_enabled,
            "quick_reply_enabled": quick_reply_enabled,
            "silent_mode": silent_mode,
            "sender_name": sender_name,
            "sender_icon": sender_icon,
            "video_ocr_enabled": video_ocr_enabled,
            "location_translate_enabled": location_translate_enabled,
            "group_flex_settings": group_flex_settings,
            "group_qr_settings": group_qr_settings,
            "group_silent_settings": group_silent_settings,
            "group_video_settings": group_video_settings,
            "group_location_settings": group_location_settings,
            "mark_read_enabled": mark_read_enabled,
            "retry_key_enabled": retry_key_enabled,
            "camera_qr_enabled": camera_qr_enabled,
            "clipboard_qr_enabled": clipboard_qr_enabled,
            "group_mark_read_settings": group_mark_read_settings,
            "group_retry_key_settings": group_retry_key_settings,
            "group_camera_qr_settings": group_camera_qr_settings,
            "group_clipboard_qr_settings": group_clipboard_qr_settings,
            "camera_roll_qr_enabled": camera_roll_qr_enabled,
            "location_qr_enabled": location_qr_enabled,
            "group_camera_roll_qr_settings": group_camera_roll_qr_settings,
            "group_location_qr_settings": group_location_qr_settings,
            "translation_tone": translation_tone,
            "translation_tone_custom": translation_tone_custom,
            "group_tone_settings": group_tone_settings,
            "model_default": model_default,
            "model_upgrade": model_upgrade,
            "model_threshold": model_threshold,
            "user_pictures": user_pictures,
            "pw1_text": pw1_text,
            "pw2_text": pw2_text,
            "scrap_text": scrap_text,
        }
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        _commit_file_to_github("bot_settings.json", json_str, "Auto-save bot settings", branch="data")
    except Exception as e:
        logger.error("Background save_settings failed: %s", e)
        # Reset timer so next call will retry instead of being throttled
        with _save_lock:
            _last_save_time = 0


def load_settings():
    """Load bot settings from GitHub on startup."""
    global dm_master_enabled, dm_whitelist, dm_known_users, dm_target_lang
    global group_settings, group_target_lang, group_img_settings, group_audio_settings
    global group_wo_settings, group_skip_users, group_tracking, group_user_names
    global admin_users, bot_stats
    global EXTRA_CUSTOMERS, group_api_usage, extra_names_by_group, user_languages
    global flex_enabled, quick_reply_enabled, silent_mode, welcome_settings, sender_name, sender_icon, user_pictures, video_ocr_enabled, location_translate_enabled
    global group_flex_settings, group_qr_settings, group_silent_settings, group_video_settings, group_location_settings, group_welcome_settings
    global group_mark_read_settings, group_retry_key_settings, group_camera_qr_settings, group_clipboard_qr_settings
    global group_camera_roll_qr_settings, group_location_qr_settings
    global mark_read_enabled, retry_key_enabled, camera_qr_enabled, clipboard_qr_enabled
    global camera_roll_qr_enabled, location_qr_enabled
    global translation_tone, translation_tone_custom
    global pw1_text, pw2_text, scrap_text, PACKAGING_LOOKUP
    data = _load_file_from_github("bot_settings.json", branch="data")
    if not data:
        logger.info("No bot_settings.json found on GitHub, starting fresh")
        return
    try:
        group_settings.update(data.get("group_settings", {}))
        group_target_lang.update(data.get("group_target_lang", {}))
        group_img_settings.update(data.get("group_img_settings", {}))
        group_audio_settings.update(data.get("group_audio_settings", {}))
        group_wo_settings.update(data.get("group_wo_settings", {}))
        group_cmd_enabled.update(data.get("group_cmd_enabled", {}))
        for k, v in data.get("group_skip_users", {}).items():
            group_skip_users[k] = set(v)
        group_tracking.update(data.get("group_tracking", {}))
        group_user_names.update(data.get("group_user_names", {}))
        dm_master_enabled = data.get("dm_master_enabled", True)
        dm_whitelist = set(data.get("dm_whitelist", []))
        dm_known_users.update(data.get("dm_known_users", {}))
        dm_target_lang.update(data.get("dm_target_lang", {}))
        admin_users.update(data.get("admin_users", {}))
        bot_stats.update(data.get("bot_stats", {}))
        if "extra_customers" in data:
            ec = data["extra_customers"]
            if isinstance(ec, dict):
                extra_names_by_group.update(ec)
            elif isinstance(ec, list):
                extra_names_by_group["__all__"] = ec
            rebuild_customer_names()
        group_api_usage.update(data.get("group_api_usage", {}))
        user_languages.update(data.get("user_languages", {}))
        if "welcome_settings" in data:
            welcome_settings.update(data.get("welcome_settings", {}))
        if "flex_enabled" in data:
            flex_enabled = data["flex_enabled"]
        if "quick_reply_enabled" in data:
            quick_reply_enabled = data["quick_reply_enabled"]
        if "silent_mode" in data:
            silent_mode = data["silent_mode"]
        if "sender_name" in data:
            sender_name = data["sender_name"]
        if "sender_icon" in data:
            sender_icon = data["sender_icon"]
        if "video_ocr_enabled" in data:
            video_ocr_enabled = data["video_ocr_enabled"]
        if "location_translate_enabled" in data:
            location_translate_enabled = data["location_translate_enabled"]
        group_flex_settings.update(data.get("group_flex_settings", {}))
        group_qr_settings.update(data.get("group_qr_settings", {}))
        group_silent_settings.update(data.get("group_silent_settings", {}))
        group_video_settings.update(data.get("group_video_settings", {}))
        group_location_settings.update(data.get("group_location_settings", {}))
        if "mark_read_enabled" in data:
            mark_read_enabled = data["mark_read_enabled"]
        if "retry_key_enabled" in data:
            retry_key_enabled = data["retry_key_enabled"]
        if "camera_qr_enabled" in data:
            camera_qr_enabled = data["camera_qr_enabled"]
        if "clipboard_qr_enabled" in data:
            clipboard_qr_enabled = data["clipboard_qr_enabled"]
        group_mark_read_settings.update(data.get("group_mark_read_settings", {}))
        group_retry_key_settings.update(data.get("group_retry_key_settings", {}))
        group_camera_qr_settings.update(data.get("group_camera_qr_settings", {}))
        group_clipboard_qr_settings.update(data.get("group_clipboard_qr_settings", {}))
        if "camera_roll_qr_enabled" in data:
            camera_roll_qr_enabled = data["camera_roll_qr_enabled"]
        if "location_qr_enabled" in data:
            location_qr_enabled = data["location_qr_enabled"]
        group_camera_roll_qr_settings.update(data.get("group_camera_roll_qr_settings", {}))
        group_location_qr_settings.update(data.get("group_location_qr_settings", {}))
        group_welcome_settings.update(data.get("group_welcome_settings", {}))
        if "translation_tone" in data:
            translation_tone = data["translation_tone"]
        if "translation_tone_custom" in data:
            translation_tone_custom = data["translation_tone_custom"]
        group_tone_settings.update(data.get("group_tone_settings", {}))
        if "model_default" in data:
            model_default = data["model_default"]
        if "model_upgrade" in data:
            model_upgrade = data["model_upgrade"]
        if "model_threshold" in data:
            model_threshold = int(data["model_threshold"])
        user_pictures.update(data.get("user_pictures", {}))
        if "pw1_text" in data:
            pw1_text = data["pw1_text"]
        if "pw2_text" in data:
            pw2_text = data["pw2_text"]
        if "scrap_text" in data:
            scrap_text = data["scrap_text"]
        logger.info("Loaded bot settings from GitHub: %d groups, %d DM users, %d protected names",
                     len(group_tracking), len(dm_known_users), len(EXTRA_CUSTOMERS))
    except Exception as e:
        logger.error("Error loading bot settings: %s", e)


# Ensure data branch exists and load settings on startup
try:
    _ensure_data_branch()
    load_settings()
except Exception as e:
    logger.error("Startup settings load failed (non-fatal): %s", e)




def check_admin_key():
    key = request.headers.get("X-Admin-Key", "")
    return key == ADMIN_KEY


@app.route("/admin")
def admin_page():
    resp = app.response_class(ADMIN_HTML, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/debug")
def debug_page():
    """Minimal debug page - no SW, no cache."""
    html = '''<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Debug</title>
<style>body{background:#000;color:#0f0;font:13px monospace;padding:12px}input,button{font:14px sans-serif;padding:8px 12px;margin:4px;border-radius:6px;border:1px solid #555}input{background:#111;color:#fff;width:200px}button{background:#7c6fef;color:#fff;border:none;cursor:pointer}#log{white-space:pre-wrap;margin-top:12px}</style>
</head><body>
<h3>🔧 Bot Debug Panel</h3>
<div>
<input id="pw" type="password" placeholder="管理密碼">
<button onclick="go()">測試登入</button>
<button onclick="testHealth()">Health</button>
<button onclick="clearSW()">清除SW快取</button>
</div>
<div id="log">等待操作...\n</div>
<script>
const L=document.getElementById('log');
const API=window.location.origin+'/api/admin';
function log(s){L.textContent+=new Date().toLocaleTimeString()+' '+s+'\\n';L.scrollTop=L.scrollHeight}

async function testHealth(){
  log('>>> GET /health');
  try{
    const r=await fetch('/health');
    const d=await r.json();
    log('<<< '+r.status+' '+JSON.stringify(d));
  }catch(e){log('!!! '+e.message)}
}

async function go(){
  const key=document.getElementById('pw').value;
  if(!key){log('請輸入密碼');return}
  log('>>> GET /api/admin/status');
  try{
    const r=await fetch(API+'/status',{headers:{'X-Admin-Key':key}});
    log('<<< status: '+r.status);
    const d=await r.json();
    log('<<< body: '+JSON.stringify(d));
    if(d.ok){
      log('✅ 登入成功! 測試其他API...');
      log('>>> GET /api/admin/stats');
      const r2=await fetch(API+'/stats',{headers:{'X-Admin-Key':key}});
      log('<<< stats: '+r2.status);
      const d2=await r2.json();
      log('<<< '+JSON.stringify(d2).substring(0,300));
      log('>>> GET /api/admin/groups');
      const r3=await fetch(API+'/groups',{headers:{'X-Admin-Key':key}});
      log('<<< groups: '+r3.status);
      const d3=await r3.json();
      log('<<< '+JSON.stringify(d3).substring(0,300));
    }
  }catch(e){log('!!! 錯誤: '+e.message)}
}

async function clearSW(){
  log('清除所有 Service Worker...');
  if('serviceWorker' in navigator){
    const regs=await navigator.serviceWorker.getRegistrations();
    log('找到 '+regs.length+' 個 SW');
    for(const r of regs){
      await r.unregister();
      log('已移除: '+r.scope);
    }
    log('清除快取...');
    const keys=await caches.keys();
    for(const k of keys){
      await caches.delete(k);
      log('已刪除快取: '+k);
    }
    log('✅ 全部清除完成! 現在可以回 /admin 了');
  }else{
    log('此瀏覽器不支援 SW');
  }
}

log('Debug頁面載入完成');
log('API: '+API);
</script></body></html>'''
    resp = app.response_class(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp


@app.route("/manifest.json")
def manifest():
    return app.response_class(MANIFEST_JSON, mimetype="application/manifest+json")


@app.route("/sw.js")
def service_worker():
    resp = app.response_class(SW_JS, mimetype="application/javascript")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


@app.route("/icon-192.png")
@app.route("/icon-512.png")
def admin_icon():
    # Generate a simple green circle PNG as icon
    import struct, zlib
    size = 192 if "192" in request.path else 512
    # 1x1 green pixel PNG, browser will scale
    png = (b'\\x89PNG\\r\\n\\x1a\\n'
           + struct.pack('>I', 13) + b'IHDR' + struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)
           + struct.pack('>I', zlib.crc32(b'IHDR' + struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0)) & 0xffffffff)
           + struct.pack('>I', 12) + b'IDAT' + zlib.compress(b'\\x00\\x06\\xc7\\x55')
           + struct.pack('>I', zlib.crc32(b'IDAT' + zlib.compress(b'\\x00\\x06\\xc7\\x55')) & 0xffffffff)
           + struct.pack('>I', 0) + b'IEND' + struct.pack('>I', zlib.crc32(b'IEND') & 0xffffffff))
    return app.response_class(png, mimetype="image/png")


# ─── Admin API ──────────────────────────────────────────

@app.route("/api/admin/status")
def api_admin_status():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"ok": True})


@app.route("/api/admin/groups", methods=["GET"])
def api_admin_groups():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    groups = []
    # Merge from group_tracking + group_settings
    all_gids = set(group_tracking.keys()) | set(group_settings.keys()) | set(group_target_lang.keys())
    for gid in all_gids:
        info = group_tracking.get(gid, {})
        skip_count = len(group_skip_users.get(gid, set()))
        groups.append({
            "id": gid,
            "name": info.get("name", ""),
            "translation_on": group_settings.get(gid, True),
            "target_lang": group_target_lang.get(gid, "id"),
            "image_on": group_img_settings.get(gid, True),
            "voice_on": group_audio_settings.get(gid, True),
            "work_order_on": group_wo_settings.get(gid, True),
            "cmd_enabled": {k: is_cmd_enabled(gid, k) for k, _, _, _ in CMD_DEFS},
            "skip_count": skip_count,
            "cost_twd": calc_group_cost_twd(gid),
            "member_count": get_group_member_count(gid),
        })
    groups.sort(key=lambda x: x["name"] or x["id"])
    return jsonify({"groups": groups})


@app.route("/api/admin/groups/leave", methods=["POST"])
def api_admin_leave_group():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    try:
        with ApiClient(configuration) as api_client:
            api = MessagingApi(api_client)
            api.leave_group(gid)
    except Exception as e:
        logger.warning("Leave group failed: %s", e)
        # Try room
        try:
            with ApiClient(configuration) as api_client:
                api = MessagingApi(api_client)
                api.leave_room(gid)
        except Exception:
            pass
    # Clean up local data
    group_tracking.pop(gid, None)
    group_settings.pop(gid, None)
    group_target_lang.pop(gid, None)
    group_img_settings.pop(gid, None)
    group_audio_settings.pop(gid, None)
    group_wo_settings.pop(gid, None)
    group_skip_users.pop(gid, None)
    group_user_names.pop(gid, None)
    save_settings()
    return jsonify({"message": "已退出群組"})


@app.route("/api/admin/stats")
def api_admin_stats():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    uptime = time.time() - bot_start_time
    # Calculate estimated cost (GPT-4o-mini pricing)
    tp = bot_stats.get("tokens_prompt", 0)
    tc = bot_stats.get("tokens_completion", 0)
    cost = (tp * 0.00000015) + (tc * 0.0000006)
    return jsonify({
        "uptime_seconds": int(uptime),
        "text_translations": bot_stats.get("text_translations", 0),
        "image_translations": bot_stats.get("image_translations", 0),
        "voice_translations": bot_stats.get("voice_translations", 0),
        "work_order_detections": bot_stats.get("work_order_detections", 0),
        "commands": bot_stats.get("commands", 0),
        "customers": len(STORAGE_LOOKUP),
        "groups": len(set(group_tracking.keys()) | set(group_settings.keys())),
        "dm_users": len(dm_known_users),
        "tokens_prompt": tp,
        "tokens_completion": tc,
        "tokens_total": tp + tc,
        "estimated_cost_usd": round(cost, 4),
        "followers": bot_stats.get("followers", 0),
        "unfollowers": bot_stats.get("unfollowers", 0),
        "line_quota": get_line_quota(),
        # Feature settings
        "welcome_enabled": welcome_settings.get("enabled", True),
        "welcome_text_zh": welcome_settings.get("text_zh", ""),
        "welcome_text_id": welcome_settings.get("text_id", ""),
        "flex_enabled": flex_enabled,
        "quick_reply_enabled": quick_reply_enabled,
        "silent_mode": silent_mode,
    })


@app.route("/api/admin/features", methods=["GET", "POST"])
def api_admin_features():
    """Get/set feature settings. Pass group_id for per-group; omit for global defaults."""
    global flex_enabled, quick_reply_enabled, silent_mode, welcome_settings
    global sender_name, sender_icon, video_ocr_enabled, location_translate_enabled
    global translation_tone, translation_tone_custom
    global mark_read_enabled, retry_key_enabled, camera_qr_enabled, clipboard_qr_enabled
    global camera_roll_qr_enabled, location_qr_enabled
    global model_default, model_upgrade, model_threshold
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    gid = request.args.get("group_id", "") if request.method == "GET" else (request.get_json() or {}).get("group_id", "")
    if request.method == "POST":
        data = request.get_json() or {}
        if gid:
            # Per-group settings
            _feat_map = {
                "flex_enabled": group_flex_settings,
                "quick_reply_enabled": group_qr_settings,
                "silent_mode": group_silent_settings,
                "video_ocr_enabled": group_video_settings,
                "location_translate_enabled": group_location_settings,
                "mark_read_enabled": group_mark_read_settings,
                "retry_key_enabled": group_retry_key_settings,
                "camera_qr_enabled": group_camera_qr_settings,
                "clipboard_qr_enabled": group_clipboard_qr_settings,
                "camera_roll_qr_enabled": group_camera_roll_qr_settings,
                "location_qr_enabled": group_location_qr_settings,
            }
            for key, d in _feat_map.items():
                if key in data:
                    d[gid] = bool(data[key])
            # Per-group welcome
            if any(k in data for k in ("welcome_enabled", "welcome_text_zh", "welcome_text_id")):
                if gid not in group_welcome_settings:
                    group_welcome_settings[gid] = {}
                if "welcome_enabled" in data:
                    group_welcome_settings[gid]["enabled"] = bool(data["welcome_enabled"])
                if "welcome_text_zh" in data:
                    group_welcome_settings[gid]["text_zh"] = str(data["welcome_text_zh"])
                if "welcome_text_id" in data:
                    group_welcome_settings[gid]["text_id"] = str(data["welcome_text_id"])
            # Per-group tone
            if "translation_tone" in data:
                if gid not in group_tone_settings:
                    group_tone_settings[gid] = {}
                group_tone_settings[gid]["tone"] = str(data["translation_tone"])
            if "translation_tone_custom" in data:
                if gid not in group_tone_settings:
                    group_tone_settings[gid] = {}
                group_tone_settings[gid]["custom"] = str(data["translation_tone_custom"])
        else:
            # Global defaults
            if "welcome_enabled" in data:
                welcome_settings["enabled"] = bool(data["welcome_enabled"])
            if "welcome_text_zh" in data:
                welcome_settings["text_zh"] = str(data["welcome_text_zh"])
            if "welcome_text_id" in data:
                welcome_settings["text_id"] = str(data["welcome_text_id"])
            if "flex_enabled" in data:
                flex_enabled = bool(data["flex_enabled"])
            if "quick_reply_enabled" in data:
                quick_reply_enabled = bool(data["quick_reply_enabled"])
            if "silent_mode" in data:
                silent_mode = bool(data["silent_mode"])
            if "video_ocr_enabled" in data:
                video_ocr_enabled = bool(data["video_ocr_enabled"])
            if "location_translate_enabled" in data:
                location_translate_enabled = bool(data["location_translate_enabled"])
            if "mark_read_enabled" in data:
                mark_read_enabled = bool(data["mark_read_enabled"])
            if "retry_key_enabled" in data:
                retry_key_enabled = bool(data["retry_key_enabled"])
            if "camera_qr_enabled" in data:
                camera_qr_enabled = bool(data["camera_qr_enabled"])
            if "clipboard_qr_enabled" in data:
                clipboard_qr_enabled = bool(data["clipboard_qr_enabled"])
            if "camera_roll_qr_enabled" in data:
                camera_roll_qr_enabled = bool(data["camera_roll_qr_enabled"])
            if "location_qr_enabled" in data:
                location_qr_enabled = bool(data["location_qr_enabled"])
            if "translation_tone" in data:
                translation_tone = str(data["translation_tone"])
            if "translation_tone_custom" in data:
                translation_tone_custom = str(data["translation_tone_custom"])
            if "model_default" in data:
                model_default = str(data["model_default"])
            if "model_upgrade" in data:
                model_upgrade = str(data["model_upgrade"])
            if "model_threshold" in data:
                model_threshold = int(data["model_threshold"])
        # Sender settings are always global
        if "sender_name" in data:
            sender_name = str(data["sender_name"])[:20]
        if "sender_icon" in data:
            sender_icon = str(data["sender_icon"])
        save_settings()
        return jsonify({"ok": True})
    # GET - return settings for specific group or global
    if gid:
        ws = get_group_welcome(gid)
        return jsonify({
            "group_id": gid,
            "welcome_enabled": ws.get("enabled", True),
            "welcome_text_zh": ws.get("text_zh", ""),
            "welcome_text_id": ws.get("text_id", ""),
            "flex_enabled": get_group_feature(gid, 'flex'),
            "quick_reply_enabled": get_group_feature(gid, 'quick_reply'),
            "silent_mode": get_group_feature(gid, 'silent'),
            "video_ocr_enabled": get_group_feature(gid, 'video_ocr'),
            "location_translate_enabled": get_group_feature(gid, 'location'),
            "mark_read_enabled": get_group_feature(gid, 'mark_read'),
            "retry_key_enabled": get_group_feature(gid, 'retry_key'),
            "camera_qr_enabled": get_group_feature(gid, 'camera_qr'),
            "clipboard_qr_enabled": get_group_feature(gid, 'clipboard_qr'),
            "camera_roll_qr_enabled": get_group_feature(gid, 'camera_roll_qr'),
            "location_qr_enabled": get_group_feature(gid, 'location_qr'),
            "translation_tone": get_group_tone(gid)[0],
            "translation_tone_custom": get_group_tone(gid)[1],
            "sender_name": sender_name,
            "sender_icon": sender_icon,
            "bot_info": get_bot_info(),
            # Include global defaults for reference
            "global_defaults": {
                "welcome_enabled": welcome_settings.get("enabled", True),
                "flex_enabled": flex_enabled,
                "quick_reply_enabled": quick_reply_enabled,
                "silent_mode": silent_mode,
                "video_ocr_enabled": video_ocr_enabled,
                "location_translate_enabled": location_translate_enabled,
                "mark_read_enabled": mark_read_enabled,
                "retry_key_enabled": retry_key_enabled,
                "camera_qr_enabled": camera_qr_enabled,
                "clipboard_qr_enabled": clipboard_qr_enabled,
                "camera_roll_qr_enabled": camera_roll_qr_enabled,
                "location_qr_enabled": location_qr_enabled,
            },
            "is_customized": gid in group_flex_settings or gid in group_qr_settings or gid in group_silent_settings or gid in group_video_settings or gid in group_location_settings or gid in group_welcome_settings or gid in group_tone_settings or gid in group_mark_read_settings or gid in group_retry_key_settings or gid in group_camera_qr_settings or gid in group_clipboard_qr_settings or gid in group_camera_roll_qr_settings or gid in group_location_qr_settings,
        })
    return jsonify({
        "welcome_enabled": welcome_settings.get("enabled", True),
        "welcome_text_zh": welcome_settings.get("text_zh", ""),
        "welcome_text_id": welcome_settings.get("text_id", ""),
        "flex_enabled": flex_enabled,
        "quick_reply_enabled": quick_reply_enabled,
        "silent_mode": silent_mode,
        "video_ocr_enabled": video_ocr_enabled,
        "location_translate_enabled": location_translate_enabled,
        "mark_read_enabled": mark_read_enabled,
        "retry_key_enabled": retry_key_enabled,
        "camera_qr_enabled": camera_qr_enabled,
        "clipboard_qr_enabled": clipboard_qr_enabled,
        "camera_roll_qr_enabled": camera_roll_qr_enabled,
        "location_qr_enabled": location_qr_enabled,
        "translation_tone": translation_tone,
        "translation_tone_custom": translation_tone_custom,
        "model_default": model_default,
        "model_upgrade": model_upgrade,
        "model_threshold": model_threshold,
        "sender_name": sender_name,
        "sender_icon": sender_icon,
        "bot_info": get_bot_info(),
    })


@app.route("/api/admin/features/reset", methods=["POST"])
def api_admin_features_reset():
    """Reset per-group feature settings to global defaults."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    group_flex_settings.pop(gid, None)
    group_qr_settings.pop(gid, None)
    group_silent_settings.pop(gid, None)
    group_video_settings.pop(gid, None)
    group_location_settings.pop(gid, None)
    group_mark_read_settings.pop(gid, None)
    group_retry_key_settings.pop(gid, None)
    group_camera_qr_settings.pop(gid, None)
    group_clipboard_qr_settings.pop(gid, None)
    group_camera_roll_qr_settings.pop(gid, None)
    group_location_qr_settings.pop(gid, None)
    group_welcome_settings.pop(gid, None)
    group_tone_settings.pop(gid, None)
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/push", methods=["POST"])
def api_admin_push():
    """Push a message to a group."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    text = data.get("text", "").strip()
    if not gid or not text:
        return jsonify({"error": "missing group_id or text"}), 400
    ok = push_message_to_group(gid, text)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu", methods=["POST"])
def api_admin_richmenu():
    """Create or delete rich menu."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    action = data.get("action", "create")
    if action == "create":
        rid = setup_rich_menu()
        return jsonify({"ok": bool(rid), "rich_menu_id": rid})
    elif action == "delete":
        count = delete_rich_menu()
        return jsonify({"ok": True, "deleted": count})
    return jsonify({"error": "invalid action"}), 400


@app.route("/api/admin/insight")
def api_admin_insight():
    """Get follower demographics and message delivery stats."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    return jsonify({
        "demographics": get_insight_followers(),
        "delivery": get_message_delivery_stats(),
    })


@app.route("/api/admin/insight/trend")
def api_admin_insight_trend():
    """Get daily follower trend data (7 days)."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    days = min(int(request.args.get("days", 7)), 30)
    trend = get_statistics_per_unit(num_days=days)
    return jsonify({"trend": trend})


@app.route("/api/admin/richmenu/upload/<rm_id>", methods=["POST"])
def api_admin_richmenu_upload(rm_id):
    """Upload a custom image to a rich menu from admin panel."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    img_b64 = data.get("image", "")
    content_type = data.get("content_type", "image/png")
    if not img_b64:
        return jsonify({"error": "missing image data"}), 400
    try:
        # Strip data URI prefix if present
        if "," in img_b64:
            header, img_b64 = img_b64.split(",", 1)
            if "jpeg" in header or "jpg" in header:
                content_type = "image/jpeg"
            elif "png" in header:
                content_type = "image/png"
        img_bytes = base64.b64decode(img_b64)
        if len(img_bytes) > 5 * 1024 * 1024:
            return jsonify({"error": "image too large (max 5MB)"}), 400
        ok = upload_rich_menu_image_custom(rm_id, img_bytes, content_type)
        return jsonify({"ok": ok})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/imagemap/send", methods=["POST"])
def api_admin_imagemap_send():
    """Send an Imagemap message to a group or user."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    to = data.get("to", "")
    base_url = data.get("base_url", "")
    alt_text = data.get("alt_text", "圖片選單")
    width = int(data.get("width", 1040))
    height = int(data.get("height", 1040))
    actions = data.get("actions", [])
    if not to or not base_url:
        return jsonify({"error": "missing to or base_url"}), 400
    if not actions:
        return jsonify({"error": "missing actions"}), 400
    ok = send_imagemap_message(to, base_url, alt_text, width, height, actions)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/list")
def api_admin_richmenu_list():
    """List all rich menus."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"menus": list_rich_menus()})


@app.route("/api/admin/richmenu/link", methods=["POST"])
def api_admin_richmenu_link():
    """Link/unlink rich menu to a specific user."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    user_id = data.get("user_id", "")
    rm_id = data.get("rich_menu_id", "")
    action = data.get("action", "link")
    if not user_id:
        return jsonify({"error": "missing user_id"}), 400
    if action == "link" and rm_id:
        ok = link_rich_menu_to_user(user_id, rm_id)
    elif action == "unlink":
        ok = unlink_rich_menu_from_user(user_id)
    else:
        return jsonify({"error": "invalid params"}), 400
    return jsonify({"ok": ok})


@app.route("/api/admin/multicast", methods=["POST"])
def api_admin_multicast():
    """Send a message to multiple users."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    user_ids = data.get("user_ids", [])
    text = data.get("text", "").strip()
    if not user_ids or not text:
        return jsonify({"error": "missing user_ids or text"}), 400
    ok = multicast_message(user_ids, text)
    return jsonify({"ok": ok})


@app.route("/api/admin/broadcast", methods=["POST"])
def api_admin_broadcast():
    """Broadcast to all followers."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"error": "missing text"}), 400
    ok = broadcast_message(text)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/alias", methods=["POST"])
def api_admin_richmenu_alias():
    """Create/delete rich menu alias."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    alias_id = data.get("alias_id", "")
    rm_id = data.get("rich_menu_id", "")
    action = data.get("action", "create")
    ok = manage_rich_menu_alias(alias_id, rm_id, action)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/batch", methods=["POST"])
def api_admin_richmenu_batch():
    """Batch link/unlink rich menu to users."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    user_ids = data.get("user_ids", [])
    rm_id = data.get("rich_menu_id", "")
    action = data.get("action", "link")
    if action == "link" and rm_id:
        ok = batch_link_rich_menu(user_ids, rm_id)
    else:
        ok = batch_unlink_rich_menu(user_ids)
    return jsonify({"ok": ok})


@app.route("/api/admin/groups/settings", methods=["POST"])
def api_admin_group_settings():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    if "target_lang" in data:
        group_target_lang[gid] = data["target_lang"]
    if "translation_on" in data:
        group_settings[gid] = bool(data["translation_on"])
    if "image_on" in data:
        group_img_settings[gid] = bool(data["image_on"])
    if "voice_on" in data:
        group_audio_settings[gid] = bool(data["voice_on"])
    if "work_order_on" in data:
        group_wo_settings[gid] = bool(data["work_order_on"])
    if "cmd_toggle" in data:
        cmd_key = data["cmd_toggle"]
        cmd_val = bool(data.get("cmd_val", True))
        if gid not in group_cmd_enabled:
            group_cmd_enabled[gid] = {}
        group_cmd_enabled[gid][cmd_key] = cmd_val
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/groups/reset-cost", methods=["POST"])
def api_admin_reset_group_cost():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    gid = data.get("group_id", "")
    if not gid:
        return jsonify({"error": "missing group_id"}), 400
    group_api_usage.pop(gid, None)
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/dm", methods=["GET", "POST"])
def api_admin_dm():
    global dm_master_enabled
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    if request.method == "POST":
        data = request.get_json() or {}
        if "master_enabled" in data:
            dm_master_enabled = bool(data["master_enabled"])
        save_settings()
        return jsonify({"ok": True})
    # Build known users list with whitelist status
    known = []
    for uid, name in dm_known_users.items():
        known.append({"user_id": uid, "name": name, "whitelisted": uid in dm_whitelist})
    return jsonify({
        "master_enabled": dm_master_enabled,
        "whitelist": list(dm_whitelist),
        "known_users": known
    })


@app.route("/api/admin/dm/whitelist", methods=["POST"])
def api_admin_dm_whitelist():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    uid = data.get("user_id", "").strip()
    action = data.get("action", "add")
    if not uid:
        return jsonify({"error": "missing user_id"}), 400
    if action == "add":
        dm_whitelist.add(uid)
    elif action == "remove":
        dm_whitelist.discard(uid)
    save_settings()
    return jsonify({"ok": True, "whitelist": list(dm_whitelist)})


@app.route("/api/admin/skip", methods=["GET", "POST"])
def api_admin_skip():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    if request.method == "POST":
        data = request.get_json() or {}
        gid = data.get("group_id", "")
        uid = data.get("user_id", "")
        action = data.get("action", "add")
        if not gid or not uid:
            return jsonify({"error": "missing group_id or user_id"}), 400
        if gid not in group_skip_users:
            group_skip_users[gid] = set()
        if action == "add":
            group_skip_users[gid].add(uid)
            save_settings()
            return jsonify({"ok": True})
        elif action == "remove":
            group_skip_users[gid].discard(uid)
            save_settings()
            return jsonify({"ok": True})
    # GET: return all known users in group with skip status
    gid = request.args.get("group_id", "")
    skipped = group_skip_users.get(gid, set())
    names_cache = group_user_names.get(gid, {})
    users = []
    for uid, dname in names_cache.items():
        users.append({"user_id": uid, "name": dname, "skipped": uid in skipped})
    users.sort(key=lambda x: x["name"])
    return jsonify({"users": users})


@app.route("/api/admin/users", methods=["GET"])
def api_admin_users():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    filter_gid = request.args.get("group_id", "")
    users = []
    if filter_gid:
        names_cache = group_user_names.get(filter_gid, {})
        for uid, name in names_cache.items():
            is_admin = admin_users.get(uid, {}).get("is_admin", False)
            users.append({
                "user_id": uid,
                "name": name,
                "is_admin": is_admin,
                "line_lang": user_languages.get(uid, ""),
                "picture_url": user_pictures.get(uid, ""),
            })
    else:
        all_users = {}
        for uid, name in dm_known_users.items():
            all_users[uid] = name
        for gid, names in group_user_names.items():
            for uid, name in names.items():
                if uid not in all_users:
                    all_users[uid] = name
        for uid, name in all_users.items():
            is_admin = admin_users.get(uid, {}).get("is_admin", False)
            users.append({
                "user_id": uid,
                "name": name,
                "is_admin": is_admin,
                "line_lang": user_languages.get(uid, ""),
                "picture_url": user_pictures.get(uid, ""),
            })
    users.sort(key=lambda x: x["name"])
    return jsonify({"users": users})


@app.route("/api/admin/users/admin", methods=["POST"])
def api_admin_users_toggle_admin():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    uid = data.get("user_id", "")
    is_admin = bool(data.get("is_admin", False))
    if not uid:
        return jsonify({"error": "missing user_id"}), 400
    if uid not in admin_users:
        admin_users[uid] = {}
    admin_users[uid]["is_admin"] = is_admin
    save_settings()
    return jsonify({"ok": True})


@app.route("/api/admin/names", methods=["GET", "POST"])
def api_admin_names():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    names_list = extra_names_by_group.setdefault("__all__", [])
    if request.method == "POST":
        data = request.get_json() or {}
        action = data.get("action", "add")
        name = data.get("name", "").strip()
        if not name:
            return jsonify({"error": "missing name"}), 400
        if action == "add":
            if name not in names_list:
                names_list.append(name)
                rebuild_customer_names()
                save_settings()
            return jsonify({"ok": True})
        elif action == "remove":
            if name in names_list:
                names_list.remove(name)
                rebuild_customer_names()
                save_settings()
            return jsonify({"ok": True})
    return jsonify({"names": names_list, "count": len(names_list)})


@app.route("/api/admin/storage/stats")
def api_admin_storage_stats():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"count": len(STORAGE_LOOKUP)})


@app.route("/api/admin/storage/upload", methods=["POST"])
def api_admin_storage_upload():
    global STORAGE_LOOKUP, CUSTOMER_NAMES
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "沒有檔案"}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "沒有檔案"}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "空的 Excel"}), 400
        # Auto-detect format: find header row
        header = [str(c).strip() if c else "" for c in rows[0]]
        new_data = {}
        # Try format: Customer | <=3200 | >3200<=4200 | >4200
        len_cols = {}
        for i, h in enumerate(header):
            hl = h.replace(" ", "")
            if "<=3200" in hl and ">3200" not in hl:
                len_cols["<=3200"] = i
            elif ">3200" in hl and "<=4200" in hl:
                len_cols[">3200<=4200"] = i
            elif ">4200" in hl:
                len_cols[">4200"] = i
        if len(len_cols) >= 2:
            # Detected column-based format
            cust_col = 0  # assume first column is customer
            for _, row in enumerate(rows[1:], 1):
                if not row or not row[cust_col]:
                    continue
                cust = str(row[cust_col]).strip()
                if not cust:
                    continue
                entries = []
                for length_key, col_idx in len_cols.items():
                    if col_idx < len(row) and row[col_idx]:
                        zone = str(row[col_idx]).strip()
                        if zone:
                            entries.append([length_key, zone])
                if entries:
                    new_data[cust] = entries
        else:
            # Try row-based format: Customer | LengthRange | Zone
            for row in rows[1:]:
                if not row or len(row) < 3:
                    continue
                cust = str(row[0]).strip() if row[0] else ""
                length_key = str(row[1]).strip() if row[1] else ""
                zone = str(row[2]).strip() if row[2] else ""
                if cust and length_key and zone:
                    if cust not in new_data:
                        new_data[cust] = []
                    new_data[cust].append([length_key, zone])
        if not new_data:
            return jsonify({"error": "無法解析 Excel，請確認格式：\n欄A=客戶 欄B=<=3200 欄C=>3200<=4200 欄D=>4200"}), 400
        # Update in-memory
        STORAGE_LOOKUP = new_data
        rebuild_customer_names()
        logger.info("Storage updated via admin: %d customers", len(new_data))
        # Auto-commit to GitHub for permanent update
        json_str = json.dumps(new_data, ensure_ascii=False, indent=2)
        gh_ok = commit_storage_to_github(json_str)
        msg = "已更新 " + str(len(new_data)) + " 筆客戶"
        if gh_ok:
            msg += "（已自動推送 GitHub，將永久生效）"
        else:
            msg += "（GitHub 推送失敗，僅暫時生效）"
        return jsonify({"ok": True, "count": len(new_data), "github": gh_ok, "message": msg})
    except Exception as e:
        logger.error("Storage upload error: %s", e)
        return jsonify({"error": "解析失敗: " + str(e)}), 400


@app.route("/api/admin/storage/json")
def api_admin_storage_json():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    json_str = json.dumps(STORAGE_LOOKUP, ensure_ascii=False, indent=2)
    return app.response_class(json_str, mimetype="application/json",
                              headers={"Content-Disposition": "attachment; filename=storage_data.json"})


# ─── Passwords API ──────────────────────────────────
@app.route("/api/admin/passwords", methods=["GET", "POST"])
def api_admin_passwords():
    global pw1_text, pw2_text
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    if request.method == "GET":
        return jsonify({"pw1": pw1_text, "pw2": pw2_text})
    data = request.get_json(force=True)
    if "pw1" in data:
        pw1_text = data["pw1"]
    if "pw2" in data:
        pw2_text = data["pw2"]
    save_settings()
    return jsonify({"ok": True, "pw1": pw1_text, "pw2": pw2_text})


# ─── Scrap Text API ─────────────────────────────────
@app.route("/api/admin/scrap", methods=["GET", "POST"])
def api_admin_scrap():
    global scrap_text
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    if request.method == "GET":
        return jsonify({"text": scrap_text})
    data = request.get_json(force=True)
    if "text" in data:
        scrap_text = data["text"]
    save_settings()
    return jsonify({"ok": True})


# ─── Packaging API ──────────────────────────────────
@app.route("/api/admin/packaging/stats")
def api_admin_packaging_stats():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    return jsonify({"count": len(PACKAGING_LOOKUP)})


@app.route("/api/admin/packaging/upload", methods=["POST"])
def api_admin_packaging_upload():
    global PACKAGING_LOOKUP
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    if 'file' not in request.files:
        return jsonify({"error": "沒有檔案"}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({"error": "沒有檔案"}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        # Use first sheet (直棒包裝 or whatever)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({"error": "空的 Excel"}), 400
        header = [str(c).strip() if c else "" for c in rows[0]]
        # Find the code column: look for header containing 碼/code
        code_col = None
        for i, h in enumerate(header):
            hl = h.lower().replace(" ", "")
            if "包裝碼" in h or "代碼" in h or "代号" in h or "code" in hl:
                code_col = i
                break
        # Fallback: if header contains just "碼" somewhere
        if code_col is None:
            for i, h in enumerate(header):
                if "碼" in h:
                    code_col = i
                    break
        # Last fallback: first non-empty header column
        if code_col is None:
            for i, h in enumerate(header):
                if h:
                    code_col = i
                    break
        if code_col is None:
            return jsonify({"error": "找不到包裝碼欄位"}), 400
        # All other columns with non-empty headers become data fields
        data_cols = []  # [(col_index, header_name), ...]
        for i, h in enumerate(header):
            if i != code_col and h:
                data_cols.append((i, h))
        # Build lookup
        new_data = {}
        for row in rows[1:]:
            if not row:
                continue
            if code_col >= len(row) or not row[code_col]:
                continue
            code = str(row[code_col]).strip()
            if not code:
                continue
            entry = {}
            for col_idx, col_name in data_cols:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if val:
                        entry[col_name] = val
            if entry:
                new_data[code] = entry
        if not new_data:
            return jsonify({"error": "無法解析 Excel，請確認第一列為標題列，含包裝碼欄位"}), 400
        PACKAGING_LOOKUP = new_data
        logger.info("Packaging updated via admin: %d codes, columns: %s",
                     len(new_data), [c[1] for c in data_cols])
        json_str = json.dumps(new_data, ensure_ascii=False, indent=2)
        gh_ok = commit_packaging_to_github(json_str)
        cols_info = "、".join([c[1] for c in data_cols])
        msg = "已更新 " + str(len(new_data)) + " 筆包裝碼（欄位：" + cols_info + "）"
        if gh_ok:
            msg += "\n已自動推送 GitHub，永久生效"
        else:
            msg += "\nGitHub 推送失敗，僅暫時生效"
        return jsonify({"ok": True, "count": len(new_data), "github": gh_ok, "message": msg})
    except Exception as e:
        logger.error("Packaging upload error: %s", e)
        return jsonify({"error": "解析失敗: " + str(e)}), 400


@app.route("/api/admin/packaging/json")
def api_admin_packaging_json():
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    json_str = json.dumps(PACKAGING_LOOKUP, ensure_ascii=False, indent=2)
    return app.response_class(json_str, mimetype="application/json",
                              headers={"Content-Disposition": "attachment; filename=packaging_data.json"})


@app.route("/api/admin/webhook", methods=["GET", "POST"])
def api_admin_webhook():
    """Get or set webhook endpoint URL."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    if request.method == "POST":
        data = request.get_json() or {}
        url = data.get("url", "").strip()
        if not url:
            return jsonify({"error": "missing url"}), 400
        ok = set_webhook_url(url)
        return jsonify({"ok": ok})
    return jsonify({"webhook": get_webhook_info()})


@app.route("/api/admin/webhook/test", methods=["POST"])
def api_admin_webhook_test():
    """Test webhook endpoint connectivity."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    endpoint = data.get("endpoint", "")
    result = test_webhook(endpoint if endpoint else None)
    return jsonify(result)


@app.route("/api/admin/validate", methods=["POST"])
def api_admin_validate():
    """Validate message objects before sending."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    text = data.get("text", "").strip()
    msg_type = data.get("type", "reply")
    if not text:
        return jsonify({"error": "missing text"}), 400
    msgs = [{"type": "text", "text": text}]
    result = validate_message_objects(msgs, msg_type)
    return jsonify(result)


@app.route("/api/admin/content/preview/<message_id>")
def api_admin_content_preview(message_id):
    """Get content preview for an image/video message."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    content = get_content_preview(message_id)
    if content:
        return app.response_class(content, mimetype="image/jpeg")
    return jsonify({"error": "not found"}), 404


@app.route("/api/admin/content/status/<message_id>")
def api_admin_content_status(message_id):
    """Check preparation status for video/audio content."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    status = check_content_preparation(message_id)
    return jsonify({"message_id": message_id, "status": status})


@app.route("/api/admin/followers")
def api_admin_followers():
    """Get all follower user IDs and resolve names."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    ids = get_all_follower_ids()
    followers = []
    for uid in ids:
        name = ""
        # Try to find in known users
        for gid, names in group_user_names.items():
            if uid in names:
                name = names[uid]
                break
        if not name:
            dm_val = dm_known_users.get(uid, "")
            name = dm_val if isinstance(dm_val, str) else (dm_val.get("name", "") if isinstance(dm_val, dict) else "")
        followers.append({"user_id": uid, "name": name})
    return jsonify({"count": len(ids), "followers": followers})


@app.route("/api/admin/interaction")
def api_admin_interaction():
    """Get message interaction stats by request_id."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    req_id = request.args.get("request_id", "")
    if not req_id:
        return jsonify({"error": "missing request_id"}), 400
    stats = get_message_interaction_stats(req_id)
    return jsonify({"stats": stats})


@app.route("/api/admin/delivery")
def api_admin_delivery():
    """Get full delivery stats (reply/push/multicast/broadcast)."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    date_str = request.args.get("date", "")
    stats = get_message_delivery_stats(date_str if date_str else None)
    return jsonify({"delivery": stats})


@app.route("/api/admin/room/members")
def api_admin_room_members():
    """Get members of a multi-person chat room."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    room_id = request.args.get("room_id", "")
    if not room_id:
        return jsonify({"error": "missing room_id"}), 400
    count = get_room_member_count(room_id)
    members = fetch_all_room_members(room_id)
    profiles = []
    for uid in members[:50]:  # limit to 50
        p = get_room_member_profile(room_id, uid)
        if p:
            profiles.append(p)
    return jsonify({"count": count, "members": profiles})


@app.route("/api/admin/richmenu/detail/<rm_id>")
def api_admin_richmenu_detail(rm_id):
    """Get detailed info of a single rich menu."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    detail = get_rich_menu_by_id(rm_id)
    return jsonify({"menu": detail})


@app.route("/api/admin/richmenu/default")
def api_admin_richmenu_default():
    """Get the current default rich menu ID."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    rm_id = get_default_rich_menu_id()
    return jsonify({"default_rich_menu_id": rm_id})


@app.route("/api/admin/richmenu/user/<user_id>")
def api_admin_richmenu_user(user_id):
    """Get which rich menu is linked to a specific user."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    rm_id = get_user_rich_menu_id(user_id)
    return jsonify({"user_id": user_id, "rich_menu_id": rm_id})


@app.route("/api/admin/richmenu/validate", methods=["POST"])
def api_admin_richmenu_validate():
    """Validate a rich menu object before creating."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    result = validate_rich_menu_obj(data)
    return jsonify(result)


@app.route("/api/admin/richmenu/image/<rm_id>")
def api_admin_richmenu_image(rm_id):
    """Download/preview the image of a rich menu."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    content = download_rich_menu_image(rm_id)
    if content:
        return app.response_class(content, mimetype="image/png")
    return jsonify({"error": "not found"}), 404


@app.route("/api/admin/richmenu/alias/list")
def api_admin_richmenu_alias_list():
    """Get all rich menu aliases."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    aliases = list_rich_menu_aliases()
    return jsonify({"aliases": aliases})


@app.route("/api/admin/richmenu/alias/update", methods=["POST"])
def api_admin_richmenu_alias_update():
    """Update a rich menu alias to point to a different menu."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json() or {}
    alias_id = data.get("alias_id", "")
    rm_id = data.get("rich_menu_id", "")
    if not alias_id or not rm_id:
        return jsonify({"error": "missing alias_id or rich_menu_id"}), 400
    ok = update_rich_menu_alias(alias_id, rm_id)
    return jsonify({"ok": ok})


@app.route("/api/admin/richmenu/alias/detail/<alias_id>")
def api_admin_richmenu_alias_detail(alias_id):
    """Get info about a specific rich menu alias."""
    if not check_admin_key():
        return jsonify({"error": "forbidden"}), 403
    info = get_rich_menu_alias(alias_id)
    return jsonify({"alias": info})


@app.route("/health", methods=["GET"])
def health():
    return {"status": "ok", "version": VERSION, "uptime": int(time.time() - bot_start_time)}


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)
